"""
Mathools 1.0  Interface Flet Premium
Itapoá Saneamento - Gestão e Processamento RPA
Flet 0.82  usa ft.Icons.NOME e ft.ConstrainedControl
"""

import warnings
# Suprimir APENAS DeprecationWarnings, não outros erros
warnings.filterwarnings("ignore", category=DeprecationWarning, module="flet.*")
# Suprimir aviso de SSL não verificado do Supabase (urllib3)
try:
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
except Exception:
    pass

# ═════════════════════════════════════════════════════════════════
# SUPRIMIR LOGS DEBUG ANTES DE IMPORTAR FLET
# ═════════════════════════════════════════════════════════════════
import logging as _logging_early
import sys as _sys_early

# Desativar todos os DEBUG logs de bibliotecas ANTES do import
_logging_early.basicConfig(level=_logging_early.WARNING)
for logger_name in ['flet', 'flet_core', 'flet_desktop', 'asyncio', 'urllib3', 'selenium', 'websockets', 'websocket', 'AUDITORIA_LGPD', 'audit_logger']:
    _logging_early.getLogger(logger_name).setLevel(_logging_early.WARNING)
    _logging_early.getLogger(logger_name).handlers = []  # Limpar handlers existentes

# Desativar propagação de logs de debug
logging_disable = _logging_early.disable
logging_disable(_logging_early.DEBUG)

try:
    import msvcrt as msvcrt
except ImportError:
    import types as _types
    msvcrt = _types.ModuleType('msvcrt')
    msvcrt.kbhit = lambda: False
    msvcrt.getch = lambda: b''
try:
    import winsound
except Exception:
    winsound = None
import flet as ft
import threading
import queue
import asyncio
import time
import os
import sys
import io
import subprocess
import logging
import builtins
import importlib
import re
import json
import base64
import tempfile
import requests
import ctypes
from pathlib import Path
from text_utils import read_json_utf8, read_text_utf8, write_json_utf8, write_text_utf8

try:
    from updater import check_update_async, apply_update, UpdateInfo as _UpdateInfo, CURRENT_VERSION as APP_VERSION
    _HAS_UPDATER = True
except ImportError:
    _HAS_UPDATER = False
    APP_VERSION = "desconhecida"

_LAST_UPDATE_INFO = None

def resource_path(relative_path: str) -> str:
    """Resolve caminho de asset  funciona como script e como .exe PyInstaller."""
    if getattr(sys, "frozen", False):
        base = sys._MEIPASS
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, relative_path)


def _get_work_area_size():
    """Retorna a area util da tela no Windows, descontando barra de tarefas."""
    try:
        if sys.platform == "win32":
            class RECT(ctypes.Structure):
                _fields_ = [
                    ("left", ctypes.c_long),
                    ("top", ctypes.c_long),
                    ("right", ctypes.c_long),
                    ("bottom", ctypes.c_long),
                ]

            rect = RECT()
            SPI_GETWORKAREA = 0x0030
            if ctypes.windll.user32.SystemParametersInfoW(SPI_GETWORKAREA, 0, ctypes.byref(rect), 0):
                return rect.right - rect.left, rect.bottom - rect.top
    except Exception:
        pass
    return 1366, 768


# Supabase  compartilhado com auth_system
try:
    from auth_system import SUPABASE_URL, get_headers as _auth_get_headers
    def get_headers():
        return _auth_get_headers()
except ImportError:
    SUPABASE_URL = ""
    def get_headers():
        return {}

# ─────────────────────────────────────────────────────────────────────────────
# NOVOS MÓDULOS INTEGRADOS (Tarefas 1-10)
# ─────────────────────────────────────────────────────────────────────────────
try:
    pass
except ImportError as e:
    # Módulos opcionais - se não existirem, continua com funcionalidade reduzida
    import sys
    print(f"Aviso: Modulo nao encontrado: {e}", file=sys.stderr)
    
    # Fallback para cache_manager se não conseguir importar
    def get_cache_manager(cache_dir=None):
        """Stub para cache_manager quando não consegue importar"""
        class NoCacheManager:
            def get_cached_data(self, *args, **kwargs):
                return None
            def get_cache_info(self, *args, **kwargs):
                return None
            def register_macro_data(self, *args, **kwargs):
                pass
            def list_cached_items(self):
                return {}
            def get_cached_8091_month(self, *args, **kwargs):
                return None
            def register_8091_month(self, *args, **kwargs):
                pass
        return NoCacheManager()
    
    # Fallback para geocoder se não conseguir importar
    def get_geocoder():
        """Stub para geocoder quando não consegue importar"""
        class NoGeocoder:
            def geocodificar_lote(self, *args, **kwargs):
                return args[0] if args else []
        return NoGeocoder()

_LAZY_MODULES = {}

def _lazy_import(module_name: str):
    if module_name not in _LAZY_MODULES:
        _LAZY_MODULES[module_name] = importlib.import_module(module_name)
    return _LAZY_MODULES[module_name]

def get_cache_manager(cache_dir=None):
    """Carrega cache_manager apenas quando necessario."""
    try:
        return _lazy_import("cache_manager").get_cache_manager(cache_dir)
    except Exception:
        class NoCacheManager:
            def get_cached_data(self, *args, **kwargs):
                return None
            def get_cache_info(self, *args, **kwargs):
                return None
            def register_macro_data(self, *args, **kwargs):
                pass
            def list_cached_items(self):
                return {}
            def get_cached_8091_month(self, *args, **kwargs):
                return None
            def register_8091_month(self, *args, **kwargs):
                pass
        return NoCacheManager()

def get_geocoder():
    """Carrega geocoder apenas quando necessario."""
    try:
        return _lazy_import("geocoder").get_geocoder()
    except Exception:
        class NoGeocoder:
            def geocodificar_lote(self, *args, **kwargs):
                return args[0] if args else []
        return NoGeocoder()

def gerar_pn_com_download(*args, **kwargs):
    """Carrega pn_generator apenas quando necessario."""
    return _lazy_import("pn_generator").gerar_pn_com_download(*args, **kwargs)

# ─────────────────────────────────────────────────────────────────────────────
#  DETECTAR API DE ÍCONES CORRETA
#  No Flet 0.82 os ícones ficam em ft.Icons (classe), não ft.icons.NOME
# ─────────────────────────────────────────────────────────────────────────────
try:
    # Tenta ft.Icons (novo padrão 0.80+)
    _I = ft.Icons
    _I.MENU  # valida
except AttributeError:
    try:
        _I = ft.icons
    except AttributeError:
        _I = None

def ic(name: str):
    """Retorna o ícone de forma segura."""
    if _I is not None:
        return getattr(_I, name, name.lower())
    return name.lower()


# ─────────────────────────────────────────────────────────────────────────────
#  CONTADORES PERSISTENTES (salvo em stats.json ao lado do .exe)
# ─────────────────────────────────────────────────────────────────────────────
import json

def _stats_path() -> str:
    """Retorna o caminho do arquivo de estatísticas."""
    base = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else os.path.abspath(os.path.dirname(__file__))
    return os.path.join(base, "stats.json")

def _app_base_dir() -> str:
    return os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else os.path.abspath(os.path.dirname(__file__))

def _user_prefs_path(username: str) -> str:
    safe = re.sub(r"[^a-zA-Z0-9_.-]+", "_", str(username or "default"))
    return os.path.join(_app_base_dir(), f"user_preferences_{safe}.json")

def _default_user_prefs(username: str = "") -> dict:
    base = _app_base_dir()
    return {
        "display_name": username or "Usuario",
        "avatar_path": "",
        "alerts_enabled": False,
        "alert_sound_path": "",
        "alert_volume": 80,
        "exact_alert_minutes": 5,
        "workdays_only": True,
        "show_clock": True,
        "yellow_warning_minutes": 30,
        "red_warning_minutes": 5,
        "time_08": "08:00",
        "time_12": "12:00",
        "time_13": "13:00",
        "time_18": "18:00",
        "friday_custom_enabled": False,
        "time_18_friday": "17:00",
        "reports_dir": os.path.join(base, "Relatorios"),
        "spreadsheets_dir": os.path.join(base, "Planilhas"),
    }

def load_user_prefs(username: str) -> dict:
    prefs = _default_user_prefs(username)
    try:
        data = read_json_utf8(_user_prefs_path(username), default={})
        if isinstance(data, dict):
            # Migração leve: se o arquivo veio de versões anteriores, inicia alertas desligados.
            if "alerts_enabled" in data and "prefs_version" not in data:
                data["alerts_enabled"] = False
            data.setdefault("prefs_version", 2)
            prefs.update(data)
    except Exception:
        pass
    return prefs

def save_user_prefs(username: str, prefs: dict):
    try:
        write_json_utf8(_user_prefs_path(username), prefs, ensure_ascii=False)
    except Exception:
        pass

def _logs_path() -> str:
    """Retorna o caminho do arquivo de log."""
    base = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else os.path.abspath(os.path.dirname(__file__))
    return os.path.join(base, "mathtools_debug.log")

def setup_logging_to_file():
    """Configura logging para arquivo (backup caso console nao funcione)."""
    try:
        # Arquivo: INFO (sem DEBUG)
        # Console: INFO (sem DEBUG)
        logging.basicConfig(
            filename=_logs_path(),
            level=logging.INFO,
            format='[%(asctime)s] %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S',
            filemode='a'
        )
        
        # Console: apenas INFO e acima
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setLevel(logging.INFO)
        formatter = logging.Formatter('[%(asctime)s] %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
        console_handler.setFormatter(formatter)
        logging.getLogger().addHandler(console_handler)
        
        # Suprimir TODOS os loggers de bibliotecas em WARNING e acima (nenhum DEBUG ou INFO)
        suppressed_loggers = ['flet', 'flet_core', 'flet_desktop', 'asyncio', 'urllib3', 'selenium', 'websocket', 'websockets', 'requests', 'AUDITORIA_LGPD', 'audit_logger']
        for logger_name in suppressed_loggers:
            logger = logging.getLogger(logger_name)
            logger.setLevel(logging.WARNING)
            logger.propagate = False
            # Remover todos os handlers desses loggers
            logger.handlers = []
        
        # Desabilitar completamente DEBUG
        logging.disable(logging.DEBUG)
        
    except Exception as e:
        print(f"Aviso: Nao foi possivel configurar logging: {e}", file=sys.stderr)

def load_stats() -> dict:
    """Carrega contadores do disco. Retorna zeros se arquivo não existir."""
    try:
        data = read_json_utf8(_stats_path(), default={}) or {}
        return {
            "csvs":       int(data.get("csvs", 0)),
            "mensagens":  int(data.get("mensagens", 0)),
            "propostas":  int(data.get("propostas", 0)),
            "paradas":    int(data.get("paradas", 0)),
        }
    except Exception:
        return {"csvs": 0, "mensagens": 0, "propostas": 0, "paradas": 0}

def save_stats(stats: dict):
    """Salva contadores no disco."""
    try:
        write_json_utf8(_stats_path(), stats)
    except Exception:
        pass

def increment_stat(key: str, amount: int = 1):
    """Incrementa um contador e salva."""
    stats = load_stats()
    stats[key] = stats.get(key, 0) + amount
    save_stats(stats)
    return stats[key]

# ─────────────────────────────────────────────────────────────────────────────
#  PALETA
# ─────────────────────────────────────────────────────────────────────────────
NAVY        = "#011F3F"
NAVY_DARK   = "#021529"
ORANGE      = "#F27D16"
ORANGE_DARK = "#D96B08"
WHITE       = "#FFFFFF"
SURFACE     = "#F4F6FA"
BORDER      = "#DDE3EE"
TEXT_MAIN   = "#0D1F35"
TEXT_MUTED  = "#7A8FAD"
GREEN       = "#22C55E"
RED         = "#EF4444"
TERM_BG     = "#0D1117"
TERM_GREEN  = "#4ade80"

# ─────────────────────────────────────────────────────────────────────────────
#  ENGINE LOADER
# ─────────────────────────────────────────────────────────────────────────────
import sys
import os

# Adiciona o diretório atual ao path para encontrar mathtools_1_0
if getattr(sys, "frozen", False):
    _base = sys._MEIPASS
else:
    _base = os.path.dirname(os.path.abspath(__file__))

if _base not in sys.path:
    sys.path.insert(0, _base)

_engine_module = None

def load_engine():
    global _engine_module
    if _engine_module:
        return _engine_module
    
    try:
        # Tenta importar diretamente como módulo
        import mathtools_1_0 as mod
        _engine_module = mod
        return mod
    except ImportError:
        pass
    
    # Fallback: tenta carregar como arquivo
    filenames = ["mathtools_1_0.py", "mathtools_main.py"]
    
    if getattr(sys, "frozen", False):
        _base = sys._MEIPASS
    else:
        _base = os.path.dirname(os.path.abspath(__file__))
    
    for filename in filenames:
        path = os.path.join(_base, filename)
        try:
            code = read_text_utf8(path)
            
            mod_dict = {'__name__': 'mathtools_main', '__file__': path}
            exec(code, mod_dict)
            
            mod = type(sys)('mathtools_main')
            for key, val in mod_dict.items():
                if not key.startswith('__'):
                    setattr(mod, key, val)
            
            _engine_module = mod
            return mod
        except Exception:
            continue
    
    return None


def _load_custom_module(module_name, filenames):
    """
    Carrega módulo customizado com fallback para arquivo.
    Funciona em desenvolvimento e no .exe bundled pelo PyInstaller.
    """
    if getattr(sys, "frozen", False):
        _base = sys._MEIPASS
    else:
        _base = os.path.dirname(os.path.abspath(__file__))
    
    # Prefere sempre o arquivo local do workspace para evitar usar
    # uma cópia antiga/cached do módulo vinda de outro path.
    for filename in filenames:
        path = os.path.join(_base, filename)
        try:
            code = read_text_utf8(path)
            
            mod_dict = {'__name__': module_name, '__file__': path}
            exec(code, mod_dict)
            
            mod = type(sys)(module_name)
            for key, val in mod_dict.items():
                if not key.startswith('__'):
                    setattr(mod, key, val)
            
            return mod
        except Exception:
            continue

    try:
        # Fallback final: importa pelo nome se não encontrou arquivo local.
        mod = __import__(module_name)
        return mod
    except (ImportError, ModuleNotFoundError):
        pass
    
    return None


def _attach_ui_logger(engine, lq):
    eng_logger = getattr(engine, "logger", None)
    if eng_logger is None:
        return None
    removed = [h for h in eng_logger.handlers
               if isinstance(h, logging.StreamHandler) and not isinstance(h, logging.FileHandler)]
    for h in removed:
        eng_logger.removeHandler(h)
    ui_handler = LogQueueHandler(lq)
    ui_handler.setLevel(logging.INFO)
    eng_logger.addHandler(ui_handler)
    return (eng_logger, removed, ui_handler)

def _detach_ui_logger(ctx):
    if ctx is None:
        return
    eng_logger, removed, ui_handler = ctx
    eng_logger.removeHandler(ui_handler)
    for h in removed:
        eng_logger.addHandler(h)

class LogCapture(io.StringIO):
    def __init__(self, lq):
        super().__init__()
        self.lq = lq
        self._lock = threading.Lock()

    # Regex compilado uma vez  remove códigos de cor/controle ANSI (\x1b[...m, \x1b[A, etc.)
    _ANSI = re.compile(r'\x1b\[[0-9;]*[A-Za-z]')

    def write(self, msg):
        if not msg:
            return 0
        # Remove códigos ANSI de cor e movimentação de cursor do tqdm
        msg = self._ANSI.sub('', msg)
        # Intercepta sinal de timer emitido pelo engine
        if msg.strip().startswith("GUI_TIMER:"):
            try:
                secs = int(msg.strip().split(":", 1)[1])
                mins, s = divmod(secs, 60)
                legivel = f"{mins:02d}:{s:02d}" if mins > 0 else f"{secs}s"
                self.lq.put(f"__MSG_SENT__\n")  # Sinal para contar mensagem
                self.lq.put(f"__TIMER__:{secs}\n")
                self.lq.put(f"Proxima mensagem sera enviada em {legivel}...\n")
            except Exception:
                pass
            return len(msg)
        # Suprime linhas que são apenas barras de progresso do tqdm (contêm |█ ou |# ou %|)
        if re.search(r'\d+%\|', msg) or re.search(r'\|\s*\d+/\d+', msg):
            return len(msg)
        # Suprime warnings de matplotlib/seaborn (thread, glyph, etc.)
        if any(x in msg for x in ("UserWarning", "will likely fail", "Starting a Matplotlib",
                                   "Glyph", "missing from font", "plt.figure", "fig = plt.figure")):
            return len(msg)
        # Suprime linhas de caminho de arquivo tipo "C:\Users\...\foo.py:123:"
        if re.search(r'[A-Za-z]:\\.*\.py:\d+', msg):
            return len(msg)
        if not msg.strip():
            return len(msg)
        # Repassa sinais de progresso do engine diretamente
        if msg.strip().startswith("__PROGRESSO__:"):
            self.lq.put(msg.strip() + "\n")
            return len(msg)
        if msg.strip().startswith("__WPP_STAGE__:"):
            self.lq.put(msg.strip() + "\n")
            return len(msg)
        with self._lock:
            self.lq.put(msg)
        return len(msg)

    def flush(self): pass


class LogQueueHandler(logging.Handler):
    """Redireciona o logger do engine para a fila da UI, limpando o prefixo."""
    def __init__(self, lq):
        super().__init__()
        self.lq = lq

    _ANSI = re.compile(r'\x1b\[[0-9;]*[A-Za-z]')

    def emit(self, record):
        msg = self._ANSI.sub('', record.getMessage())
        if not msg.strip():
            return
        # Suprime barras de progresso do tqdm
        if re.search(r'\d+%\|', msg) or re.search(r'\|\s*\d+/\d+', msg):
            return
        # Suprime warnings internos do matplotlib/seaborn
        if any(x in msg for x in ("UserWarning", "will likely fail", "Starting a Matplotlib", "Glyph", "missing from font")):
            return
        if msg.strip().startswith("__PROGRESSO__:"):
            self.lq.put(msg.strip() + "\n")
            return
        if msg.strip().startswith("__WPP_STAGE__:"):
            self.lq.put(msg.strip() + "\n")
            return
        self.lq.put(msg + "\n")

# ─────────────────────────────────────────────────────────────────────────────
#  HELPERS DE UI
# ─────────────────────────────────────────────────────────────────────────────
def card(content, pad=24, mb=16):
    return ft.Container(
        content=content,
        bgcolor=WHITE,
        border_radius=12,
        padding=pad,
        margin=ft.Margin.only(bottom=mb),
        border=ft.Border.all(1, BORDER),
        shadow=ft.BoxShadow(
            spread_radius=0, blur_radius=16,
            color=ft.Colors.with_opacity(0.06, ft.Colors.BLACK),
            offset=ft.Offset(0, 4),
        ),
    )

def sec_title(text, icon_name):
    return ft.Row([
        ft.Icon(ic(icon_name), color=ORANGE, size=20),
        ft.Text(text, size=15, weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
    ], spacing=10)

def flabel(text):
    return ft.Text(text, size=11, color=TEXT_MUTED, weight=ft.FontWeight.W_600)

def tfield(hint, expand=True, width=None, password=False, value=""):
    return ft.TextField(
        hint_text=hint, value=value,
        password=password, can_reveal_password=password,
        expand=expand, width=width, height=44, text_size=13,
        border_radius=8, filled=True, fill_color="#F8FAFD",
        border_color=BORDER, focused_border_color=ORANGE,
        hint_style=ft.TextStyle(color=TEXT_MUTED, size=12),
        text_style=ft.TextStyle(color=TEXT_MAIN, size=13),
        content_padding=ft.Padding.symmetric(horizontal=14, vertical=10),
    )

def pbtn(text, icon_name, on_click=None, width=None):
    row_items = [
        ft.Icon(ic(icon_name), size=16, color=WHITE),
        ft.Container(width=8),
        ft.Text(text, size=13, weight=ft.FontWeight.BOLD, color=WHITE),
    ]
    return ft.Container(
        content=ft.Row(row_items, spacing=0,
                       alignment=ft.MainAxisAlignment.CENTER,
                       vertical_alignment=ft.CrossAxisAlignment.CENTER),
        height=44,
        width=width,
        bgcolor=ORANGE,
        border_radius=8,
        padding=ft.Padding.symmetric(horizontal=20, vertical=0),
        on_click=on_click,
        ink=True,
        shadow=ft.BoxShadow(spread_radius=0, blur_radius=6,
                            color=ft.Colors.with_opacity(0.3, ft.Colors.BLACK),
                            offset=ft.Offset(0, 2)),
    )

def sbtn(text, icon_name=None, on_click=None, width=None):
    row_items = []
    if icon_name:
        row_items.append(ft.Icon(ic(icon_name), size=16, color=WHITE))
        row_items.append(ft.Container(width=6))
    row_items.append(ft.Text(text, size=13, weight=ft.FontWeight.W_600, color=WHITE))
    btn_kw = dict(
        content=ft.Row(row_items, spacing=0,
                       alignment=ft.MainAxisAlignment.CENTER,
                       vertical_alignment=ft.CrossAxisAlignment.CENTER),
        height=44,
        bgcolor=NAVY,
        border_radius=8,
        padding=ft.Padding.symmetric(horizontal=18, vertical=0),
        on_click=on_click,
        ink=True,
        animate=ft.Animation(80, ft.AnimationCurve.EASE_OUT),
        shadow=ft.BoxShadow(spread_radius=0, blur_radius=6,
                            color=ft.Colors.with_opacity(0.25, ft.Colors.BLACK),
                            offset=ft.Offset(0, 2)),
    )
    if width is not None:
        btn_kw["width"] = width
    btn = ft.Container(**btn_kw)
    normal_bg = NAVY
    hover_bg = "#143A61"
    pressed_bg = "#0A223D"

    def _apply_state(bg, blur, offset_y):
        btn.bgcolor = bg
        btn.shadow = ft.BoxShadow(
            spread_radius=0,
            blur_radius=blur,
            color=ft.Colors.with_opacity(0.25, ft.Colors.BLACK),
            offset=ft.Offset(0, offset_y),
        )
        try:
            btn.update()
        except Exception:
            pass

    def _on_enter(e):
        _apply_state(hover_bg, 4, 1)

    def _on_exit(e):
        _apply_state(normal_bg, 6, 2)

    btn.on_hover = lambda e: _on_enter(e) if str(getattr(e, "data", "")).lower() == "true" else _on_exit(e)
    return btn

def check_row(text):
    return ft.Row([
        ft.Icon(ic("CHECK_CIRCLE"), color=GREEN, size=16),
        ft.Text(text, size=13, color=TEXT_MAIN),
    ], spacing=8)

def info_box(text):
    return ft.Container(
        content=ft.Row([
            ft.Icon(ic("INFO_OUTLINE"), color=ORANGE, size=16),
            ft.Text(text, size=12, color="#7A4F10"),
        ], spacing=10),
        bgcolor="#FFF7ED",
        border=ft.Border.all(1, "#F6C28B"),
        border_radius=8,
        padding=ft.Padding.symmetric(horizontal=14, vertical=10),
    )

def role_badge(role):
    cfg = {
        "ADM":            (RED,       "#FEECEC"),
        "Atendimento":    (ORANGE,    "#FEF3E8"),
        "Call center":    ("#3B82F6", "#EFF6FF"),
        "Administrativo": (GREEN,     "#ECFDF5"),
    }
    color, bg = cfg.get(role, (TEXT_MUTED, SURFACE))
    return ft.Container(
        content=ft.Text(role, size=11, color=color, weight=ft.FontWeight.W_600),
        bgcolor=bg, border_radius=20,
        padding=ft.Padding.symmetric(horizontal=10, vertical=4),
    )


# ─────────────────────────────────────────────────────────────────────────────
#  SIDEBAR ITEM
#  GestureDetector individual por item (on_enter/on_exit/on_tap).
#  Não conflita com o GestureDetector da sidebar inteira (expansão).
# ─────────────────────────────────────────────────────────────────────────────
class SidebarItem:
    """Item de sidebar com clique rápido e visual estável.

    O hover explícito foi removido daqui para evitar filas de update quando o
    mouse passa muito rápido sobre vários itens. O feedback de seleção continua
    sendo controlado por este componente.
    """

    def __init__(self, icon_name: str, label: str, index: int, on_select, page=None):
        self.index     = index
        self.label     = label
        self.on_select = on_select
        self.selected  = False
        self._last_visual = None
        self._page     = page

        self._ind = ft.Container(
            width=3, height=28, bgcolor=ORANGE,
            border_radius=ft.BorderRadius.only(top_right=4, bottom_right=4),
            opacity=0,
        )
        self._ico = ft.Icon(ic(icon_name),
                            color=ft.Colors.with_opacity(0.55, WHITE), size=20)
        self.badge = ft.Container(
            content=ft.Text("", size=9, color=WHITE, weight=ft.FontWeight.BOLD,
                            text_align=ft.TextAlign.CENTER),
            bgcolor=RED, border_radius=8,
            width=16, height=16,
            alignment=ft.Alignment(0, 0),
            visible=False,
        )
        _ico_stack = ft.Stack([
            self._ico,
            ft.Container(content=self.badge, alignment=ft.Alignment(1, -1)),
        ], width=24, height=24)

        self._inner = ft.Container(
            content=ft.Stack([
                ft.Container(
                    content=_ico_stack,
                    alignment=ft.Alignment(0, 0),
                    expand=True,
                ),
                ft.Container(
                    content=self._ind,
                    alignment=ft.Alignment(-1, 0),
                    expand=True,
                    padding=ft.Padding.only(left=0),
                ),
            ], expand=True),
            width=62,
            height=56,
            border_radius=12,
            bgcolor=ft.Colors.with_opacity(0.015, WHITE),
            border=ft.Border.all(1, ft.Colors.with_opacity(0.02, WHITE)),
            alignment=ft.Alignment(0, 0),
            shadow=ft.BoxShadow(
                spread_radius=0,
                blur_radius=10,
                color=ft.Colors.with_opacity(0.08, ft.Colors.BLACK),
                offset=ft.Offset(0, 3),
            ),
            on_click=self._on_tap,
            ink=True,
        )

        self.control = ft.Container(
            content=self._inner,
            width=62,
            height=56,
            alignment=ft.Alignment(0, 0),
            on_click=self._on_tap,
            tooltip=label,
            ink=True,
        )

    def _on_tap(self, e):
        self.on_select(self.index)

    def set_selected(self, v: bool):
        """Atualiza estado de seleção. NÃO faz update()  o chamador deve."""
        self.selected = v
        self._refresh_props()

    def _refresh_props(self) -> bool:
        """Aplica as propriedades visuais corretas SEM chamar update().
        Retorna True se houve mudança real (para o chamador decidir se atualiza)."""
        visual = "selected" if self.selected else "idle"
        if visual == self._last_visual:
            return False
        self._last_visual = visual

        if self.selected:
            self._inner.bgcolor = ft.Colors.with_opacity(0.18, WHITE)
            self._inner.border  = ft.Border.all(1, ft.Colors.with_opacity(0.10, "#FFD3A8"))
            self._inner.shadow  = ft.BoxShadow(
                spread_radius=0,
                blur_radius=14,
                color=ft.Colors.with_opacity(0.16, "#000000"),
                offset=ft.Offset(0, 4),
            )
            self._ind.opacity   = 1
            self._ico.color     = ORANGE
        else:
            self._inner.bgcolor = ft.Colors.with_opacity(0.015, WHITE)
            self._inner.border  = ft.Border.all(1, ft.Colors.with_opacity(0.02, WHITE))
            self._inner.shadow  = ft.BoxShadow(
                spread_radius=0,
                blur_radius=10,
                color=ft.Colors.with_opacity(0.08, ft.Colors.BLACK),
                offset=ft.Offset(0, 3),
            )
            self._ind.opacity   = 0
            self._ico.color     = ft.Colors.with_opacity(0.72, WHITE)
        return True

    # Mantido por compatibilidade  mas não usado internamente no hover path
    def _refresh(self):
        self._refresh_props()


def _sidebar_tooltip(message: str):
    return message


# ─────────────────────────────────────────────────────────────────────────────
#  PÁGINA 0  DASHBOARD
# ─────────────────────────────────────────────────────────────────────────────


# ── CATÁLOGO DE MACROS  adicione novas macros aqui ──────────────────────────
# Cada entrada define: id, nome, descrição, ícone, cor e gráficos gerados.
MACROS_CATALOGO = [
    {
        "id":       "8117",
        "nome":     "Macro 8117",
        "tipo":     "diaria",
        "desc":     "Boletim diário de arrecadação por data de pagamento. Consolida os recebimentos do período por canal e forma de pagamento.",
        "icone":    "BAR_CHART",
        "cor":      "#1565C0",
        "bg":       "#E3F2FD",
        "graficos": [
            "Arrecadação diária por canal (barras)",
            "Curva acumulada do período (linha)",
            "Participação por forma de pagamento (pizza)",
            "Ranking de arrecadadores (barras horiz.)",
        ],
    },
    {
        "id":       "8121",
        "nome":     "Macro 8121",
        "tipo":     "diaria",
        "desc":     "Confronto diário e acumulado entre arrecadação (data de pagamento) e faturamento (data de emissão). Identifica gaps de inadimplência no período.",
        "icone":    "SHOW_CHART",
        "cor":      "#FF8F00",
        "bg":       "#FFF8E1",
        "graficos": [
            "Arrecadação vs Faturamento por dia (barras agrupadas)",
            "Curva acumulada Arr × Fat (linha dupla)",
            "Índice de eficiência de cobrança (%)",
        ],
    },
    # ── Macros mensais  adicione aqui ───────────────────────────────────────
    {
        "id":       "8091",
        "nome":     "Macro 8091",
        "tipo":     "mensal",
        "desc":     "Faturamento mensal por competência. Detalha receita por categoria, grupo, situação de cobrança e volume medido, com série histórica multi-mês.",
        "icone":    "CALENDAR_MONTH",
        "cor":      "#2E7D32",
        "bg":       "#E8F5E9",
        "graficos": [
            "Receita por competência (barras)",
            "Evolução mensal acumulada (linha)",
            "Composição por categoria/grupo (pizza)",
            "Faturamento por bairro  Top 10",
            "Faixas de valor de fatura (histograma)",
            "Situação de cobrança e inadimplência",
        ],
    },
    # ── Macros mensais geográficas  análise com Lat/Long ──────────────────────
    {
        "id":       "50012",
        "uid":      "50012-mensal",
        "nome":     "Macro 50012",
        "tipo":     "mensal",
        "desc":     "Relatório analítico de leitura e medição com coordenadas geográficas. Consolida volume, críticas de medição, faturamento e cobertura de lat/long por ligação.",
        "icone":    "MAP",
        "cor":      "#7B1FA2",
        "bg":       "#F3E5F5",
        "graficos": [
            "KPIs: ligações, volume total, faturamento e % normais",
            "Críticas de medição por tipo (doughnut)",
            "Faixas de consumo em m³ (barras horiz.)",
            "Desempenho por leiturista (combo barra + taxa)",
            "Volume e receita por grupo de leitura",
            "Desvio de consumo vs média histórica",
            "Leituras por dia do mês (linha + volume)",
            "Faturamento por bairro  Top 12",
            "Cobertura geográfica e centróide calculado",
        ],
    },
    # ── Macros geográficas  análises com Ordens de Serviço ──────────────────
    {
        "id":       "50012",
        "uid":      "50012-atemporal",
        "nome":     "Mapeamento das Equipes de Campo",
        "tipo":     "atemporal",
        "desc":     "Baixa a macro 50012, cruza lat/long por matrícula e distribui os dados em arquivos CSV individuais por funcionário, prontos para uso na aba de Logística.",
        "icone":    "PERSON_PIN_CIRCLE",
        "cor":      "#00695C",
        "bg":       "#E0F2F1",
        "graficos": [
            "CSV do dia por funcionário (Equipe-COM/COM-Nome/)",
            "CSV histórico geral por funcionário",
            "Bases do mapeamento em Equipe-COM/Bases Geral/",
        ],
    },
    {
        "id":       "PN",
        "uid":      "PN",
        "nome":     "Gerar PN",
        "tipo":     "diaria",
        "desc":     "Painel de Negócios  Consolida macros 8117 e 8121 no período selecionado, com opção de incluir a 8091 para enriquecer o HTML final.",
        "icone":    "ASSESSMENT",
        "cor":      "#00ACC1",
        "bg":       "#E0F2F1",
        "graficos": [
            "Consolidação 8117  Arrecadação diária por canal",
            "Consolidação 8121  Arrecadação vs Faturamento",
            "(Opcional) Série mensal 8091 no contexto do PN",
            "Análise PN combinada no período",
            "HTML final em Dashboard_Waterfy/dashboard_PN_*.html",
        ],
    },
]


def build_dashboard(num_operadores: int = 0, lq=None, page: ft.Page = None):
    import threading, os, sys, webbrowser
    from datetime import date, datetime

    _s = load_stats()

    def stat(icon_name, value, label, color, bg, ref=None):
        txt = ft.Text(str(value), size=26, weight=ft.FontWeight.BOLD, color=TEXT_MAIN)
        if ref is not None:
            ref[0] = txt
        return ft.Container(
            content=ft.Column([
                ft.Container(
                    content=ft.Icon(ic(icon_name), color=color, size=22),
                    width=44, height=44, bgcolor=bg,
                    border_radius=10, alignment=ft.Alignment(0, 0),
                ),
                ft.Container(height=10),
                txt,
                ft.Text(label, size=12, color=TEXT_MUTED),
            ], spacing=2),
            bgcolor=WHITE, border_radius=12, padding=20, expand=True,
            border=ft.Border.all(1, BORDER),
            shadow=ft.BoxShadow(
                spread_radius=0, blur_radius=12,
                color=ft.Colors.with_opacity(0.05, ft.Colors.BLACK),
                offset=ft.Offset(0, 4),
            ),
        )

    ref_csvs      = [None]
    ref_mensagens = [None]
    ref_propostas = [None]
    ref_paradas   = [None]

    def _dash_log(msg: str):
        if lq is not None:
            try:
                lq.put(f"[Dashboard] {msg}\n")
            except Exception:
                pass

    def reset_stats(e):
        # Log técnico para auditoria
        save_stats({"csvs": 0, "mensagens": 0, "propostas": 0, "paradas": 0})
        for ref in [ref_csvs, ref_mensagens, ref_propostas, ref_paradas]:
            if ref[0]:
                try: ref[0].value = "0"; ref[0].update()
                except: pass

    # ── Campos de login ───────────────────────────────────────────────────────
    _hoje = date.today()
    e_user = ft.TextField(
        label="Usuário Waterfy", hint_text="ex: joao.silva_COM",
        prefix_icon=ic("PERSON"), height=52, border_radius=10,
        content_padding=ft.Padding.symmetric(horizontal=14, vertical=10),
        border_color=BORDER, focused_border_color=NAVY, width=250,
    )
    e_pass = ft.TextField(
        label="Senha", password=True, can_reveal_password=True,
        prefix_icon=ic("LOCK"), height=52, border_radius=10,
        content_padding=ft.Padding.symmetric(horizontal=14, vertical=10),
        border_color=BORDER, focused_border_color=NAVY, width=250,
    )
    # ── Campos de período diário ──────────────────────────────────────────────
    e_di = ft.TextField(
        label="Data Inicial", value=f"01/{_hoje.month:02d}/{_hoje.year}",
        prefix_icon=ic("CALENDAR_TODAY"), height=52, border_radius=10,
        content_padding=ft.Padding.symmetric(horizontal=14, vertical=10),
        border_color=BORDER, focused_border_color=NAVY, width=250,
    )
    e_df = ft.TextField(
        label="Data Final", value=_hoje.strftime("%d/%m/%Y"),
        prefix_icon=ic("EVENT"), height=52, border_radius=10,
        content_padding=ft.Padding.symmetric(horizontal=14, vertical=10),
        border_color=BORDER, focused_border_color=NAVY, width=250,
    )

    # ── Campos de período mensal (MM/AAAA) ────────────────────────────────────
    e_mes_ini = ft.TextField(
        label="Mês/Ano Inicial", value=_hoje.strftime("%m/%Y"),
        hint_text="ex: 01/2026",
        prefix_icon=ic("CALENDAR_MONTH"), height=52, border_radius=10,
        content_padding=ft.Padding.symmetric(horizontal=14, vertical=10),
        border_color=BORDER, focused_border_color=NAVY, width=250,
    )
    e_mes_fim = ft.TextField(
        label="Mês/Ano Final", value=_hoje.strftime("%m/%Y"),
        hint_text="ex: 03/2026",
        prefix_icon=ic("EVENT"), height=52, border_radius=10,
        content_padding=ft.Padding.symmetric(horizontal=14, vertical=10),
        border_color=BORDER, focused_border_color=NAVY, width=250,
    )

    row_periodo_diario = ft.Row(
        [e_di, ft.Container(width=12), e_df, ft.Container(expand=True)],
        visible=False,
    )
    row_periodo_mensal = ft.Row(
        [e_mes_ini, ft.Container(width=12), e_mes_fim, ft.Container(expand=True)],
        visible=False,
    )

    # Opção específica do PN: incluir ou não a 8091 no pacote do PN
    chk_pn_incluir_8091 = ft.Checkbox(
        value=False,
        active_color="#2E7D32",
        label="Incluir Macro 8091 no Gerar PN (opcional)",
    )
    row_pn_8091 = ft.Row(
        [chk_pn_incluir_8091, ft.Container(expand=True)],
        visible=False,
    )

    def _atualizar_periodo(e=None):
        """Mostra campos corretos conforme macros selecionadas."""
        tem_diaria = any(
            macro_checks.get(m.get("uid", m["id"])) and macro_checks[m.get("uid", m["id"])].value
            for m in MACROS_CATALOGO if m.get("tipo") == "diaria"
        )
        tem_mensal = any(
            macro_checks.get(m.get("uid", m["id"])) and macro_checks[m.get("uid", m["id"])].value
            for m in MACROS_CATALOGO if m.get("tipo") == "mensal"
        )
        tem_pn = bool(macro_checks.get("PN") and macro_checks["PN"].value)
        row_periodo_diario.visible = tem_diaria
        row_periodo_mensal.visible = tem_mensal
        row_pn_8091.visible = tem_pn
        try:
            row_periodo_diario.update()
            row_periodo_mensal.update()
            row_pn_8091.update()
        except: pass

    status_txt = ft.Text("", size=12, color=TEXT_MUTED, italic=True)
    btn_gerar  = sbtn("Gerar Dashboard", "DASHBOARD")
    btn_limpar_cache = ft.TextButton(
        content=ft.Row([
            ft.Icon(ic("DELETE_SWEEP"), size=13, color=TEXT_MUTED),
            ft.Text("Limpar cache", size=11, color=TEXT_MUTED),
        ], spacing=4),
        tooltip="Remove os dados em cache do dashboard para forcar novo download.",
    )
    lgpd_lbl   = ft.Text(
        "Credenciais usadas apenas em memória  não armazenadas (LGPD Art. 6º, III)",
        size=10, color=TEXT_MUTED, italic=True,
    )

    # ── Seletor de Macros  gerado dinamicamente a partir do catálogo ─────────
    macro_checks = {}   # id -> ft.Checkbox
    macro_cards  = {}   # id -> ft.Container (card visual)

    def _make_macro_card(m):
        cb = ft.Checkbox(value=False, active_color=m["cor"])
        uid = m.get("uid", m["id"])  # Use uid if available, else id
        macro_checks[uid] = cb

        graficos_col = ft.Column(
            [ft.Row([
                ft.Icon(ic("SHOW_CHART"), size=12,
                        color=ft.Colors.with_opacity(0.5, m["cor"])),
                ft.Text(g, size=11, color=TEXT_MUTED),
            ], spacing=6) for g in m["graficos"]],
            spacing=5,
        )

        def _on_toggle(e, uid=uid):
            ativo = macro_checks[uid].value
            card  = macro_cards[uid]
            card.opacity = 1.0 if ativo else 0.4
            try: card.update()
            except: pass
            _atualizar_periodo()

        cb.on_change = _on_toggle

        inner = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Container(
                        content=ft.Icon(ic(m["icone"]), color=m["cor"], size=16),
                        width=32, height=32, bgcolor=m["bg"],
                        border_radius=8, alignment=ft.Alignment(0, 0),
                    ),
                    ft.Container(width=10),
                    ft.Text(m["nome"], size=13, weight=ft.FontWeight.BOLD,
                            color=TEXT_MAIN, expand=True),
                    cb,
                ], vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=0),
                ft.Text(m["desc"], size=11, color=TEXT_MUTED,
                        max_lines=2, overflow=ft.TextOverflow.ELLIPSIS),
                ft.Divider(height=12, color=BORDER),
                ft.Text("Gráficos incluídos:", size=11,
                        color=ft.Colors.with_opacity(0.55, TEXT_MAIN),
                        weight=ft.FontWeight.W_500),
                ft.Container(height=4),
                graficos_col,
            ], spacing=6),
            bgcolor=WHITE,
            border_radius=12,
            padding=ft.Padding.symmetric(horizontal=16, vertical=14),
            width=300,
            opacity=0.4,
            border=ft.Border.all(1, BORDER),
            shadow=ft.BoxShadow(
                spread_radius=0, blur_radius=8,
                color=ft.Colors.with_opacity(0.05, ft.Colors.BLACK),
                offset=ft.Offset(0, 2),
            ),
        )
        macro_cards[uid] = inner
        return inner

    macros_diarias = [m for m in MACROS_CATALOGO if m.get("tipo") == "diaria"]
    macros_mensais = [m for m in MACROS_CATALOGO if m.get("tipo") == "mensal"]
    macros_atemporais = [m for m in MACROS_CATALOGO if m.get("tipo") == "atemporal"]

    row_diarias = ft.Row(
        [_make_macro_card(m) for m in macros_diarias],
        spacing=14, wrap=True, run_spacing=14,
        vertical_alignment=ft.CrossAxisAlignment.START,
    )
    row_mensais = ft.Row(
        [_make_macro_card(m) for m in macros_mensais],
        spacing=14, wrap=True, run_spacing=14,
        vertical_alignment=ft.CrossAxisAlignment.START,
    ) if macros_mensais else ft.Text(
        "Nenhuma macro mensal cadastrada ainda.",
        size=12, color=TEXT_MUTED, italic=True,
    )
    row_atemporais = ft.Row(
        [_make_macro_card(m) for m in macros_atemporais],
        spacing=14, wrap=True, run_spacing=14,
        vertical_alignment=ft.CrossAxisAlignment.START,
    ) if macros_atemporais else ft.Text(
        "Nenhuma macro atemporal cadastrada.",
        size=12, color=TEXT_MUTED, italic=True,
    )

    def _get_base():
        if getattr(sys, "frozen", False):
            return os.path.dirname(sys.executable)
        return os.path.dirname(os.path.abspath(__file__))

    def _limpar_cache_dashboard(e):
        base = _get_base()
        cache_dir = os.path.join(base, "Dashboard_Waterfy")
        removidos = 0

        try:
            if os.path.isdir(cache_dir):
                for nome in os.listdir(cache_dir):
                    if nome == ".cache_index.json" or nome.startswith(".cache_data_"):
                        caminho = os.path.join(cache_dir, nome)
                        try:
                            if os.path.isfile(caminho):
                                os.remove(caminho)
                                removidos += 1
                        except Exception:
                            pass

            if removidos > 0:
                status_txt.value = f"Cache do dashboard limpo ({removidos} arquivo(s))."
                status_txt.color = GREEN
                if lq is not None:
                    try:
                        lq.put(f"Cache do dashboard limpo ({removidos} arquivo(s)).\n")
                    except Exception:
                        pass
            else:
                status_txt.value = "Nenhum cache do dashboard encontrado para limpar."
                status_txt.color = TEXT_MUTED
                if lq is not None:
                    try:
                        lq.put("Nenhum cache do dashboard encontrado para limpar.\n")
                    except Exception:
                        pass
        except Exception as ex:
            status_txt.value = f"Erro ao limpar cache: {ex}"
            status_txt.color = RED
            if lq is not None:
                try:
                    lq.put(f"Erro ao limpar cache do dashboard: {ex}\n")
                except Exception:
                    pass
        finally:
            try:
                status_txt.update()
            except Exception:
                pass

    def on_gerar(e):
        import calendar
        _dash_log("Gerando dashboard...")

        def _meses_intervalo_ddmmyyyy(di_txt: str, df_txt: str):
            """Converte DD/MM/YYYY..DD/MM/YYYY para lista MM/YYYY (inclusive)."""
            dt_i = datetime.strptime(di_txt, "%d/%m/%Y")
            dt_f = datetime.strptime(df_txt, "%d/%m/%Y")
            cur = datetime(dt_i.year, dt_i.month, 1)
            fim = datetime(dt_f.year, dt_f.month, 1)
            meses = []
            while cur <= fim:
                meses.append(cur.strftime("%m/%Y"))
                nm = cur.month + 1
                na = cur.year + (1 if nm > 12 else 0)
                nm = 1 if nm > 12 else nm
                cur = datetime(na, nm, 1)
            return meses

        usuario = e_user.value.strip()
        senha   = e_pass.value.strip()

        macros_sel = [m for m in MACROS_CATALOGO
                      if macro_checks.get(m.get("uid", m["id"])) and macro_checks[m.get("uid", m["id"])].value]

        if not usuario or not senha:
            status_txt.value = "Preencha usuário e senha."
            status_txt.color = RED; status_txt.update()
            if lq is not None:
                try: lq.put("Validacao: preencha usuario e senha.\n")
                except Exception: pass
            return
        if not macros_sel:
            status_txt.value = "Selecione ao menos uma Macro."
            status_txt.color = RED; status_txt.update()
            if lq is not None:
                try: lq.put("Validacao: selecione ao menos uma macro.\n")
                except Exception: pass
            return

        def _macro_runtime_key(macro_cfg):
            """Normaliza a chave enviada ao motor/cache."""
            uid = macro_cfg.get("uid", macro_cfg["id"])
            if uid in ("50012-atemporal", "PN"):
                return uid
            if uid == "50012-mensal":
                return "50012"
            return uid

        # Monta dict de períodos por macro
        periodos = {}
        for m in macros_sel:
            mid = _macro_runtime_key(m)
            tipo = m.get("tipo", "diaria")
            if tipo == "diaria":
                di = e_di.value.strip()
                df = e_df.value.strip()
                if not di or not df:
                    status_txt.value = "Preencha Data Inicial e Final para macros diárias."
                    status_txt.color = RED; status_txt.update()
                    if lq is not None:
                        try: lq.put("Validacao: preencha data inicial e final para macros diarias.\n")
                        except Exception: pass
                    return
                periodos[mid] = (di, df)
            elif tipo == "mensal":
                mi_str = e_mes_ini.value.strip()
                mf_str = e_mes_fim.value.strip()
                if not mi_str or len(mi_str) != 7 or not mf_str or len(mf_str) != 7:
                    status_txt.value = "Preencha Mês/Ano Inicial e Final no formato MM/AAAA."
                    status_txt.color = RED; status_txt.update()
                    if lq is not None:
                        try: lq.put("Validacao: preencha mes/ano inicial e final no formato MM/AAAA.\n")
                        except Exception: pass
                    return
                try:
                    data_ini = datetime.strptime(f"01/{mi_str}", "%d/%m/%Y")
                    data_fim = datetime.strptime(f"01/{mf_str}", "%d/%m/%Y")
                    if data_ini > data_fim:
                        status_txt.value = "Mês/Ano Inicial não pode ser maior que o Final."
                        status_txt.color = RED; status_txt.update()
                        if lq is not None:
                            try: lq.put("Validacao: mes/ano inicial nao pode ser maior que o final.\n")
                            except Exception: pass
                        return
                    # Gera lista de meses MM/AAAA no intervalo
                    meses = []
                    cur = data_ini
                    while cur <= data_fim:
                        meses.append(cur.strftime("%m/%Y"))
                        nm = cur.month + 1
                        na = cur.year + (1 if nm > 12 else 0)
                        nm = 1 if nm > 12 else nm
                        cur = datetime(na, nm, 1)
                    # periodos para macro mensal: lista de strings MM/AAAA
                    periodos[mid] = meses
                except Exception:
                    status_txt.value = "Formato inválido. Use MM/AAAA (ex: 01/2026)."
                    status_txt.color = RED; status_txt.update()
                    if lq is not None:
                        try: lq.put("Validacao: formato invalido. Use MM/AAAA.\n")
                        except Exception: pass
                    return
            else:  # atemporal (50012)
                periodos[mid] = None  # Não requer período específico

        btn_gerar.disabled = True
        status_txt.value   = "Conectando ao Waterfy..."
        status_txt.color   = TEXT_MUTED
        try: btn_gerar.update(); status_txt.update()
        except: pass

        _usr = usuario
        _pwd = senha

        # Verifica se o mapeamento (50012-atemporal) foi selecionado
        _tem_mapeamento = any(m.get("uid") == "50012-atemporal" for m in macros_sel)
        # Verifica se Gerar PN foi selecionado
        _tem_pn = any(m.get("uid") == "PN" for m in macros_sel)
        _pn_incluir_8091 = bool(chk_pn_incluir_8091.value) if _tem_pn else False
        # Macros normais: exclui 50012-atemporal e PN, deduplica por chave de runtime
        _macros_normais = []
        _seen = set()
        for m in macros_sel:
            if m.get("uid") in ["50012-atemporal", "PN"]:
                continue  # Skip special handling macros (processed separately)
            macro_key = _macro_runtime_key(m)
            if macro_key not in _seen:
                _macros_normais.append(macro_key)
                _seen.add(macro_key)
        
        # Detectar se 8117 e 8121 foram selecionadas ANTES de remover
        _selecionou_8117 = '8117' in _macros_normais
        _selecionou_8121 = '8121' in _macros_normais
        _selecionou_8091 = '8091' in _macros_normais
        
        # Se PN foi selecionado, remover 8117 e 8121 de _macros_normais
        # APENAS se também foram selecionadas pelo usuário (para reutilizar do PN)
        if _tem_pn:
            if _selecionou_8117:
                _macros_normais.remove('8117')
            if _selecionou_8121:
                _macros_normais.remove('8121')
            # Só remove 8091 das macros normais se PN também for baixá-la
            if _selecionou_8091 and _pn_incluir_8091:
                _macros_normais.remove('8091')
        
        # Sistema de sincronização para decisão de cache
        _cache_decision = {'use_cache': None}  # None = aguardando, True = usar cache, False = baixar novo
        _cache_confirmation_event = threading.Event()
        
        def _show_cache_confirmation_dialog(macros_list: list, cache_info: dict = None):
            """Mostra diálogo perguntando se usuário quer usar cache ou baixar novamente."""
            def _on_use_cache(_):
                _cache_decision['use_cache'] = True
                dlg_cache.open = False
                _cache_confirmation_event.set()
                try: page.update()
                except: pass
            
            def _on_download_new(_):
                _cache_decision['use_cache'] = False
                dlg_cache.open = False
                _cache_confirmation_event.set()
                try: page.update()
                except: pass
            
            macros_str = ", ".join(macros_list)
            
            # Monta informações detalhadas do cache
            cache_details = f"Os dados das macros {macros_str} já foram baixados hoje."
            
            if cache_info:
                file_name = cache_info.get('file_name', 'dados.json')
                download_date = cache_info.get('download_date', '?')
                periodo = cache_info.get('periodo', '')
                
                # Formata a data para ficar mais legível
                try:
                    from datetime import datetime as _dt_fmt
                    data_fmt = _dt_fmt.fromisoformat(download_date).strftime('%d/%m/%Y às %H:%M')
                except Exception:
                    data_fmt = download_date
                
                cache_details += (
                    f"\n\n Arquivo: {file_name}"
                    f"\n Data: {data_fmt}"
                    f"\n Período: {periodo}"
                )
            
            # Conteúdo com detalhes do cache
            content_elements = [
                ft.Text(cache_details, color=TEXT_MUTED, size=13),
                ft.Divider(height=16, color=BORDER),
                ft.Text("O que você deseja fazer?", color=TEXT_MUTED, size=12, weight=ft.FontWeight.W_500),
            ]
            
            dlg_cache = ft.AlertDialog(
                modal=True,
                bgcolor=WHITE,
                title=ft.Text(" Dados em Cache Detectados", weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
                content=ft.Column(content_elements, spacing=8),
                actions=[
                    ft.Container(
                        content=ft.Row([
                            ft.Icon(ic("CACHED"), size=14, color=WHITE),
                            ft.Container(width=6),
                            ft.Text("Usar Cache", size=13, color=WHITE, weight=ft.FontWeight.W_500),
                        ], spacing=0, alignment=ft.MainAxisAlignment.CENTER,
                           vertical_alignment=ft.CrossAxisAlignment.CENTER),
                        height=44, bgcolor="#2E7D32", border_radius=8,
                        padding=ft.Padding.symmetric(horizontal=16, vertical=0),
                        ink=True,
                        on_click=_on_use_cache,
                    ),
                    ft.Container(
                        content=ft.Row([
                            ft.Icon(ic("REFRESH"), size=14, color=WHITE),
                            ft.Container(width=6),
                            ft.Text("Baixar Novamente", size=13, color=WHITE, weight=ft.FontWeight.W_500),
                        ], spacing=0, alignment=ft.MainAxisAlignment.CENTER,
                           vertical_alignment=ft.CrossAxisAlignment.CENTER),
                        height=44, bgcolor=NAVY, border_radius=8,
                        padding=ft.Padding.symmetric(horizontal=16, vertical=0),
                        ink=True,
                        on_click=_on_download_new,
                    ),
                ],
                actions_alignment=ft.MainAxisAlignment.END,
            )
            page.dialog = dlg_cache
            dlg_cache.open = True
            try: page.update()
            except: pass

        def worker(_u=_usr, _p=_pwd, _macros=_macros_normais, _periodos=periodos,
                   _mapeamento=_tem_mapeamento, _pn_selecionado=_tem_pn,
                   _pn_incluir_8091=_pn_incluir_8091,
                   _sel_8117=_selecionou_8117, _sel_8121=_selecionou_8121, _sel_8091=_selecionou_8091):
            try:
                base = _get_base()

                def _log(msg):
                    status_txt.value = msg
                    try:
                        status_txt.update()
                    except: pass
                    if lq is not None:
                        from datetime import datetime as _dt
                        ts = _dt.now().strftime('%H:%M:%S')
                        try: lq.put(f"[{ts}] {msg}\n")
                        except: pass

                _log("Conectando ao Waterfy...")

                eng = _load_custom_module("waterfy_engine", ["waterfy_engine.py"])
                if not eng:
                    _log("Erro: nao foi possivel carregar modulos internos")
                    raise Exception("Não foi possível carregar waterfy_engine.py")

                gen = _load_custom_module("dashboard_html", ["dashboard_html.py"])
                if not gen:
                    _log("Erro: nao foi possivel carregar modulos internos")
                    raise Exception("Não foi possível carregar dashboard_html.py")

                # ── Mapeamento das Equipes (50012-atemporal) ─────────────────
                _dados_combo_mapa = None
                if _mapeamento:
                    _log("Processando dados geográficos...")
                    res_mapa = eng.gerar_combo_mapa(_u, _p, log_fn=_log, base_dir=base)
                    if res_mapa.get("status") == "error":
                        _log(f"Não foi possível processar dados geográficos. Continuando...")
                    elif res_mapa.get("status") == "partial":
                        sem = res_mapa.get("sem_correspondencia", [])
                        _log(f"Dados geográficos processados (alguns registros sem correspondência).")
                        _dados_combo_mapa = res_mapa
                    else:
                        pessoas = res_mapa.get("pessoas", [])
                        sem_coord = res_mapa.get("sem_coord", 0)
                        total_registros = res_mapa.get('total_registros', 0)
                        msg_mapa = f"Dados geográficos processados ({total_registros} registros)."
                        _log(msg_mapa)
                        _dados_combo_mapa = res_mapa

                # ── GERAR PN (8117 + 8121) ───────────────────────────────────
                _pn_dados = None
                if _pn_selecionado:
                    _log("Processando dados...")
                    
                    # Pega o período diário fornecido
                    data_ini = e_di.value.strip()
                    data_fim = e_df.value.strip()
                    
                    if not data_ini or not data_fim:
                        _log("Defina o período para processar")
                        status_txt.value = "Defina o período (Data Inicial e Final)"
                        status_txt.color = RED
                        try: status_txt.update()
                        except: pass
                        return
                    
                    # Se PN é realmente a ÚNICA seleção (sem 8117/8121/outras macros),
                    # gera HTML separado. Se 8117/8121 também foram marcadas, segue para
                    # o dashboard geral com abas reutilizando os dados do PN.
                    if not _macros and not _tem_mapeamento and not (_sel_8117 or _sel_8121):
                        try:
                            success, message, filepath = gerar_pn_com_download(
                                usuario=_u,
                                senha=_p,
                                data_inicial=data_ini,
                                data_final=data_fim,
                                output_dir=os.path.join(base, "Dashboard_Waterfy"),
                                log_fn=_log,
                                base_dir=base
                            )
                            
                            if success:
                                _log(f"Dashboard gerado com sucesso.")
                                _log("Abrindo no navegador...")
                                webbrowser.open(f"file:///{filepath.replace(chr(92), '/')}")
                                status_txt.value = "Dashboard gerado com sucesso!"
                                status_txt.color = GREEN
                            else:
                                _log(f"Não foi possível gerar o dashboard: {message}")
                                status_txt.value = f"Erro: {message}"
                                status_txt.color = RED
                            
                            try: status_txt.update()
                            except: pass
                            return
                        
                        except Exception as pn_ex:
                            _log(f"Não foi possível gerar: {str(pn_ex)}")
                            status_txt.value = f"Erro ao gerar: {str(pn_ex)}"
                            status_txt.color = RED
                            try: status_txt.update()
                            except: pass
                            return
                    
                    # Senão, PN será integrado ao dashboard como aba
                    # Precisamos buscar dados 8117+8121 para o PN
                    if _sel_8117 or _sel_8121:
                        _log("   ℹ  8117/8121 serão reutilizados do PN (sem novo download)")
                    
                    try:
                        _macros_pn = ['8117', '8121']
                        if _pn_incluir_8091:
                            _macros_pn.append('8091')

                        _cache_mgr_pn = get_cache_manager(os.path.join(base, "Dashboard_Waterfy"))

                        _log(f"   Baixando macros {', '.join(_macros_pn)} para PN...")
                        _meses_8091 = _meses_intervalo_ddmmyyyy(data_ini, data_fim)
                        _periodos_pn = {
                            '8117': [data_ini, data_fim] if ' a ' not in data_ini else data_ini.split(' a '),
                            '8121': [data_ini, data_fim] if ' a ' not in data_ini else data_ini.split(' a '),
                        }
                        if _pn_incluir_8091:
                            _periodos_pn['8091'] = _meses_8091

                        _pn_dados = eng.buscar_macros(_u, _p,
                                                      macros=_macros_pn,
                                                      periodos=_periodos_pn,
                                                      log_fn=_log,
                                                      base_dir=base,
                                                      cache_mgr=_cache_mgr_pn)
                        _log("   OK PN dados carregados")
                    except Exception as pn_dl_ex:
                        _log(f"Não foi possível carregar Painel de Negócios, continuando...")
                        _pn_dados = None

                # ── Macros normais → Dashboard ────────────────────────────────
                if not _macros and not _pn_selecionado:
                    # Só mapeamento foi selecionado  gera dashboard só com o mapa
                    if _dados_combo_mapa:
                        _df_nome = datetime.now().strftime('%d-%m-%Y_%Hh%M')
                        saida = os.path.join(base, "Dashboard_Waterfy",
                                             f"dashboard_mapeamento_{_df_nome}.html")
                        os.makedirs(os.path.dirname(saida), exist_ok=True)
                        dados_mapa_only = {
                            'periodo': {'inicial': datetime.now().strftime('%d/%m/%Y'),
                                        'final':   datetime.now().strftime('%d/%m/%Y')},
                            'cidade': 'ITAPOÁ',
                            'combo_mapa': _dados_combo_mapa,
                        }
                        _log(" Gerando Dashboard de Mapeamento...")
                        gen.gerar_html(dados_mapa_only, saida)
                        _log("OK Dashboard de Mapeamento criado! Abrindo no navegador...")
                        webbrowser.open(f"file:///{saida.replace(chr(92), '/')}")
                    status_txt.value = "OK Mapeamento das Equipes concluído!"
                    status_txt.color = GREEN
                    try: status_txt.update()
                    except: pass
                    return

                # Se não há macros normais MAS há PN, continua para gerar PN integrado
                # Se há macros normais, fecha o dict de dados
                if _macros:
                    _log(f"Macros selecionadas: {', '.join(_macros)}")
                    for mid, pval in _periodos.items():
                        if pval is None:
                            _log(f"   {mid}: (sem período  atemporal)")
                        elif isinstance(pval, list):
                            _log(f"   {mid}: {pval[0]} → {pval[-1]} ({len(pval)} mes(es))")
                        else:
                            _log(f"   {mid}: {pval[0]} → {pval[1]}")

                    # Verifica cache de dados do mesmo dia
                    dados = None
                    _cached_dados = None
                    try:
                        _cache_mgr = get_cache_manager(os.path.join(base, "Dashboard_Waterfy"))
                        _periodo_cache = _periodos.get(_macros[0]) if _macros else None
                        if isinstance(_periodo_cache, (list, tuple)) and _periodo_cache:
                            _data_ini_cache = _periodo_cache[0]
                            _data_fim_cache = _periodo_cache[-1]
                        else:
                            _data_ini_cache = None
                            _data_fim_cache = None
                        
                        _cached_dados = _cache_mgr.get_cached_data(_u, _data_ini_cache, _data_fim_cache, _macros)
                        if _cached_dados:
                            _tem_50012_macro = any(str(m).startswith('50012') for m in _macros)
                            _cache_50012_erro = (
                                _tem_50012_macro and
                                isinstance(_cached_dados.get('macro_50012'), dict) and
                                _cached_dados['macro_50012'].get('status') in ('erro', 'error')
                            )
                            if _cache_50012_erro:
                                _log("   Cache da 50012 contém erro anterior; ignorando cache e tentando novo download.")
                                _cached_dados = None
                            else:
                                _log(f"  Dados das macros {', '.join(_macros)} encontrados em cache do dia!")

                                # Obtém informações do cache para mostrar no diálogo
                                _cache_info = _cache_mgr.get_cache_info(_u, _data_ini_cache, _data_fim_cache, _macros)

                                # Mostra diálogo perguntando ao usuário
                                _cache_decision['use_cache'] = None
                                _cache_confirmation_event.clear()
                                _show_cache_confirmation_dialog(_macros, _cache_info)

                                # Aguarda resposta do usuário (máximo 60 segundos)
                                _user_confirmed = _cache_confirmation_event.wait(timeout=60)

                                if _user_confirmed and _cache_decision['use_cache'] is True:
                                    _log("   OK Usando dados em cache (sem novo download do Waterfy)")
                                    dados = _cached_dados
                                else:
                                    _log("    Usuário optou por baixar novamente do Waterfy")
                                    _cached_dados = None
                    except Exception as cache_ex:
                        _log(f"   ℹ  Cache não disponível: {str(cache_ex)[:50]}  fazendo download do Waterfy...")
                        _cached_dados = None
                    
                    # Se não tem cache (ou usuário escolheu baixar novamente), faz download
                    if not _cached_dados:
                        if _cached_dados is None:
                            _log(f"   Nenhum cache para hoje. Fazendo download do Waterfy...")
                        try:
                            dados = eng.buscar_macros(_u, _p,
                                                      macros=_macros,
                                                      periodos=_periodos,
                                                      log_fn=_log,
                                                      base_dir=base,
                                                      cache_mgr=_cache_mgr)

                            _tem_50012_macro = any(str(m).startswith('50012') for m in _macros)
                            _50012_erro = (
                                _tem_50012_macro and
                                isinstance(dados.get('macro_50012'), dict) and
                                dados['macro_50012'].get('status') in ('erro', 'error')
                            )
                            if _50012_erro:
                                raise Exception(dados['macro_50012'].get('mensagem') or "Falha ao carregar a macro 50012.")
                            
                            # Registra download no cache apenas se não houver falhas parciais/estruturais.
                            try:
                                _8091_falhas = (
                                    dados and
                                    isinstance(dados.get('macro_8091'), dict) and
                                    bool(dados['macro_8091'].get('meses_falha'))
                                )
                                _50012_falha = (
                                    dados and
                                    _tem_50012_macro and
                                    isinstance(dados.get('macro_50012'), dict) and
                                    dados['macro_50012'].get('status') in ('erro', 'error')
                                )
                                if _8091_falhas or _50012_falha:
                                    _log(f"Alguns períodos falharam - cache não foi registrado")
                                else:
                                    _log(f" Registrando cache: usuário={_u}, data_ini={_data_ini_cache}, data_fim={_data_fim_cache}, macros={_macros}")
                                    _cache_mgr.register_macro_data(_u, _data_ini_cache, _data_fim_cache, _macros, dados)
                            except Exception as cache_reg_ex:
                                _log(f"   ℹ  Não foi possível registrar cache: {str(cache_reg_ex)[:50]}")
                        
                        except Exception as login_ex:
                            msg = str(login_ex)

                            _ERROS = {
                                '400': "ERRO HTTP 400  Requisição inválida. O payload enviado ao Waterfy está incorreto.",
                                '401': "ERRO HTTP 401  Não autorizado. Sessão expirada ou credenciais inválidas. Tente novamente.",
                                '403': "ERRO HTTP 403  Acesso negado. Seu usuário não tem permissão para acessar esta macro.",
                                '404': "ERRO HTTP 404  Recurso não encontrado. A URL do Waterfy pode ter mudado.",
                                '408': " HTTP 408  Timeout do servidor. O Waterfy demorou para responder. Tente novamente.",
                                '429': " HTTP 429  Muitas requisições. Aguarde alguns minutos antes de tentar novamente.",
                                '500': "Erro no servidor Waterfy. Tente novamente em alguns minutos.",
                                '502': "Servidor Waterfy com problemas. Tente mais tarde.",
                                '503': "Serviço indisponível. Waterfy pode estar em manutenção.",
                                '504': "Servidor não respondeu a tempo. Tente novamente.",
                                'Login falhou': "Usuário ou senha incorretos. Verifique as credenciais.",
                                'Sessão expirada': "Sessão encerrada. Gere novamente.",
                                'Verifique usuário': "Verifique usuário e senha.",
                                'CSV não encontrado': "Nenhum dado diário para este período.",
                                'CSV da macro': "Arquivo recebido está vazio.",
                                'Nenhuma linha válida': "Nenhum registro para esta cidade.",
                                'Colunas essenciais': "Formato de arquivo inesperado.",
                                'não é um ZIP': "Arquivo inválido. Tente novamente.",
                                'timeout': "Operação demorada. Tente com um período menor.",
                                'Timeout': "Operação demorada. Tente com um período menor.",
                                'ConnectionError': "Sem conexão. Verifique sua internet.",
                                'connection': "Sem conexão. Verifique sua internet.",
                            }

                            mensagem = None
                            for chave, texto in _ERROS.items():
                                if chave in msg:
                                    mensagem = texto
                                    break
                            if not mensagem:
                                mensagem = f"Erro ao buscar dados: {msg}"

                            status_txt.value = mensagem
                            status_txt.color = RED
                            try: status_txt.update()
                            except: pass
                            if lq is not None:
                                try: lq.put(f"ERRO: {msg}\n")
                                except: pass
                            return
                else:
                    # Sem macros normais  inicializa dados vazio
                    dados = {}

                # Determina nome do arquivo baseado no período disponível
                _pval_first = None
                if _macros and _periodos:
                    _pval_first = list(_periodos.values())[0]
                elif _pn_selecionado:
                    _pval_first = (e_di.value.strip(), e_df.value.strip())  # Período do PN
                
                if _pval_first is None or (isinstance(_pval_first, tuple) and not _pval_first[0]):
                    # Fallback para data/hora atual
                    _df_nome = datetime.now().strftime('%d-%m-%Y_%Hh%M')
                elif isinstance(_pval_first, list):
                    _df_nome = _pval_first[-1].replace('/', '-')
                elif isinstance(_pval_first, tuple):
                    _df_nome = _pval_first[1].replace('/', '-')  # data_final
                else:
                    _df_nome = str(_pval_first[1]).replace('/', '-')

                saida = os.path.join(base, "Dashboard_Waterfy",
                                     f"dashboard_{_df_nome}.html")
                _pn_nome_arquivo = f"PN_Financeiro_{_df_nome}.html"
                os.makedirs(os.path.dirname(saida), exist_ok=True)

                # Garante que 'dados' tem as chaves necessárias
                if 'periodo' not in dados:
                    data_ini = e_di.value.strip()  
                    data_fim = e_df.value.strip()
                    if not data_ini:
                        data_ini = datetime.now().strftime('%d/%m/%Y')
                    if not data_fim:
                        data_fim = datetime.now().strftime('%d/%m/%Y')
                    dados['periodo'] = {'inicial': data_ini, 'final': data_fim}
                
                if 'cidade' not in dados:
                    dados['cidade'] = 'ITAPOÁ'

                if _pn_selecionado:
                    dados['pn_html_arquivo'] = _pn_nome_arquivo

                # Injeta dados do combo_mapa se mapeamento também foi executado
                if _dados_combo_mapa:
                    dados['combo_mapa'] = _dados_combo_mapa
                
                # Reutiliza dados do PN se disponível
                if _pn_dados:
                    if _sel_8117 and '8117' not in _macros_normais and 'macro_8117' not in dados:
                        _log("     Reutilizando dados 8117 do PN (sem novo download)")
                        dados['macro_8117'] = _pn_dados.get('macro_8117')
                    
                    if _sel_8121 and '8121' not in _macros_normais and 'macro_8121' not in dados:
                        _log("     Reutilizando dados 8121 do PN (sem novo download)")
                        dados['macro_8121'] = _pn_dados.get('macro_8121')

                    if _sel_8091 and '8091' not in _macros_normais and 'macro_8091' not in dados:
                        _log("     Reutilizando dados 8091 do PN (sem novo download)")
                        dados['macro_8091'] = _pn_dados.get('macro_8091')
                
                # Injeta dados do PN se disponível
                if _pn_dados:
                    dados['macro_8117_pn'] = _pn_dados.get('macro_8117')
                    dados['macro_8121_pn'] = _pn_dados.get('macro_8121')
                    dados['macro_8091_pn'] = _pn_dados.get('macro_8091')

                _log(f"Gerando relatório...")
                gen.gerar_html(dados, saida, incluir_pn=_pn_selecionado)
                _log(f"Relatório criado")

                # Quando PN vier junto com outras macros, também gera o arquivo PN dedicado.
                if _pn_selecionado:
                    if not hasattr(gen, "gerar_pn_html"):
                        _log("Aviso: PN dedicado não foi gerado")
                    else:
                        try:
                            saida_pn = os.path.join(base, "Dashboard_Waterfy", _pn_nome_arquivo)

                            # Prioriza payload próprio do PN; se indisponível, usa fallback
                            # com as macros já presentes no dashboard geral.
                            dados_pn = dict(_pn_dados) if _pn_dados else {}
                            if 'macro_8117' not in dados_pn and 'macro_8117' in dados:
                                dados_pn['macro_8117'] = dados.get('macro_8117')
                            if 'macro_8121' not in dados_pn and 'macro_8121' in dados:
                                dados_pn['macro_8121'] = dados.get('macro_8121')
                            if 'macro_8091' not in dados_pn and 'macro_8091' in dados:
                                dados_pn['macro_8091'] = dados.get('macro_8091')

                            dados_pn.setdefault('periodo', dados.get('periodo', {
                                'inicial': datetime.now().strftime('%d/%m/%Y'),
                                'final': datetime.now().strftime('%d/%m/%Y')
                            }))
                            dados_pn.setdefault('cidade', dados.get('cidade', 'ITAPOÁ'))

                            if dados_pn.get('macro_8117') and dados_pn.get('macro_8121'):
                                gen.gerar_pn_html(dados_pn, saida_pn)
                                _log(f"PN também gerado")
                                if lq is not None:
                                    try: lq.put(f"PN salvo em: {saida_pn}\n")
                                    except: pass
                            else:
                                _log("Aviso: PN selecionado, mas faltam dados para montar o arquivo dedicado")
                        except Exception as pn_mix_ex:
                            _log(f"Aviso  Dashboard gerado, mas PN dedicado falhou: {pn_mix_ex}")

                # ── Distribui CSVs por Equipe em Equipe-COM ──────────────────
                # Usa dados já prontos do _dados_combo_mapa (OSs cruzadas com lat/long)
                # Isso garante que TODOS os funcionários COM-* apareçam corretamente,
                # independente de serem leituristas de água na macro 50012.
                _log(f"Organizando dados por Equipe...")
                if _dados_combo_mapa and _dados_combo_mapa.get('status') != 'error':
                    try:
                        from datetime import datetime as _dt_eq
                        hoje_eq_fmt = _dt_eq.now().strftime('%d-%m-%Y')

                        pontos = _dados_combo_mapa.get('pontos_mapa', [])
                        pessoas_info = _dados_combo_mapa.get('pessoas', [])

                        # Header dos CSVs por pessoa
                        header_pessoa = [
                            'MATRICULA', 'LEITURISTA', 'BAIRRO', 'SETOR',
                            'LATITUDE', 'LONGITUDE', 'LOGRADOURO', 'NR_IMOVEL',
                            'NUM_OS', 'TIPO_SERVICO', 'STATUS_OS', 'DATA_OS',
                            'PRAZO_OS', 'PRIORIDADE', 'CRITICA', 'VOLUME',
                            'MEDIA_CONSUMO', 'VALOR_FATURA', 'CATEGORIA',
                            'GRUPO', 'COBRANCA', 'DATA_LEITURA', 'HIDROMETRO',
                        ]
                        _campo_map = {
                            'MATRICULA': 'matricula', 'LEITURISTA': 'equipe',
                            'BAIRRO': 'bairro', 'SETOR': 'setor',
                            'LATITUDE': 'lat', 'LONGITUDE': 'lng',
                            'LOGRADOURO': 'logradouro', 'NR_IMOVEL': 'nr_imovel',
                            'NUM_OS': 'num_os', 'TIPO_SERVICO': 'tipo_servico',
                            'STATUS_OS': 'status_os', 'DATA_OS': 'data_os',
                            'PRAZO_OS': 'prazo_os', 'PRIORIDADE': 'prioridade',
                            'CRITICA': 'critica', 'VOLUME': 'volume',
                            'MEDIA_CONSUMO': 'media_consumo', 'VALOR_FATURA': 'valor_fatura',
                            'CATEGORIA': 'categoria', 'GRUPO': 'grupo',
                            'COBRANCA': 'cobranca', 'DATA_LEITURA': 'data_leitura',
                            'HIDROMETRO': 'hidrometro',
                        }

                        def _row_to_csv(p):
                            return [str(p.get(_campo_map.get(c, c.lower()), '')) for c in header_pessoa]

                        # Agrupa pontos por equipe
                        by_equipe = {}
                        for p in pontos:
                            eq = (p.get('equipe') or '').strip()
                            by_equipe.setdefault(eq, []).append(p)

                        equipe_com_path = os.path.join(base, "Equipe-COM")
                        os.makedirs(equipe_com_path, exist_ok=True)

                        if pessoas_info:
                            _log(f"   ├─  {len(pessoas_info)} funcionário(s) mapeados:")
                            for p_info in pessoas_info:
                                nome   = p_info['nome']
                                pasta  = f"COM-{nome}"
                                pessoa_dir = os.path.join(equipe_com_path, pasta)
                                os.makedirs(pessoa_dir, exist_ok=True)

                                linhas_pessoa = []
                                for eq_name, rows_eq in by_equipe.items():
                                    if nome.lower() in eq_name.lower() or eq_name.lower() in f"com-{nome.lower()}":
                                        linhas_pessoa.extend(rows_eq)

                                _log(f"OK {nome}: {len(linhas_pessoa)} registros")

                                import openpyxl as _oxl
                                def _save_xlsx(caminho, linhas):
                                    wb = _oxl.Workbook()
                                    ws = wb.active
                                    ws.append(header_pessoa)
                                    for p in linhas:
                                        ws.append(_row_to_csv(p))
                                    wb.save(caminho)

                                _save_xlsx(os.path.join(pessoa_dir, f"Equipe_{nome}_{hoje_eq_fmt}.xlsx"), linhas_pessoa)
                                _save_xlsx(os.path.join(pessoa_dir, f"Equipe_{nome}_Historico.xlsx"), linhas_pessoa)

                            _log(f"Dados organizados por Equipe")
                        else:
                            _log(f"Sem funcionários no mapeamento")

                    except Exception as eq_ex:
                        import traceback
                        _log(f"Aviso: Erro ao organizar dados por Equipe: {eq_ex}")
                        if lq:
                            try: lq.put(f"[DEBUG] {traceback.format_exc()}\n")
                            except: pass
                else:
                    _log(f"Organize as Equipes para distribuir dados")

                status_txt.value = "Relatório gerado com sucesso! Abrindo..."
                status_txt.color = GREEN
                try: status_txt.update()
                except: pass
                if lq is not None:
                    try: lq.put(f"Dashboard salvo em: {saida}\n")
                    except: pass

                webbrowser.open(f"file:///{saida.replace(chr(92), '/')}")

            except Exception as ex:
                import traceback
                tb = traceback.format_exc()
                status_txt.value = f"Ocorreu um erro: {ex}"
                status_txt.color = RED
                try: status_txt.update()
                except: pass
                if lq is not None:
                    try:
                        lq.put(f"ERRO: {ex}\n")
                        lq.put(f"{tb}\n")
                    except: pass
            finally:
                btn_gerar.disabled = False
                btn_gerar.on_click = on_gerar
                try: btn_gerar.update()
                except: pass

        btn_gerar.on_click = None
        threading.Thread(target=worker, daemon=True).start()

    btn_gerar.on_click = on_gerar
    btn_limpar_cache.on_click = _limpar_cache_dashboard

    # ── MODO LOCAL: seletores de arquivo e estado ────────────────────────────
    MACROS_LOCAL = [
        {"id": "8091", "nome": "Macro 8091", "cor": "#2E7D32", "bg": "#E8F5E9"},
        {"id": "8117", "nome": "Macro 8117", "cor": "#1565C0", "bg": "#E3F2FD"},
        {"id": "8121", "nome": "Macro 8121", "cor": "#FF8F00", "bg": "#FFF8E1"},
    ]
    _lc_paths  = {m["id"]: {"files": [], "periodo": None} for m in MACROS_LOCAL}
    _lc_labels = {}
    _modo_local = {"v": False}

    # Pill toggle "Gerar Localmente"
    _pill_icon = ft.Icon(ic("FOLDER_SPECIAL"), size=14, color=ORANGE)
    _pill_txt  = ft.Text("Gerar Localmente", size=12,
                         weight=ft.FontWeight.W_600, color=ORANGE)
    _pill_active_dot = ft.Container(
        width=6, height=6, bgcolor=WHITE, border_radius=3, visible=False,
    )
    _pill = ft.Container(
        content=ft.Row(
            [_pill_icon, _pill_txt, _pill_active_dot],
            spacing=6, vertical_alignment=ft.CrossAxisAlignment.CENTER,
        ),
        padding=ft.Padding.symmetric(horizontal=14, vertical=8),
        border_radius=20,
        border=ft.Border.all(1.5, ORANGE),
        bgcolor=WHITE,
        ink=True,
        tooltip="Gerar dashboard a partir de CSVs locais (sem login)\n\n"
                "Formatos aceitos:\n"
                " .csv (UTF-8/Latin-1) e .zip com CSV interno\n"
                " Nome deve conter: 8091, 8117 ou 8121\n\n"
                "Padrões de competência (8091): MM-AAAA, DD-MM-AAAA\n\n"
                "Exemplos: Macro_8091_03-2026.csv | MACRO 8121_20-03-2026.csv",
        animate=ft.Animation(180, ft.AnimationCurve.EASE_IN_OUT),
    )

    # Labels dinâmicos da seção
    _sec_label = ft.Text(
        "Credenciais e Período", size=13,
        weight=ft.FontWeight.W_600, color=TEXT_MAIN,
    )
    _sec_label_sub = ft.Text(
        "Informe suas credenciais Waterfy e o período desejado.",
        size=12, color=TEXT_MUTED,
    )

    # Painel online (já existente, só embrulhado)
    _painel_online = ft.Column([
        ft.Row([e_user, ft.Container(width=12), e_pass,
                ft.Container(expand=True)], spacing=0),
        ft.Container(height=8),
        row_periodo_diario,
        row_periodo_mensal,
        row_pn_8091,
        ft.Container(height=6),
        ft.Row([btn_gerar, ft.Container(width=8), btn_limpar_cache, ft.Container(expand=True)]),
        ft.Container(height=6),
        status_txt,
        ft.Container(height=8),
        lgpd_lbl,
    ], spacing=0, visible=True)

    # Painel local
    _lc_status = ft.Text("", size=12, color=TEXT_MUTED, italic=True)
    _lc_btn    = sbtn("Gerar Dashboard", "DASHBOARD")
    _lc_btn_pn = sbtn("Gerar Relatório PN", "TABLE_VIEW")
    _lc_btn_pn.bgcolor = "#2E7D32"
    _lc_help_text = (
        "Formatos aceitos no modo local:\n"
        " Arquivos: .csv (UTF-8/Latin-1) e .zip com CSV interno.\n"
        " Nome do arquivo deve conter a macro: 8091, 8117 ou 8121.\n"
        " Competência 8091 no nome (recomendado): MM-AAAA, MM_AAAA, DD-MM-AAAA ou DD/MM/AAAA.\n"
        "Exemplos: Macro_8091_03-2026.csv, MACRO 8121_20-03-2026.csv, macro8117.csv"
    )

    def _lc_log(msg, cor=TEXT_MUTED):
        _lc_status.value = msg
        _lc_status.color = cor
        try: _lc_status.update()
        except: pass
        if lq:
            from datetime import datetime as _dt
            try: lq.put(f"[{_dt.now().strftime('%H:%M:%S')}] {msg}\n")
            except: pass

    def _abrir_ajuda_local(e=None):
        if not page:
            _lc_log("Erro: página não disponível", RED)
            return
        
        def _fechar_ajuda_local():
            try:
                if page.dialog:
                    page.dialog.open = False
                page.update()
            except Exception as ex:
                print(f"Erro ao fechar ajuda: {ex}")
        
        dlg = ft.AlertDialog(
            bgcolor=WHITE,
            title=ft.Text("Formatos aceitos no modo local", weight=ft.FontWeight.W_700, color=TEXT_MAIN),
            content=ft.Container(
                width=520,
                content=ft.Column([
                    ft.Text("Arquivos aceitos", weight=ft.FontWeight.W_600, color=TEXT_MAIN),
                    ft.Text(".csv (UTF-8/Latin-1) e .zip com CSV interno.", size=12, color=TEXT_MUTED),
                    ft.Container(height=8),
                    ft.Text("Identificação da macro no nome", weight=ft.FontWeight.W_600, color=TEXT_MAIN),
                    ft.Text("O nome do arquivo deve conter 8091, 8117 ou 8121.", size=12, color=TEXT_MUTED),
                    ft.Container(height=8),
                    ft.Text("Competência da 8091", weight=ft.FontWeight.W_600, color=TEXT_MAIN),
                    ft.Text("Padrões aceitos: MM-AAAA, MM_AAAA, DD-MM-AAAA ou DD/MM/AAAA.", size=12, color=TEXT_MUTED),
                    ft.Container(height=8),
                    ft.Text("Exemplos", weight=ft.FontWeight.W_600, color=TEXT_MAIN),
                    ft.Text("Macro_8091_03-2026.csv\nMACRO 8121_20-03-2026.csv\nmacro8117.csv", size=12, color=TEXT_MUTED),
                ], tight=True, spacing=0),
            ),
            actions=[
                ft.TextButton("Fechar", on_click=lambda ev: _fechar_ajuda_local()),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )

        try:
            page.dialog = dlg
            dlg.open = True
            page.update()
        except Exception as ex:
            _lc_log("Erro ao abrir ajuda", RED)
            print(f"Erro ao abrir ajuda: {ex}")

    # Labels de status por macro (mostram o arquivo selecionado)
    for _m in MACROS_LOCAL:
        _lbl = ft.Text("", size=11, color=TEXT_MUTED,
                       overflow=ft.TextOverflow.ELLIPSIS, expand=True)
        _lc_labels[_m["id"]] = _lbl

    def _detectar_macro(nome_arquivo):
        """Detecta qual macro é o arquivo pelo número no nome.
        Suporta padrões como: 'Macro 8091_09-2025.csv', 'macro8117.csv', 'MACRO 8121_01-2026.csv'
        """
        import re
        nome_upper = nome_arquivo.upper().replace(" ", "")
        for mid in ("8091", "8117", "8121"):
            if re.search(rf'(?:MACRO[\s_-]*)?{mid}', nome_upper):
                return mid
        return None

    def _extrair_periodo(nome_arquivo, mid=None):
        """
        Extrai o período do nome do arquivo.
        Apenas para macros mensais (8091)  padrão MM-AAAA.
        Para macros diárias (8117, 8121) retorna None  período vem do conteúdo.
        """
        # Macros diárias não têm data no nome
        if mid in ('8117', '8121'):
            return None

        import re
        from datetime import datetime
        import calendar

        # Tenta DD-MM-AAAA ou DD/MM/AAAA
        m = re.search(r'(\d{2})[-/](\d{2})[-/](\d{4})(?!\d)', nome_arquivo)
        if m:
            dia, mes, ano = m.group(1), m.group(2), m.group(3)
            try:
                dt = datetime(int(ano), int(mes), int(dia))
                data_fmt = dt.strftime('%d/%m/%Y')
                return {'inicial': data_fmt, 'final': data_fmt}
            except ValueError:
                pass

        # Tenta MM-AAAA ou MM_AAAA
        m = re.search(r'(\d{2})[-_](\d{4})(?!\d)', nome_arquivo)
        if m:
            mes, ano = int(m.group(1)), int(m.group(2))
            try:
                ultimo_dia = calendar.monthrange(ano, mes)[1]
                inicio = f'01/{mes:02d}/{ano}'
                fim    = f'{ultimo_dia:02d}/{mes:02d}/{ano}'
                return {'inicial': inicio, 'final': fim}
            except ValueError:
                pass

        return None

    def _ler_csv_local(path, eng):
        """Lê CSV ou ZIP e retorna texto do CSV."""
        with open(path, "rb") as f:
            header = f.read(2)
        if header == b'PK':
            with open(path, "rb") as f:
                return eng._extrair_csv_do_zip(f.read(), "8091")
        for enc in ("utf-8-sig", "utf-8", "latin-1", "ISO-8859-1"):
            try:
                with open(path, "r", encoding=enc) as f:
                    txt = f.read()
                    if hasattr(eng, '_normalizar_csv_texto'):
                        return eng._normalizar_csv_texto(txt)
                    return txt.replace('\r\n', '\n').replace('\r', '\n')
            except UnicodeDecodeError:
                continue
        return None

    def _make_pick_btn(m):
        mid = m["id"]
        cor = m["cor"]

        def _abrir(e):
            def _thread():
                import tkinter as tk
                from tkinter import filedialog
                from datetime import datetime as _dtp
                root = tk.Tk()
                root.withdraw()
                root.attributes('-topmost', True)
                paths = filedialog.askopenfilenames(
                    title=f"Selecione arquivo(s) para {m['nome']} (Ctrl+clique para múltiplos)",
                    filetypes=[("Arquivos CSV", "*.csv"), ("Todos os arquivos", "*.*")],
                )
                root.destroy()
                if not paths:
                    return

                # Filtra apenas arquivos que batem com esta macro
                correspondentes = []
                nao_bat = []
                for path in paths:
                    nome = os.path.basename(path)
                    mid_det = _detectar_macro(nome)
                    if mid_det == mid or (len(paths) == 1 and not mid_det):
                        correspondentes.append(path)
                    else:
                        nao_bat.append(nome)

                if not correspondentes:
                    _lc_log(
                        f"Nenhum arquivo válido para {m['nome']} (nome deve conter {mid})",
                        ORANGE,
                    )
                    return

                # Sobrescreve a seleção anterior desta macro
                _lc_paths[mid]["files"] = correspondentes

                # Calcula período consolidado (mais antigo → mais recente)
                periodos = [_extrair_periodo(os.path.basename(p), mid) for p in correspondentes]
                periodos = [p for p in periodos if p]
                qtd_sem_periodo = 0 if mid != '8091' else (
                    len(correspondentes) - len(periodos)
                )

                def _parse(s):
                    try: return _dtp.strptime(s, '%d/%m/%Y')
                    except: return _dtp.min

                if periodos:
                    inicio = min(periodos, key=lambda p: _parse(p['inicial']))['inicial']
                    fim    = max(periodos, key=lambda p: _parse(p['final']))['final']
                    _lc_paths[mid]["periodo"] = {'inicial': inicio, 'final': fim}
                    periodo_txt = f"  ({inicio} → {fim})"
                else:
                    _lc_paths[mid]["periodo"] = None
                    periodo_txt = ""

                n = len(correspondentes)
                nome_display = (
                    os.path.basename(correspondentes[0]) if n == 1
                    else f"{n} arquivos selecionados"
                )
                _lc_labels[mid].value = f"{nome_display}{periodo_txt}"
                _lc_labels[mid].color = TEXT_MAIN

                if nao_bat:
                    _lc_log(f"Ignorado(s): {', '.join(nao_bat)} (não é {mid})", ORANGE)
                else:
                    _lc_log(f"{m['nome']}: {n} arquivo(s) carregado(s).")

                if mid == '8091' and qtd_sem_periodo > 0:
                    _lc_log(
                        f"{qtd_sem_periodo} arquivo(s) sem mês/ano no nome - use MM-AAAA (ex.: 03-2026)",
                        ORANGE,
                    )

                try: _lc_labels[mid].update()
                except: pass

            _thread()

        return ft.Container(
            content=ft.Row([
                ft.Container(
                    content=ft.Text(m["nome"], size=11,
                                    weight=ft.FontWeight.W_700, color=cor),
                    bgcolor=m["bg"], border_radius=6,
                    padding=ft.Padding.symmetric(horizontal=10, vertical=6),
                    width=100,
                ),
                ft.Container(
                    content=ft.Row([
                        ft.Icon(ic("UPLOAD_FILE"), size=13, color=cor),
                        ft.Text("Selecionar", size=11, color=cor,
                                weight=ft.FontWeight.W_500),
                    ], spacing=5),
                    padding=ft.Padding.symmetric(horizontal=10, vertical=6),
                    border=ft.Border.all(1, cor),
                    border_radius=8,
                    ink=True,
                    on_click=_abrir,
                    tooltip="Ctrl+clique para selecionar múltiplos meses de uma vez",
                ),
                ft.Icon(ic("ATTACH_FILE"), size=13,
                        color=ft.Colors.with_opacity(0.35, TEXT_MUTED)),
                _lc_labels[mid],
            ], spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            bgcolor=WHITE,
            border=ft.Border.all(1, BORDER),
            border_radius=10,
            padding=ft.Padding.symmetric(horizontal=12, vertical=10),
        )

    _lc_rows = [_make_pick_btn(m) for m in MACROS_LOCAL]

    def _lc_on_gerar(e):
        selecionados = [mid for mid, p in _lc_paths.items() if p["files"]]
        if not selecionados:
            _lc_log("Aviso Selecione ao menos um arquivo CSV.", RED); return

        _lc_btn.disabled = True
        _lc_log("Processando arquivos locais...")
        try: _lc_btn.update()
        except: pass

        def _worker():
            try:
                eng = _load_custom_module("waterfy_engine", ["waterfy_engine.py"])
                gen = _load_custom_module("dashboard_html", ["dashboard_html.py"])
                if not eng or not gen:
                    _lc_log("Módulos não encontrados.", RED); return

                from datetime import datetime as _dt, datetime as _dtp

                def _parse_dt(s):
                    try: return _dtp.strptime(s, '%d/%m/%Y')
                    except: return _dtp.min

                # Coleta todos os períodos para montar o período global do dashboard
                todos_periodos = []
                for mid in selecionados:
                    p = _lc_paths[mid].get("periodo")
                    if p:
                        todos_periodos.append(p)

                if todos_periodos:
                    _periodo = {
                        'inicial': min(todos_periodos, key=lambda p: _parse_dt(p['inicial']))['inicial'],
                        'final':   max(todos_periodos, key=lambda p: _parse_dt(p['final']))['final'],
                    }
                else:
                    _hoje = _dt.now().strftime("%d/%m/%Y")
                    _periodo = {'inicial': _hoje, 'final': _hoje}

                dados = {"periodo": _periodo, "cidade": "ITAPOÁ"}

                for mid in selecionados:
                    files = _lc_paths[mid]["files"]
                    _lc_log(f" Lendo {mid} ({len(files)} arquivo(s))...")

                    # Ordena cronologicamente:
                    # 8091 → pelo nome do arquivo (MM-AAAA)
                    # 8117/8121 → pela primeira data encontrada no conteúdo do CSV
                    def _sort_key(path):
                        p = _extrair_periodo(os.path.basename(path), mid)
                        if p:
                            return _parse_dt(p['inicial'])
                        # Para diárias: lê a primeira data do CSV
                        try:
                            import csv as _csv
                            csv_txt = _ler_csv_local(path, eng)
                            if csv_txt:
                                reader = _csv.DictReader(
                                    __import__('io').StringIO(csv_txt.strip()), delimiter=';')
                                for row in reader:
                                    for col in ('DATA PAGTO','DATA PAGAMENTO','DATA','DT'):
                                        val = row.get(col, '').strip()
                                        if val:
                                            try: return _dtp.strptime(val, '%d/%m/%Y')
                                            except: pass
                                    break
                        except: pass
                        return _dtp.min

                    files_sorted = sorted(files, key=_sort_key)

                    if mid == "8091":
                        parciais = []
                        for path in files_sorted:
                            csv_texto = _ler_csv_local(path, eng)
                            if not csv_texto:
                                _lc_log(f"Não foi possível ler: {os.path.basename(path)}", RED)
                                return
                            mes_ano = ''
                            p = _extrair_periodo(os.path.basename(path), '8091')
                            if p and p.get('inicial'):
                                try:
                                    mes_ano = _dtp.strptime(p['inicial'], "%d/%m/%Y").strftime("%m/%Y")
                                except Exception:
                                    mes_ano = ''
                            else:
                                _lc_log(
                                    f"Aviso Competência não detectada para 8091 em {os.path.basename(path)}. "
                                    "Usando processamento sem mes_ano.",
                                    ORANGE,
                                )
                            parciais.append(eng.processar_8091(csv_texto, mes_ano=mes_ano))
                        dados["macro_8091"] = eng.merge_8091(parciais)

                    elif mid == "8117":
                        parciais = []
                        for path in files_sorted:
                            csv_texto = _ler_csv_local(path, eng)
                            if not csv_texto:
                                _lc_log(f"Não foi possível ler: {os.path.basename(path)}", RED)
                                return
                            parciais.append(eng.processar_8117(csv_texto))
                        dados["macro_8117"] = eng.merge_8117(parciais)

                    elif mid == "8121":
                        if len(files_sorted) >= 2:
                            # Mais recente = atual, demais = anterior
                            csv_atual = _ler_csv_local(files_sorted[-1], eng)
                            if not csv_atual:
                                _lc_log(f"Não foi possível ler {mid}.", RED); return
                            dados["macro_8121"] = eng.reordenar_8121(eng.processar_8121(csv_atual))

                            parciais_ant = []
                            for path in files_sorted[:-1]:
                                csv_texto = _ler_csv_local(path, eng)
                                if not csv_texto: continue
                                parciais_ant.append(eng.processar_8121(csv_texto))
                            if parciais_ant:
                                dados["macro_8121_anterior"] = eng.merge_8121(parciais_ant)
                        else:
                            csv_texto = _ler_csv_local(files_sorted[0], eng)
                            if not csv_texto:
                                _lc_log(f"Não foi possível ler {mid}.", RED); return
                            dados["macro_8121"] = eng.reordenar_8121(eng.processar_8121(csv_texto))

                    _lc_log(f"{mid} processada.")

                _base = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) \
                        else os.path.dirname(os.path.abspath(__file__))
                saida = os.path.join(_base, "Dashboard_Waterfy",
                                     f"dashboard_local_{_dt.now().strftime('%Y-%m-%d_%H%M%S')}.html")
                os.makedirs(os.path.dirname(saida), exist_ok=True)
                _lc_log("Gerando relatório...")
                gen.gerar_html(dados, saida)
                _lc_log("Relatório pronto! Abrindo...", GREEN)
                webbrowser.open(f"file:///{saida.replace(chr(92), '/')}")
            except Exception as ex:
                import traceback
                _lc_log(f"Ocorreu um erro: {ex}", RED)
                if lq:
                    try: lq.put(f"ERRO LOCAL: {traceback.format_exc()}\n")
                    except: pass
            finally:
                _lc_btn.disabled = False
                try: _lc_btn.update()
                except: pass

        threading.Thread(target=_worker, daemon=True).start()

    def _lc_on_gerar_pn(e):
        selecionados = [mid for mid, v in _lc_paths.items() if v["files"]]

        if "8117" not in selecionados or "8121" not in selecionados:
            _lc_log("Para gerar o relatório, carregue dados de Arrecadação e Faturamento.", RED)
            return

        _lc_btn_pn.disabled = True
        _lc_btn.disabled = True
        try:
            _lc_btn_pn.update()
            _lc_btn.update()
        except: pass
        _lc_log("Processando dados para o Painel de Negócios...")

        def _worker_pn():
            try:
                base = _get_base()
                eng = _load_custom_module("waterfy_engine", ["waterfy_engine.py"])
                gen = _load_custom_module("dashboard_html", ["dashboard_html.py"])
                if not eng or not gen:
                    _lc_log("ERRO Módulos não encontrados.", RED)
                    return
                if not hasattr(gen, "gerar_pn_html"):
                    _lc_log("ERRO Função gerar_pn_html não encontrada em dashboard_html.py.", RED)
                    return

                from datetime import datetime as _dt

                periodo_base = _lc_paths["8121"].get("periodo")
                _hoje = _dt.now().strftime("%d/%m/%Y")
                _periodo = periodo_base if periodo_base else {'inicial': _hoje, 'final': _hoje}
                dados = {"periodo": _periodo, "cidade": "ITAPOÁ"}

                _lc_log(" Processando 8117...")
                arquivos_8117 = [_ler_csv_local(p, eng) for p in _lc_paths["8117"]["files"]]
                proc_8117 = [eng.processar_8117(txt) for txt in arquivos_8117 if txt]
                if not proc_8117:
                    _lc_log("ERRO Não foi possível ler/processar os arquivos da macro 8117.", RED)
                    return
                dados["macro_8117"] = eng.merge_8117(proc_8117)

                _lc_log(" Processando 8121...")
                arquivos_8121 = [_ler_csv_local(p, eng) for p in _lc_paths["8121"]["files"]]
                proc_8121 = [eng.processar_8121(txt) for txt in arquivos_8121 if txt]
                if not proc_8121:
                    _lc_log("ERRO Não foi possível ler/processar os arquivos da macro 8121.", RED)
                    return
                dados["macro_8121"] = eng.merge_8121(proc_8121)

                # 8091 é opcional no PN local: processa apenas se houver arquivos selecionados
                if _lc_paths["8091"]["files"]:
                    _lc_log(" Processando 8091 (opcional)...")

                    def _sort_8091(path):
                        try:
                            p = _extrair_periodo(os.path.basename(path), '8091')
                            if p and p.get('inicial'):
                                return _dt.strptime(p['inicial'], "%d/%m/%Y")
                        except Exception:
                            pass
                        return _dt.min

                    proc_8091 = []
                    for path_8091 in sorted(_lc_paths["8091"]["files"], key=_sort_8091):
                        csv_8091 = _ler_csv_local(path_8091, eng)
                        if not csv_8091:
                            _lc_log(f"ERRO Não foi possível ler o arquivo 8091: {os.path.basename(path_8091)}.", RED)
                            return

                        mes_ano = ''
                        try:
                            p = _extrair_periodo(os.path.basename(path_8091), '8091')
                            if p and p.get('inicial'):
                                mes_ano = _dt.strptime(p['inicial'], "%d/%m/%Y").strftime("%m/%Y")
                        except Exception:
                            mes_ano = ''

                        if not mes_ano:
                            _lc_log(
                                f"Aviso Competência não detectada para 8091 em {os.path.basename(path_8091)}. "
                                "Usando processamento sem mes_ano.",
                                ORANGE,
                            )

                        proc_8091.append(eng.processar_8091(csv_8091, mes_ano=mes_ano))

                    if not proc_8091:
                        _lc_log("ERRO Não foi possível ler/processar os arquivos da macro 8091.", RED)
                        return

                    dados["macro_8091"] = eng.merge_8091(proc_8091)
                    _lc_log(f"OK 8091 opcional integrada ({len(proc_8091)} arquivo(s)).")
                else:
                    _lc_log("ℹ 8091 não selecionada no PN local (opcional).")

                # Ajusta período do PN com base nas datas reais da macro 8121
                try:
                    datas_pn = dados["macro_8121"].get("datas", [])
                    datas_dt = []
                    for d in datas_pn:
                        try:
                            datas_dt.append(_dt.strptime(str(d), "%d/%m/%Y"))
                        except Exception:
                            pass
                    if datas_dt:
                        _periodo = {
                            'inicial': min(datas_dt).strftime('%d/%m/%Y'),
                            'final': max(datas_dt).strftime('%d/%m/%Y'),
                        }
                        dados["periodo"] = _periodo
                except Exception:
                    pass

                _df_nome = _periodo['final'].replace('/', '-')
                saida = os.path.join(base, "Dashboard_Waterfy", f"PN_Financeiro_{_df_nome}.html")
                os.makedirs(os.path.dirname(saida), exist_ok=True)

                _lc_log(" Montando dashboard...")
                gen.gerar_pn_html(dados, saida)

                _lc_log("Dashboard pronto!", GREEN)
                webbrowser.open(f"file:///{saida.replace(chr(92), '/')}")

            except Exception as ex:
                import traceback
                _lc_log(f"Erro ao gerar o relatório: {ex}", RED)
                if lq:
                    try: lq.put(f"ERRO PN LOCAL: {traceback.format_exc()}\n")
                    except: pass
            finally:
                _lc_btn_pn.disabled = False
                _lc_btn.disabled = False
                try:
                    _lc_btn_pn.update()
                    _lc_btn.update()
                except: pass

        threading.Thread(target=_worker_pn, daemon=True).start()

    _lc_btn.on_click = _lc_on_gerar
    _lc_btn_pn.on_click = _lc_on_gerar_pn

    _painel_local = ft.Column([
        ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Text("Formatos e nomes aceitos", size=11, color=TEXT_MUTED),
                    ft.IconButton(
                        icon=ic("HELP_OUTLINE"),
                        icon_size=15,
                        icon_color=TEXT_MUTED,
                        tooltip="Seguindo os padrões de arquivos e nomes aceitos, o processo de geração local é mais fluido e com menos chances de erros.",
                        on_click=_abrir_ajuda_local,
                        style=ft.ButtonStyle(padding=0),
                    ),
                ], spacing=6, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                ft.Container(height=8),
                ft.Text(_lc_help_text, size=10, color=ft.Colors.with_opacity(0.7, TEXT_MUTED),
                        selectable=True),
                ft.Container(height=8),
                *_lc_rows,
            ], spacing=8),
            bgcolor=SURFACE,
            border_radius=10,
            padding=ft.Padding.symmetric(horizontal=14, vertical=12),
            border=ft.Border.all(1, BORDER),
        ),
        ft.Container(height=14),
        ft.Row([_lc_btn, ft.Container(width=10), _lc_btn_pn, ft.Container(expand=True)]),
        ft.Container(height=6),
        _lc_status,
    ], spacing=0, visible=False)

    # Toggle entre os dois painéis
    def _toggle_modo(e):
        _modo_local["v"] = not _modo_local["v"]
        ativo = _modo_local["v"]
        _dash_log("Modo local ativado." if ativo else "Modo online ativado.")
        _painel_online.visible  = not ativo
        _painel_local.visible   = ativo
        _pill.bgcolor           = ORANGE if ativo else WHITE
        _pill_txt.color         = WHITE  if ativo else ORANGE
        _pill_icon.color        = WHITE  if ativo else ORANGE
        _pill_active_dot.visible = ativo
        _sec_label.value = (
            "Modo Local  Sem Login"       if ativo else "Credenciais e Período"
        )
        _sec_label_sub.value = (
            "Usando arquivos CSV locais, nenhuma credencial necessária."
            if ativo else
            "Informe suas credenciais Waterfy e o período desejado."
        )
        try:
            page.update()
        except: pass

    _pill.on_click = _toggle_modo

    col = ft.Column([
        ft.Text("Painel de Controle", size=22, weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
        ft.Text("Bem-vindo ao Mathools 1.0 - Itapoá Saneamento.", size=13, color=TEXT_MUTED),
        ft.Container(height=16),
        ft.Row([
            stat("TABLE_CHART",  _s["csvs"],       "CSVs Processados",   NAVY,      "#EEF2FF", ref_csvs),
            ft.Container(width=12),
            stat("CHAT",         _s["mensagens"],  "Mensagens Enviadas", "#0277BD", "#E1F5FE", ref_mensagens),
            ft.Container(width=12),
            stat("ATTACH_MONEY", _s["propostas"],  "Propostas Geradas",  GREEN,     "#F0FDF4", ref_propostas),
            ft.Container(width=12),
            stat("LOCAL_SHIPPING", _s["paradas"],  "Paradas Calculadas", "#FF6F00", "#FFF8E1", ref_paradas),
        ]),
        ft.Container(height=4),
        ft.Row([
            ft.TextButton(
                content=ft.Row([
                    ft.Icon(ic("RESTART_ALT"), size=13, color=TEXT_MUTED),
                    ft.Text("Zerar contadores", size=11, color=TEXT_MUTED),
                ], spacing=4),
                on_click=reset_stats,
            ),
        ], alignment=ft.MainAxisAlignment.END),
        ft.Container(height=12),
        card(ft.Column([
            ft.Row([
                sec_title("Dashboard Financeiro Waterfy", "DASHBOARD"),
                ft.Container(expand=True),
                _pill,
            ], vertical_alignment=ft.CrossAxisAlignment.CENTER),
            ft.Container(height=4),
            ft.Text(
                "Selecione as macros desejadas, informe as credenciais e o período.",
                size=12, color=TEXT_MUTED,
            ),
            ft.Container(height=16),
            ft.Divider(color=BORDER, height=1),
            ft.Container(height=14),
            ft.Row([
                ft.Icon(ic("LOCK"), size=14,
                        color=ft.Colors.with_opacity(0.5, TEXT_MAIN)),
                _sec_label,
            ], spacing=6, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            ft.Container(height=2),
            _sec_label_sub,
            ft.Container(height=12),
            _painel_online,
            _painel_local,
            ft.Container(height=18),
            ft.Divider(color=BORDER, height=1),
            ft.Container(height=14),
            sec_title("Macros Diárias", "TODAY"),
            ft.Container(height=8),
            row_diarias,
            ft.Container(height=16),
            ft.Divider(color=BORDER, height=1),
            ft.Container(height=14),
            sec_title("Macros Mensais", "CALENDAR_MONTH"),
            ft.Container(height=8),
            row_mensais,
            ft.Container(height=16),
            ft.Divider(color=BORDER, height=1),
            ft.Container(height=14),
            sec_title("Análise Geográfica", "MAP"),
            ft.Container(height=8),
            row_atemporais,
        ])),
    ], spacing=0, scroll=ft.ScrollMode.AUTO, expand=True)

    return col, ref_csvs, ref_mensagens, ref_propostas, ref_paradas


def build_csv_page(page: ft.Page, lq: queue.Queue, ref_csvs=None):
    e_main  = tfield("Nenhum arquivo selecionado")
    e_macro = tfield("Opcional  Macro 8162 para cruzamento de dados")
    is_running = {"v": False}

    # ── Spinner numérico de dias ───────────────────────────────────────────────
    corte_val = {"n": 30}
    corte_num_lbl = ft.Text(
        str(corte_val["n"]), size=16, weight=ft.FontWeight.BOLD,
        color=TEXT_MAIN, text_align=ft.TextAlign.CENTER,
    )

    def _update_corte_display():
        corte_num_lbl.value = str(corte_val["n"])
        try: corte_num_lbl.update()
        except: pass

    def corte_up(e):
        corte_val["n"] = min(corte_val["n"] + 1, 999)
        _update_corte_display()

    def corte_down(e):
        corte_val["n"] = max(corte_val["n"] - 1, 1)
        _update_corte_display()

    def _spin_btn(icon_name, handler):
        return ft.Container(
            content=ft.Icon(ic(icon_name), size=16, color=WHITE),
            width=30, height=26,
            bgcolor=NAVY, border_radius=6,
            alignment=ft.Alignment(0, 0),
            ink=True, on_click=handler,
            shadow=ft.BoxShadow(spread_radius=0, blur_radius=4,
                                color=ft.Colors.with_opacity(0.2, ft.Colors.BLACK),
                                offset=ft.Offset(0, 1)),
        )

    spinner = ft.Container(
        content=ft.Row([
            ft.Container(
                content=corte_num_lbl,
                width=56, height=44,
                bgcolor=SURFACE,
                border=ft.Border.all(1.5, BORDER),
                border_radius=8,
                alignment=ft.Alignment(0, 0),
            ),
            ft.Container(width=6),
            ft.Column([
                _spin_btn("KEYBOARD_ARROW_UP",   corte_up),
                ft.Container(height=2),
                _spin_btn("KEYBOARD_ARROW_DOWN", corte_down),
            ], spacing=0),
        ], spacing=0, vertical_alignment=ft.CrossAxisAlignment.CENTER),
    )

    # ── Barra de progresso do processamento ───────────────────────────────────
    csv_prog      = ft.ProgressBar(value=0, bgcolor=BORDER, color=ORANGE,
                                   height=7, border_radius=4)
    csv_prog_lbl  = ft.Text("Aguardando execução...", size=11,
                             color=TEXT_MUTED, italic=True)
    csv_prog_card = card(ft.Column([
        ft.Row([
            ft.Icon(ic("HOURGLASS_EMPTY"), color=ORANGE, size=16),
            ft.Text("Progresso do Processamento", size=13,
                    weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
            ft.Container(expand=True),
            csv_prog_lbl,
        ], spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER),
        ft.Container(height=10),
        csv_prog,
    ]), mb=0)

    def _set_prog(val, msg, color=TEXT_MUTED):
        csv_prog.value    = val        # None = indeterminate, 01 = fixed
        csv_prog_lbl.value = msg
        csv_prog_lbl.color = color
        try: csv_prog.update()
        except: pass
        try: csv_prog_lbl.update()
        except: pass

    # ── Seletor de arquivo ────────────────────────────────────────────────────
    def pick(entry, e):
        def do_pick():
            try:
                import tkinter as tk
                from tkinter import filedialog
                root = tk.Tk(); root.withdraw(); root.attributes('-topmost', True)
                path = filedialog.askopenfilename(filetypes=[("Excel / CSV", "*.xlsx *.csv"), ("Todos", "*.*")])
                root.destroy()
                if not path:
                    return

                entry.value = path
                try:
                    entry.update()
                except Exception:
                    pass

                lq.put(f"Arquivo selecionado: {os.path.basename(path)}\n")

            except Exception as ex:
                lq.put(f"Erro ao abrir seletor: {ex}\n")
        do_pick()

    # ── Executar ──────────────────────────────────────────────────────────────
    def run(e):
        if not e_main.value or not os.path.exists(e_main.value):
            page.snack_bar = ft.SnackBar(ft.Text("Selecione a base Macro 8104!", color=WHITE), bgcolor=RED, open=True)
            page.update()
            return
        if e_macro.value and os.path.exists(e_macro.value):
            try:
                if os.path.abspath(e_main.value.strip()) == os.path.abspath(e_macro.value.strip()):
                    page.snack_bar = ft.SnackBar(
                        ft.Text("A base 8104 e a 8162 não podem ser o mesmo arquivo.", color=WHITE),
                        bgcolor=RED,
                        open=True,
                    )
                    page.update()
                    return
            except Exception:
                pass
        if is_running["v"]:
            page.snack_bar = ft.SnackBar(ft.Text("Processamento já em andamento!", color=WHITE), bgcolor=ORANGE, open=True)
            page.update()
            return

        def worker():
            is_running["v"] = True
            _set_prog(None, "Preparando processamento...", ORANGE)
            engine = load_engine()
            if not engine:
                lq.put("Componente de processamento não encontrado.\n")
                _set_prog(0, "Componente não encontrado.", RED)
                is_running["v"] = False
                return

            orig_out   = sys.stdout
            _log_ctx   = None
            sys.stdout = sys.stderr = LogCapture(lq)
            _log_ctx   = _attach_ui_logger(engine, lq)

            _arq_8104 = e_main.value.strip()
            _arq_8162 = e_macro.value.strip() if e_macro.value else None
            _corte    = corte_val["n"]

            _csv_ok = False
            try:
                _set_prog(None, "Processando dados...", ORANGE)
                lq.put("Carregando dados...\n")
                # Chama diretamente com parâmetros  sem Tkinter nem input() bloqueante
                engine.processar_csv(
                    arquivo_8104=_arq_8104,
                    arquivo_8162=_arq_8162,
                    corte_dias=_corte,
                )
                lq.put("Planilha gerada com sucesso.\n")
                _set_prog(1, "Concluído com sucesso!", GREEN)
                _csv_ok = True
            except Exception as ex:
                msg_amigavel = str(ex)
                if "PermissionError" in type(ex).__name__ or "Permission denied" in str(ex):
                    msg_amigavel = "A planilha de destino está aberta em outro programa. Por favor, feche o Excel e tente novamente."
                
                lq.put(f"Não conseguimos concluir o processamento: {msg_amigavel}\n")
                _set_prog(0, "Atenção: verifique a planilha e tente novamente.", RED)
            finally:
                if _log_ctx is not None:
                    _detach_ui_logger(_log_ctx)
                sys.stdout = orig_out
                is_running["v"] = False
                # Incrementa contador APÓS restaurar stdout (evita conflito com LogCapture)
                if _csv_ok:
                    _v = increment_stat("csvs")
                    if ref_csvs and ref_csvs[0]:
                        try: ref_csvs[0].value = str(_v); ref_csvs[0].update()
                        except: pass

        threading.Thread(target=worker, daemon=True).start()
        lq.put("Iniciando processamento...\n")
        lq.put(f"   Base principal  : {e_main.value}\n")
        if e_macro.value: lq.put(f"   Cruzamento     : {e_macro.value}\n")
        lq.put(f"   Corte de atraso : {corte_val['n']} dias\n")

    return ft.Column([
        ft.Text("Processar CSV", size=22, weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
        ft.Text("Processamento e cruzamento de bases de inadimplência.", size=13, color=TEXT_MUTED),
        ft.Container(height=8),
        card(ft.Column([
            sec_title("Macro 8104  Base Principal *", "TABLE_CHART"),
            ft.Container(height=10),
            ft.Row([e_main, ft.Container(width=8),
                    sbtn("Procurar", "FOLDER", on_click=lambda e: pick(e_main, e))],
                   vertical_alignment=ft.CrossAxisAlignment.CENTER),
        ])),
        card(ft.Column([
            sec_title("Macro 8162  Cruzamento (Opcional)", "COMPARE_ARROWS"),
            ft.Container(height=10),
            ft.Row([e_macro, ft.Container(width=8),
                    sbtn("Procurar", "FOLDER", on_click=lambda e: pick(e_macro, e))],
                   vertical_alignment=ft.CrossAxisAlignment.CENTER),
        ])),
        # ── Barra de progresso ────────────────────────────────────────────────
        csv_prog_card,
        ft.Container(height=16),
        # ── Critério de filtragem com spinner ─────────────────────────────────
        card(ft.Column([
            sec_title("Critério de Filtragem", "TUNE"),
            ft.Container(height=12),
            ft.Row([
                ft.Column([
                    ft.Text("Dias mínimos de atraso", size=12,
                            color=TEXT_MUTED, weight=ft.FontWeight.W_600),
                    ft.Container(height=6),
                    ft.Row([
                        spinner,
                        ft.Container(width=10),
                        ft.Text("dias", size=13, color=TEXT_MUTED),
                    ], vertical_alignment=ft.CrossAxisAlignment.CENTER),
                ], spacing=0),
                ft.Container(width=28),
                ft.Container(
                    content=ft.Column([
                        ft.Icon(ic("INFO_OUTLINE"), color=ORANGE, size=18),
                        ft.Container(height=4),
                        ft.Text("Use botões +/- para\najustar o valor",
                                size=11, color=TEXT_MUTED,
                                text_align=ft.TextAlign.CENTER),
                    ], spacing=0,
                       horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                    padding=ft.Padding.symmetric(horizontal=14, vertical=10),
                    bgcolor=SURFACE, border_radius=8,
                    border=ft.Border.all(1, BORDER),
                ),
            ], vertical_alignment=ft.CrossAxisAlignment.CENTER),
        ])),
        pbtn("Executar Processamento", "PLAY_ARROW", on_click=run, width=480),
    ], spacing=0, scroll=ft.ScrollMode.AUTO, expand=True)


# ─────────────────────────────────────────────────────────────────────────────
#  PÁGINA 2  DISPARO WHATSAPP
# ─────────────────────────────────────────────────────────────────────────────
def build_send_page(page: ft.Page, lq: queue.Queue, ref_msgs=None):
    e_xlsx  = tfield("Nenhuma planilha selecionada")
    e_rate  = tfield("", expand=False, width=100, value="20")
    # Novos controles para download de faturas
    cb_baixar_faturas = ft.Switch(label="Baixar faturas PDF individuais", value=False)
    e_usuario_waterfy = tfield("Usuário Waterfy", password=False)
    e_senha_waterfy = tfield("Senha Waterfy", password=True)
    credenciais_container = ft.Container(
        content=ft.Column([
            ft.Container(height=10),
            ft.Row([e_usuario_waterfy, ft.Container(width=10), e_senha_waterfy]),
        ]),
        visible=False,
    )
    
    sel        = {"sheets": [], "all_names": [], "checks": {}, "rows": {}}
    sheets_col = ft.Column(spacing=4)

    # Contador badge
    badge_count = ft.Container(
        content=ft.Text("0", size=10, color=WHITE, weight=ft.FontWeight.BOLD,
                        text_align=ft.TextAlign.CENTER),
        bgcolor=TEXT_MUTED, border_radius=10,
        width=20, height=20, alignment=ft.Alignment(0, 0),
        visible=False,
    )
    lbl_sh = ft.Text("Nenhuma planilha carregada ainda.", size=12, color=TEXT_MUTED, italic=True)

    btn_sel_all = ft.Container(
        content=ft.Row([ft.Icon(ic("DONE_ALL"), size=13, color=NAVY),
                        ft.Text("Todas", size=11, color=NAVY, weight=ft.FontWeight.W_600)],
                       spacing=4),
        bgcolor="#EEF2FF", border_radius=6,
        padding=ft.Padding.symmetric(horizontal=10, vertical=5),
        ink=True, visible=False,
    )
    btn_desel_all = ft.Container(
        content=ft.Row([ft.Icon(ic("CLOSE"), size=13, color=TEXT_MUTED),
                        ft.Text("Limpar", size=11, color=TEXT_MUTED, weight=ft.FontWeight.W_600)],
                       spacing=4),
        bgcolor=SURFACE, border_radius=6,
        padding=ft.Padding.symmetric(horizontal=10, vertical=5),
        ink=True, visible=False,
        border=ft.Border.all(1, BORDER),
    )

    def _update_sel_sheets():
        sel["sheets"] = [
            name for name in sel["all_names"]
            if sel["checks"].get(name) and sel["checks"][name].value
        ]
        n     = len(sel["sheets"])
        total = len(sel["all_names"])
        # Atualiza badge
        badge_count.content.value = str(n)
        badge_count.bgcolor       = ORANGE if n > 0 else TEXT_MUTED
        badge_count.visible       = total > 0
        # Atualiza label status
        if total == 0:
            lbl_sh.value  = "Nenhuma planilha carregada ainda."
            lbl_sh.color  = TEXT_MUTED
            lbl_sh.italic = True
        elif n == 0:
            lbl_sh.value  = "Nenhuma aba selecionada. Por favor, marque as que deseja usar."
            lbl_sh.color  = RED
            lbl_sh.italic = True
        else:
            lbl_sh.value  = f"Tudo certo! {n} de {total} aba(s) selecionada(s)."
            lbl_sh.color  = GREEN
            lbl_sh.italic = False
        # Destaca visualmente cada row
        for name, row_c in sel["rows"].items():
            marcada = sel["checks"].get(name) and sel["checks"][name].value
            row_c.bgcolor = ft.Colors.with_opacity(0.08, ORANGE) if marcada else WHITE
            row_c.border  = ft.Border.all(1.5, ORANGE) if marcada else ft.Border.all(1, BORDER)
            try: row_c.update()
            except: pass
        try: badge_count.update(); lbl_sh.update()
        except: pass

    def _toggle_sheet(name: str):
        cb = sel["checks"].get(name)
        if cb:
            cb.value = not cb.value
            try: cb.update()
            except: pass
            _update_sel_sheets()

    def _toggle_all(marcar: bool, e=None):
        for cb in sel["checks"].values():
            cb.value = marcar
            try: cb.update()
            except: pass
        _update_sel_sheets()

    btn_sel_all.on_click   = lambda e: _toggle_all(True,  e)
    btn_desel_all.on_click = lambda e: _toggle_all(False, e)
    
    # Handler para o switch de baixar faturas
    def _on_baixar_faturas_change(e):
        credenciais_container.visible = cb_baixar_faturas.value
        try: credenciais_container.update()
        except: pass
    
    cb_baixar_faturas.on_change = _on_baixar_faturas_change
    
    state   = {"running": False, "key": None, "paused": False}
    _contagem = {"enviados": 0, "falhas": 0, "total": 0}

    prog   = ft.ProgressBar(value=0, bgcolor=BORDER, color=ORANGE, height=10, border_radius=5)
    lbl_st = ft.Text("Pronto para começar! Aguardando o seu comando...", size=12, color=TEXT_MUTED, italic=True)
    lbl_count = ft.Text("", size=12, color=TEXT_MUTED)
    _wpp_stage = {"pct": 0.0, "label": "Pronto para começar! Aguardando o seu comando...", "detail": ""}

    # ── Contador em tempo real ────────────────────────────────────────────────
    lbl_enviados = ft.Text("0", size=18, weight=ft.FontWeight.BOLD, color=GREEN)
    lbl_falhas   = ft.Text("0", size=18, weight=ft.FontWeight.BOLD, color=RED)
    lbl_total    = ft.Text("0", size=18, weight=ft.FontWeight.BOLD, color=TEXT_MUTED)

    contador_card = ft.Container(
        content=ft.Row([
            ft.Column([
                ft.Text("Enviadas", size=11, color=TEXT_MUTED, weight=ft.FontWeight.W_600),
                lbl_enviados,
            ], spacing=2, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            ft.Container(width=1, height=40, bgcolor=BORDER),
            ft.Column([
                ft.Text("Falhas", size=11, color=TEXT_MUTED, weight=ft.FontWeight.W_600),
                lbl_falhas,
            ], spacing=2, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            ft.Container(width=1, height=40, bgcolor=BORDER),
            ft.Column([
                ft.Text("Total", size=11, color=TEXT_MUTED, weight=ft.FontWeight.W_600),
                lbl_total,
            ], spacing=2, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
        ], spacing=24, alignment=ft.MainAxisAlignment.CENTER),
        bgcolor=SURFACE,
        border_radius=10,
        padding=ft.Padding.symmetric(horizontal=24, vertical=12),
        border=ft.Border.all(1, BORDER),
        visible=False,
    )

    def _reset_contagem():
        _contagem["enviados"] = 0
        _contagem["falhas"]   = 0
        _contagem["total"]    = 0
        _wpp_stage["pct"]     = 0.0
        _wpp_stage["label"]   = "Pronto para começar! Aguardando o seu comando..."
        _wpp_stage["detail"]  = ""
        lbl_enviados.value = "0"
        lbl_falhas.value   = "0"
        lbl_total.value    = "0"
        lbl_count.value    = ""
        prog.value         = 0
        prog.color         = ORANGE
        contador_card.visible = False
        try:
            lbl_enviados.update(); lbl_falhas.update()
            lbl_total.update(); contador_card.update(); lbl_count.update(); prog.update()
        except: pass

    def _refresh_wpp_progress():
        total = max(0, int(_contagem.get("total", 0) or 0))
        enviados = max(0, int(_contagem.get("enviados", 0) or 0))
        falhas = max(0, int(_contagem.get("falhas", 0) or 0))
        concluidos = min(total, enviados + falhas) if total > 0 else 0
        etapa_pct = min(1.0, max(0.0, float(_wpp_stage.get("pct", 0.0) or 0.0)))
        if total > 0:
            valor = min(1.0, (concluidos + etapa_pct) / total)
        else:
            valor = etapa_pct
        prog.value = valor
        try:
            prog.update()
        except Exception:
            pass

    def _atualizar_contagem():
        lbl_enviados.value = str(_contagem["enviados"])
        lbl_falhas.value   = str(_contagem["falhas"])
        lbl_total.value    = str(_contagem["total"])
        lbl_falhas.color   = RED if _contagem["falhas"] > 0 else TEXT_MUTED
        contador_card.visible = True
        _refresh_wpp_progress()
        try:
            lbl_enviados.update(); lbl_falhas.update()
            lbl_total.update(); contador_card.update()
        except: pass

    def _atualizar_etapa_wpp(pct: float, label: str, detail: str = ""):
        _wpp_stage["pct"] = min(1.0, max(0.0, pct))
        _wpp_stage["label"] = label or "Processando..."
        _wpp_stage["detail"] = detail or ""
        lbl_st.value = _wpp_stage["label"]
        lbl_st.color = ORANGE if pct < 1 else GREEN
        if detail:
            lbl_count.value = f"Etapa atual: {detail}"
        else:
            lbl_count.value = ""
        _refresh_wpp_progress()
        try:
            lbl_st.update()
            lbl_count.update()
        except Exception:
            pass

    # ── Timer de próxima mensagem ─────────────────────────────────────────────
    _timer_state = {"total": 0, "restante": 0, "ativo": False, "firing": False, "sid": 0}

    # ── Componentes do timer ──────────────────────────────────────────────────
    timer_countdown = ft.Text(
        "00:00", size=26, weight=ft.FontWeight.BOLD,
        color=ORANGE, font_family="monospace",
        text_align=ft.TextAlign.CENTER,
    )
    timer_label = ft.Text(
        "Próxima mensagem em", size=11,
        color=ft.Colors.with_opacity(0.6, WHITE),
        text_align=ft.TextAlign.CENTER,
        weight=ft.FontWeight.W_500,
    )
    timer_bar = ft.ProgressBar(
        value=1.0,
        bgcolor=ft.Colors.with_opacity(0.12, WHITE),
        color=ORANGE,
        height=4,
        border_radius=2,
    )
    timer_divider = ft.Container(
        width=1, height=40,
        bgcolor=ft.Colors.with_opacity(0.15, WHITE),
    )
    timer_icon_container = ft.Container(
        content=ft.Icon(ic("SCHEDULE"), color=ORANGE, size=20),
        width=40, height=40,
        bgcolor=ft.Colors.with_opacity(0.12, ORANGE),
        border_radius=20,
        alignment=ft.Alignment(0, 0),
    )
    timer_card = ft.Container(
        content=ft.Column([
            ft.Row([
                timer_icon_container,
                ft.Container(width=14),
                ft.Column([
                    timer_label,
                    ft.Container(height=2),
                    timer_countdown,
                ], spacing=0, horizontal_alignment=ft.CrossAxisAlignment.START),
                ft.Container(width=14),
                timer_divider,
                ft.Container(width=14),
                ft.Column([
                    ft.Text("Barra de espera", size=10,
                            color=ft.Colors.with_opacity(0.45, WHITE),
                            weight=ft.FontWeight.W_500),
                    ft.Container(height=6),
                    ft.Container(content=timer_bar, width=180),
                ], spacing=0, horizontal_alignment=ft.CrossAxisAlignment.START),
            ], vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=0),
        ], spacing=0),
        bgcolor="#0B1D33",
        border_radius=12,
        padding=ft.Padding.symmetric(horizontal=20, vertical=14),
        border=ft.Border.all(1, ft.Colors.with_opacity(0.18, ORANGE)),
        shadow=ft.BoxShadow(
            spread_radius=0, blur_radius=16,
            color=ft.Colors.with_opacity(0.22, ORANGE),
            offset=ft.Offset(0, 2),
        ),
        visible=False,
        opacity=0,
        animate_opacity=ft.Animation(350, ft.AnimationCurve.EASE_IN_OUT),
    )

    def _iniciar_timer(segundos: int):
        if segundos <= 0:
            return
        _timer_state["sid"] += 1
        meu_sid = _timer_state["sid"]
        _timer_state["ativo"] = True
        timer_label.value = "Próxima mensagem em"
        timer_label.color = ft.Colors.with_opacity(0.6, WHITE)
        timer_icon_container.content = ft.Icon(ic("SCHEDULE"), color=ORANGE, size=20)
        timer_icon_container.bgcolor = ft.Colors.with_opacity(0.12, ORANGE)
        timer_countdown.color = ORANGE
        timer_bar.color       = ORANGE
        timer_card.visible    = True
        timer_card.opacity    = 1.0
        _atualizar_etapa_wpp(0.0, "Preparando próxima mensagem...", f"Intervalo de {segundos}s")
        try: timer_card.update()
        except: pass

        async def _tick():
            inicio = time.time()
            pausado_em = None  # timestamp de quando pausou
            ultimo_seg = -1
            while _timer_state["sid"] == meu_sid:
                # Detecta pausa  congela o início para o tempo não avançar
                if state["paused"]:
                    if pausado_em is None:
                        pausado_em = time.time()
                    await asyncio.sleep(0.1)
                    continue

                # Retomou  reajusta o início descontando o tempo pausado
                if pausado_em is not None:
                    inicio += time.time() - pausado_em
                    pausado_em = None

                elapsed = time.time() - inicio
                r = max(0, segundos - elapsed)
                seg_atual = int(r)
                if seg_atual != ultimo_seg:
                    ultimo_seg = seg_atual
                    mins, secs = divmod(seg_atual, 60)
                    timer_countdown.value = f"{mins:02d}:{secs:02d}"
                    timer_bar.value = r / segundos
                    if seg_atual <= 5:
                        timer_bar.color       = GREEN
                        timer_countdown.color = GREEN
                    else:
                        timer_bar.color       = ORANGE
                        timer_countdown.color = ORANGE
                    try:
                        timer_countdown.update()
                        timer_bar.update()
                    except: pass
                if r <= 0:
                    break
                await asyncio.sleep(0.1)

        page.run_task(_tick)

    def _parar_timer():
        _timer_state["ativo"]    = False
        _timer_state["sid"]     += 1
        timer_card.opacity  = 0
        timer_card.visible  = False
        try: timer_card.update()
        except: pass

        # Variáveis auxiliares
    _tqdm_obj = None
    orig_tw = None

    def mkib(icon_name, bg_color, label, tip, dis=True):
        return ft.Container(
            content=ft.Column([
                ft.Icon(ic(icon_name), size=20, color=WHITE),
                ft.Text(label, size=10, color=WHITE, weight=ft.FontWeight.W_600,
                        text_align=ft.TextAlign.CENTER),
            ], spacing=4, horizontal_alignment=ft.CrossAxisAlignment.CENTER,
               alignment=ft.MainAxisAlignment.CENTER),
            width=80, height=58, border_radius=10, ink=True,
            alignment=ft.Alignment(0, 0), tooltip=tip,
            bgcolor=bg_color,
            opacity=0.35 if dis else 1.0,
            shadow=ft.BoxShadow(spread_radius=0, blur_radius=6,
                                color=ft.Colors.with_opacity(0.25, ft.Colors.BLACK),
                                offset=ft.Offset(0, 2)),
        )

    btn_pause  = mkib("PAUSE",      NAVY,   "Pausar",  "Pausar Robô",                          dis=True)
    btn_resume = mkib("PLAY_ARROW", GREEN,  "Retomar", "Retomar Robô",                         dis=True)
    btn_skip   = mkib("SKIP_NEXT",  ORANGE, "Pular",   "Pular Registro (só quando pausado)",   dis=True)
    btn_abort  = mkib("STOP",       RED,    "Abortar", "Abortar Tudo",                         dis=True)

    def _refresh_btns():
        """Atualiza opacidade dos botões de acordo com o estado atual."""
        rodando = state["running"]
        pausado = state["paused"]
        btn_pause.opacity  = 1.0 if rodando and not pausado else 0.35
        btn_resume.opacity = 1.0 if rodando and pausado      else 0.35
        btn_skip.opacity   = 1.0 if rodando and pausado      else 0.35
        btn_abort.opacity  = 1.0 if rodando                  else 0.35
        try:
            btn_pause.update(); btn_resume.update()
            btn_skip.update();  btn_abort.update()
        except: pass

    def on_pause(e):
        if not state["running"] or state["paused"]:
            return
        state["key"] = "p"
        state["paused"] = True
        lbl_st.value = "Envios pausados. Clique em Retomar ou Pular para continuar."
        lbl_st.color = ORANGE
        lbl_count.value = "Etapa atual: Pausa solicitada por você"
        try: lbl_st.update()
        except: pass
        _refresh_btns()

    def on_resume(e):
        if not state["running"] or not state["paused"]:
            return
        state["key"] = "c"
        state["paused"] = False
        lbl_st.value = "Voltando ao trabalho! Retomando os envios..."
        lbl_st.color = GREEN
        lbl_count.value = "Etapa atual: Retomando a fila de envios"
        try: lbl_st.update()
        except: pass
        _refresh_btns()

    def on_skip(e):
        # Skip só tem efeito quando o engine está no loop de pausa aguardando 'c' ou 's'
        if not state["running"] or not state["paused"]:
            return
        state["key"] = "s"
        state["paused"] = False
        lbl_st.value = "Entendido! Pulando este cliente..."
        lbl_st.color = ORANGE
        lbl_count.value = "Etapa atual: O registro será ignorado"
        try: lbl_st.update()
        except: pass
        _refresh_btns()
        lq.put("Registro pulado pelo operador.\n")

    def on_abort(e):
        if not state["running"]:
            return
        state["key"] = "m"
        state["paused"] = False
        lbl_st.value = "OPERAÇÃO CANCELADA"
        lbl_st.color = RED
        lbl_count.value = "Etapa atual: Encerrando os disparos"
        try: lbl_st.update()
        except: pass
        _refresh_btns()

    btn_pause.on_click  = on_pause
    btn_resume.on_click = on_resume
    btn_skip.on_click   = on_skip
    btn_abort.on_click  = on_abort

    def _salvar_progresso():
        """Salva progresso dos disparos."""
        if ref_msgs:
            try:
                _v = increment_stat("mensagens")
                if ref_msgs[0]:
                    ref_msgs[0].value = str(_v)
                    ref_msgs[0].update()
            except Exception:
                pass

    def procurar_arquivo(e):
        """Abre seletor de arquivo XLSX."""
        def do_pick():
            try:
                import tkinter as tk
                from tkinter import filedialog
                root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
                path = filedialog.askopenfilename(
                    filetypes=[("Excel / CSV", "*.xlsx *.xls *.csv"), ("Todos", "*.*")]
                )
                root.destroy()
                if not path:
                    return
                e_xlsx.value = path
                try:
                    e_xlsx.update()
                except Exception:
                    pass

            except Exception as ex:
                lq.put(f"Erro ao abrir seletor: {ex}\n")
        do_pick()

    def load_sheets(e):
        path = e_xlsx.value

        def show_snack(msg, color=RED):
            sb = ft.SnackBar(ft.Text(msg, color=WHITE), bgcolor=color, open=True)
            page.overlay.append(sb)
            page.update()

        if not path or not os.path.exists(path):
            lq.put("Aviso: Selecione uma planilha do Excel primeiro!\n")
            show_snack("Por favor, selecione uma planilha do Excel válida primeiro!")
            return

        try:
            import openpyxl
            lq.put(f"Lendo abas: {os.path.basename(path)}...\n")
            wb = openpyxl.load_workbook(path, read_only=True)
            names = list(wb.sheetnames)
            wb.close()

            # Limpa estado anterior
            sel["all_names"] = names
            sel["sheets"]    = []
            sel["checks"]    = {}
            sheets_col.controls.clear()

            # Gera card clicável por aba
            for name in names:
                cb = ft.Checkbox(value=False, active_color=ORANGE)

                def _on_change(ev, n=name):
                    _update_sel_sheets()
                cb.on_change = _on_change
                sel["checks"][name] = cb

                # Ícone de tipo de aba
                is_invalido = "invalid" in name.lower() or "inválid" in name.lower()
                is_maior    = "maior" in name.lower()
                icon_name   = "WARNING" if is_invalido else ("ARROW_UPWARD" if is_maior else "TABLE_ROWS")
                icon_color  = RED if is_invalido else (ORANGE if is_maior else "#1565C0")

                row_card = ft.Container(
                    content=ft.Row([
                        ft.Container(
                            content=ft.Icon(ic(icon_name), size=14, color=icon_color),
                            width=30, height=30,
                            bgcolor=ft.Colors.with_opacity(0.1, icon_color),
                            border_radius=6,
                            alignment=ft.Alignment(0, 0),
                        ),
                        ft.Container(width=10),
                        ft.Text(name, size=13, color=TEXT_MAIN,
                                weight=ft.FontWeight.W_500, expand=True),
                        cb,
                    ], vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=0),
                    bgcolor=WHITE,
                    border_radius=8,
                    padding=ft.Padding.symmetric(horizontal=12, vertical=8),
                    border=ft.Border.all(1, BORDER),
                    ink=True,
                    on_click=lambda ev, n=name: _toggle_sheet(n),
                )
                sel["rows"][name] = row_card
                sheets_col.controls.append(row_card)

            btn_sel_all.visible   = True
            btn_desel_all.visible = True
            lbl_sh.value  = f"0 de {len(names)} aba(s)  marque as que deseja disparar."
            lbl_sh.color  = RED
            lbl_sh.italic = True

            try:
                sheets_col.update()
                lbl_sh.update()
                btn_sel_all.update()
                btn_desel_all.update()
                page.update()
            except: pass

            lq.put(f"Temos { len(names) } aba(s). Marque as que deseja usar.\n")

        except ImportError:
            lq.put("Erro: openpyxl não instalado.\n")
            show_snack("Ops! Falta instalar um componente necessário (openpyxl).")
        except Exception as ex:
            msg_erro = str(ex)
            if "BadZipFile" in type(ex).__name__ or "Invalid file" in msg_erro:
                msg_amigavel = "O arquivo selecionado parece estar corrompido ou não é um Excel válido. Tente salvá-lo novamente."
            else:
                msg_amigavel = "Não conseguimos ler a planilha. Verifique se ela está salva corretamente e se não está com senha."
            lq.put(f"Não foi possível abrir: {msg_amigavel}\n")
            show_snack(msg_amigavel)

    def start_send(e):
        if not e_xlsx.value or not os.path.exists(e_xlsx.value):
            page.snack_bar = ft.SnackBar(ft.Text("A planilha selecionada não foi encontrada!", color=WHITE), bgcolor=RED, open=True)
            page.update(); return
        if not sel["sheets"]:
            page.snack_bar = ft.SnackBar(ft.Text("Lembre-se de carregar a planilha e escolher pelo menos uma aba!", color=WHITE), bgcolor=RED, open=True)
            page.update(); return
        if state["running"]:
            page.snack_bar = ft.SnackBar(ft.Text("Já existe um envio acontecendo agora. Aguarde!", color=WHITE), bgcolor=ORANGE, open=True)
            page.update(); return

        try:
            msgs_por_hora = int(e_rate.value or "20")
        except ValueError:
            msgs_por_hora = 20

        state["running"] = True; state["key"] = None; state["paused"] = False
        lbl_st.value = "Iniciando os envios... Por favor, não feche o navegador!"; lbl_st.color = ORANGE
        lbl_count.value = "Etapa atual: Preparando o disparo"
        prog.value = 0
        _wpp_stage["pct"] = 0.0
        _wpp_stage["label"] = lbl_st.value
        _wpp_stage["detail"] = "Preparando disparo"
        page._wpp_timer_fn = _iniciar_timer      # expõe para o poll global
        # Calcula total de registros e reseta contagem
        try:
            import openpyxl
            wb = openpyxl.load_workbook(e_xlsx.value, read_only=True)
            total_rows = sum(
                max(0, wb[s].max_row - 1)
                for s in sel["sheets"] if s in wb.sheetnames
            )
            wb.close()
        except Exception:
            total_rows = 0
        _reset_contagem()
        _contagem["total"] = total_rows
        lbl_total.value = str(total_rows)
        contador_card.visible = True
        try: contador_card.update()
        except: pass
        page._wpp_counter_fn = _atualizar_contagem  # expõe pro poll
        page._wpp_contagem   = _contagem            # expõe estado pro poll
        page._wpp_stage_fn   = _atualizar_etapa_wpp
        try:
            lbl_st.update(); lbl_count.update(); prog.update()
        except: pass
        _refresh_btns()

        sent   = {"n": 0}
        total  = {"n": 0}
        msgs_enviadas = {"count": 0}

        def worker():
            lq.put("Iniciando assistente de envio automático...\n")
            
            engine = load_engine()
            if not engine:
                lq.put("Componente de automação não encontrado.\n")
                state["running"] = False
                return

            # Importa msvcrt diretamente
            import msvcrt
            orig_kbhit = msvcrt.kbhit
            orig_getch = msvcrt.getch
            orig_out   = sys.stdout
            _log_ctx   = None

            def _kbhit(): return state["key"] is not None
            def _getch():
                k = state["key"]
                state["key"] = None
                return k.encode("utf-8") if k else b""

            # Aplicando o "sequestro" GLOBALMENTE direto no módulo do sistema
            msvcrt.kbhit = _kbhit
            msvcrt.getch = _getch

            orig_out = sys.stdout
            sys.stdout = sys.stderr = LogCapture(lq)
            _log_ctx = _attach_ui_logger(engine, lq)
            _wpp_ok = False
            try:
                lq.put("Preparando o Chrome e o WhatsApp Web para você...\n")
                lq.put("Os envios começaram! Por favor, não feche o navegador.\n")
                
                # Parâmetros para download de faturas
                baixar_faturas = cb_baixar_faturas.value
                usuario_waterfy = e_usuario_waterfy.value if baixar_faturas else ""
                senha_waterfy = e_senha_waterfy.value if baixar_faturas else ""
                
                if baixar_faturas:
                    lq.put(f"Modo: Download de faturas PDF ativado\n")
                else:
                    lq.put(f"Modo: Imagem padrão\n")
                
                engine.enviar_whatsapp_sheets(sel["sheets"], e_xlsx.value, msgs_por_hora, baixar_faturas, usuario_waterfy, senha_waterfy)
                lq.put("Prontinho! Finalizamos os envios desta lista.\n")
                _wpp_ok = True
            except Exception as ex:
                lq.put(f"Tivemos um problema inesperado com os envios. Verifique a planilha ou sua conexão de internet.\n(Detalhe técnico para o suporte: {type(ex).__name__})\n")
            finally:
                if _log_ctx is not None:
                    _detach_ui_logger(_log_ctx)
                sys.stdout = orig_out
                # Restaurando o tqdm.write original  evita encadeamento em disparos futuros
                if _tqdm_obj is not None and orig_tw is not None:
                    _tqdm_obj.write = orig_tw
                # Restaurando o msvcrt globalmente (sem "engine.")
                msvcrt.kbhit = orig_kbhit
                msvcrt.getch = orig_getch
                state["running"] = False
                state["paused"]  = False
                _parar_timer()
                page._wpp_timer_fn    = None
                page._wpp_counter_fn  = None
                page._wpp_contagem    = None
                page._wpp_stage_fn    = None
                _refresh_btns()

            if _wpp_ok:
                try:
                    lbl_st.value = "Disparo concluído."; lbl_st.color = GREEN
                    lbl_count.value = "Etapa atual: lote finalizado"
                    prog.value = 1
                    lbl_st.update(); lbl_count.update(); prog.update()
                except: pass

       
        lq.put(f"   Abas      : {', '.join(sel['sheets'])}\n")
        lq.put(f"   Velocidade : {msgs_por_hora} msgs/hora\n")
        lq.put(f"   Modo      : {'PDF individual' if cb_baixar_faturas.value else 'Imagem padrão'}\n")
        threading.Thread(target=worker, daemon=True).start()

    return ft.Column([
        ft.Text("Disparo WhatsApp", size=22, weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
        ft.Text("Comunicação automatizada via WhatsApp Web.", size=13, color=TEXT_MUTED),
        ft.Container(height=8),
        card(ft.Column([
            sec_title("Selecionar Arquivo XLSX *", "DESCRIPTION"),
            ft.Container(height=10),
            ft.Row([
                e_xlsx,
                ft.Container(width=8),
                sbtn("Procurar", "FOLDER", on_click=procurar_arquivo),
            ], vertical_alignment=ft.CrossAxisAlignment.CENTER),
        ])),
        card(ft.Column([
            sec_title("Configurações de Envio", "SETTINGS"),
            ft.Container(height=12),
            # Botão de carregar + badge contador
            ft.Row([
                pbtn("Carregar Abas do Arquivo", "LIST", on_click=load_sheets, width=300),
                ft.Container(width=10),
                badge_count,
            ], vertical_alignment=ft.CrossAxisAlignment.CENTER),
            ft.Container(height=14),
            # Área de seleção de abas
            ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Text("Abas disponíveis", size=12,
                                color=TEXT_MUTED, weight=ft.FontWeight.W_600),
                        ft.Container(expand=True),
                        btn_sel_all,
                        ft.Container(width=6),
                        btn_desel_all,
                    ], vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    ft.Container(height=10),
                    sheets_col,
                    ft.Container(height=10),
                    ft.Row([
                        ft.Icon(ic("INFO_OUTLINE"), size=13, color=TEXT_MUTED),
                        ft.Container(width=4),
                        lbl_sh,
                    ], vertical_alignment=ft.CrossAxisAlignment.CENTER),
                ]),
                bgcolor=SURFACE,
                border_radius=10,
                padding=ft.Padding.symmetric(horizontal=16, vertical=14),
                border=ft.Border.all(1, BORDER),
            ),
            ft.Divider(height=20, color=BORDER),
            ft.Row([
                ft.Text("Velocidade (msgs/hora):", size=13, color=TEXT_MAIN),
                ft.Container(width=16), e_rate,
                ft.Text("msgs/hora", size=12, color=TEXT_MUTED),
            ], vertical_alignment=ft.CrossAxisAlignment.CENTER),
            ft.Divider(height=20, color=BORDER),
            cb_baixar_faturas,
            credenciais_container,
        ])),
        card(ft.Column([
            sec_title("Controles da Automação", "SETTINGS_APPLICATIONS"),
            ft.Container(height=10),
            ft.Row([btn_pause, btn_resume, btn_skip, ft.Container(expand=True), btn_abort]),
            ft.Container(height=10),
            prog,
            ft.Container(height=6),
            ft.Row([lbl_st, ft.Container(expand=True)]),
            ft.Container(height=4),
            ft.Row([lbl_count, ft.Container(expand=True)]),
            ft.Container(height=10),
            contador_card,
            ft.Container(height=10),
            timer_card,
        ])),
        pbtn("Iniciar Disparo WhatsApp", "SEND", on_click=start_send, width=450),
    ], spacing=0, scroll=ft.ScrollMode.AUTO, expand=True)


# ─────────────────────────────────────────────────────────────────────────────
#  PÁGINA 3  NEGOCIAÇÃO
# ─────────────────────────────────────────────────────────────────────────────
def build_neg_page(page: ft.Page, lq: queue.Queue, ref_props=None):
    e_neg = tfield("Nenhum arquivo selecionado")
    is_running = {"v": False}
    state = {"key": None}  # Estado local para esta página

    def pick(e):
        def do_pick():
            try:
                import tkinter as tk
                from tkinter import filedialog
                root = tk.Tk(); root.withdraw(); root.attributes('-topmost', True)
                path = filedialog.askopenfilename(filetypes=[("CSV","*.csv")])
                root.destroy()
                if not path:
                    return
                e_neg.value = path; e_neg.update()
            except Exception as ex:
                lq.put(f"Erro ao abrir seletor: {ex}\n")
        do_pick()

    def run(e):
        if not e_neg.value or not os.path.exists(e_neg.value):
            page.snack_bar = ft.SnackBar(ft.Text("Selecione a Macro 8104!", color=WHITE), bgcolor=RED, open=True)
            page.update(); return
        if is_running["v"]:
            page.snack_bar = ft.SnackBar(ft.Text("Já em andamento!", color=WHITE), bgcolor=ORANGE, open=True)
            page.update(); return

        def worker():
            is_running["v"] = True
            engine = load_engine()
            orig_kbhit = msvcrt.kbhit
            orig_getch = msvcrt.getch
            orig_out   = sys.stdout

            # 2. Cria as funções falsas baseadas nos botões da UI
            def _kbhit(): return state["key"] is not None
            def _getch():
                k = state["key"]
                state["key"] = None
                return k.encode("utf-8") if k else b""

            # 3. Aplica o "sequestro" GLOBALMENTE direto no módulo do sistema
            msvcrt.kbhit = _kbhit
            msvcrt.getch = _getch

            if not engine:
                lq.put("Motor RPA não encontrado (mathtools_1_0.py)\n"); is_running["v"] = False; return
            orig_out = sys.stdout
            _log_ctx = None
            sys.stdout = sys.stderr = LogCapture(lq)
            _log_ctx = _attach_ui_logger(engine, lq)
            _neg_ok = False
            try:
                lq.put("Calculando propostas...\n")
                engine.gerar_planilha_negociacao(e_neg.value, engine.DIR_NEGOCIACAO)
                lq.put("Planilha de negociação gerada em Relatorios_Negociacao/\n")
                _neg_ok = True
            except Exception as ex:
                lq.put(f"ERRO: {ex}\n")
            finally:
                if _log_ctx is not None:
                    _detach_ui_logger(_log_ctx)
                sys.stdout = orig_out
                msvcrt.kbhit = orig_kbhit
                msvcrt.getch = orig_getch
                is_running["v"] = False
                
            if _neg_ok:
                _v = increment_stat("propostas")
                if ref_props and ref_props[0]:
                    try: ref_props[0].value = str(_v); ref_props[0].update()
                    except: pass

        threading.Thread(target=worker, daemon=True).start()

    return ft.Column([
        ft.Text("Negociação", size=22, weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
        ft.Text("Gerador inteligente de propostas financeiras com descontos automáticos.", size=13, color=TEXT_MUTED),
        ft.Container(height=8),
        card(ft.Column([
            sec_title("Macro 8104 de Referência *", "DESCRIPTION"),
            ft.Container(height=10),
            ft.Row([
                e_neg,
                ft.Container(width=8),
                sbtn("Procurar", "FOLDER", on_click=pick),
            ], vertical_alignment=ft.CrossAxisAlignment.CENTER),
        ])),
        card(ft.Column([
            sec_title("O que será gerado", "STAR"),
            ft.Container(height=10),
            ft.Column([
                check_row("Planilhas com descontos automáticos por faixa de atraso"),
                check_row("Relatórios segmentados por categoria de cliente"),
                check_row("Arquivos exportados em Relatorios_Negociacao/"),
            ], spacing=8),
        ])),
        pbtn("Gerar Planilhas de Propostas Financeiras", "ATTACH_MONEY", on_click=run, width=450),
    ], spacing=0, scroll=ft.ScrollMode.AUTO, expand=True)



# ─────────────────────────────────────────────────────────────────────────────
#  PÁGINA 4  LOGÍSTICA (100% Flet  sem Tkinter)
# ─────────────────────────────────────────────────────────────────────────────
def build_routes_page(page: ft.Page, lq: queue.Queue, ref_paradas=None):

    # Estado interno
    state = {
        "paradas":          [],
        "paradas_com_coord": [],
        "paradas_sem_coord": [],
        "map_file":         None,
        "app_dir":          None,
        "running":          False,
        "recalc_needed":    False,
    }

    SEDE = {"id": "SEDE", "coord": (-26.08401244163725, -48.61613419448409)}

    # Widgets reativos
    lbl_arquivo = ft.Text("Nenhum arquivo selecionado.", size=13,
                          color=TEXT_MUTED, italic=True)
    btn_recalcular = None

    def set_recalc_needed(flag: bool):
        state["recalc_needed"] = flag
        if btn_recalcular is not None:
            btn_recalcular.visible = flag
            try:
                btn_recalcular.update()
            except Exception:
                pass

    def _metric_card(icon_name, label, color, bg):
        value_txt = ft.Text("---", size=22, weight=ft.FontWeight.BOLD, color=color)
        label_txt = ft.Text(label, size=11, color=TEXT_MUTED)
        box = ft.Container(
            content=ft.Column([
                ft.Container(
                    content=ft.Icon(ic(icon_name), color=color, size=18),
                    width=36, height=36, bgcolor=bg,
                    border_radius=8, alignment=ft.Alignment(0, 0),
                ),
                ft.Container(height=6),
                value_txt,
                label_txt,
            ], spacing=2, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            bgcolor=WHITE, border_radius=10, padding=16, expand=True,
            border=ft.Border.all(1, BORDER),
        )
        return box, value_txt

    card_entregas,  txt_entregas  = _metric_card("PLACE", "Entregas",   NAVY,      "#EEF2FF")
    card_distancia, txt_distancia = _metric_card("ROUTE", "Distancia",  RED,       "#FEF2F2")
    card_tempo,     txt_tempo     = _metric_card("TIMER", "Tempo Est.", "#0891B2", "#EFF6FF")

    metrics_row = ft.Row([
        card_entregas,  ft.Container(width=10),
        card_distancia, ft.Container(width=10),
        card_tempo,
    ], expand=False)

    def refresh_metricas(n=None, dist_km=None, min_val=None):
        txt_entregas.value  = str(n)                if n        is not None else "---"
        txt_distancia.value = f"{dist_km:.1f} km"   if dist_km  is not None else "---"
        txt_tempo.value     = f"{min_val} min"       if min_val  is not None else "---"
        try:
            txt_entregas.update()
            txt_distancia.update()
            txt_tempo.update()
        except Exception:
            pass
    
    def atualizar_contador_sem_coord():
        """Atualiza o contador de OSs sem coordenada."""
        qtd = len(state.get("paradas_sem_coord", []))
        lbl_sem_coord_count.value = str(qtd)
        lbl_sem_coord_count.color = ORANGE if qtd > 0 else TEXT_MUTED
        try:
            lbl_sem_coord_count.update()
        except Exception:
            pass

    lbl_status = ft.Text("Motor de logistica nao iniciado.",
                         size=13, color=TEXT_MUTED, italic=True)
    lbl_sem_coord_count = ft.Text("0", size=11, color=ORANGE, weight=ft.FontWeight.BOLD)
    prog = ft.ProgressBar(value=0, bgcolor=BORDER, color=ORANGE,
                          height=6, border_radius=3)
    
    def set_status(msg, color=TEXT_MUTED, prog_val=0):
        lbl_status.value = msg
        lbl_status.color = color
        prog.value = prog_val
        try:
            lbl_status.update()
            prog.update()
        except Exception:
            pass


    # Lista de paradas com Drag & Drop
    lista_col = ft.Column(spacing=0, scroll=ft.ScrollMode.AUTO)

    drag_src = {"idx": None}  # Flet 0.82: e.data chega None, rastreamos manualmente

    def refresh_lista():
        lista_col.controls.clear()

        for i, p in enumerate(state["paradas"]):
            def make_item(idx, parada):
                def on_drag_start(e):
                    drag_src["idx"] = idx

                def on_accept(e):
                    src_idx = drag_src["idx"]
                    if src_idx is None or src_idx == idx:
                        drag_src["idx"] = None
                        return
                    drag_src["idx"] = None
                    lst = state["paradas"]
                    lst.insert(idx, lst.pop(src_idx))
                    refresh_lista()
                    set_recalc_needed(True)
                    set_status("Ordem alterada manualmente. Clique em Recalcular Rota.", ORANGE, 0)
                    lq.put("Parada reordenada. Recalcule a rota para atualizar mapa e metricas.\n")

                row_content = ft.Row([
                    ft.Icon(ic("DRAG_INDICATOR"), size=18, color=TEXT_MUTED),
                    ft.Container(width=6),
                    ft.Container(
                        content=ft.Text(str(idx + 1), size=11,
                                        weight=ft.FontWeight.BOLD, color=WHITE),
                        width=26, height=26, bgcolor=ORANGE,
                        border_radius=13, alignment=ft.Alignment(0, 0),
                    ),
                    ft.Container(width=8),
                    ft.Column([
                        ft.Row([
                            ft.Text(f"Parada #{idx + 1} - {parada['id']}", 
                                   size=13, color=TEXT_MAIN, weight=ft.FontWeight.W_600),
                            # Indicador de coordenada aproximada
                            ft.Container(
                                content=ft.Text("GPS APROX.", size=8, color=WHITE, weight=ft.FontWeight.BOLD),
                                bgcolor=ORANGE,
                                padding=ft.Padding(4, 2, 4, 2),
                                border_radius=3,
                            ) if parada.get("coordenada_aproximada") else ft.Container(width=0, height=0),
                        ], spacing=8),
                        ft.Text(
                            f"({parada['coord'][0]:.5f}, {parada['coord'][1]:.5f})" if parada['coord'] else "Sem coordenada",
                            size=9, color=TEXT_MUTED),
                    ], spacing=2, expand=True),
                ], vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=0)

                draggable = ft.Draggable(
                    group="parada",
                    data=str(idx),
                    on_drag_start=on_drag_start,
                    content=ft.Container(
                        content=row_content,
                        padding=ft.Padding.symmetric(horizontal=12, vertical=9),
                        border=ft.Border.only(bottom=ft.BorderSide(1, BORDER)),
                        bgcolor=WHITE,
                    ),
                    content_when_dragging=ft.Container(
                        content=row_content,
                        padding=ft.Padding.symmetric(horizontal=12, vertical=9),
                        bgcolor=SURFACE,
                        border=ft.Border.all(1.5, ORANGE),
                        border_radius=8,
                        opacity=0.55,
                    ),
                )

                return ft.DragTarget(
                    group="parada",
                    content=draggable,
                    on_accept=on_accept,
                )

            lista_col.controls.append(make_item(i, p))

        if not state["paradas"]:
            lista_col.controls.append(
                ft.Container(
                    content=ft.Text("Nenhuma parada carregada.",
                                    size=13, color=TEXT_MUTED, italic=True),
                    padding=20,
                )
            )
        try:
            lista_col.update()
        except Exception:
            pass

    refresh_lista()

    # ── Lista de OSs sem coordenada ───────────────────────────────────────────
    lista_sem_coord_col = ft.Column(spacing=0, scroll=ft.ScrollMode.AUTO)

    def refresh_lista_sem_coord():
        lista_sem_coord_col.controls.clear()
        sem_coord = state.get("paradas_sem_coord", [])

        if not sem_coord:
            lista_sem_coord_col.controls.append(
                ft.Container(
                    content=ft.Row([
                        ft.Icon(ic("CHECK_CIRCLE"), color=GREEN, size=15),
                        ft.Container(width=8),
                        ft.Text("Todas as OSs possuem coordenadas!", size=12, color=GREEN),
                    ], vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    padding=ft.Padding.symmetric(horizontal=14, vertical=12),
                )
            )
        else:
            for p in sem_coord:
                endereco_display = p.get("endereco", "").strip() or "Endereço não disponível"
                logradouro = p.get("logradouro", "")
                nr_imovel  = p.get("nr_imovel", "")
                bairro     = p.get("bairro", "")

                linha_end  = ", ".join(x for x in [logradouro, nr_imovel] if x and x.lower() != "nan")
                linha_bairro = bairro if bairro and bairro.lower() != "nan" else ""

                lista_sem_coord_col.controls.append(
                    ft.Container(
                        content=ft.Row([
                            ft.Container(
                                content=ft.Icon(ic("LOCATION_OFF"), color=ORANGE, size=14),
                                width=28, height=28, bgcolor="#FFF3E8",
                                border_radius=6, alignment=ft.Alignment(0, 0),
                            ),
                            ft.Container(width=10),
                            ft.Column([
                                ft.Text(
                                    f"OS #{p['id']}",
                                    size=12, color=TEXT_MAIN, weight=ft.FontWeight.W_600,
                                ),
                                ft.Text(
                                    linha_end if linha_end else "Logradouro não informado",
                                    size=11, color=TEXT_MUTED,
                                    overflow=ft.TextOverflow.ELLIPSIS,
                                ),
                                ft.Text(
                                    linha_bairro,
                                    size=10, color=ft.Colors.with_opacity(0.5, TEXT_MUTED),
                                    overflow=ft.TextOverflow.ELLIPSIS,
                                ) if linha_bairro else ft.Container(height=0),
                            ], spacing=1, expand=True),
                        ], vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=0),
                        padding=ft.Padding.symmetric(horizontal=12, vertical=8),
                        border=ft.Border.only(bottom=ft.BorderSide(1, BORDER)),
                        bgcolor=WHITE,
                    )
                )

        try:
            lista_sem_coord_col.update()
        except Exception:
            pass

    refresh_lista_sem_coord()


    def calcular_e_atualizar(silencioso=False):
        if not state["paradas"] or state["running"]:
            return

        # Filtra apenas paradas com coordenadas válidas
        paradas_com_coord = [p for p in state["paradas"] if p.get("coord") is not None]
        
        if not paradas_com_coord:
            if not silencioso:
                set_status("Aviso Nenhuma parada com coordenada válida. Geocodifique as OSs pendentes.", ORANGE, 0)
                lq.put("Nenhuma parada com coordenada válida.\n")
            return

        state["running"] = True
        set_status("Calculando a melhor rota...", ORANGE, None)
        lq.put("Traçando rota inteligente e atualizando o mapa...\n")

        try:
            import roteiricacao as rot
            import folium
            from datetime import datetime

            coords = [SEDE["coord"]] + [p["coord"] for p in paradas_com_coord]
            geometria, tempo, dist = rot.RouteEngine.get_full_road_route(tuple(coords))

            refresh_metricas(
                n=len(paradas_com_coord),
                dist_km=dist / 1000,
                min_val=int(tempo / 60),
            )

            m = folium.Map(tiles="Esri.WorldImagery")
            if coords:
                m.fit_bounds(coords)

            folium.Marker(
                location=SEDE["coord"], tooltip="SEDE",
                popup="Ponto de Partida",
                icon=folium.Icon(color="black", icon="home"),
            ).add_to(m)

            cores = rot.gerar_cores_gradiente(max(5, len(paradas_com_coord)))
            for i, p in enumerate(paradas_com_coord):
                aproximada   = p.get("coordenada_aproximada", False)
                endereco_geo = p.get("endereco_geocodificado", "")
                tipo_geo     = p.get("tipo_geocodificacao", "")
                logradouro   = p.get("logradouro", "")
                nr_imovel    = p.get("nr_imovel", "")
                bairro       = p.get("bairro", "")
                hidrometro   = p.get("hidrometro", "")
                num_os       = p.get("num_os", "")
                tipo_servico = p.get("tipo_servico", "")
                status_os    = p.get("status_os", "")
                prazo_os     = p.get("prazo_os", "")

                # Monta endereço legível: remove código numérico do início do logradouro
                import re as _re
                m_rua = _re.match(r'^\d+\s+(.+)$', logradouro)
                nome_rua = m_rua.group(1) if m_rua else logradouro
                end_display = f"{nome_rua}, {nr_imovel}" if nr_imovel else nome_rua
                if bairro:
                    end_display += f"  {bairro}"

                # Tooltip curto (aparece no hover)
                if aproximada:
                    tooltip_txt = f"Aviso #{i+1} | {p['id']} | {end_display} | GPS APROX."
                else:
                    tooltip_txt = f"#{i+1} | {p['id']} | {end_display}"

                # Popup completo (aparece no clique)
                def _row(label, value):
                    if not value:
                        return ""
                    return (f"<tr><td style='color:#888;padding-right:8px;white-space:nowrap'>"
                            f"<b>{label}</b></td>"
                            f"<td>{value}</td></tr>")

                aviso_aprox = ""
                if aproximada:
                    aviso_aprox = (
                        "<tr><td colspan='2' style='color:red;font-weight:bold;"
                        "padding-top:6px'>Aviso Coordenada aproximada</td></tr>"
                        f"<tr><td colspan='2' style='color:#B91C1C;font-size:11px;font-weight:700'>"
                        f"{tipo_geo or 'LOCALIZACAO APROXIMADA'}</td></tr>"
                        f"<tr><td colspan='2' style='color:#888;font-size:11px'>"
                        f"Localizada via busca: {endereco_geo}</td></tr>"
                    )

                lat_dest, lng_dest = p["coord"]
                google_route_url = (
                    "https://www.google.com/maps/dir/?api=1"
                    f"&destination={lat_dest:.6f},{lng_dest:.6f}"
                    "&travelmode=driving"
                )
                rota_cta = (
                    "<div style='margin-top:12px;padding-top:10px;border-top:1px solid #E5E7EB;'>"
                    "<div style='font-size:11px;color:#6B7280;margin-bottom:8px'>"
                    "Abrir navegacao externa para esta parada"
                    "</div>"
                    f"<a href='{google_route_url}' target='_blank' rel='noopener noreferrer' "
                    "style='display:inline-block;padding:9px 12px;border-radius:8px;"
                    "background:#1565C0;color:#FFFFFF;text-decoration:none;font-weight:700;"
                    "font-size:12px;box-shadow:0 1px 4px rgba(0,0,0,0.18)'>"
                    "Tracar rota no Google Maps"
                    "</a>"
                    "</div>"
                )

                popup_html = (
                    f"<div style='font-family:sans-serif;font-size:13px;min-width:220px'>"
                    f"<b style='font-size:14px'>Parada #{i+1}</b><br>"
                    f"<table style='margin-top:6px;border-collapse:collapse'>"
                    + _row("Matrícula",    p['id'])
                    + _row("OS nº",        num_os)
                    + _row("Serviço",      tipo_servico)
                    + _row("Status",       status_os)
                    + _row("Prazo",        prazo_os)
                    + _row("Endereço",     end_display)
                    + _row("Hidrômetro",   hidrometro)
                    + aviso_aprox
                    + "</table>"
                    + rota_cta
                    + "</div>"
                )

                if aproximada:
                    marker_color = "red"
                    marker_icon  = "exclamation-triangle"
                else:
                    marker_color = "blue"
                    marker_icon  = "circle"

                folium.Marker(
                    location=p["coord"],
                    tooltip=tooltip_txt,
                    popup=folium.Popup(popup_html, max_width=320),
                    icon=folium.Icon(color=marker_color, icon_color="white",
                                     prefix="fa", icon=marker_icon),
                ).add_to(m)

            if geometria and len(geometria) > 1:
                n_seg   = len(cores)
                seg_tam = max(1, len(geometria) // n_seg)
                for idx, cor in enumerate(cores):
                    ini = idx * seg_tam
                    fim = (idx + 1) * seg_tam if idx < n_seg - 1 else len(geometria)
                    if fim - ini > 1:
                        folium.PolyLine(geometria[ini:fim],
                                        weight=4, opacity=0.85, color=cor).add_to(m)

            # Salva sempre no mesmo arquivo (mapa_atual.html)
            mapas_dir = os.path.join(state["app_dir"] or ".", "Mapas")
            os.makedirs(mapas_dir, exist_ok=True)
            map_path = os.path.join(mapas_dir, "mapa_atual.html")
            m.save(map_path)
            state["map_file"] = map_path

            set_status(
                f"OK: {len(state['paradas'])} entregas | "
                f"{dist / 1000:.1f} km | {int(tempo / 60)} min",
                GREEN, 1)
            set_recalc_needed(False)
            lq.put(f"OK Mapa atualizado com sucesso!\n")
            lq.put(f"Arquivo: {map_path}\n")
            lq.put(f"Distância: {dist/1000:.1f} km\n")
            lq.put(f"Tempo deslocamento: {int(tempo/60)} min ({int(tempo/3600)}h {int((tempo%3600)/60)}min)\n")

        except Exception as ex:
            import traceback
            tb = traceback.format_exc()
            set_status(f"ERRO Erro: {ex}", RED, 0)
            lq.put(f"ERRO ERRO ao gerar mapa: {ex}\n")
            lq.put(f"{tb}\n")
        finally:
            state["running"] = False

    # Seletor de arquivo
    def pick_xlsx(e):
        def do_pick():
            try:
                import tkinter as tk
                from tkinter import filedialog
                root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
                path = filedialog.askopenfilename(
                    filetypes=[("Excel / CSV", "*.xlsx *.xls *.csv"), ("Todos", "*.*")]
                )
                root.destroy()
                if not path:
                    return

                lbl_arquivo.value = os.path.basename(path)
                lbl_arquivo.color = GREEN
                try:
                    lbl_arquivo.update()
                except Exception:
                    pass

                set_status("Carregando arquivo...", ORANGE, None)
                lq.put(f"Carregando: {os.path.basename(path)}\n")

                try:
                    import pandas as pd
                    import roteiricacao as rot

                    if path.endswith(".csv"):
                        # Detecta separador automaticamente (vírgula ou ponto e vírgula)
                        with open(path, encoding="utf-8-sig", errors="ignore") as _f:
                            _primeira = _f.readline()
                        _sep = ";" if _primeira.count(";") > _primeira.count(",") else ","
                        try:
                            df = pd.read_csv(path, sep=_sep, encoding="utf-8-sig", encoding_errors="replace")
                        except TypeError:
                            df = pd.read_csv(path, sep=_sep, encoding="utf-8-sig")
                    else:
                        df = pd.read_excel(path)
                    df.columns = [str(c).strip().upper() for c in df.columns]

                    paradas = []
                    paradas_com_coord = []
                    paradas_sem_coord = []

                    for idx, row in df.iterrows():
                        id_val = str(row.get(
                            "MATRICULA",
                            row.get("MATRICULA ", f"P{idx + 1}")))
                        if id_val.strip().upper() == "SEDE":
                            continue

                        lat_raw = row.get("LATITUDE", row.get("LATITUDE "))
                        lon_raw = row.get("LONGITUDE", row.get("LONGITUDE "))
                        lat = rot.DataMaster.fix_coord(lat_raw)
                        lon = rot.DataMaster.fix_coord(lon_raw)

                        def _str(v):
                            s = str(v).strip() if v is not None else ""
                            return "" if s.lower() == "nan" else s

                        logradouro   = _str(row.get("LOGRADOURO",   row.get("ENDEREÇO", "")))
                        nr_imovel    = _str(row.get("NR_IMOVEL",    row.get("NUMERO", "")))
                        bairro       = _str(row.get("BAIRRO", ""))
                        hidrometro   = _str(row.get("HIDROMETRO", ""))
                        num_os       = _str(row.get("NUM_OS", ""))
                        tipo_servico = _str(row.get("TIPO_SERVICO", ""))
                        status_os    = _str(row.get("STATUS_OS", ""))
                        prazo_os     = _str(row.get("PRAZO_OS", ""))

                        partes_end = [p for p in [logradouro, nr_imovel] if p]
                        endereco_completo = ", ".join(partes_end)
                        if bairro:
                            endereco_completo += f", {bairro}"

                        if lat is not None and lon is not None and -90 <= lat <= 90 and -180 <= lon <= 180:
                            parada_info = {
                                "id":           id_val,
                                "coord":        (lat, lon),
                                "coordenada_aproximada": False,
                                "logradouro":   logradouro,
                                "nr_imovel":    nr_imovel,
                                "bairro":       bairro,
                                "hidrometro":   hidrometro,
                                "num_os":       num_os,
                                "tipo_servico": tipo_servico,
                                "status_os":    status_os,
                                "prazo_os":     prazo_os,
                                "endereco":     endereco_completo,
                            }
                            paradas.append(parada_info)
                            paradas_com_coord.append(parada_info)
                        else:
                            parada_info = {
                                "id":           id_val,
                                "coord":        None,
                                "coordenada_aproximada": False,
                                "logradouro":   logradouro,
                                "nr_imovel":    nr_imovel,
                                "bairro":       bairro,
                                "hidrometro":   hidrometro,
                                "num_os":       num_os,
                                "tipo_servico": tipo_servico,
                                "status_os":    status_os,
                                "prazo_os":     prazo_os,
                                "endereco":     endereco_completo,
                            }
                            paradas.append(parada_info)
                            paradas_sem_coord.append(parada_info)

                    if not paradas:
                        set_status("Nenhum dado encontrado no arquivo.", RED, 0)
                        lq.put("Nenhum dado encontrado no arquivo.\n")
                        return

                    state["paradas"]           = paradas
                    state["paradas_com_coord"] = paradas_com_coord
                    state["paradas_sem_coord"] = paradas_sem_coord
                    state["map_file"]          = None
                    state["app_dir"]           = os.path.dirname(path)
                    set_recalc_needed(False)

                    refresh_lista()
                    refresh_metricas(n=len(paradas_com_coord))
                    atualizar_contador_sem_coord()
                    refresh_lista_sem_coord()

                    msg_status = f"{len(paradas_com_coord)} paradas carregadas"
                    if paradas_sem_coord:
                        msg_status += f" + {len(paradas_sem_coord)} sem coordenada"
                    msg_status += ". Calculando rota..."
                    set_status(msg_status, GREEN, 0)
                    lq.put(f"{len(paradas_com_coord)} paradas com coordenada.\n")
                    if paradas_sem_coord:
                        lq.put(f"{len(paradas_sem_coord)} OS(s) sem coordenada  clique 'Adicionar Localização' para geocodificar!\n")

                    # Contabiliza paradas no dashboard
                    _v = increment_stat("paradas", len(paradas_com_coord))
                    if ref_paradas and ref_paradas[0]:
                        try: ref_paradas[0].value = str(_v); ref_paradas[0].update()
                        except: pass

                    # Calcula automaticamente apos carregar (só usa as com coordenada)
                    if paradas_com_coord:
                        calcular_e_atualizar()
                    else:
                        set_status(
                            f"Aviso Nenhuma OS com coordenada. Clique 'Adicionar Localização' para geocodificar {len(paradas_sem_coord)} OS(s).",
                            ORANGE, 0)

                except ImportError:
                    set_status("roteiricacao.py nao encontrado.", RED, 0)
                    lq.put("ERRO: roteiricacao.py nao encontrado.\n")
                except Exception as ex:
                    import traceback
                    set_status(f"Erro ao carregar: {ex}", RED, 0)
                    lq.put(f"ERRO interno: {ex}\n")
                    lq.put(f"{traceback.format_exc()}\n")

            except Exception as ex:
                lq.put(f"Erro ao abrir seletor: {ex}\n")

        do_pick()

    def btn_calcular(e):
        if not state["paradas"]:
            page.snack_bar = ft.SnackBar(
                ft.Text("Carregue um arquivo primeiro!", color=WHITE),
                bgcolor=RED, open=True)
            page.update(); return
        
        if state["running"]:
            page.snack_bar = ft.SnackBar(
                ft.Text("Cálculo já em andamento  aguarde!", color=WHITE),
                bgcolor=ORANGE, open=True)
            page.update(); return
        
        threading.Thread(target=calcular_e_atualizar, daemon=True).start()

    def otimizar_rota(e):
        if not state["paradas"]:
            page.snack_bar = ft.SnackBar(
                ft.Text("Carregue um arquivo primeiro!", color=WHITE),
                bgcolor=RED, open=True)
            page.update(); return
        
        if state["running"]:
            page.snack_bar = ft.SnackBar(
                ft.Text("Operação em andamento  aguarde!", color=WHITE),
                bgcolor=ORANGE, open=True)
            page.update(); return

        def worker():
            try:
                set_status("Otimizando sequência de entregas...", ORANGE, None)
                lq.put("Otimizando rota...\n")
                
                import roteiricacao as rot
                ordem = rot.RouteEngine.auto_optimize([SEDE] + state["paradas"])
                state["paradas"] = [state["paradas"][i - 1]
                                    for i in ordem if i != 0]
                refresh_lista()
                lq.put("OK Sequência otimizada! Recalculando rota...\n")
                calcular_e_atualizar(silencioso=True)
            except Exception as ex:
                set_status(f"Erro ao otimizar: {ex}", RED, 0)
                lq.put(f"ERRO Erro: {ex}\n")

        threading.Thread(target=worker, daemon=True).start()

    def reverter_rota(e):
        if not state["paradas"]:
            page.snack_bar = ft.SnackBar(
                ft.Text("Carregue um arquivo primeiro!", color=WHITE),
                bgcolor=RED, open=True)
            page.update(); return
        
        if state["running"]:
            page.snack_bar = ft.SnackBar(
                ft.Text("Operação em andamento  aguarde!", color=WHITE),
                bgcolor=ORANGE, open=True)
            page.update(); return
        
        state["paradas"] = state["paradas"][::-1]
        refresh_lista()
        lq.put("OK Sequência revertida. Recalculando...\n")
        threading.Thread(target=lambda: calcular_e_atualizar(silencioso=True),
                         daemon=True).start()

    def exportar_csv(e):
        if not state["paradas"]:
            page.snack_bar = ft.SnackBar(
                ft.Text("Carregue um arquivo primeiro!", color=WHITE),
                bgcolor=RED, open=True)
            page.update(); return

        def worker():
            set_status("Exportando CSV...", ORANGE, None)
            try:
                import pandas as pd
                from datetime import datetime
                csv_dir  = os.path.join(state["app_dir"] or ".", "CSV_Exportados")
                os.makedirs(csv_dir, exist_ok=True)
                ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
                csv_path = os.path.join(csv_dir, f"rota_{ts}.csv")
                if not state["paradas"]:
                    set_status("Aviso Nenhuma parada para exportar.", ORANGE, 0)
                    lq.put("Aviso Nenhuma parada para exportar CSV.\n")
                    page.snack_bar = ft.SnackBar(
                        ft.Text("Nenhuma parada para exportar!", color=WHITE),
                        bgcolor=ORANGE, open=True)
                    try: page.update()
                    except Exception: pass
                    return

                paradas_sem_coord = [p for p in state["paradas"] if p.get("coord") is None]

                def _coord_csv(v):
                    if v is None:
                        return ""
                    return f"{v:.8f}".replace(".", ",")

                rows = [{
                    "PARADA": 0,
                    "MATRICULA": "SEDE",
                    "NUM_OS": "",
                    "LOGRADOURO": "SEDE",
                    "NR_IMOVEL": "",
                    "BAIRRO": "",
                    "HIDROMETRO": "",
                    "TIPO_SERVICO": "",
                    "STATUS_OS": "",
                    "PRAZO_OS": "",
                    "ENDERECO": "Ponto de Partida",
                    "LATITUDE": _coord_csv(SEDE['coord'][0]),
                    "LONGITUDE": _coord_csv(SEDE['coord'][1]),
                    "STATUS_COORDENADA": "SEDE",
                }]
                for i, p in enumerate(state["paradas"]):
                    coord = p.get("coord")
                    if coord is None:
                        status_coord = "SEM_COORDENADA"
                    elif p.get("coordenada_aproximada"):
                        status_coord = "ADICIONADA"
                    else:
                        status_coord = "ORIGINAL"
                    rows.append({
                        "PARADA": i + 1,
                        "MATRICULA": p["id"],
                        "NUM_OS": p.get("num_os", ""),
                        "LOGRADOURO": p.get("logradouro", ""),
                        "NR_IMOVEL": p.get("nr_imovel", ""),
                        "BAIRRO": p.get("bairro", ""),
                        "HIDROMETRO": p.get("hidrometro", ""),
                        "TIPO_SERVICO": p.get("tipo_servico", ""),
                        "STATUS_OS": p.get("status_os", ""),
                        "PRAZO_OS": p.get("prazo_os", ""),
                        "ENDERECO": p.get("endereco", ""),
                        "LATITUDE": _coord_csv(coord[0]) if coord is not None else "",
                        "LONGITUDE": _coord_csv(coord[1]) if coord is not None else "",
                        "STATUS_COORDENADA": status_coord,
                    })
                pd.DataFrame(rows).to_csv(csv_path, index=False, sep=";", encoding="utf-8-sig")
                set_status("OK CSV exportado com sucesso!", GREEN, 1)
                lq.put(f"OK CSV salvo: {csv_path}\n")
                lq.put("Formato preparado para reimportação na aba de Logística após preenchimento manual de LATITUDE/LONGITUDE.\n")
                if paradas_sem_coord:
                    lq.put(f"Aviso {len(paradas_sem_coord)} OS(s) foram exportadas sem coordenada, com LATITUDE/LONGITUDE vazias.\n")
                page.snack_bar = ft.SnackBar(
                    ft.Text("CSV exportado com sucesso!", color=WHITE),
                    bgcolor=GREEN, open=True)
                try: page.update()
                except Exception: pass
            except Exception as ex:
                lq.put("ERRO Erro ao salvar CSV.\n")

        threading.Thread(target=worker, daemon=True).start()

    def gerar_google_maps(e):
        if not state["paradas"]:
            page.snack_bar = ft.SnackBar(
                ft.Text("Carregue um arquivo primeiro!", color=WHITE),
                bgcolor=RED, open=True)
            page.update(); return

        def worker():
            set_status("Gerando links Google Maps...", ORANGE, None)
            lq.put("Processando rotas para Google Maps...\n")
            try:
                from datetime import datetime

                paradas_validas = [p for p in state["paradas"] if p.get("coord") is not None]
                paradas_sem_coord = [p for p in state["paradas"] if p.get("coord") is None]

                if not paradas_validas:
                    set_status("Aviso Nenhuma parada com coordenada para gerar links.", ORANGE, 0)
                    lq.put("Aviso Nenhuma parada com coordenada para gerar links do Google Maps.\n")
                    page.snack_bar = ft.SnackBar(
                        ft.Text("Nenhuma parada com coordenada para gerar links!", color=WHITE),
                        bgcolor=ORANGE, open=True)
                    try: page.update()
                    except Exception: pass
                    return

                todas = [SEDE] + paradas_validas
                rotas, i = [], 0
                while i < len(todas):
                    fim = min(i + 13, len(todas))
                    rotas.append(todas[i:fim])
                    if fim >= len(todas): break
                    i = fim - 1
                log  = "\n" + "=" * 70 + "\n"
                log += f"ROTAS GOOGLE MAPS - {len(rotas)} rota(s)\n"
                log += f"{datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n"
                log += "=" * 70 + "\n\n"
                for num, rota in enumerate(rotas, 1):
                    waypts = "/".join(f"{p['coord'][0]},{p['coord'][1]}" for p in rota)
                    url    = f"https://www.google.com/maps/dir/{waypts}"
                    ids    = [p["id"] for p in rota]
                    log += f"ROTA {num} ({len(rota)} paradas)\n"
                    log += "-" * 70 + "\n"
                    log += f"Paradas: {' > '.join(ids)}\n"
                    log += f"Link: {url}\n\n"
                if paradas_sem_coord:
                    ids_ignoradas = ", ".join(str(p.get("id", "?")) for p in paradas_sem_coord[:20])
                    log += f"Aviso {len(paradas_sem_coord)} OS(s) sem coordenada foram ignoradas na geração dos links.\n"
                    if ids_ignoradas:
                        log += f"IDs ignorados: {ids_ignoradas}"
                        if len(paradas_sem_coord) > 20:
                            log += ", ..."
                        log += "\n\n"
                try:
                    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
                    txt = os.path.join(state["app_dir"] or ".", f"rotas_googlemaps_{ts}.txt")
                    write_text_utf8(txt, log)
                    log = f"OK Arquivo salvo: {txt}\n" + log
                except Exception as ex:
                    log += f"Aviso  Aviso: não foi possível salvar arquivo: {ex}\n"
                lq.put(log)
                set_status(f"OK {len(rotas)} rota(s) Google Maps gerada(s) - veja o terminal!", GREEN, 1)
                page.snack_bar = ft.SnackBar(
                    ft.Text(f"{len(rotas)} rota(s) gerada(s)!", color=WHITE),
                    bgcolor=GREEN, open=True)
                try: page.update()
                except: pass
            except Exception as ex:
                set_status(f"ERRO Erro ao gerar rotas: {ex}", RED, 0)
                lq.put(f"ERRO ERRO: {ex}\n")

        threading.Thread(target=worker, daemon=True).start()

    def abrir_mapa(e):
        if not state["map_file"] or not os.path.exists(state["map_file"]):
            page.snack_bar = ft.SnackBar(
                ft.Text("Gere um mapa primeiro!", color=WHITE),
                bgcolor=RED, open=True)
            page.update(); return
        
        import webbrowser
        try:
            webbrowser.open(f"file:///{state['map_file'].replace(chr(92), '/')}")
            lq.put(f"OK Abrindo mapa: {state['map_file']}\n")
            page.snack_bar = ft.SnackBar(
                ft.Text("Mapa aberto no navegador!", color=WHITE),
                bgcolor=GREEN, open=True)
            page.update()
        except Exception as ex:
            lq.put(f"ERRO Erro ao abrir mapa: {ex}\n")
            page.snack_bar = ft.SnackBar(
                ft.Text(f"Erro: {ex}", color=WHITE),
                bgcolor=RED, open=True)
            page.update()

    def geocodificar_oss_pendentes(e):
        """Geocodifica OSs sem coordenada usando Nominatim com contexto de Itapoá/SC."""
        paradas_sem_coord = state.get("paradas_sem_coord", [])

        if not paradas_sem_coord:
            page.snack_bar = ft.SnackBar(
                ft.Text("Nenhuma OS sem coordenada!", color=WHITE),
                bgcolor=ORANGE, open=True)
            page.update()
            return

        if state["running"]:
            page.snack_bar = ft.SnackBar(
                ft.Text("Operação em andamento!", color=WHITE),
                bgcolor=ORANGE, open=True)
            page.update()
            return

        state["running"] = True
        set_status(f"Geocodificando {len(paradas_sem_coord)} OSs...", ORANGE, None)

        def worker():
            # ── Contexto geográfico fixo: Itapoá / SC ────────────────────────
            # Bounding box generosa ao redor de Itapoá e arredores
            # (lat: -26.30 a -25.80 | lon: -48.80 a -48.40)
            CIDADE_SUFIXO = "Itapoá, Santa Catarina, Brasil"
            LAT_MIN, LAT_MAX = -26.30, -25.80
            LON_MIN, LON_MAX = -48.80, -48.40

            def _dentro_da_regiao(lat, lon):
                return LAT_MIN <= lat <= LAT_MAX and LON_MIN <= lon <= LON_MAX

            def _parse_logradouro(logradouro):
                """
                O campo LOGRADOURO vem no formato "{cod} {NOME_DA_RUA}",
                ex: "643 POTIPEMA", "760 FRANCISCO QUINTINO CORREIA", "S. FRANCISCO".
                Separa o código numérico inicial do nome real da rua.
                Retorna (nome_rua, codigo_numerico_ou_None).
                """
                import re
                if not logradouro:
                    return "", None
                m = re.match(r'^(\d+)\s+(.+)$', logradouro.strip())
                if m:
                    return m.group(2).strip(), m.group(1)
                return logradouro.strip(), None

            def _normalizar_bairro(bairro):
                """
                Bairros vêm no formato "ITAPEMA DO NORTE - BALN. BRASÍLIA".
                Extrai apenas o nome do balneário (após o ' - ') para busca mais limpa.
                """
                if not bairro:
                    return []
                partes = [p.strip() for p in bairro.split(" - ") if p.strip()]
                # Remove prefixos como "BALN." para ficar só o nome
                resultado = []
                for p in partes:
                    limpo = p.replace("BALN.", "").replace("BALN", "").strip()
                    if limpo:
                        resultado.append(limpo)
                return resultado

            def _geocodificar_endereco(logradouro, nr_imovel, bairro):
                """
                Tenta várias combinações de endereço + contexto de cidade,
                do mais específico ao mais genérico.
                O formato do LOGRADOURO é "{cod_rua} {NOME_RUA}"  extrai e
                testa com prefixos (Rua, Avenida, etc.) e com/sem número do imóvel.
                Valida que o resultado cai dentro da bounding box de Itapoá.
                """
                import time as _time

                # ── Instancia o geocoder ──────────────────────────────────────
                try:
                    from geopy.geocoders import Nominatim
                    from geopy.exc import GeocoderTimedOut, GeocoderUnavailable

                    geolocator = Nominatim(user_agent="Mathools_Itapoa_1.0", timeout=10)

                    def _buscar(query):
                        _time.sleep(1.1)
                        try:
                            loc = geolocator.geocode(query)
                            if loc:
                                return loc.latitude, loc.longitude
                        except (GeocoderTimedOut, GeocoderUnavailable):
                            _time.sleep(2)
                        except Exception:
                            pass
                        return None, None

                except ImportError:
                    # Fallback HTTP puro via urllib
                    import urllib.request, urllib.parse, json as _json

                    def _buscar(query):
                        _time.sleep(1.1)
                        try:
                            params = urllib.parse.urlencode({
                                "q": query, "format": "json",
                                "limit": 1, "countrycodes": "br",
                            })
                            url = f"https://nominatim.openstreetmap.org/search?{params}"
                            req = urllib.request.Request(
                                url, headers={"User-Agent": "Mathools/1.0"})
                            with urllib.request.urlopen(req, timeout=10) as resp:
                                data = _json.loads(resp.read())
                            if data:
                                return float(data[0]["lat"]), float(data[0]["lon"])
                        except Exception:
                            pass
                        return None, None

                # ── Prepara os componentes ────────────────────────────────────
                nome_rua, _ = _parse_logradouro(logradouro)

                # Prefixos de logradouro para tentar (Nominatim às vezes precisa do tipo)
                PREFIXOS = ["", "Rua ", "Avenida ", "Av. ", "Travessa ", "Alameda "]

                # Monta lista de tentativas apenas com combinações seguras.
                tentativas = []
                numero_valido = bool(nr_imovel and nr_imovel not in {"0", "S/N", "SN", "s/n", "sn"})

                if nome_rua and numero_valido:
                    for pref in PREFIXOS:
                        tentativas.append((
                            f"{pref}{nome_rua}, {nr_imovel}, {CIDADE_SUFIXO}",
                            "RUA + NUMERO"
                        ))
                    for pref in PREFIXOS:
                        tentativas.append((
                            f"{nr_imovel} {pref}{nome_rua}, {CIDADE_SUFIXO}",
                            "NUMERO + RUA"
                        ))

                # Sem número da casa  ainda na rua
                if nome_rua:
                    for pref in PREFIXOS:
                        tentativas.append((
                            f"{pref}{nome_rua}, {CIDADE_SUFIXO}",
                            "APENAS RUA ENCONTRADA"
                        ))

                # Remove duplicatas mantendo ordem
                seen = set()
                tentativas_unicas = []
                for query, rotulo in tentativas:
                    if query not in seen:
                        seen.add(query)
                        tentativas_unicas.append((query, rotulo))

                # ── Executa as tentativas ─────────────────────────────────────
                for query, rotulo in tentativas_unicas:
                    lat, lon = _buscar(query)
                    if lat is not None and _dentro_da_regiao(lat, lon):
                        return lat, lon, query, rotulo
                    elif lat is not None:
                        lq.put(f"   ↳ Descartado (fora da região): {query} → ({lat:.4f}, {lon:.4f})\n")

                return None, None, None, None

            try:
                sucesso = 0
                nao_encontrado = 0
                total = len(paradas_sem_coord)

                # Trabalha sobre cópia para evitar modificar a lista enquanto itera
                pendentes = list(paradas_sem_coord)

                for i, parada in enumerate(pendentes):
                    logradouro = parada.get("logradouro", "").strip()
                    nr_imovel  = parada.get("nr_imovel",  "").strip()
                    bairro     = parada.get("bairro",     "").strip()

                    # Ignora valores "nan" que vêm do pandas
                    if logradouro.lower() == "nan": logradouro = ""
                    if nr_imovel.lower()  == "nan": nr_imovel  = ""
                    if bairro.lower()     == "nan": bairro     = ""

                    desc = logradouro or bairro or parada["id"]
                    set_status(
                        f"[{i+1}/{total}] Buscando: {desc[:50]}...",
                        ORANGE, (i + 1) / total,
                    )
                    lq.put(f" [{i+1}/{total}] OS {parada['id']}  {desc}\n")

                    lat, lon, query_usada, tipo_geo = _geocodificar_endereco(logradouro, nr_imovel, bairro)

                    if lat is not None and lon is not None:
                        # Atualiza a parada no state
                        for p in state["paradas"]:
                            if p["id"] == parada["id"]:
                                p["coord"]                  = (lat, lon)
                                p["coordenada_aproximada"]  = True
                                p["endereco_geocodificado"] = query_usada or ""
                                p["tipo_geocodificacao"]    = tipo_geo or "LOCALIZACAO APROXIMADA"
                                if p in state["paradas_sem_coord"]:
                                    state["paradas_sem_coord"].remove(p)
                                    state["paradas_com_coord"].append(p)
                                break
                        lq.put(f"   OK ({lat:.5f}, {lon:.5f}) via: {query_usada} [{tipo_geo or 'LOCALIZACAO APROXIMADA'}]\n")
                        sucesso += 1
                    else:
                        lq.put(f"   Aviso Não encontrado dentro da região de Itapoá\n")
                        nao_encontrado += 1

                # Atualiza UI
                refresh_lista()
                atualizar_contador_sem_coord()
                refresh_lista_sem_coord()

                msg = f"OK Geocodificação concluída! {sucesso} encontradas"
                if nao_encontrado > 0:
                    msg += f", {nao_encontrado} fora da região ou sem endereço"
                if sucesso > 0:
                    msg += ". Clique em Recalcular Rota para atualizar o mapa."
                    set_recalc_needed(True)
                set_status(msg, GREEN if sucesso > 0 else ORANGE, 1)
                lq.put(msg + "\n")

                page.snack_bar = ft.SnackBar(
                    ft.Text(f"{sucesso} OSs geocodificadas!", color=WHITE),
                    bgcolor=GREEN if sucesso > 0 else ORANGE, open=True)
                page.update()

            except Exception as ex:
                import traceback
                set_status(f"ERRO Erro ao geocodificar: {ex}", RED, 0)
                lq.put(f"ERRO ERRO: {ex}\n")
                lq.put(f"{traceback.format_exc()}\n")
                page.snack_bar = ft.SnackBar(
                    ft.Text(f"Erro: {str(ex)[:50]}", color=WHITE),
                    bgcolor=RED, open=True)
                page.update()
            finally:
                state["running"] = False

        threading.Thread(target=worker, daemon=True).start()

    btn_recalcular = pbtn("Recalcular Rota", "MAP", on_click=btn_calcular, width=220)
    btn_recalcular.visible = False

    # Montagem da UI
    return ft.Column([

        ft.Text("Logistica & Rotas", size=22,
                weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
        ft.Text("Roteirização inteligente com mapeamento geográfico interativo.", 
                size=13, color=TEXT_MUTED),
        ft.Container(height=8),

        card(ft.Column([
            sec_title("Arquivo de Enderecos (XLSX/CSV)", "TABLE_CHART"),
            ft.Container(height=10),
            ft.Row([
                ft.Container(content=lbl_arquivo, expand=True),
                ft.Container(width=8),
                pbtn("Procurar", "FOLDER", on_click=pick_xlsx),
            ], vertical_alignment=ft.CrossAxisAlignment.CENTER),
            ft.Container(height=8),
            info_box("Selecione o arquivo com colunas LATITUDE, LONGITUDE e MATRICULA."),
        ])),

        card(ft.Column([
            sec_title("Metricas da Rota", "SPEED"),
            ft.Container(height=12),
            metrics_row,
        ])),

        card(ft.Column([
            sec_title("Status", "MONITOR"),
            ft.Container(height=8),
            lbl_status,
            ft.Container(height=6),
            ft.Row([
                ft.Icon(ic("LOCATION_OFF"), size=14, color=ORANGE),
                ft.Container(width=6),
                ft.Text("OSs sem coordenada: ", size=11, color=TEXT_MUTED, weight=ft.FontWeight.W_600),
                lbl_sem_coord_count,
            ], spacing=0),
            ft.Container(height=8),
            prog,
        ])),

        card(ft.Column([
            sec_title("OSs Sem Coordenada", "LOCATION_SEARCHING"),
            ft.Container(height=4),
            ft.Text(
                "Estas OSs não possuem LATITUDE/LONGITUDE no arquivo. "
                "O endereço (logradouro + nº) será usado para geocodificação aproximada ao clicar em 'Adicionar Localização'.",
                size=11, color=TEXT_MUTED, italic=True,
            ),
            ft.Container(height=10),
            ft.Container(
                content=lista_sem_coord_col,
                border=ft.Border.all(1, BORDER),
                border_radius=8,
                clip_behavior=ft.ClipBehavior.ANTI_ALIAS,
                height=200,
                width=550,
            ),
        ])),

        card(ft.Column([
            sec_title("Sequencia de Entregas", "DRAG_INDICATOR"),
            ft.Container(height=4),
            ft.Text("Arraste os itens para reordenar  metricas atualizam automaticamente.",
                    size=11, color=TEXT_MUTED, italic=True),
            ft.Container(height=10),
            ft.Container(
                content=lista_col,
                border=ft.Border.all(1, BORDER),
                border_radius=8,
                clip_behavior=ft.ClipBehavior.ANTI_ALIAS,
                height=260,
                width=550,
            ),
            ft.Container(height=12),
            ft.Row([
                ft.Container(
                    content=ft.Row([
                        ft.Icon(ic("AUTO_AWESOME"), size=16, color=WHITE),
                        ft.Container(width=6),
                        ft.Text("Otimizar", size=13, weight=ft.FontWeight.BOLD, color=WHITE),
                    ], spacing=0, alignment=ft.MainAxisAlignment.CENTER,
                       vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    height=44, bgcolor=NAVY, border_radius=8,
                    padding=ft.Padding.symmetric(horizontal=20, vertical=0),
                    on_click=otimizar_rota, ink=True,
                    tooltip="Otimiza a sequência de entregas com algoritmo inteligente",
                    shadow=ft.BoxShadow(spread_radius=0, blur_radius=6,
                                        color=ft.Colors.with_opacity(0.25, ft.Colors.BLACK),
                                        offset=ft.Offset(0, 2)),
                ),
                ft.Container(width=8),
                ft.Container(
                    content=ft.Row([
                        ft.Icon(ic("SWAP_VERT"), size=16, color=WHITE),
                        ft.Container(width=6),
                        ft.Text("Reverter", size=13, weight=ft.FontWeight.W_600, color=WHITE),
                    ], spacing=0, alignment=ft.MainAxisAlignment.CENTER,
                       vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    height=44, bgcolor=ORANGE, border_radius=8,
                    padding=ft.Padding.symmetric(horizontal=20, vertical=0),
                    on_click=reverter_rota, ink=True,
                    tooltip="Reverte a sequência de entregas",
                    shadow=ft.BoxShadow(spread_radius=0, blur_radius=6,
                                        color=ft.Colors.with_opacity(0.3, ft.Colors.BLACK),
                                        offset=ft.Offset(0, 2)),
                ),
            ]),
        ])),

        card(ft.Column([
            sec_title("Acoes", "BOLT"),
            ft.Container(height=12),
            ft.Row([
                btn_recalcular,
                ft.Container(width=10),
                ft.Container(
                    content=ft.Row([
                        ft.Icon(ic("DOWNLOAD"), size=16, color=NAVY),
                        ft.Container(width=6),
                        ft.Text("Exportar CSV", size=13, weight=ft.FontWeight.BOLD, color=NAVY),
                    ], spacing=0, alignment=ft.MainAxisAlignment.CENTER,
                       vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    height=44, bgcolor="#E8F0FF", border_radius=8,
                    padding=ft.Padding.symmetric(horizontal=20, vertical=0),
                    on_click=exportar_csv, 
                    tooltip="Exporta a rota otimizada em arquivo CSV",
                    ink=True,
                    shadow=ft.BoxShadow(spread_radius=0, blur_radius=6,
                                        color=ft.Colors.with_opacity(0.25, ft.Colors.BLACK),
                                        offset=ft.Offset(0, 2)),
                ),
                ft.Container(width=8),
                ft.Container(
                    content=ft.Row([
                        ft.Icon(ic("DIRECTIONS"), size=16, color=WHITE),
                        ft.Container(width=6),
                        ft.Text("Links Google Maps", size=13, weight=ft.FontWeight.BOLD, color=WHITE),
                    ], spacing=0, alignment=ft.MainAxisAlignment.CENTER,
                       vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    height=44, bgcolor="#16A34A", border_radius=8,
                    padding=ft.Padding.symmetric(horizontal=20, vertical=0),
                    on_click=gerar_google_maps, 
                    tooltip="Gera links do Google Maps para cada segmento de rota",
                    ink=True,
                    shadow=ft.BoxShadow(spread_radius=0, blur_radius=6,
                                        color=ft.Colors.with_opacity(0.25, ft.Colors.BLACK),
                                        offset=ft.Offset(0, 2)),
                ),
            ]),
            ft.Container(height=12),
            ft.Row([
                ft.Container(
                    content=ft.Row([
                        ft.Icon(ic("MAP"), size=16, color=WHITE),
                        ft.Container(width=6),
                        ft.Text("Abrir Mapa Interativo", size=13, weight=ft.FontWeight.BOLD, color=WHITE),
                    ], spacing=0, alignment=ft.MainAxisAlignment.CENTER,
                       vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    height=44, bgcolor="#E74C3C", border_radius=8,
                    padding=ft.Padding.symmetric(horizontal=20, vertical=0),
                    on_click=abrir_mapa, 
                    tooltip="Abre o mapa gerado em tempo real com satélite, marcadores de todas as paradas e a rota colorida no navegador. Atualiza cada vez que você otimiza a rota.",
                    ink=True,
                    shadow=ft.BoxShadow(spread_radius=0, blur_radius=6,
                                        color=ft.Colors.with_opacity(0.3, ft.Colors.BLACK),
                                        offset=ft.Offset(0, 2)),
                ),
                ft.Container(width=8),
                ft.Container(
                    content=ft.Row([
                        ft.Icon(ic("LOCATION_SEARCHING"), size=16, color=WHITE),
                        ft.Container(width=6),
                        ft.Text("Adicionar Localização", size=13, weight=ft.FontWeight.BOLD, color=WHITE),
                    ], spacing=0, alignment=ft.MainAxisAlignment.CENTER,
                       vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    height=44, bgcolor="#9333EA", border_radius=8,
                    padding=ft.Padding.symmetric(horizontal=20, vertical=0),
                    on_click=geocodificar_oss_pendentes,
                    tooltip="Geocodifica OSs sem coordenadas usando endereço (rua + número).\n\n"
                            "Tenta buscar coordenadas aproximadas via OpenStreetMap.\n"
                            "Aviso IMPORTANTE: As coordenadas são APROXIMADAS - podem levar apenas até a rua.\n"
                            "Use para OSs que não têm dados de GPS.",
                    ink=True,
                    shadow=ft.BoxShadow(spread_radius=0, blur_radius=6,
                                        color=ft.Colors.with_opacity(0.3, ft.Colors.BLACK),
                                        offset=ft.Offset(0, 2)),
                ),
            ]),
        ])),

        card(ft.Column([
            sec_title("O que cada botão faz", "HELP_OUTLINE"),
            ft.Container(height=14),
            ft.Column([
                # Procurar
                ft.Container(
                    content=ft.Row([
                        ft.Container(
                            content=ft.Icon(ic("FOLDER"), color=WHITE, size=15),
                            width=32, height=32, bgcolor=NAVY, border_radius=7,
                            alignment=ft.Alignment(0, 0),
                        ),
                        ft.Container(width=12),
                        ft.Column([
                            ft.Text("Procurar", size=12, weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
                            ft.Text("Abre o explorador de arquivos para você selecionar o XLSX ou CSV com as coordenadas de entrega.",
                                    size=11, color=TEXT_MUTED),
                        ], spacing=2, expand=True),
                    ], vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=0),
                    padding=ft.Padding.symmetric(horizontal=14, vertical=12),
                    bgcolor=SURFACE, border_radius=8,
                    border=ft.Border.all(1, BORDER),
                ),
                # Otimizar
                ft.Container(
                    content=ft.Row([
                        ft.Container(
                            content=ft.Icon(ic("AUTO_AWESOME"), color=WHITE, size=15),
                            width=32, height=32, bgcolor=NAVY, border_radius=7,
                            alignment=ft.Alignment(0, 0),
                        ),
                        ft.Container(width=12),
                        ft.Column([
                            ft.Text("Otimizar", size=12, weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
                            ft.Text("Reordena automaticamente as paradas para o menor caminho possível, economizando tempo e combustível.",
                                    size=11, color=TEXT_MUTED),
                        ], spacing=2, expand=True),
                    ], vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=0),
                    padding=ft.Padding.symmetric(horizontal=14, vertical=12),
                    bgcolor=SURFACE, border_radius=8,
                    border=ft.Border.all(1, BORDER),
                ),
                # Reverter
                ft.Container(
                    content=ft.Row([
                        ft.Container(
                            content=ft.Icon(ic("SWAP_VERT"), color=WHITE, size=15),
                            width=32, height=32, bgcolor=ORANGE, border_radius=7,
                            alignment=ft.Alignment(0, 0),
                        ),
                        ft.Container(width=12),
                        ft.Column([
                            ft.Text("Reverter", size=12, weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
                            ft.Text("Inverte a ordem atual das paradas  útil para planejar o retorno ou testar rotas ao contrário.",
                                    size=11, color=TEXT_MUTED),
                        ], spacing=2, expand=True),
                    ], vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=0),
                    padding=ft.Padding.symmetric(horizontal=14, vertical=12),
                    bgcolor=SURFACE, border_radius=8,
                    border=ft.Border.all(1, BORDER),
                ),
                # Recalcular Rota
                ft.Container(
                    content=ft.Row([
                        ft.Container(
                            content=ft.Icon(ic("MAP"), color=WHITE, size=15),
                            width=32, height=32, bgcolor=ORANGE, border_radius=7,
                            alignment=ft.Alignment(0, 0),
                        ),
                        ft.Container(width=12),
                        ft.Column([
                            ft.Text("Recalcular Rota", size=12, weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
                            ft.Text("Aparece quando voce muda a ordem manualmente. Recalcula a rota atual pelas estradas reais e atualiza mapa e metricas.",
                                    size=11, color=TEXT_MUTED),
                        ], spacing=2, expand=True),
                    ], vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=0),
                    padding=ft.Padding.symmetric(horizontal=14, vertical=12),
                    bgcolor=SURFACE, border_radius=8,
                    border=ft.Border.all(1, BORDER),
                ),
                # Exportar CSV
                ft.Container(
                    content=ft.Row([
                        ft.Container(
                            content=ft.Icon(ic("DOWNLOAD"), color=WHITE, size=15),
                            width=32, height=32, bgcolor=NAVY, border_radius=7,
                            alignment=ft.Alignment(0, 0),
                        ),
                        ft.Container(width=12),
                        ft.Column([
                            ft.Text("Exportar CSV", size=12, weight=ft.FontWeight.BOLD, color=WHITE),
                            ft.Text("Salva a sequência atual de entregas em um arquivo CSV na pasta CSV_Exportados/, com ordem, matrícula e coordenadas.",
                                    size=11, color=TEXT_MUTED),
                        ], spacing=2, expand=True),
                    ], vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=0),
                    padding=ft.Padding.symmetric(horizontal=14, vertical=12),
                    bgcolor=SURFACE, border_radius=8,
                    border=ft.Border.all(1, BORDER),
                ),
                # Links Google Maps
                ft.Container(
                    content=ft.Row([
                        ft.Container(
                            content=ft.Icon(ic("DIRECTIONS"), color=WHITE, size=15),
                            width=32, height=32, bgcolor=NAVY, border_radius=7,
                            alignment=ft.Alignment(0, 0),
                        ),
                        ft.Container(width=12),
                        ft.Column([
                            ft.Text("Links Google Maps", size=12, weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
                            ft.Text("Gera links prontos do Google Maps com até 13 paradas por rota e salva em arquivo .txt para compartilhar com o motorista.",
                                    size=11, color=TEXT_MUTED),
                        ], spacing=2, expand=True),
                    ], vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=0),
                    padding=ft.Padding.symmetric(horizontal=14, vertical=12),
                    bgcolor=SURFACE, border_radius=8,
                    border=ft.Border.all(1, BORDER),
                ),
                # Abrir Mapa Interativo
                ft.Container(
                    content=ft.Row([
                        ft.Container(
                            content=ft.Icon(ic("MAP"), color=WHITE, size=15),
                            width=32, height=32, bgcolor="#E74C3C", border_radius=7,
                            alignment=ft.Alignment(0, 0),
                        ),
                        ft.Container(width=12),
                        ft.Column([
                            ft.Text("Abrir Mapa Interativo", size=12, weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
                            ft.Text("Abre o mapa gerado em tempo real com satélite, marcadores de todas as paradas e a rota colorida no navegador. Atualiza cada vez que você otimiza a rota.",
                                    size=11, color=TEXT_MUTED),
                        ], spacing=2, expand=True),
                    ], vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=0),
                    padding=ft.Padding.symmetric(horizontal=14, vertical=12),
                    bgcolor=SURFACE, border_radius=8,
                    border=ft.Border.all(1, BORDER),
                ),
            ], spacing=8),
        ])),
    ], spacing=0, scroll=ft.ScrollMode.AUTO, expand=True)


# ─────────────────────────────────────────────────────────────────────────────
#  PÁGINA 5  GESTÃO DE USUÁRIOS
# ─────────────────────────────────────────────────────────────────────────────
def build_users_page(page: ft.Page):
    try:
        from auth_system import Authentication
        auth = Authentication()
    except ImportError:
        auth = None

    editing_state = {"user_id": None}

    e_user  = tfield("Nome de usuário corporativo", expand=False, width=250)
    e_pwd   = tfield("Senha de acesso", password=True, expand=False, width=250)
    dd_role = ft.Dropdown(
        value="Atendimento",
        options=[ft.dropdown.Option(r) for r in ["ADM","Atendimento","Call center","Administrativo"]],
        width=200, height=44, text_size=13, border_radius=8,
        filled=True, fill_color="#F8FAFD", border_color=BORDER,
        focused_border_color=ORANGE,
        content_padding=ft.Padding.symmetric(horizontal=14, vertical=10),
    )
    mod_cbs = {
        n: ft.Checkbox(value=True, label=n, active_color=ORANGE,
                       label_style=ft.TextStyle(size=12, color=TEXT_MAIN))
        for n in ["Análise de CSV","Disparo WhatsApp","Gestão de Negociação","Logística e Rotas"]
    }

    btn_add   = [None]  # mutable ref
    table_col = ft.Column(spacing=0)
    form_title = ft.Text("Configurar Perfil do Operador", size=15,
                         weight=ft.FontWeight.BOLD, color=TEXT_MAIN)

    def get_perms():
        if dd_role.value == "ADM":
            return "Todos"
        return ",".join([m for m, cb in mod_cbs.items() if cb.value])

    def snack(msg, color=GREEN):
        page.snack_bar = ft.SnackBar(ft.Text(msg, color=WHITE), bgcolor=color, open=True)
        page.update()

    def reset_form():
        editing_state["user_id"] = None
        e_user.value = ""; e_pwd.value = ""
        dd_role.value = "Atendimento"
        for cb in mod_cbs.values(): cb.value = True
        form_title.value = "Configurar Perfil do Operador"
        if btn_add[0]: btn_add[0].content.controls[-1].value = "Adicionar Usuário"
        e_user.update(); e_pwd.update(); dd_role.update()
        for cb in mod_cbs.values(): cb.update()
        form_title.update()
        if btn_add[0]: btn_add[0].update()

    def sync_table():
        table_col.controls.clear()
        users = auth.get_all_users() if auth else []
        hdr = ft.Container(
            content=ft.Row([
                ft.Text("ID",       size=11, weight=ft.FontWeight.BOLD, color=TEXT_MUTED, width=40),
                ft.Text("USUÁRIO",  size=11, weight=ft.FontWeight.BOLD, color=TEXT_MUTED, expand=1),
                ft.Text("CARGO",    size=11, weight=ft.FontWeight.BOLD, color=TEXT_MUTED, width=140),
                ft.Text("MÓDULOS",  size=11, weight=ft.FontWeight.BOLD, color=TEXT_MUTED, expand=2),
                ft.Text("CADASTRO", size=11, weight=ft.FontWeight.BOLD, color=TEXT_MUTED, width=90),
                ft.Text("AÇÕES",    size=11, weight=ft.FontWeight.BOLD, color=TEXT_MUTED, width=80),
            ]),
            bgcolor=SURFACE,
            padding=ft.Padding.symmetric(horizontal=16, vertical=10),
            border_radius=ft.BorderRadius.only(top_left=8, top_right=8),
            border=ft.Border.only(bottom=ft.BorderSide(1, BORDER)),
        )
        table_col.controls.append(hdr)
        for u in users:
            uid      = u.get("id", "")
            uname    = u.get("username", "")
            urole    = u.get("role", "")
            umods    = u.get("allowed_modules", "")
            ucreated = str(u.get("created_at", ""))[:10]
            table_col.controls.append(ft.Container(
                content=ft.Row([
                    ft.Text(str(uid), size=13, color=TEXT_MUTED, width=40),
                    ft.Row([
                        ft.Icon(ic("PERSON"), size=15, color=NAVY),
                        ft.Text(uname, size=13, color=TEXT_MAIN, weight=ft.FontWeight.W_500),
                    ], spacing=6, expand=1),
                    ft.Container(content=role_badge(urole), width=140),
                    ft.Text(umods, size=12, color=TEXT_MUTED, expand=2),
                    ft.Text(ucreated, size=12, color=TEXT_MUTED, width=90),
                    ft.Row([
                        ft.Container(content=ft.Icon(ic("EDIT"),   size=15, color=NAVY),
                                     width=34, height=34, border_radius=6, ink=True,
                                     alignment=ft.Alignment(0,0),
                                     on_click=lambda e, u=u: start_edit(u)),
                        ft.Container(content=ft.Icon(ic("DELETE"), size=15, color=RED),
                                     width=34, height=34, border_radius=6, ink=True,
                                     alignment=ft.Alignment(0,0),
                                     on_click=lambda e, uid=uid, un=uname: confirm_delete(uid, un)),
                    ], spacing=0, width=80),
                ], vertical_alignment=ft.CrossAxisAlignment.CENTER),
                padding=ft.Padding.symmetric(horizontal=16, vertical=12),
                border=ft.Border.only(bottom=ft.BorderSide(1, BORDER)),
                bgcolor=WHITE,
            ))
        try: table_col.update()
        except: pass

    def start_edit(u):
        editing_state["user_id"] = u.get("id")
        e_user.value = u.get("username", "")
        e_pwd.value  = ""
        dd_role.value = u.get("role", "Atendimento")
        perms = u.get("allowed_modules", "")
        for m, cb in mod_cbs.items():
            cb.value = (perms == "Todos" or m in perms)
        form_title.value = f"Editando: {u.get('username')}"
        if btn_add[0]: btn_add[0].content.controls[-1].value = "Salvar Alterações"
        e_user.update(); e_pwd.update(); dd_role.update()
        for cb in mod_cbs.values(): cb.update()
        form_title.update()
        if btn_add[0]: btn_add[0].update()

    def confirm_delete(uid, uname):
        def do_delete(e):
            dlg.open = False; page.update()
            if auth:
                ok = auth.delete_user(uid)
                if ok:
                    snack(f"Acesso de '{uname}' removido.")
                    sync_table()
                else:
                    snack("Erro ao remover usuário.", RED)

        dlg = ft.AlertDialog(
            modal=True,
            bgcolor=WHITE,
            title=ft.Text(f"Remover '{uname}'?", weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
            content=ft.Text("Esta ação não pode ser desfeita.", color=TEXT_MUTED),
            actions=[
                ft.Container(
                    content=ft.Row([
                        ft.Icon(ic("CLOSE"), size=14, color=WHITE),
                        ft.Container(width=6),
                        ft.Text("Cancelar", size=13, color=WHITE, weight=ft.FontWeight.W_500),
                    ], spacing=0, alignment=ft.MainAxisAlignment.CENTER,
                       vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    height=44, bgcolor=NAVY, border_radius=8,
                    padding=ft.Padding.symmetric(horizontal=16, vertical=0),
                    ink=True,
                    on_click=lambda e: [setattr(dlg, "open", False), page.update()],
                ),
                pbtn("Remover", "DELETE", on_click=do_delete),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        page.dialog = dlg; dlg.open = True; page.update()

    def submit(e):
        uname = e_user.value.strip()
        upwd  = e_pwd.value
        urole = dd_role.value or "Atendimento"
        uperms = get_perms()

        if not uname:
            snack("Nome de usuário obrigatório.", RED); return

        if not auth:
            snack("auth_system não disponível.", RED); return

        if editing_state["user_id"] is None:
            if not upwd:
                snack("Defina uma senha de acesso.", RED); return
            ok, msg = auth.add_user(uname, upwd, urole, uperms)
        else:
            ok, msg = auth.update_user(editing_state["user_id"], uname, upwd, urole, uperms)

        if ok:
            snack(f"{msg}")
            reset_form()
            sync_table()
        else:
            snack(f"Erro: {msg}", RED)

    def clear_form(e):
        reset_form()
        e_user.update(); e_pwd.update()

    add_btn = pbtn("Adicionar Usuário", "PERSON_ADD", on_click=submit, width=200)
    btn_add[0] = add_btn

    sync_table()

    return ft.Column([
        ft.Text("Gestão de Acesso", size=22, weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
        ft.Text("Controle de usuários, cargos e permissões por módulo (RBAC).", size=13, color=TEXT_MUTED),
        ft.Container(height=8),
        card(ft.Column([
            ft.Row([ft.Icon(ic("MANAGE_ACCOUNTS"), color=ORANGE, size=20), form_title], spacing=10),
            ft.Container(height=14),
            ft.Row([
                ft.Column([flabel("Nome de Usuário"), ft.Container(height=4), e_user],  spacing=0),
                ft.Container(width=14),
                ft.Column([flabel("Senha"),           ft.Container(height=4), e_pwd],   spacing=0),
                ft.Container(width=14),
                ft.Column([flabel("Cargo"),           ft.Container(height=4), dd_role], spacing=0),
            ], wrap=True, spacing=0),
            ft.Container(height=14),
            flabel("Módulos Habilitados:"),
            ft.Container(height=8),
            ft.Container(
                content=ft.Row(list(mod_cbs.values()), wrap=True, spacing=16),
                bgcolor=SURFACE, border_radius=8,
                padding=ft.Padding.symmetric(horizontal=16, vertical=12),
            ),
            ft.Container(height=14),
            ft.Row([
                add_btn,
                ft.Container(width=10),
                sbtn("Limpar", "CLEAR", on_click=clear_form),
            ]),
        ])),
        card(ft.Column([
            sec_title("Operadores Vinculados", "GROUP"),
            ft.Container(height=12),
            ft.Container(
                content=table_col,
                border=ft.Border.all(1, BORDER),
                border_radius=8,
                clip_behavior=ft.ClipBehavior.ANTI_ALIAS,
            ),
        ])),
    ], spacing=0, scroll=ft.ScrollMode.AUTO, expand=True)


# ─────────────────────────────────────────────────────────────────────────────
#  PÁGINA 6  AJUDA
# ─────────────────────────────────────────────────────────────────────────────
def build_help_page(user=None):
    MODULES = [
        {
            "icon": "TABLE_CHART",
            "access_module": "Análise de CSV",
            "title": "Processar CSV",
            "summary": "Tratamento da Macro 8104 com cruzamento opcional da 8162 e geracao da planilha final.",
            "purpose": [
                "Ajuda a separar clientes por faixa de vencimento para facilitar comunicacoes diferentes conforme os dias de atraso.",
                "Permite trabalhar, por exemplo, com uma abordagem para clientes acima de X dias e outra para clientes abaixo ou igual a X dias.",
                "Quando a Macro 8162 e usada junto, o sistema cruza a matricula com os dados cadastrais e traz informacoes de endereco para concentrar tudo em uma unica planilha.",
            ],
            "steps": [
                "Selecione a Macro 8104 em 'Base Principal'.",
                "Se desejar enriquecer os dados, selecione tambem a Macro 8162 em 'Cruzamento'.",
                "Defina a quantidade minima de dias em atraso no controle lateral.",
                "Clique em 'Executar Processamento' e acompanhe o log no console.",
                "Ao final, valide o resumo do processamento e a planilha Excel gerada.",
            ],
        },
        {
            "icon": "CHAT",
            "access_module": "Disparo WhatsApp",
            "title": "Disparo WhatsApp",
            "summary": "Envio rápido e automatizado de mensagens para seus clientes usando o WhatsApp Web.",
            "purpose": [
                "Ideal para enviar comunicados, lembretes de cobrança ou avisos de forma ágil.",
                "Você economiza tempo disparando mensagens padronizadas direto da sua planilha.",
            ],
            "steps": [
                "Selecione a sua planilha do Excel com os dados dos clientes.",
                "Clique em 'Carregar Abas' e marque quais listas deseja usar.",
                "Ajuste a velocidade de envio (recomendado manter o padrão para evitar bloqueios).",
                "Clique em 'Iniciar Disparo WhatsApp' e deixe o sistema trabalhar por você.",
            ],
        },
        {
            "icon": "ATTACH_MONEY",
            "access_module": "Gestão de Negociação",
            "title": "Gestao de Negociacao",
            "summary": "Geracao de propostas com faixas de desconto automaticas para clientes inadimplentes.",
            "purpose": [
                "Ajuda a montar propostas padronizadas para clientes em atraso, reduzindo o trabalho manual.",
                "Organiza os valores de desconto e as condicoes por faixa para acelerar o atendimento e a negociacao.",
            ],
            "steps": [
                "Selecione a Macro 8104 com os debitos que serao usados na negociacao.",
                "Defina os percentuais de desconto e os limites por faixa de atraso.",
                "Execute a rotina para calcular valores, parcelas e condicoes elegiveis.",
                "Revise a planilha final para confirmar clientes aprovados e restricoes aplicadas.",
                "Use a saida gerada como base para atendimento, proposta ou disparo posterior.",
            ],
        },
        {
            "icon": "TRAVEL_EXPLORE",
            "access_module": "Logística e Rotas",
            "title": "Logistica & Rotas",
            "summary": "Importa enderecos, organiza paradas e monta rotas otimizadas com apoio visual.",
            "purpose": [
                "Ajuda a planejar a ordem de atendimento das equipes de campo com mais eficiencia.",
                "Facilita a visualizacao dos enderecos, a distribuicao das paradas e a geracao de rotas para execucao operacional.",
            ],
            "steps": [
                "Carregue o arquivo da equipe ou da macro geografica com os enderecos das OSs.",
                "Revise as paradas carregadas e identifique itens sem coordenada.",
                "Use os recursos de geolocalizacao e otimizacao para ordenar a rota.",
                "Gere links do Google Maps ou abra o mapa consolidado para conferencia.",
                "Exporte a rota final para uso operacional da equipe em campo.",
            ],
        },
        {
            "icon": "MANAGE_ACCOUNTS",
            "access_module": "Gestão de acessos",
            "title": "Gestao de Acesso",
            "summary": "Cadastro de usuarios, cargos e permissoes por modulo com controle RBAC.",
            "purpose": [
                "Serve para controlar quem pode acessar cada parte do sistema.",
                "Ajuda a limitar modulos por perfil e manter a operacao organizada e segura.",
            ],
            "steps": [
                "Cadastre nome de usuario, senha e cargo do operador.",
                "Marque quais modulos ficarao disponiveis para esse perfil.",
                "Salve o cadastro e confirme a presenca do usuario na tabela inferior.",
                "Para editar permissoes, selecione o usuario correspondente e atualize os campos.",
                "Use esse painel para limitar acessos e manter a operacao segmentada por funcao.",
            ],
        },
        {
            "icon": "EMAIL",
            "access_module": "Disparo de E-mails",
            "title": "Disparo de Emails",
            "summary": "Envio rapido e automatizado de e-mails para multiplos clientes, diretamente pelo seu Outlook.",
            "purpose": [
                "Ideal para enviar comunicados, lembretes de cobranca ou avisos de forma agil para uma lista de clientes.",
                "Voce escreve a mensagem uma unica vez e o sistema preenche os dados de cada cliente (como nome e valor) automaticamente.",
            ],
            "steps": [
                "Certifique-se de que o Outlook esta instalado e com conta configurada na maquina.",
                "Selecione a planilha (CSV ou Excel) com os dados dos destinatarios.",
                "Escolha a aba e visualize as colunas disponiveis para mapeamento.",
                "Selecione um template de email e insira os tokens correspondentes aos campos da planilha.",
                "Inicie o disparo e acompanhe o progresso, contadores e logs de envio.",
            ],
        },
        {
            "icon": "HELP",
            "access_module": None,
            "title": "Ajuda & Suporte",
            "summary": "Consulta rapida dos requisitos, dos responsaveis pelo sistema e das orientacoes de uso.",
            "purpose": [
                "Centraliza orientacoes de uso e a finalidade de cada modulo para consulta rapida.",
                "Ajuda novos usuarios a entenderem o sistema sem depender de explicacao externa o tempo todo.",
            ],
            "steps": [
                "Use esta tela para revisar como cada modulo funciona antes de operar.",
                "Confira os requisitos minimos do ambiente logo abaixo.",
                "Em duvidas operacionais, valide primeiro o log do console ao final de cada rotina.",
                "Se houver erro recorrente, registre a mensagem exibida para facilitar o suporte.",
            ],
        },
    ]

    user = user or {"role": "ADM", "allowed_modules": "Todos"}
    _role = user.get("role", "")
    _mods = user.get("allowed_modules", "Todos")

    def _can_view_help_module(mod):
        if mod.get("icon") == "HELP":
            return True
        if mod.get("icon") == "MANAGE_ACCOUNTS":
            return _role == "ADM"
        access_module = mod.get("access_module")
        return _role == "ADM" or _mods == "Todos" or access_module in _mods

    visible_modules = [mod for mod in MODULES if _can_view_help_module(mod)]

    def help_module(mod):
        expanded = {"v": False}
        chevron = ft.Icon(ic("KEYBOARD_ARROW_DOWN"), color=TEXT_MUTED, size=18)
        details = ft.Container(
            visible=False,
            opacity=0,
            animate_opacity=ft.Animation(180, ft.AnimationCurve.EASE_IN_OUT),
            padding=ft.Padding.only(left=16, right=16, bottom=16),
        )
        details.content = ft.Column([
            ft.Text("Como usar este modulo:", size=12, weight=ft.FontWeight.BOLD, color=NAVY),
            *[
                ft.Row([
                    ft.Container(
                        content=ft.Text(str(i + 1), size=10, color=WHITE, weight=ft.FontWeight.BOLD),
                        width=18, height=18, border_radius=9, bgcolor=ORANGE,
                        alignment=ft.Alignment(0, 0),
                    ),
                    ft.Text(step_txt, size=12, color=TEXT_MAIN, expand=True),
                ], spacing=8, vertical_alignment=ft.CrossAxisAlignment.START)
                for i, step_txt in enumerate(mod["steps"])
            ],
            ft.Container(height=6),
            ft.Text("Para que serve este modulo:", size=12, weight=ft.FontWeight.BOLD, color=NAVY),
            *[
                ft.Row([
                    ft.Icon(ic("INFO_OUTLINE"), size=14, color=ORANGE),
                    ft.Text(purpose_txt, size=12, color=TEXT_MAIN, expand=True),
                ], spacing=8, vertical_alignment=ft.CrossAxisAlignment.START)
                for purpose_txt in mod["purpose"]
            ],
        ], spacing=8)

        def toggle(e):
            expanded["v"] = not expanded["v"]
            details.visible = expanded["v"]
            details.opacity = 1 if expanded["v"] else 0
            chevron.name = ic("KEYBOARD_ARROW_UP" if expanded["v"] else "KEYBOARD_ARROW_DOWN")
            try:
                chevron.update()
                details.update()
            except Exception:
                pass

        return ft.Container(
            content=ft.Column([
                ft.Container(
                    content=ft.Row([
                        ft.Container(
                            content=ft.Icon(ic(mod["icon"]), color=ORANGE, size=22),
                            width=44, height=44, bgcolor="#FFF3E8",
                            border_radius=10, alignment=ft.Alignment(0, 0),
                        ),
                        ft.Container(width=14),
                        ft.Column([
                            ft.Text(mod["title"], size=13, weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
                            ft.Text(mod["summary"], size=12, color=TEXT_MUTED),
                        ], spacing=3, expand=True),
                        chevron,
                    ], vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    padding=ft.Padding.symmetric(horizontal=16, vertical=14),
                    ink=True,
                    on_click=toggle,
                    border_radius=10,
                ),
                details,
            ], spacing=0),
            bgcolor=WHITE,
            border=ft.Border.only(bottom=ft.BorderSide(1, BORDER)),
        )

    return ft.Column([
        ft.Text("Ajuda & Creditos", size=22, weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
        ft.Text("Clique no modulo desejado para abrir as instrucoes detalhadas de uso.", size=13, color=TEXT_MUTED),
        ft.Container(height=8),
        card(ft.Column([
            sec_title("Guia por Modulo", "HELP"),
            ft.Container(height=10),
            ft.Container(
                content=ft.Column([help_module(mod) for mod in visible_modules], spacing=0),
                border=ft.Border.all(1, BORDER),
                border_radius=8,
                clip_behavior=ft.ClipBehavior.ANTI_ALIAS,
            ),
        ])),
        card(ft.Column([
            sec_title("Requisitos do Sistema", "COMPUTER"),
            ft.Container(height=10),
            ft.Column([
                check_row("Executavel autossuficiente - nao requer Python instalado"),
                check_row("Google Chrome atualizado - obrigatorio para a rotina de envio"),
                check_row("Conexao HTTPS (porta 443) - autenticacao online do sistema"),
            ], spacing=8),
        ])),
        card(ft.Column([
            sec_title("Creditos & Desenvolvimento", "CODE"),
            ft.Container(height=10),
            ft.Row([
                ft.Column([
                    ft.Text("Mathools v1.0", size=18, weight=ft.FontWeight.BOLD, color=NAVY),
                    ft.Text("Automacao e Processamento Inteligente de Dados", size=12, color=TEXT_MUTED),
                    ft.Container(height=12),
                    ft.Row([ft.Icon(ic("PERSON"), size=14, color=ORANGE), ft.Text("Desenvolvido por: Matheus Marques", size=13, color=TEXT_MAIN)], spacing=6),
                    ft.Row([ft.Icon(ic("BRUSH"), size=14, color=ORANGE), ft.Text("Design UX/UI: Eduarda Souza", size=13, color=TEXT_MAIN)], spacing=6),
                    ft.Row([ft.Icon(ic("CALENDAR_TODAY"), size=14, color=ORANGE), ft.Text("Lancamento: Marco de 2026", size=13, color=TEXT_MAIN)], spacing=6),
                    ft.Row([ft.Icon(ic("EMAIL"), size=14, color=ORANGE), ft.Text("Matheuskelvinm@gmail.com", size=13, color=TEXT_MAIN)], spacing=6),
                ], spacing=8, expand=True),
                ft.Column([
                    ft.Container(
                        content=ft.Image(
                            src=resource_path("Logo mathey tk 1.png"),
                            width=52, height=52, fit="contain",
                        ),
                        width=72, height=72, bgcolor=NAVY,
                        border_radius=16, alignment=ft.Alignment(0, 0),
                    ),
                    ft.Container(height=8),
                    ft.Text("Itapoa\nSaneamento", size=11, color=TEXT_MUTED,
                            text_align=ft.TextAlign.CENTER),
                ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            ], vertical_alignment=ft.CrossAxisAlignment.START),
        ])),
    ], spacing=0, scroll=ft.ScrollMode.AUTO, expand=True)

# ─────────────────────────────────────────────────────────────────────────────
#  PÁGINA 7  DISPARO DE E-MAILS
# ─────────────────────────────────────────────────────────────────────────────
def build_email_page(page: ft.Page, lq: queue.Queue):
    """
    Página de Mala Direta / Email Merge
    Sistema de envio de emails em massa com templates reutilizáveis
    - Envio via Outlook (pywin32) — sem configuração SMTP necessária
    - Delay configurável entre envios
    - Preview inline (na tela, não só no terminal)
    - Botão de Enviar com progress bar e controle de parada
    - Modelos HTML em email_templates/ (Itapoá) ou arquivo .html à parte
    - Merge {{CAMPO}} com colunas da planilha (nomes case-insensitive)
    - Inserção de campos no assunto (ou corpo no modo texto simples)
    """
    import threading
    import json as _json

    _TMPL_CFG_FILE = os.path.join(_app_base_dir(), "email_template_last.json")

    # ── Estado local ──────────────────────────────────────────────────────────
    state = {
        "csv_path": None,
        "data_source": None,
        "columns": [],
        "preview_data": [],
        "preview_html_path": None,  # último .html gerado para reabrir no navegador
        "preview_html_via_browser": False,
        "running": False,
        "stop_flag": False,
        "last_focused": "assunto",   # "assunto" ou "corpo"
        "custom_html_path": None,
        "custom_html_mtime": 0,
    }

    # ── Componentes UI ────────────────────────────────────────────────────────
    lbl_arquivo = ft.Text("Nenhum arquivo selecionado", size=13, color=TEXT_MUTED, italic=True)
    lbl_colunas = ft.Text("Campos disponíveis: nenhum arquivo carregado", size=11, color=TEXT_MUTED)

    # Delay entre envios
    txt_delay = ft.TextField(
        label="Delay entre envios (segundos)",
        hint_text="Ex: 2",
        value="2",
        width=200, height=44, border_radius=8, border_color=BORDER,
        focused_border_color=ORANGE, filled=True, fill_color="#F8FAFD",
    )
    txt_lote = ft.TextField(
        label="Pausa a cada N emails (0 = desligar)",
        hint_text="Ex: 50",
        value="0",
        width=240, height=44, border_radius=8, border_color=BORDER,
        focused_border_color=ORANGE, filled=True, fill_color="#F8FAFD",
    )
    txt_pausa_lote = ft.TextField(
        label="Pausa de lote (segundos)",
        hint_text="Ex: 60",
        value="60",
        width=200, height=44, border_radius=8, border_color=BORDER,
        focused_border_color=ORANGE, filled=True, fill_color="#F8FAFD",
    )

    # Seletor de sheets para XLSX multi-abas
    def on_sheet_change(e):
        """Carrega dados da sheet selecionada."""
        if not state["data_source"]:
            return
        
        selected_sheet = e.data if e and e.data else ddl_sheet.value
        if not selected_sheet:
            return
        ddl_sheet.value = selected_sheet
        
        try:
            lq.put(f"Carregando sheet '{selected_sheet}'...\n")
            success, msg = state["data_source"].select_sheet(selected_sheet)
            
            if not success:
                lq.put(f"ERRO ao carregar sheet: {msg}\n")
                _snack(f"Erro: {msg}", ok=False)
                return
            
            # Atualiza o estado
            state["columns"] = state["data_source"].data.columns.tolist() if state["data_source"].data is not None else []
            
            try:
                state["preview_data"] = state["data_source"].get_preview(limit=10)
            except Exception:
                state["preview_data"] = []

            guessed_email_col = _infer_email_column(state["columns"])
            _refresh_email_column_dropdown(preferred_value=guessed_email_col)
            
            total = len(state["data_source"].data) if state["data_source"].data is not None else 0
            
            lbl_arquivo.value = f"✓ {os.path.basename(state['csv_path'])}  ({total} linhas · {len(state['columns'])} campos) [Aba: {selected_sheet}]"
            try:
                lbl_arquivo.update()
            except Exception:
                pass

            cols_str = ", ".join(state["columns"][:10])
            if len(state["columns"]) > 10:
                cols_str += f", ... (+{len(state['columns']) - 10} mais)"
            lbl_colunas.value = f"Campos: {cols_str}"
            lbl_colunas.color = TEXT_MAIN
            lbl_colunas.update()
            
            refresh_col_buttons()
            
            if container_preview.visible:
                if txt_assunto.value and txt_assunto.value.strip():
                    gerar_preview(None)

            lq.put(f"✓ Sheet carregada: {total} linhas, {len(state['columns'])} colunas\n")
            _snack(f"Sheet '{selected_sheet}' carregada com sucesso!")
            if guessed_email_col:
                _set_status(f"Sheet carregada. Coluna de e-mail detectada automaticamente: {guessed_email_col}", GREEN)
            
        except Exception as ex:
            import traceback
            lq.put(f"ERRO ao carregar sheet: {ex}\n{traceback.format_exc()}\n")
            _snack(f"Erro ao carregar sheet: {ex}", ok=False)
    
    ddl_sheet = ft.Dropdown(
        label="Selecione a aba (Sheet)",
        options=[],
        visible=False,
        width=250,
        height=44,
        border_radius=8,
        border_color=BORDER,
        focused_border_color=ORANGE,
    )
    ddl_sheet.on_change = on_sheet_change

    # Template
    txt_assunto = ft.TextField(
        label="Assunto do Email",
        hint_text="Ex: Aviso de {{ENDERECO}}",
        multiline=False,
        height=44, border_radius=8, border_color=BORDER,
        focused_border_color=ORANGE, filled=True, fill_color="#F8FAFD",
        on_focus=lambda e: state.update({"last_focused": "assunto"}),
    )
    txt_corpo = ft.TextField(
        label="Corpo do Email",
        hint_text="Ex: Prezado {{NOME}},\n\nSeu endereço: {{ENDERECO}}",
        multiline=True, min_lines=6, max_lines=14,
        border_radius=8, border_color=BORDER,
        focused_border_color=ORANGE, filled=True, fill_color="#F8FAFD",
        on_focus=lambda e: state.update({"last_focused": "corpo"}),
    )
    txt_destinatario = ft.Dropdown(
        label="Coluna de Email do Destinatário *",
        hint_text="Selecione a coluna com o e-mail do cliente",
        options=[],
        width=300, height=44, border_radius=8, border_color=BORDER,
        focused_border_color=ORANGE, filled=True, fill_color="#F8FAFD",
    )

    cb_baixar_faturas_email = ft.Switch(
        label="Baixar faturas automaticamente do Waterfy em PDF",
        value=False,
    )
    e_usuario_wf_email = ft.TextField(
        label="Usuário Waterfy",
        width=280, height=44, border_radius=8, border_color=BORDER,
        focused_border_color=ORANGE, filled=True, fill_color="#F8FAFD",
    )
    e_senha_wf_email = ft.TextField(
        label="Senha Waterfy",
        password=True,
        width=280, height=44, border_radius=8, border_color=BORDER,
        focused_border_color=ORANGE, filled=True, fill_color="#F8FAFD",
    )
    credencias_fatura_email = ft.Container(
        content=ft.Column([
            ft.Container(height=8),
            ft.Row([e_usuario_wf_email, ft.Container(width=12), e_senha_wf_email]),
            ft.Container(height=6),
            ft.Text(
                "Para o download funcionar, a sua planilha precisa ter as colunas MATRICULA e ID_FATURA.",
                size=11,
                color=TEXT_MUTED,
            ),
        ]),
        visible=False,
    )
    cb_senha_pdf_fatura = ft.Checkbox(
        label="Proteger PDFs com senha (requer coluna com CPF ou CNPJ do cliente na planilha)",
        value=False,
    )

    def _on_baixar_faturas_email_change(_e=None):
        credencias_fatura_email.visible = cb_baixar_faturas_email.value
        cb_senha_pdf_fatura.disabled = not cb_baixar_faturas_email.value
        if not cb_baixar_faturas_email.value:
            cb_senha_pdf_fatura.value = False
        try:
            credencias_fatura_email.update()
            cb_senha_pdf_fatura.update()
        except Exception:
            pass

    cb_baixar_faturas_email.on_change = _on_baixar_faturas_email_change
    _on_baixar_faturas_email_change(None)

    lbl_tpl_status = ft.Text(
        "Modelo: Itapoá — menos de 30 dias",
        size=12,
        color=TEXT_MUTED,
    )

    def _email_tpl_builtin_options():
        from email_merge_module import list_builtin_email_template_choices
        opts = [ft.dropdown.Option(tid, lab) for tid, lab in list_builtin_email_template_choices()]
        opts.append(ft.dropdown.Option("custom", "Outro arquivo HTML…"))
        return opts

    ddl_template = ft.Dropdown(
        label="Modelo de e-mail (HTML)",
        options=_email_tpl_builtin_options(),
        value="itapoa_lt30",
        width=420,
        height=44,
        border_radius=8,
        border_color=BORDER,
        focused_border_color=ORANGE,
        filled=True,
        fill_color="#F8FAFD",
    )

    def _update_tpl_status_text():
        if sw_manual_body.value:
            lbl_tpl_status.value = "Modo manual: o texto do corpo abaixo será enviado (sem arquivo HTML)."
            lbl_tpl_status.color = TEXT_MAIN
            return
        if ddl_template.value == "custom":
            p = state.get("custom_html_path")
            lbl_tpl_status.value = f"HTML personalizado: {p or '(use Escolher arquivo…)'}"
            lbl_tpl_status.color = ORANGE if not p else TEXT_MAIN
            return
        from email_merge_module import list_builtin_email_template_choices
        for tid, lab in list_builtin_email_template_choices():
            if tid == ddl_template.value:
                lbl_tpl_status.value = f"Modelo: {lab}"
                lbl_tpl_status.color = TEXT_MUTED
                return

    def on_ddl_template_change(e):
        if ddl_template.value != "custom":
            state["custom_html_path"] = None
            state["custom_html_mtime"] = 0
        btn_pick_html.visible = ddl_template.value == "custom" and not sw_manual_body.value
        _update_tpl_status_text()
        try:
            lbl_tpl_status.update()
            btn_pick_html.update()
        except Exception:
            pass
        
        if ddl_template.value == "custom" and not state.get("custom_html_path"):
            pick_html_template_file(None)

    ddl_template.on_change = on_ddl_template_change

    def pick_html_template_file(e=None):
        def do_pick():
            try:
                import tkinter as tk
                from tkinter import filedialog
                root = tk.Tk()
                root.withdraw()
                root.attributes("-topmost", True)
                path = filedialog.askopenfilename(
                    filetypes=[("HTML", "*.html *.htm"), ("Todos", "*.*")],
                )
                root.destroy()
                if path:
                    state["custom_html_path"] = path
                    state["custom_html_mtime"] = 0
                    ddl_template.value = "custom"
                    _update_tpl_status_text()
                    try:
                        ddl_template.update()
                        lbl_tpl_status.update()
                        btn_pick_html.update()
                    except Exception:
                        pass
                    _snack(f"Template: {os.path.basename(path)}")
            except Exception as ex:
                lq.put(f"Erro ao selecionar HTML: {ex}\n")
                _snack(str(ex), ok=False)
        do_pick()

    btn_pick_html = sbtn("Escolher arquivo…", "FOLDER", on_click=pick_html_template_file, width=160)
    btn_pick_html.visible = False

    sw_manual_body = ft.Switch(
        label="Escrever a mensagem manualmente (desativa os modelos prontos)",
        value=False,
    )

    def on_sw_manual_body(e):
        manual = sw_manual_body.value
        txt_corpo.visible = manual
        txt_corpo.disabled = not manual
        ddl_template.disabled = manual
        btn_pick_html.visible = (ddl_template.value == "custom") and not manual
        if manual:
            state["last_focused"] = "corpo"
        else:
            state["last_focused"] = "assunto"
        _update_tpl_status_text()
        try:
            txt_corpo.update()
            ddl_template.update()
            btn_pick_html.update()
            lbl_tpl_status.update()
            sw_manual_body.update()
        except Exception:
            pass

    sw_manual_body.on_change = on_sw_manual_body
    txt_corpo.visible = False
    txt_corpo.disabled = True
    _update_tpl_status_text()

    def _get_merge_body_html() -> tuple:
        """
        Retorna (corpo_str, erro).
        Usa modelo HTML interno, arquivo custom ou texto do campo corpo (modo manual).
        """
        import re as _re
        from email_merge_module import load_builtin_email_template, load_template_html_file

        if sw_manual_body.value:
            body = (txt_corpo.value or "").strip()
            if not body:
                return "", "Preencha o corpo do e-mail ou desligue o modo texto simples e escolha um modelo HTML."
            try:
                from email_dispatch_module import CorporateEmailHtmlBuilder
                builder = CorporateEmailHtmlBuilder()
                return builder.build_html(txt_assunto.value.strip() or "Mensagem", body), ""
            except ImportError:
                return body, ""

        if ddl_template.value == "custom":
            p = state.get("custom_html_path")
            if not p:
                return "", "Selecione um arquivo HTML (modelo «Outro arquivo HTML…»)."
            html, err = load_template_html_file(p)
            return html, err

        html, err = load_builtin_email_template(ddl_template.value)
        return html, err

    def _strip_html_for_preview(html: str, limit: int = 900) -> str:
        import re as _re
        if not html:
            return ""
        t = _re.sub(r"(?is)<script.*?>.*?</script>", " ", html)
        t = _re.sub(r"(?is)<style.*?>.*?</style>", " ", t)
        t = _re.sub(r"<[^>]+>", " ", t)
        t = _re.sub(r"\s+", " ", t).strip()
        return t[:limit] + ("…" if len(t) > limit else "")

    def _preview_html_corpo_ui(previews_list: list) -> ft.Control:
        """
        Flet 0.82 no Windows não expõe ft.Html. Prévia fiel = mesmo HTML do Outlook
        gravado em arquivo e aberto no navegador padrão (Edge/Chrome).
        Se existir ft.Html em versões futuras, usa painel embutido.
        """
        import tempfile
        import webbrowser
        import json
        from pathlib import Path as _Path

        state["preview_html_via_browser"] = False
        Html = getattr(ft, "Html", None)
        
        # Gera um arquivo HTML separado para cada preview — sem JS dinamico, sem iframe,
        # sem innerHTML. Cada arquivo e um documento completo e estatico, o que elimina
        # qualquer restricao de cross-origin do Chrome ao abrir via file://.
        import re as _re
        tmp_dir = _Path(tempfile.gettempdir())
        preview_files = []
        total_p = len(previews_list)

        for idx_p, p in enumerate(previews_list):
            body_html = p.get("body", "")
            subject_p = p.get("subject", "")
            # Se nao e documento completo, envolve num wrapper minimo
            if not body_html.strip().lower().startswith(("<!doctype", "<html")):
                body_html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"></head>
<body style="font-family:sans-serif;padding:15px;color:#333;">{body_html}</body></html>"""

            # Injeta barra de navegacao no topo do <body>
            nav_style = (
                "position:fixed;top:0;left:0;right:0;z-index:9999;"
                "background:#011F3F;color:#fff;padding:10px 20px;"
                "display:flex;align-items:center;justify-content:space-between;"
                "font-family:'Segoe UI',sans-serif;font-size:14px;font-weight:600;"
                "box-shadow:0 2px 6px rgba(0,0,0,0.3);"
            )
            btn_style = (
                "cursor:pointer;padding:7px 16px;background:#F27D16;color:#fff;"
                "border:none;border-radius:6px;font-weight:bold;font-size:13px;"
            )
            btn_dis = btn_style + "background:#4A5D78;color:#A3B4CC;cursor:not-allowed;"

            prev_file = f"mathtools_email_preview_{idx_p - 1}.html" if idx_p > 0 else ""
            next_file = f"mathtools_email_preview_{idx_p + 1}.html" if idx_p < total_p - 1 else ""

            prev_btn = (f'<a href="{prev_file}"><button style="{btn_style}">&laquo; Anterior</button></a>'
                        if prev_file else f'<button style="{btn_dis}" disabled>&laquo; Anterior</button>')
            next_btn = (f'<a href="{next_file}"><button style="{btn_style}">Próxima &raquo;</button></a>'
                        if next_file else f'<button style="{btn_dis}" disabled>Próxima &raquo;</button>')

            nav_bar = (
                f'<div style="{nav_style}">' +
                prev_btn +
                f'<span>Prévia {idx_p + 1} de {total_p}</span>' +
                next_btn +
                f'</div><div style="height:52px"></div>'
            )

            # Insere a nav_bar logo apos o <body ...>
            if _re.search(r"<body", body_html, _re.IGNORECASE):
                body_html = _re.sub(r"(<body[^>]*>)", r"\1" + nav_bar, body_html, count=1, flags=_re.IGNORECASE)
            else:
                body_html = nav_bar + body_html

            fname = tmp_dir / f"mathtools_email_preview_{idx_p}.html"
            try:
                fname.write_text(body_html, encoding="utf-8")
                preview_files.append(fname)
            except Exception:
                pass

        tmp = preview_files[0] if preview_files else tmp_dir / "mathtools_email_preview_0.html"
        state["preview_html_path"] = str(tmp)

        try:
            pass  # arquivos ja escritos acima
        except Exception as ex:
            state["preview_html_path"] = None
            return ft.Text(
                f"Não foi possível gerar prévia externa: {ex}\n\n"
                + _strip_html_for_preview(previews_list[0]["body"], 1200),
                size=12,
                color=TEXT_MAIN,
                selectable=True,
            )

        Html = None  # força abertura no browser (arquivos separados nao funcionam com ft.Html)
        if Html is not None:
            pass  # bloco mantido por compatibilidade

        if state["preview_html_path"]:
            state["preview_html_via_browser"] = True
            webbrowser.open(tmp.as_uri())

        def _reopen(_):
            p = state.get("preview_html_path")
            if p and os.path.isfile(p):
                webbrowser.open(_Path(p).as_uri())

        return ft.Column(
            tight=True,
            spacing=10,
            controls=[
                ft.Text(
                    "Para uma melhor visualização do layout do e-mail, abrimos uma prévia no seu navegador de internet padrão.",
                    size=11,
                    color=TEXT_MUTED,
                ),
                ft.OutlinedButton(
                    "Abrir preview no navegador novamente",
                    on_click=_reopen,
                ),
                ft.Text(
                    str(tmp),
                    size=10,
                    color=TEXT_MUTED,
                    selectable=True,
                ),
            ],
        )

    # Preview inline (HTML: navegador ou ft.Html se existir na versão do Flet)
    lbl_preview_assunto = ft.Text("", size=13, color=TEXT_MAIN, selectable=True)
    preview_corpo_host = ft.Container(
        content=ft.Text(
            "(clique em «Gerar Preview» para ver o corpo)",
            size=12,
            color=TEXT_MUTED,
            italic=True,
        ),
    )
    container_preview = ft.Container(
        visible=False,
        bgcolor="#F0FFF4",
        border_radius=8,
        border=ft.Border.all(1, GREEN),
        padding=16,
        content=ft.Column([
            ft.Text("Preview — Primeiros E-mails da Planilha", size=11, color=GREEN, weight=ft.FontWeight.W_600),
            ft.Divider(height=8, color=ft.Colors.with_opacity(0.3, GREEN)),
            ft.Text("Assunto (1º e-mail):", size=11, color=TEXT_MUTED, weight=ft.FontWeight.W_600),
            lbl_preview_assunto,
            ft.Container(height=6),
            ft.Text("Corpo:", size=11, color=TEXT_MUTED, weight=ft.FontWeight.W_600),
            preview_corpo_host,
        ], spacing=4),
    )

    col_campos = ft.Row(wrap=True, spacing=6, run_spacing=6)

    lbl_status = ft.Text("Aguardando arquivo CSV...", size=12, color=TEXT_MUTED, italic=True)
    prog = ft.ProgressBar(value=0, bgcolor=BORDER, color=ORANGE, height=6, border_radius=3, visible=False)
    lbl_contagem = ft.Text("", size=12, color=TEXT_MUTED)

    # Botão enviar / parar (referência mutável via lista)
    _btn_enviar_ref = [None]
    _btn_parar_ref  = [None]

    # ── Funções auxiliares ────────────────────────────────────────────────────
    def _set_status(msg, color=TEXT_MUTED, log_message=None):
        lbl_status.value = msg
        lbl_status.color = color
        try:
            lbl_status.update()
        except Exception:
            pass
        if log_message:
            try:
                lq.put(log_message if log_message.endswith("\n") else f"{log_message}\n")
            except Exception:
                pass

    def _open_overlay(control):
        try:
            if hasattr(page, "open"):
                page.open(control)
            else:
                if control not in page.overlay:
                    page.overlay.append(control)
                control.open = True
                page.update()
        except Exception:
            try:
                if control not in page.overlay:
                    page.overlay.append(control)
                control.open = True
                page.update()
            except Exception:
                pass

    def _close_overlay(control):
        try:
            if hasattr(page, "close"):
                page.close(control)
            else:
                control.open = False
                page.update()
        except Exception:
            try:
                control.open = False
                page.update()
            except Exception:
                pass

    def _snack(msg, ok=True):
        _set_status(msg, GREEN if ok else RED)
        page.snack_bar = ft.SnackBar(ft.Text(msg, color=WHITE), bgcolor=GREEN if ok else RED)
        _open_overlay(page.snack_bar)

    def _infer_email_column(columns):
        if not columns:
            return ""
        normalized = {str(col).strip().upper(): str(col) for col in columns}
        for candidate in ("EMAIL", "E-MAIL", "MAIL", "CORREIO", "EMAIL_DESTINATARIO", "DESTINATARIO"):
            if candidate in normalized:
                return normalized[candidate]
        for original in columns:
            up = str(original).strip().upper()
            if "EMAIL" in up or "E-MAIL" in up or up.endswith("_MAIL"):
                return str(original)
        return ""

    def _refresh_email_column_dropdown(preferred_value=None):
        txt_destinatario.options = [ft.dropdown.Option(str(col)) for col in state["columns"]]

        selected = preferred_value if preferred_value in state["columns"] else txt_destinatario.value
        if selected not in state["columns"]:
            selected = _infer_email_column(state["columns"])

        txt_destinatario.value = selected if selected in state["columns"] else None
        try:
            txt_destinatario.update()
        except Exception:
            pass

    def pick_csv(e):
        def do_pick():
            try:
                import tkinter as tk
                from tkinter import filedialog
                root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
                path = filedialog.askopenfilename(filetypes=[("CSV / Excel", "*.csv *.xlsx"), ("Todos", "*.*")])
                root.destroy()
                if path:
                    load_csv(path)
            except Exception as ex:
                lq.put(f"Erro ao abrir seletor: {ex}\n")
        do_pick()

    def load_csv(path: str):
        try:
            lbl_arquivo.value = f"Carregando {os.path.basename(path)}..."
            lbl_arquivo.color = ORANGE
            lbl_arquivo.update()

            try:
                from email_merge_module import MergeDataSource
            except ImportError:
                lq.put("ERRO: email_merge_module.py não encontrado!\n")
                lbl_arquivo.value = "Erro: módulo não encontrado"
                lbl_arquivo.color = RED
                lbl_arquivo.update()
                return

            ds = MergeDataSource()
            success, msg = ds.load_csv(path)
            if not success:
                lq.put(f"ERRO ao carregar CSV: {msg}\n")
                lbl_arquivo.value = f"Erro: {msg}"
                lbl_arquivo.color = RED
                lbl_arquivo.update()
                return

            state["csv_path"] = path
            state["data_source"] = ds
            state["columns"] = ds.data.columns.tolist() if ds.data is not None else []

            try:
                state["preview_data"] = ds.get_preview(limit=10)
            except Exception:
                state["preview_data"] = []

            guessed_email_col = _infer_email_column(state["columns"])
            _refresh_email_column_dropdown(preferred_value=guessed_email_col)

            # Se é arquivo XLSX, lista as sheets disponíveis
            if path.lower().endswith(('.xlsx', '.xls', '.xlsm')):
                sheets = ds.get_excel_sheets(path)
                if len(sheets) > 1:
                    # Cria opções para o dropdown
                    ddl_sheet.options = [ft.dropdown.Option(sheet) for sheet in sheets]
                    ddl_sheet.value = sheets[0]  # Seleciona a primeira por padrão
                    ddl_sheet.visible = True
                    lq.put(f"ℹ Encontramos {len(sheets)} abas na sua planilha: {', '.join(sheets)}\n")
                    # Garante que o data_source carrega os dados da sheet exibida no dropdown
                    on_sheet_change(None)
                else:
                    ddl_sheet.visible = False
                try: ddl_sheet.update()
                except Exception: pass
            else:
                ddl_sheet.visible = False
                try: ddl_sheet.update()
                except Exception: pass

            total = len(ds.data) if ds.data is not None else 0
            aba_atual = f" [Aba: {ddl_sheet.value}]" if ddl_sheet.visible else ""
            lbl_arquivo.value = f"✓ {os.path.basename(path)}  ({total} linhas · {len(state['columns'])} campos){aba_atual}"
            lbl_arquivo.color = GREEN

            cols_str = ", ".join(state["columns"][:10])
            if len(state["columns"]) > 10:
                cols_str += f", ... (+{len(state['columns']) - 10} mais)"
            lbl_colunas.value = f"Campos: {cols_str}"
            lbl_colunas.color = TEXT_MAIN

            lbl_arquivo.update()
            lbl_colunas.update()

            refresh_col_buttons()
            lq.put(f"✓ CSV carregado: {total} linhas, {len(state['columns'])} colunas\n")
            if guessed_email_col:
                _set_status(f"Arquivo carregado. Coluna de e-mail detectada automaticamente: {guessed_email_col}", GREEN)
            else:
                _set_status("Arquivo carregado. Confira a coluna de e-mail antes de enviar.", ORANGE)

        except Exception as ex:
            lq.put("Problema ao ler a planilha: O arquivo pode estar corrompido ou salvo em um formato incompatível.\n")
            lbl_arquivo.value = "Arquivo inválido ou corrompido"
            lbl_arquivo.color = RED
            lbl_arquivo.update()

    def refresh_col_buttons():
        col_campos.controls.clear()
        if not state["columns"]:
            txt_destinatario.options = []
            txt_destinatario.value = None
            try:
                txt_destinatario.update()
            except Exception:
                pass
            col_campos.controls.append(
                ft.Text("Carregue um CSV primeiro.", size=11, color=TEXT_MUTED, italic=True)
            )
        else:
            _refresh_email_column_dropdown()
            for col in state["columns"]:
                def make_handler(field_name):
                    def insert_field(e):
                        token = f"{{{{{field_name}}}}}"
                        use_corpo = bool(sw_manual_body.value) and state["last_focused"] == "corpo"
                        if use_corpo:
                            txt_corpo.value = (txt_corpo.value or "") + f" {token}"
                            try: txt_corpo.update()
                            except Exception: pass
                        else:
                            txt_assunto.value = (txt_assunto.value or "") + f" {token}"
                            try: txt_assunto.update()
                            except Exception: pass
                    return insert_field

                col_campos.controls.append(
                    ft.Container(
                        content=ft.Row([
                            ft.Icon(ic("ADD_CIRCLE_OUTLINE"), size=12, color=WHITE),
                            ft.Container(width=5),
                            ft.Text(col, size=11, color=WHITE, weight=ft.FontWeight.W_600),
                        ], spacing=0),
                        height=30,
                        bgcolor=NAVY,
                        border_radius=6,
                        padding=ft.Padding.symmetric(horizontal=10, vertical=0),
                        ink=True,
                        on_click=make_handler(col),
                        alignment=ft.Alignment(0, 0),
                    )
                )
        try: col_campos.update()
        except Exception: pass

    def gerar_preview(e):
        if not state["data_source"]:
            _snack("Carregue um CSV primeiro!", ok=False); return

        subject = txt_assunto.value.strip()
        dest    = (txt_destinatario.value or "").strip()
        body, body_err = _get_merge_body_html()
        if not subject:
            _snack("Preencha o assunto!", ok=False); return
        if body_err:
            _snack(body_err, ok=False); return
        if not body:
            _snack("Configure o modelo HTML ou o corpo manual.", ok=False); return

        try:
            # Garante que o preview sempre usa a aba selecionada no dropdown
            if ddl_sheet.visible and ddl_sheet.value:
                success, msg_sheet = state["data_source"].select_sheet(ddl_sheet.value)
                if success:
                    state["columns"] = state["data_source"].data.columns.tolist() if state["data_source"].data is not None else []
                    state["preview_data"] = state["data_source"].get_preview(limit=10)
                    lq.put(f"   Preview usando aba: {ddl_sheet.value} ({len(state['data_source'].data)} linhas)\n")
                else:
                    lq.put(f"   Aviso: nao foi possivel carregar aba '{ddl_sheet.value}': {msg_sheet}\n")

            state["preview_html_via_browser"] = False
            from email_merge_module import MergeTemplate, MergeEngine, is_probably_html_document
            template = MergeTemplate(name="Preview", subject=subject, body=body)
            engine   = MergeEngine()
            engine.template    = template
            engine.data_source = state["data_source"]

            valid, msg = engine.validate_template()
            if not valid:
                lq.put(f"Erro de validação: {msg}\n")
                _snack(f"Erro: {msg}", ok=False); return

            preview_data = state["preview_data"]
            if not preview_data:
                lbl_preview_assunto.value = "Nenhum dado disponível nesta aba."
                preview_corpo_host.content = ft.Text("Aba vazia ou sem registros para gerar a prévia.", size=12, color=TEXT_MUTED, italic=True)
                try: container_preview.update()
                except Exception: pass
                _snack("Nenhuma linha para preview!", ok=False); return

            previews_list = []
            for row in preview_data:
                try:
                    previews_list.append(engine.render_record(row))
                except Exception:
                    pass
                    
            if not previews_list:
                _snack("Erro ao renderizar previews.", ok=False); return

            lbl_preview_assunto.value = f"{previews_list[0]['subject']} (veja os demais no preview abaixo/navegador)"
            preview_corpo_host.content = _preview_html_corpo_ui(previews_list)
            container_preview.visible = True

            try: container_preview.update()
            except Exception: pass
            try: page.update()
            except Exception: pass

            lq.put(f"Preview gerado com sucesso para o assunto: {previews_list[0]['subject']}\n")
            if state.get("preview_html_via_browser"):
                _snack("Preview HTML aberto no navegador (mesmo conteúdo do e-mail).")
            else:
                _snack("Preview gerado com sucesso!")

        except Exception as ex:
            import traceback
            lq.put(f"ERRO ao gerar preview: {ex}\n{traceback.format_exc()}\n")
            _snack(f"Erro: {ex}", ok=False)

    def testar_outlook(e):
        """Verifica se o Outlook está disponível via pywin32."""
        def do_test():
            lq.put("Verificando integração com o Outlook...\n")
            pythoncom = None
            _OUTLOOK_TIMEOUT = 15  # segundos
            try:
                import pythoncom  # type: ignore
                import win32com.client  # type: ignore
                pythoncom.CoInitialize()
                pythoncom_ref = pythoncom
                lq.put("Preparando ambiente para envio seguro...\n")
                lq.put("Abrindo comunicação com a sua conta de e-mail...\n")

                _result: dict = {}

                def _dispatch():
                    try:
                        import pythoncom
                        pythoncom.CoInitialize()
                        dummy = win32com.client.Dispatch("Outlook.Application")
                    except Exception as _ex:
                        _result["error"] = _ex
                    finally:
                        try:
                            import pythoncom
                            pythoncom.CoUninitialize()
                        except Exception:
                            pass

                _t = threading.Thread(target=_dispatch, daemon=True)
                _t.start()
                _t.join(timeout=_OUTLOOK_TIMEOUT)

                if _t.is_alive():
                    raise RuntimeError(
                        f"Outlook não respondeu em {_OUTLOOK_TIMEOUT}s. "
                        "Verifique se está aberto e sem janelas de diálogo bloqueando."
                    )
                if "error" in _result:
                    raise _result["error"]

                # Apenas instancia — não envia nada
                lq.put("✓ Tudo certo! Outlook conectado com sucesso.\n")
                page.snack_bar = ft.SnackBar(
                    ft.Text("✓ Outlook disponível e pronto para envio.", color=WHITE),
                    bgcolor=GREEN, open=True,
                )
                try: page.update()
                except Exception: pass
            except Exception as ex:
                lq.put(f"ERRO ao conectar no Outlook: {ex}\n")
                page.snack_bar = ft.SnackBar(
                    ft.Text(f"Falha: {ex}", color=WHITE),
                    bgcolor=RED, open=True,
                )
                try: page.update()
                except Exception: pass
            finally:
                if pythoncom is not None:
                    try:
                        pythoncom.CoUninitialize()
                    except Exception:
                        pass

        threading.Thread(target=do_test, daemon=True).start()

    def iniciar_envio(e):
        if state["running"]:
            return

        # Validações
        if not state["data_source"]:
            _snack("Carregue um CSV primeiro!", ok=False); return

        subject  = txt_assunto.value.strip()
        dest_col = (txt_destinatario.value or "").strip()
        body, body_err = _get_merge_body_html()

        if not subject:
            _snack("Preencha o assunto!", ok=False); return
        if body_err:
            _snack(body_err, ok=False); return
        if not body:
            _snack("Configure o modelo HTML ou o corpo manual.", ok=False); return
        if not dest_col:
            guessed_email_col = _infer_email_column(state["columns"])
            if guessed_email_col:
                dest_col = guessed_email_col
                txt_destinatario.value = guessed_email_col
                try:
                    txt_destinatario.update()
                except Exception:
                    pass
            else:
                _snack("Especifique a coluna de email!", ok=False); return

        if not dest_col:
            _snack("Especifique a coluna de email!", ok=False); return

        if cb_baixar_faturas_email.value:
            if not e_usuario_wf_email.value.strip() or not e_senha_wf_email.value.strip():
                _snack("Preencha usuário e senha Waterfy para baixar faturas.", ok=False)
                return

        try:
            delay_s     = float(txt_delay.value.strip() or "2")
            batch_size  = int(txt_lote.value.strip() or "0")
            batch_pause = float(txt_pausa_lote.value.strip() or "60")
        except ValueError:
            _snack("Delay / lote inválido!", ok=False); return

        # Confirma — garante que a sheet selecionada no dropdown está carregada
        if ddl_sheet.visible and ddl_sheet.value:
            try:
                state["data_source"].select_sheet(ddl_sheet.value)
                state["columns"] = state["data_source"].data.columns.tolist() if state["data_source"].data is not None else []
            except Exception:
                pass
        ds = state["data_source"]
        total = len(ds.data) if ds.data is not None else 0

        anexo_txt = (
            "\n\nAnexos: PDF da planilha (ex.: ARQUIVO_PDF) se houver caminho válido."
            + (
                "\nDownload Waterfy ativo (MATRICULA + ID_FATURA)."
                + (
                    "\nPDF com senha do consumidor quando a opção estiver marcada."
                    if cb_senha_pdf_fatura.value
                    else ""
                )
                if cb_baixar_faturas_email.value
                else ""
            )
        )

        def _confirmar(ev):
            _close_overlay(dlg)
            _executar_envio(
                subject,
                body,
                dest_col,
                delay_s,
                batch_size,
                batch_pause,
                total,
                baixar_fatura_wf=cb_baixar_faturas_email.value,
                wf_user=e_usuario_wf_email.value.strip(),
                wf_pass=e_senha_wf_email.value,
                proteger_pdf=bool(cb_senha_pdf_fatura.value),
            )

        def _cancelar_dlg(ev):
            _close_overlay(dlg)

        dlg = ft.AlertDialog(
            modal=True,
            bgcolor=WHITE,
            title=ft.Row([
                ft.Icon(ic("SEND"), color=ORANGE, size=20),
                ft.Container(width=8),
                ft.Text("Confirmar Disparo", size=15, weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
            ], spacing=0),
            content=ft.Text(
                f"Serão preparados até {total} e-mails\npara os contatos da coluna '{dest_col}'."
                f"{anexo_txt}\n\nPodemos iniciar os envios?",
                size=13, color=TEXT_MAIN,
            ),
            actions=[
                ft.TextButton("Cancelar", on_click=_cancelar_dlg,
                              style=ft.ButtonStyle(color=NAVY)),
                ft.Container(
                    content=ft.Text("Enviar", size=13, weight=ft.FontWeight.BOLD, color=WHITE),
                    bgcolor=ORANGE, border_radius=8,
                    padding=ft.Padding.symmetric(horizontal=20, vertical=8),
                    on_click=_confirmar, ink=True,
                ),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        _open_overlay(dlg)

    def parar_envio(e):
        state["stop_flag"] = True
        lq.put("⚠ Sinal de parada enviado — aguardando email atual terminar...\n")

    def _executar_envio(
        subject,
        body,
        dest_col,
        delay_s,
        batch_size,
        batch_pause,
        total,
        baixar_fatura_wf=False,
        wf_user="",
        wf_pass="",
        proteger_pdf=False,
    ):
        state["running"]   = True
        state["stop_flag"] = False

        # Atualiza botões
        if _btn_enviar_ref[0]:
            _btn_enviar_ref[0].disabled = True
            try: _btn_enviar_ref[0].update()
            except Exception: pass
        if _btn_parar_ref[0]:
            _btn_parar_ref[0].disabled = False
            try: _btn_parar_ref[0].update()
            except Exception: pass

        prog.visible = True
        prog.value   = 0
        lbl_status.value = "Iniciando envio..."
        lbl_status.color = ORANGE
        try: prog.update(); lbl_status.update()
        except Exception: pass

        def _thread():
            enviados  = 0
            erros     = 0
            pulados   = 0
            pythoncom = None

            try:
                # Garante que a sheet selecionada no dropdown é a que será disparada
                # (necessário pois _thread roda em background e pode pegar estado desatualizado)
                if ddl_sheet.visible and ddl_sheet.value:
                    try:
                        state["data_source"].select_sheet(ddl_sheet.value)
                        lq.put(f"   Aba confirmada para disparo: {ddl_sheet.value}\n")
                    except Exception as _se:
                        lq.put(f"   Aviso: não foi possível reconfirmar aba ({_se})\n")

                df_dados = state["data_source"].data
                if "STATUS" not in df_dados.columns:
                    df_dados["STATUS"] = ""

                from email_merge_module import MergeTemplate, MergeEngine, is_probably_html_document
                template = MergeTemplate(name="Disparo", subject=subject, body=body)
                engine   = MergeEngine()
                engine.template    = template
                engine.data_source = state["data_source"]
                engine.email_column = dest_col

                # Tenta conectar ao Outlook, mas se falhar usa SMTP como fallback
                _OUTLOOK_TIMEOUT = 15  # segundos
                usar_outlook = False
                outlook = None
                try:
                    import pythoncom  # type: ignore
                    import win32com.client  # type: ignore
                    pythoncom.CoInitialize()
                    lq.put("Preparando integração com o sistema...\n")
                    lq.put("Abrindo comunicação com o Outlook...\n")

                    _ol_result: dict = {}

                    def _connect_outlook():
                        try:
                            import pythoncom
                            pythoncom.CoInitialize()
                            dummy = win32com.client.Dispatch("Outlook.Application")
                        except Exception as _ex:
                            _ol_result["error"] = _ex
                        finally:
                            try:
                                import pythoncom
                                pythoncom.CoUninitialize()
                            except Exception:
                                pass

                    _ol_t = threading.Thread(target=_connect_outlook, daemon=True)
                    _ol_t.start()
                    _ol_t.join(timeout=_OUTLOOK_TIMEOUT)

                    if _ol_t.is_alive():
                        raise RuntimeError(
                            f"Outlook não respondeu em {_OUTLOOK_TIMEOUT}s. "
                            "Verifique se está aberto e sem janelas de diálogo bloqueando."
                        )
                    if "error" in _ol_result:
                        raise _ol_result["error"]

                    outlook = win32com.client.Dispatch("Outlook.Application")
                    lq.put("✓ Outlook conectado! Usando Outlook para envio.\n")
                    usar_outlook = True
                except Exception as e_outlook:
                    lq.put(f"⚠ Outlook não disponível ({str(e_outlook)[:50]}).\n")
                    lq.put("   Usando servidor de e-mail alternativo.\n")
                    usar_outlook = False

                rows_iter = list(df_dados.iterrows())
                _anexo_cache = os.path.join(_app_base_dir(), "faturas_email_cache")
                os.makedirs(_anexo_cache, exist_ok=True)
                _anexo_cache_abs = os.path.abspath(_anexo_cache)

                _bloco_inicio = None
                _bloco_fim = None
                _bloco_count = 0

                for pos, (idx, row_series) in enumerate(rows_iter):
                    row = row_series.to_dict()
                    if state["stop_flag"]:
                        lq.put("⚠ Disparo interrompido pelo usuário.\n")
                        break
                        
                    status_atual = str(row.get("STATUS", "")).strip().upper()
                    if status_atual.startswith("ENVIADO") or status_atual == "PULADO MANUALMENTE":
                        pulados += 1
                        if _bloco_inicio is None:
                            _bloco_inicio = pos + 1
                        _bloco_fim = pos + 1
                        _bloco_count += 1
                        continue
                    else:
                        if _bloco_count > 0:
                            lq.put(f"  Linhas de {_bloco_inicio} até {_bloco_fim} foram puladas por já terem sido enviadas\n")
                            _bloco_inicio = None
                            _bloco_fim = None
                            _bloco_count = 0

                    dest_email = str(row.get(dest_col, "")).strip()
                    if not dest_email or "@" not in dest_email:
                        pulados += 1
                        lq.put(f"  Pulado linha {pos+1}: email inválido '{dest_email}'\n")
                        df_dados.at[idx, "STATUS"] = "Falhou: Email inválido"
                        continue

                    try:
                        from fatura_anexo_email import resolve_fatura_anexo_para_email

                        path_anexo, aviso_anexo = resolve_fatura_anexo_para_email(
                            row,
                            linha=pos + 1,
                            baixar_faturas=baixar_fatura_wf,
                            usuario_waterfy=wf_user,
                            senha_waterfy=wf_pass,
                            proteger_pdf=proteger_pdf,
                            destino_dir=_anexo_cache,
                        )
                        if aviso_anexo:
                            lq.put(f"  ⚠ [{pos+1}] {aviso_anexo}\n")

                        rendered  = engine.render_record(row)
                        r_subject = rendered["subject"]
                        r_body    = rendered["body"]

                        if usar_outlook:
                            # Envia via Outlook (HTML se o corpo parecer documento HTML)
                            mail = outlook.CreateItem(0)  # 0 = olMailItem
                            mail.To      = dest_email
                            mail.Subject = r_subject
                            if is_probably_html_document(r_body):
                                try:
                                    mail.BodyFormat = 2  # olFormatHTML
                                except Exception:
                                    pass
                                mail.HTMLBody = r_body
                            else:
                                mail.Body = r_body
                            if path_anexo and os.path.isfile(path_anexo):
                                mail.Attachments.Add(path_anexo)
                            mail.Send()
                            df_dados.at[idx, "STATUS"] = "Enviado"
                            if path_anexo and os.path.isfile(path_anexo):
                                ap = os.path.abspath(path_anexo)
                                if ap.startswith(_anexo_cache_abs + os.sep) or ap == _anexo_cache_abs:
                                    try:
                                        os.remove(path_anexo)
                                    except Exception:
                                        pass
                        else:
                            # SMTP fallback: apenas simula (não envia de verdade sem config)
                            if path_anexo and os.path.isfile(path_anexo):
                                lq.put(f"  [SIM] anexo: {path_anexo}\n")
                                ap = os.path.abspath(path_anexo)
                                if ap.startswith(_anexo_cache_abs + os.sep) or ap == _anexo_cache_abs:
                                    try:
                                        os.remove(path_anexo)
                                    except Exception:
                                        pass
                            lq.put(f"  [SIMULAÇÃO] {dest_email} (servidor de e-mail não configurado)\n")
                            df_dados.at[idx, "STATUS"] = "Enviado (Simulação)"

                        enviados += 1
                        lq.put(f"  ✓ [{pos+1}/{total}] {dest_email}\n")

                    except Exception as ex_send:
                        erros += 1
                        lq.put(f"  ✗ [{pos+1}/{total}] {dest_email} — {ex_send}\n")
                        df_dados.at[idx, "STATUS"] = f"Falhou: {ex_send}"

                    # Atualiza progress
                    pct = (pos + 1) / total if total > 0 else 1
                    prog.value = pct
                    lbl_status.value   = f"Enviando... {pos+1}/{total} ({enviados} ok, {erros} erros)"
                    lbl_contagem.value = f"Enviados: {enviados}  |  Erros: {erros}  |  Pulados: {pulados}"
                    try: prog.update(); lbl_status.update(); lbl_contagem.update()
                    except Exception: pass

                    # Delay entre envios
                    if delay_s > 0 and pos < len(rows_iter) - 1 and not state["stop_flag"]:
                        time.sleep(delay_s)

                    # Pausa de lote
                    if batch_size > 0 and (pos + 1) % batch_size == 0 and pos < len(rows_iter) - 1 and not state["stop_flag"]:
                        lq.put(f"  ⏸ Pausa de lote ({batch_pause}s)...\n")
                        time.sleep(batch_pause)

                if _bloco_count > 0:
                    lq.put(f"  Linhas de {_bloco_inicio} até {_bloco_fim} foram puladas por já terem sido enviadas\n")
                    _bloco_inicio = None
                    _bloco_fim = None
                    _bloco_count = 0

            except Exception as ex:
                import traceback
                lq.put(f"ERRO no disparo: {ex}\n{traceback.format_exc()}\n")

            finally:
                if pythoncom is not None:
                    try:
                        pythoncom.CoUninitialize()
                    except Exception:
                        pass

                # Salvar a planilha com os status atualizados
                try:
                    import pandas as pd
                    import shutil
                    file_path = state.get("csv_path")
                    if file_path and os.path.exists(file_path):
                        p = Path(file_path)
                        if p.suffix.lower() in [".xlsx", ".xls", ".xlsm"]:
                            all_sheets = pd.read_excel(file_path, sheet_name=None, engine="openpyxl", dtype=str)
                            sheet_to_update = ddl_sheet.value if ddl_sheet.visible and ddl_sheet.value else list(all_sheets.keys())[0]
                            all_sheets[sheet_to_update] = df_dados
                            
                            temp_caminho = str(p).replace(p.suffix, f"_TEMP{p.suffix}")
                            with pd.ExcelWriter(temp_caminho, engine="openpyxl") as writer:
                                for s_name, s_df in all_sheets.items():
                                    s_df.to_excel(writer, sheet_name=s_name, index=False)
                            
                            sucesso = False
                            for _ in range(5):
                                try:
                                    shutil.move(temp_caminho, file_path)
                                    sucesso = True
                                    break
                                except PermissionError:
                                    time.sleep(2)
                            if not sucesso:
                                raise PermissionError("Arquivo aberto em outro programa.")
                        else:
                            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                                header = f.readline()
                            sep = ";" if header.count(";") > header.count(",") else ","
                            
                            temp_caminho = str(p).replace(p.suffix, f"_TEMP{p.suffix}")
                            df_dados.to_csv(temp_caminho, sep=sep, index=False, encoding="utf-8-sig")
                            
                            sucesso = False
                            for _ in range(5):
                                try:
                                    shutil.move(temp_caminho, file_path)
                                    sucesso = True
                                    break
                                except PermissionError:
                                    time.sleep(2)
                            if not sucesso:
                                raise PermissionError("Arquivo aberto em outro programa.")
                                
                        lq.put("✓ Planilha atualizada com os status de envio.\n")
                except Exception as e_save:
                    lq.put(f"⚠ Não foi possível salvar o status na planilha: {e_save}\n")

                state["running"]   = False
                state["stop_flag"] = False

                prog.value         = 1 if enviados > 0 else 0
                prog.visible       = False
                lbl_status.value   = f"Concluído: {enviados} enviados, {erros} erros, {pulados} pulados."
                lbl_status.color   = GREEN if erros == 0 else ORANGE
                lbl_contagem.value = f"Enviados: {enviados}  |  Erros: {erros}  |  Pulados: {pulados}"

                if _btn_enviar_ref[0]:
                    _btn_enviar_ref[0].disabled = False
                    try: _btn_enviar_ref[0].update()
                    except Exception: pass
                if _btn_parar_ref[0]:
                    _btn_parar_ref[0].disabled = True
                    try: _btn_parar_ref[0].update()
                    except Exception: pass

                try: prog.update(); lbl_status.update(); lbl_contagem.update()
                except Exception: pass

                lq.put(f"\n{'═'*40}\n✓ Disparo finalizado: {enviados} enviados · {erros} erros · {pulados} pulados\n{'═'*40}\n")

                _snack(f"Disparo concluído! {enviados} emails enviados.", ok=erros == 0)

        threading.Thread(target=_thread, daemon=True).start()

    # ── Construção dos botões referenciados ───────────────────────────────────
    refresh_col_buttons()

    btn_enviar = pbtn("Enviar Emails", "SEND", on_click=iniciar_envio, width=200)
    btn_parar  = ft.Container(
        content=ft.Row([
            ft.Icon(ic("STOP_CIRCLE"), size=16, color=WHITE),
            ft.Container(width=8),
            ft.Text("Parar", size=13, weight=ft.FontWeight.BOLD, color=WHITE),
        ], spacing=0, alignment=ft.MainAxisAlignment.CENTER),
        height=44, width=130,
        bgcolor=RED, border_radius=8,
        padding=ft.Padding.symmetric(horizontal=20, vertical=0),
        on_click=parar_envio, ink=True, disabled=True,
        opacity=0.45,
        shadow=ft.BoxShadow(spread_radius=0, blur_radius=6,
                            color=ft.Colors.with_opacity(0.3, ft.Colors.BLACK),
                            offset=ft.Offset(0, 2)),
    )

    def _on_parar_disabled_change(e=None):
        btn_parar.opacity = 1.0 if not btn_parar.disabled else 0.45
        try: btn_parar.update()
        except Exception: pass

    _btn_enviar_ref[0] = btn_enviar
    _btn_parar_ref[0]  = btn_parar

    async def _watch_custom_html():
        while True:
            await asyncio.sleep(1.0)
            try:
                p = state.get("custom_html_path")
                if p and ddl_template.value == "custom" and not sw_manual_body.value and os.path.isfile(p):
                    current_mtime = os.path.getmtime(p)
                    last_mtime = state.get("custom_html_mtime", 0)
                    if last_mtime != 0 and current_mtime > last_mtime:
                        state["custom_html_mtime"] = current_mtime
                        if container_preview.visible:
                            gerar_preview(None)
                            _snack(f"Arquivo {os.path.basename(p)} salvo. Prévia atualizada!")
                    elif last_mtime == 0:
                        state["custom_html_mtime"] = current_mtime
            except Exception:
                pass

    page.run_task(_watch_custom_html)

    # ── Layout final ──────────────────────────────────────────────────────────
    return ft.Column([
        ft.Text("Mala Direta / Envio de Emails", size=22, weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
        ft.Text(
            "Envie mensagens personalizadas em massa. Use as colunas da sua planilha como variáveis (ex: {{NOME}}) para preencher os textos automaticamente.",
            size=13,
            color=TEXT_MUTED,
        ),
        ft.Container(height=8),

        # ── 1. Dados CSV ─────────────────────────────────────────────────────
        card(ft.Column([
            sec_title("1. Fonte de Dados (Planilha)", "TABLE_CHART"),
            ft.Container(height=10),
            ft.Row([
                ft.Column([lbl_arquivo, ft.Container(height=4), lbl_colunas], spacing=0, expand=True),
                ft.Container(width=10),
                sbtn("Selecionar CSV", "FOLDER", on_click=pick_csv),
            ], vertical_alignment=ft.CrossAxisAlignment.START),
            ft.Container(height=10),
            ft.Row([ddl_sheet], vertical_alignment=ft.CrossAxisAlignment.CENTER),
        ])),

        # ── 2. Outlook ───────────────────────────────────────────────────────
        card(ft.Column([
            sec_title("2. Integração com Outlook", "MAIL"),
            ft.Container(height=10),
            ft.Text(
                "O disparo é feito de forma segura através do Microsoft Outlook do seu computador. "
                "Certifique-se de que o aplicativo esteja aberto e conectado à sua conta.",
                size=12, color=TEXT_MUTED,
            ),
            ft.Container(height=10),
            ft.Row([
                sbtn("Verificar Conexão", "WIFI_TETHERING", on_click=testar_outlook),
            ]),
        ])),

        # ── 3. Controle de Envio ─────────────────────────────────────────────
        card(ft.Column([
            sec_title("3. Configurações de Disparo", "TUNE"),
            ft.Container(height=10),
            ft.Column([
                flabel("Qual coluna da planilha contém o e-mail do cliente? *"),
                txt_destinatario,
            ], spacing=4),
            ft.Container(height=10),
            ft.Row([txt_delay, ft.Container(width=12), txt_lote, ft.Container(width=12), txt_pausa_lote], spacing=0),
            ft.Container(height=4),
            ft.Text(
                "💡 Dica: Mantenha um intervalo de pelo menos 2 segundos para evitar bloqueios por excesso de envios no Outlook.",
                size=11, color=TEXT_MUTED, italic=True,
            ),
        ])),

        card(ft.Column([
            sec_title("Anexos e Faturas PDF", "ATTACH_FILE"),
            ft.Container(height=8),
            cb_baixar_faturas_email,
            credencias_fatura_email,
            ft.Container(height=6),
            cb_senha_pdf_fatura,
            ft.Container(height=4),
            ft.Text(
                "💡 Você também pode anexar arquivos salvos no seu computador: basta ter na planilha uma coluna chamada 'ARQUIVO_PDF' contendo o caminho do arquivo de cada cliente.",
                size=11,
                color=TEXT_MUTED,
            ),
        ])),

        # ── 4. Template ──────────────────────────────────────────────────────
        card(ft.Column([
            sec_title("4. Composição da Mensagem", "EDIT"),
            ft.Container(height=10),
            ft.Row([ddl_template, ft.Container(width=10), btn_pick_html], wrap=True, spacing=8),
            ft.Container(height=6),
            lbl_tpl_status,
            ft.Container(height=8),
            sw_manual_body,
            ft.Container(height=10),
            ft.Column([
                flabel("Assunto do e-mail *"),
                txt_assunto,
                ft.Container(height=6),
                flabel("Mensagem do e-mail"),
                txt_corpo,
            ], spacing=2),
            ft.Container(height=10),
            flabel("📌 Variáveis da sua planilha (clique em um botão abaixo para inserir a informação no texto):"),
            ft.Container(height=6),
            col_campos,
        ])),

        # ── 5. Preview ───────────────────────────────────────────────────────
        card(ft.Column([
            sec_title("5. Preview", "PREVIEW"),
            ft.Container(height=10),
            ft.Row([
                sbtn("Gerar Preview (1ª linha)", "VISIBILITY", on_click=gerar_preview),
            ]),
            ft.Container(height=10),
            container_preview,
        ])),

        # ── Botões de disparo ─────────────────────────────────────────────────
        ft.Row([btn_enviar, ft.Container(width=12), btn_parar], spacing=0),
        ft.Container(height=10),
        prog,
        ft.Container(height=4),
        lbl_status,
        lbl_contagem,
        ft.Container(height=24),

    ], spacing=0, scroll=ft.ScrollMode.AUTO, expand=True)


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main(page: ft.Page):
    page.title   = "Mathools 1.0 - Painel Corporativo | Itapoá Saneamento"
    page.bgcolor = SURFACE
    page.padding = 0
    page.theme   = ft.Theme(
        color_scheme_seed=ORANGE,
        tooltip_theme=ft.TooltipTheme(
            text_style=ft.TextStyle(
                color=WHITE,
                size=12,
                weight=ft.FontWeight.W_600,
            ),
            padding=ft.Padding.symmetric(horizontal=12, vertical=8),
            margin=ft.Margin.all(0),
            wait_duration=360,
            exit_duration=80,
            prefer_below=True,
            vertical_offset=26,
            decoration=ft.BoxDecoration(
                bgcolor="#163252",
                border_radius=ft.BorderRadius.all(10),
                border=ft.Border.all(1, ft.Colors.with_opacity(0.18, "#BFD8FF")),
                shadows=[
                    ft.BoxShadow(
                        spread_radius=0,
                        blur_radius=16,
                        color=ft.Colors.with_opacity(0.24, ft.Colors.BLACK),
                        offset=ft.Offset(0, 4),
                    )
                ],
            ),
        ),
    )

    # Expõe a page para o engine usar (save_fn)
    import sys as _sys
    _main_mod = _sys.modules.get("__main__")
    if _main_mod is not None:
        _main_mod._FLET_PAGE = page

    _ico_path = resource_path("Logo-mathey-tk-3.ico")

    # Ícone da janela e barra de tarefas (Flet 0.82)
    if os.path.exists(_ico_path):
        try:
            page.window.icon = _ico_path
        except Exception:
            try:
                page.window_icon = _ico_path
            except Exception:
                pass

    target_w = 900
    target_h = 650

    # API nova (Flet 0.80+)
    try:
        page.window.maximized  = False
        page.window.width      = target_w
        page.window.height     = target_h
        page.window.min_width  = 900
        page.window.min_height = 650
        page.update()
        try:
            page.window.center()
        except Exception:
            pass
    except Exception:
        # Fallback API antiga
        page.window_maximized  = False
        page.window_width      = target_w
        page.window_height     = target_h
        page.window_min_width  = 900
        page.window_min_height = 650
        page.update()

    lq: queue.Queue = queue.Queue()
    global _LOGIN_RESULT
    if _LOGIN_RESULT and isinstance(_LOGIN_RESULT, dict):
        USER = {
            "username":        _LOGIN_RESULT.get("username", "admin"),
            "role":            _LOGIN_RESULT.get("role", "ADM"),
            "allowed_modules": _LOGIN_RESULT.get("allowed_modules", "Todos"),
        }
    else:
        USER = {"username": "admin", "role": "ADM", "allowed_modules": "Todos"}

    user_prefs = load_user_prefs(USER["username"])
    if not user_prefs.get("display_name"):
        user_prefs["display_name"] = USER["username"]
        save_user_prefs(USER["username"], user_prefs)

    def _display_name() -> str:
        return str(user_prefs.get("display_name") or USER["username"])

    def _reports_dir() -> str:
        return str(user_prefs.get("reports_dir") or os.path.join(_app_base_dir(), "Relatorios"))

    def _spreadsheets_dir() -> str:
        return str(user_prefs.get("spreadsheets_dir") or os.path.join(_app_base_dir(), "Planilhas"))

    def _time_value(key: str, fallback: str) -> str:
        return str(user_prefs.get(key) or fallback)

    def _valid_file_or_empty(path_value: str) -> str:
        try:
            return path_value if path_value and os.path.exists(path_value) else ""
        except Exception:
            return ""

    def _build_avatar_content(size: int = 32):
        avatar_path = _valid_file_or_empty(str(user_prefs.get("avatar_path", "")))
        if avatar_path:
            return ft.Image(src=avatar_path, width=size, height=size, fit="cover")
        initial = (_display_name().strip()[:1] or USER["username"][:1] or "U").upper()
        return ft.Text(initial, size=max(12, int(size * 0.4)), weight=ft.FontWeight.BOLD, color=WHITE)

    def _parse_clock_hhmm(raw_value: str):
        try:
            hh, mm = str(raw_value).strip().split(":")
            hh_i = int(hh); mm_i = int(mm)
            if 0 <= hh_i <= 23 and 0 <= mm_i <= 59:
                return hh_i, mm_i
        except Exception:
            pass
        return None

    def _clock_targets():
        targets = []
        friday_enabled = bool(user_prefs.get("friday_custom_enabled", False))
        is_friday = time.localtime().tm_wday == 4
        for key, label, default in (
            ("time_08", "Entrada", "08:00"),
            ("time_12", "Almoco", "12:00"),
            ("time_13", "Retorno", "13:00"),
            ("time_18", "Saida", "18:00"),
        ):
            raw_value = _time_value(key, default)
            if key == "time_18" and is_friday and friday_enabled:
                raw_value = _time_value("time_18_friday", raw_value)
            parsed = _parse_clock_hhmm(raw_value)
            if parsed:
                targets.append((key, label, parsed[0], parsed[1]))
        return targets

    def _play_alert_sound():
        if winsound is None:
            return False
        sound_path = _valid_file_or_empty(str(user_prefs.get("alert_sound_path", "")))
        try:
            if sound_path:
                if not sound_path.lower().endswith(".wav"):
                    return False
                winsound.PlaySound(sound_path, winsound.SND_FILENAME | winsound.SND_ASYNC)
            else:
                winsound.MessageBeep(winsound.MB_ICONEXCLAMATION)
            return True
        except Exception:
            return False

    def _logout_to_login():
        try:
            if getattr(sys, "frozen", False):
                subprocess.Popen([sys.executable], cwd=_app_base_dir())
            else:
                subprocess.Popen([sys.executable, os.path.abspath(__file__)], cwd=_app_base_dir())
        except Exception:
            pass
        try:
            page.window_destroy()
        except Exception:
            try:
                page.window.destroy()
            except Exception:
                pass

    _ui_refs = {}
    _clock_alert_state = {"last_trigger": ""}

    def _show_snack(msg: str, color=GREEN):
        try:
            page.snack_bar = ft.SnackBar(ft.Text(msg, color=WHITE), bgcolor=color, open=True)
            page.update()
        except Exception:
            pass

    def _apply_user_preferences(show_feedback: bool = False):
        save_user_prefs(USER["username"], user_prefs)

        display_name = _display_name()
        if _ui_refs.get("sidebar_name"):
            _ui_refs["sidebar_name"].value = display_name
        if _ui_refs.get("topbar_user"):
            _ui_refs["topbar_user"].value = display_name
        if _ui_refs.get("avatar_container"):
            avatar_size = _ui_refs.get("avatar_size", 32)
            _ui_refs["avatar_container"].content = _build_avatar_content(avatar_size)
        if _ui_refs.get("clock_chip"):
            _ui_refs["clock_chip"].visible = bool(user_prefs.get("show_clock", True))

        for key in ("sidebar_name", "topbar_user", "avatar_container", "clock_chip"):
            control = _ui_refs.get(key)
            if control:
                try:
                    control.update()
                except Exception:
                    pass

        if show_feedback:
            _show_snack("Configuracoes salvas.", GREEN)

    # Busca número de operadores ativos no Supabase para o dashboard
    try:
        from auth_system import Authentication
        _auth_tmp = Authentication()
        _users = _auth_tmp.get_all_users()
        _num_op = len([u for u in _users if u.get("role") not in ("", None)])
    except Exception:
        _num_op = 0

    dash_col, _ref_csvs, _ref_msgs, _ref_props, _ref_paradas = build_dashboard(num_operadores=_num_op, lq=lq, page=page)

    def build_settings_page():
        def _make_settings_status():
            return ft.Container(
                visible=False,
                bgcolor=ft.Colors.with_opacity(0.08, GREEN),
                border_radius=10,
                border=ft.Border.all(1, ft.Colors.with_opacity(0.18, GREEN)),
                padding=ft.Padding.symmetric(horizontal=12, vertical=10),
                content=ft.Text("", size=12, color=TEXT_MAIN, weight=ft.FontWeight.W_500),
            )

        settings_status = _make_settings_status()
        settings_status_bottom = _make_settings_status()

        def _set_settings_status(message: str, color=GREEN):
            for target in (settings_status, settings_status_bottom):
                target.visible = True
                target.bgcolor = ft.Colors.with_opacity(0.08, color)
                target.border = ft.Border.all(1, ft.Colors.with_opacity(0.18, color))
                target.content.value = message
                target.content.color = color if color in (GREEN, RED, ORANGE) else TEXT_MAIN
                try:
                    target.update()
                except Exception:
                    pass

        def _log_settings(message: str):
            try:
                lq.put(f"[Configuracoes] {message}\n")
            except Exception:
                pass

        e_display_name = tfield("Nome exibido", value=_display_name())
        e_username = tfield("Usuario", value=USER["username"])
        e_username.read_only = True
        e_role = tfield("Cargo", value=USER["role"])
        e_role.read_only = True

        sw_alerts = ft.Switch(value=bool(user_prefs.get("alerts_enabled", False)), active_color=ORANGE)
        sw_workdays = ft.Switch(value=bool(user_prefs.get("workdays_only", True)), active_color=ORANGE)
        sw_show_clock = ft.Switch(value=bool(user_prefs.get("show_clock", True)), active_color=ORANGE)
        e_sound = tfield("Som de alarme", value=str(user_prefs.get("alert_sound_path", "")))
        e_reports = tfield("Pasta de relatorios", value=str(user_prefs.get("reports_dir", "")))
        e_spreadsheets = tfield("Pasta de planilhas", value=str(user_prefs.get("spreadsheets_dir", "")))

        sl_volume = ft.Slider(min=0, max=100, divisions=20, label="{value}%", value=float(user_prefs.get("alert_volume", 80)))
        sl_yellow = ft.Slider(min=5, max=60, divisions=11, label="{value} min", value=float(user_prefs.get("yellow_warning_minutes", 30)))
        sl_red = ft.Slider(min=1, max=30, divisions=29, label="{value} min", value=float(user_prefs.get("red_warning_minutes", 5)))

        e_time_08 = tfield("08:00", value=_time_value("time_08", "08:00"), width=120, expand=False)
        e_time_12 = tfield("12:00", value=_time_value("time_12", "12:00"), width=120, expand=False)
        e_time_13 = tfield("13:00", value=_time_value("time_13", "13:00"), width=120, expand=False)
        e_time_18 = tfield("18:00", value=_time_value("time_18", "18:00"), width=120, expand=False)
        cb_friday_custom = ft.Checkbox(
            label="Horario diferente na sexta",
            value=bool(user_prefs.get("friday_custom_enabled", False)),
            active_color=ORANGE,
        )
        e_time_18_friday = tfield("Saida de sexta", value=_time_value("time_18_friday", "17:00"), width=140, expand=False)
        friday_time_row = ft.Container(
            content=ft.Row([
                ft.Text("Saida de sexta", size=13, color=TEXT_MAIN, width=120),
                e_time_18_friday,
            ], spacing=12, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            visible=bool(cb_friday_custom.value),
        )

        avatar_preview = ft.Container(
            content=_build_avatar_content(72),
            width=72, height=72, bgcolor=ORANGE,
            border_radius=36, alignment=ft.Alignment(0, 0), clip_behavior=ft.ClipBehavior.ANTI_ALIAS,
        )

        def _collect_and_save(show_feedback: bool = True):
            invalid_fields = []
            for key, control, fallback in (
                ("time_08", e_time_08, "08:00"),
                ("time_12", e_time_12, "12:00"),
                ("time_13", e_time_13, "13:00"),
                ("time_18", e_time_18, "18:00"),
            ):
                raw_value = str(control.value or fallback).strip()
                if not _parse_clock_hhmm(raw_value):
                    invalid_fields.append(raw_value or fallback)
                else:
                    user_prefs[key] = raw_value

            if invalid_fields:
                _log_settings("Falha ao salvar: horario invalido. Use HH:MM.")
                _set_settings_status("Use horarios no formato HH:MM.", RED)
                _show_snack("Use horarios no formato HH:MM.", RED)

            if cb_friday_custom.value:
                friday_value = str(e_time_18_friday.value or "17:00").strip()
                if not _parse_clock_hhmm(friday_value):
                    _log_settings("Falha ao salvar: horario de sexta invalido. Use HH:MM.")
                    _set_settings_status("Use o horario de sexta no formato HH:MM.", RED)
                    _show_snack("Use o horario de sexta no formato HH:MM.", RED)
                    return
                user_prefs["time_18_friday"] = friday_value

            user_prefs["display_name"] = (e_display_name.value or USER["username"]).strip() or USER["username"]
            user_prefs["alerts_enabled"] = bool(sw_alerts.value)
            user_prefs["workdays_only"] = bool(sw_workdays.value)
            user_prefs["show_clock"] = bool(sw_show_clock.value)
            user_prefs["friday_custom_enabled"] = bool(cb_friday_custom.value)
            user_prefs["alert_sound_path"] = str(e_sound.value or "").strip()
            user_prefs["reports_dir"] = str(e_reports.value or _reports_dir()).strip()
            user_prefs["spreadsheets_dir"] = str(e_spreadsheets.value or _spreadsheets_dir()).strip()
            user_prefs["alert_volume"] = int(sl_volume.value or 80)
            user_prefs["yellow_warning_minutes"] = int(sl_yellow.value or 30)
            user_prefs["red_warning_minutes"] = int(sl_red.value or 5)
            _apply_user_preferences(show_feedback=show_feedback)
            if show_feedback:
                _log_settings("Configuracoes salvas com sucesso.")
                _set_settings_status("Configuracoes salvas com sucesso.", GREEN)

        def _update_avatar_preview():
            avatar_preview.content = _build_avatar_content(72)
            try:
                avatar_preview.update()
            except Exception:
                pass

        def _pick_file(filetypes):
            try:
                import tkinter as tk
                from tkinter import filedialog
                root = tk.Tk()
                root.withdraw()
                root.attributes("-topmost", True)
                path = filedialog.askopenfilename(filetypes=filetypes)
                root.destroy()
                return path
            except Exception:
                return ""

        def _pick_directory():
            try:
                import tkinter as tk
                from tkinter import filedialog
                root = tk.Tk()
                root.withdraw()
                root.attributes("-topmost", True)
                path = filedialog.askdirectory()
                root.destroy()
                return path
            except Exception:
                return ""

        def _choose_avatar(e=None):
            _log_settings("Abrindo seletor de foto do perfil...")
            path = _pick_file([("Imagens", "*.png *.jpg *.jpeg *.bmp *.webp"), ("Todos", "*.*")])
            if path:
                user_prefs["avatar_path"] = path
                _update_avatar_preview()
                _collect_and_save(show_feedback=True)

        def _choose_sound(e=None):
            _log_settings("Abrindo seletor de som do alarme...")
            path = _pick_file([("Audio WAV", "*.wav"), ("Todos", "*.*")])
            if path:
                if not path.lower().endswith(".wav"):
                    _log_settings("Som rejeitado: formato invalido. Use .wav.")
                    _set_settings_status("Formato aceito para alarme: .wav", ORANGE)
                    _show_snack("Use um arquivo .wav para o alarme.", ORANGE)
                    return
                e_sound.value = path
                try: e_sound.update()
                except Exception: pass
                _log_settings(f"Som selecionado: {os.path.basename(path)}")
                _collect_and_save(show_feedback=True)

        def _choose_reports_dir(e=None):
            _log_settings("Abrindo seletor de pasta de relatorios...")
            path = _pick_directory()
            if path:
                e_reports.value = path
                try: e_reports.update()
                except Exception: pass
                _log_settings(f"Pasta de relatorios selecionada: {path}")
                _collect_and_save(show_feedback=True)

        def _choose_spreadsheets_dir(e=None):
            _log_settings("Abrindo seletor de pasta de planilhas...")
            path = _pick_directory()
            if path:
                e_spreadsheets.value = path
                try: e_spreadsheets.update()
                except Exception: pass
                _log_settings(f"Pasta de planilhas selecionada: {path}")
                _collect_and_save(show_feedback=True)

        def _test_sound(e=None):
            _log_settings("Testando som do alarme...")
            _set_settings_status("Testando som do alarme...", ORANGE)
            _collect_and_save(show_feedback=False)
            if _play_alert_sound():
                _log_settings("Som reproduzido com sucesso.")
                _set_settings_status("Som reproduzido com sucesso.", GREEN)
                _show_snack("Som reproduzido.", GREEN)
            else:
                _log_settings("Falha ao reproduzir som. Use um arquivo .wav valido.")
                _set_settings_status("Nao foi possivel reproduzir o som. Use um arquivo .wav valido.", RED)
                _show_snack("Nao foi possivel reproduzir o som. Use um arquivo .wav.", RED)

        def _toggle_friday_custom(e=None):
            friday_time_row.visible = bool(cb_friday_custom.value)
            try:
                friday_time_row.update()
            except Exception:
                pass

        cb_friday_custom.on_change = _toggle_friday_custom

        def _check_updates_now(e=None):
            if not _HAS_UPDATER:
                _log_settings("Atualizador nao disponivel nesta versao.")
                _set_settings_status("Atualizador nao disponivel nesta versao.", ORANGE)
                _show_snack("Atualizador nao disponivel.", ORANGE)

            _log_settings("Verificando atualizacoes...")
            _set_settings_status("Verificando atualizacoes...", ORANGE)

            def _cb(info):
                def _finish():
                    if info and getattr(info, "has_update", False):
                        _pending_update_info[0] = info
                        _refresh_update_chip()
                        _log_settings(f"Atualizacao encontrada: v{info.latest_version}.")
                        _set_settings_status(f"Atualizacao disponivel: v{info.latest_version}. Clique no chip da topbar para atualizar.", ORANGE)
                        _show_snack(f"Atualizacao disponivel: v{info.latest_version}", ORANGE)
                    else:
                        _log_settings("Nenhuma atualizacao encontrada. Sistema ja atualizado.")
                        _set_settings_status("Nenhuma atualizacao encontrada. O sistema ja esta atualizado.", GREEN)
                        _show_snack("Sistema ja esta atualizado.", GREEN)
                try:
                    page.run_thread(_finish)
                except Exception:
                    _finish()
            check_update_async(_cb)

        return ft.Column([
            ft.Text("Configuracoes", size=22, weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
            ft.Text("Personalize perfil, alertas, horarios e caminhos padrao do sistema.", size=13, color=TEXT_MUTED),
            ft.Container(height=8),
            settings_status,
            ft.Container(height=8),
            card(ft.Column([
                sec_title("Perfil", "PERSON"),
                ft.Container(height=14),
                ft.Row([
                    avatar_preview,
                    ft.Container(width=16),
                    ft.Column([
                        flabel("Nome exibido"),
                        e_display_name,
                        ft.Container(height=8),
                        ft.Row([
                            sbtn("Trocar foto", "IMAGE", on_click=_choose_avatar),
                            ft.Container(width=8),
                            sbtn("Salvar perfil", "SAVE", on_click=lambda e: _collect_and_save(show_feedback=True)),
                        ], spacing=0),
                    ], expand=True, spacing=0),
                ], vertical_alignment=ft.CrossAxisAlignment.START),
                ft.Container(height=16),
                ft.Row([e_username, ft.Container(width=12), e_role], spacing=0),
            ])),
            card(ft.Column([
                sec_title("Alertas de ponto", "ALARM"),
                ft.Container(height=14),
                ft.Row([
                    ft.Text("Ativar alertas", size=13, color=TEXT_MAIN, expand=True),
                    sw_alerts,
                ]),
                ft.Row([
                    ft.Text("Somente dias uteis", size=13, color=TEXT_MAIN, expand=True),
                    sw_workdays,
                ]),
                ft.Row([
                    ft.Text("Mostrar relogio/data na topbar", size=13, color=TEXT_MAIN, expand=True),
                    sw_show_clock,
                ]),
                ft.Container(height=10),
                flabel("Som do alarme"),
                ft.Row([
                    e_sound,
                    ft.Container(width=8),
                    sbtn("Escolher som", "MUSIC_NOTE", on_click=_choose_sound),
                    ft.Container(width=8),
                    sbtn("Testar som", "PLAY_ARROW", on_click=_test_sound),
                ], spacing=0),
                ft.Text("Formato aceito para o alarme: .wav", size=11, color=TEXT_MUTED),
                ft.Container(height=12),
                flabel("Volume do alerta"),
                sl_volume,
                ft.Text("Observacao: a reproducao usa o volume do sistema no Windows; este controle fica salvo para personalizacao futura.", size=11, color=TEXT_MUTED),
                ft.Container(height=12),
                flabel("Horario de bater ponto"),
                ft.Text(
                    "O aviso 'Hora de bater o ponto' aparece apenas no minuto exato do horario escolhido.",
                    size=11,
                    color=TEXT_MUTED,
                ),
            ])),
            card(ft.Column([
                sec_title("Horarios", "SCHEDULE"),
                ft.Container(height=14),
                ft.Row([e_time_08, ft.Container(width=12), e_time_12, ft.Container(width=12), e_time_13, ft.Container(width=12), e_time_18], spacing=0),
                ft.Container(height=12),
                cb_friday_custom,
                ft.Container(height=8),
                friday_time_row,
                ft.Container(height=12),
                flabel("Antecedencia amarela"),
                sl_yellow,
                ft.Container(height=8),
                flabel("Antecedencia vermelha"),
                sl_red,
            ])),
            card(ft.Column([
                sec_title("Sistema e armazenamento", "SETTINGS"),
                ft.Container(height=14),
                check_row(f"Usuario logado: {USER['username']}"),
                check_row(f"Versao atual: {APP_VERSION}"),
                ft.Container(height=12),
                flabel("Pasta padrao para relatorios"),
                ft.Row([
                    e_reports,
                    ft.Container(width=8),
                    sbtn("Escolher pasta", "FOLDER", on_click=_choose_reports_dir),
                ], spacing=0),
                ft.Container(height=12),
                flabel("Pasta padrao para planilhas"),
                ft.Row([
                    e_spreadsheets,
                    ft.Container(width=8),
                    sbtn("Escolher pasta", "FOLDER_OPEN", on_click=_choose_spreadsheets_dir),
                ], spacing=0),
                ft.Container(height=16),
                settings_status_bottom,
                ft.Container(height=12),
                ft.Row([
                    sbtn("Salvar tudo", "SAVE", on_click=lambda e: _collect_and_save(show_feedback=True)),
                    ft.Container(width=8),
                    sbtn("Verificar atualizacao", "SYSTEM_UPDATE", on_click=_check_updates_now),
                    ft.Container(width=8),
                    sbtn("Sair da conta", "LOGOUT", on_click=lambda e: _logout_to_login()),
                ], spacing=0),
            ])),
        ], spacing=0, scroll=ft.ScrollMode.AUTO, expand=True)

    pages = [
        dash_col,
        build_csv_page(page, lq, ref_csvs=_ref_csvs),
        build_send_page(page, lq, ref_msgs=_ref_msgs),
        build_neg_page(page, lq, ref_props=_ref_props),
        build_routes_page(page, lq, ref_paradas=_ref_paradas),
        build_users_page(page),
        None,
        build_help_page(USER),
        build_settings_page(),
    ]
    SETTINGS_PAGE_INDEX = len(pages) - 1
    settings_panel = pages[SETTINGS_PAGE_INDEX]

    content_area = ft.Container(
        content=pages[0], expand=True,
        padding=ft.Padding.only(left=28, right=28, top=24, bottom=0),
    )

    settings_overlay = ft.Container(visible=False, expand=True)

    def close_settings_overlay(e=None):
        settings_overlay.visible = False
        try:
            page.update()
        except Exception:
            pass

    settings_overlay.content = ft.Row([
        ft.Container(expand=True, ink=True, on_click=close_settings_overlay),
        ft.Container(
            content=ft.Column([
                ft.Container(
                    content=ft.Row([
                        ft.Text("Configuracoes", size=18, weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
                        ft.Container(expand=True),
                        ft.IconButton(
                            icon=ic("CLOSE"),
                            icon_color=TEXT_MUTED,
                            tooltip="Fechar",
                            on_click=close_settings_overlay,
                        ),
                    ], spacing=0, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    padding=ft.Padding.symmetric(horizontal=20, vertical=16),
                    border=ft.Border.only(bottom=ft.BorderSide(1, BORDER)),
                ),
                ft.Container(
                    content=settings_panel,
                    expand=True,
                    padding=ft.Padding.only(left=20, right=20, top=16, bottom=20),
                ),
            ], spacing=0, expand=True),
            width=560,
            expand=False,
            bgcolor=ft.Colors.with_opacity(0.98, WHITE),
            border=ft.Border.only(left=ft.BorderSide(1, BORDER)),
            shadow=ft.BoxShadow(
                spread_radius=0, blur_radius=20,
                color=ft.Colors.with_opacity(0.18, ft.Colors.BLACK),
                offset=ft.Offset(-4, 0),
            ),
        ),
    ], expand=True, spacing=0)
    settings_overlay.bgcolor = ft.Colors.with_opacity(0.28, ft.Colors.BLACK)

    # ── terminal ──────────────────────────────────────────────────────────────
    _MAX_LOG_LINES  = 200   # linhas visíveis na tela
    _MAX_LOG_BUFFER = 2000  # buffer completo para o botão copiar
    _log_buffer: list[str] = []  # fonte de verdade para o copiar  nunca truncado pela UI

    log_list = ft.ListView(expand=True, spacing=1, auto_scroll=True)

    # ── filtro amigável ───────────────────────────────────────────────────────
    # Uma única regex OR compilada é muito mais rápida que iterar 30 padrões
    _SUPRIMIR_RE = re.compile(
        r"\bHTTP[/ ]\d+\b"
        r"|status[_\s]?code\b.*\d{3}"
        r"|\b(GET|POST|PUT|DELETE|PATCH)\b.*http"
        r"|requests\.get|requests\.post|response\."
        r"|<Response \["
        r"|(colunas|columns|dtype|dtypes|Index|RangeIndex)\s*[\[:\(\']"
        r"|\bDataFrame\b.*\bshape\b"
        r"|\bshape\s*[:=\(]\s*\(?\d+"
        r"|^Traceback \(most recent call last\)"
        r"|^\s*File \".*\", line \d+"
        r"|^\s{4,}[A-Za-z_]\w+\("
        r"|^\w+Error:|^\w+Exception:"
        r"|^\[DEBUG\]"
        r"|FutureWarning|DeprecationWarning|UserWarning"
        r"|pandas|numpy|openpyxl|zipfile|urllib|ssl"
        r"|\.py:\d+:"
        r"|[A-Za-z]:\\.*\.py"
        r"|/[\w/\.\-]+\.py\b"
        r"|^\s*[-=_\.]{4,}\s*$"
        r"|\bencoding\b.*\butf"
        r"|\bsep\s*=|delimiter\s*="
        r"|\bskiprows?\b|\bheader\s*="
        r"|\bchunksize\b|\bnrows\b"
        r"|^\s*\d+\s+\d+\s*$"
        r"|^\[UPDATER_LOG\]"
        r"|· URL:|/api/downloadFile/"
        r"|· Baixando arquivo|· Tipo de arquivo|· Arquivos no ZIP"
        r"|· CSV extraído|· Tentando decodif|· Decodificado|falhou, tentando"
        r"|Geração solicitada|Download conclu[ií]do \["
        r"|└─ Tamanho:|matrículas de OSs extraídas"
        r"|CSV salvo:|CSV Geral salvo:| CSV"
        r"|Equipes no cruzamento|Centro geográfico calculado:"
        r"|Matrícula: col \d+|col \d+ \|",
        re.I | re.MULTILINE,
    )

    _SUBSTITUIR = [
        (re.compile(r"^\[?\d{2}:\d{2}:\d{2}\]?\s*"), ""),
        (re.compile(r"^(INFO|DEBUG|WARNING|ERROR|CRITICAL)\s*[:\-]\s*", re.I), ""),
        ("ERRO LOCAL:", "ERRO Erro:"),
        ("ERRO:", "ERRO Erro:"),
        ("Erro:", "ERRO Erro:"),
        ("Erro ao", "ERRO Erro ao"),
        ("Motor RPA não encontrado", "Aviso Módulo de processamento não encontrado"),
        ("mathtools_1_0.py", ""),
        ("waterfy_engine.py", ""),
        ("Dashboard salvo em:", "OK Dashboard salvo"),
        ("[DEBUG]", ""),
    ]
    _log_ctx = {
        "secao_grafico": None,
        "resumo_final": False,
        "resumo_acima": "-",
        "resumo_abaixo": "-",
    }

    def _humanize(linha: str) -> str | None:
        m = linha.strip()
        if not m:
            return None
        # Uma única chamada de regex para suprimir  muito mais rápido
        if _SUPRIMIR_RE.search(m):
            return None
        if re.fullmatch(r"[─-]{10,}", m):
            return None
        for pat, rep in _SUBSTITUIR:
            if isinstance(pat, re.Pattern):
                m = pat.sub(rep, m).strip()
            elif pat in m:
                m = m.replace(pat, rep).strip()
        if "Registros carregados" in m:
            _log_ctx["resumo_final"] = True
            return "\nResumo final do processamento:"
        if _log_ctx["resumo_final"] and "limpeza e filtros" in m:
            return m
        if _log_ctx["resumo_final"] and "Acima do corte (" in m:
            return m
        if _log_ctx["resumo_final"] and "Abaixo do corte (" in m:
            return m
        if _log_ctx["resumo_final"] and "WhatsApp" in m:
            _log_ctx["resumo_final"] = False
            return f"{m}\n"
        if "Registros carregados" in m:
            _log_ctx["resumo_final"] = True
            return (
                "\n"
                "+------ RESUMO FINAL -------+\n"
                f"| Carregados  : {m.split(':', 1)[1].strip()} |"
            )
        if "Após limpeza e filtros" in m:
            return f"| Limpos      : {m.split(':', 1)[1].strip()} |"
        if _log_ctx["resumo_final"] and "Acima do corte (" in m:
            return f"| > corte     : {m.split(':', 1)[1].strip()} |"
        if _log_ctx["resumo_final"] and "Abaixo do corte (" in m:
            return f"| <= corte    : {m.split(':', 1)[1].strip()} |"
        if _log_ctx["resumo_final"] and "Só pelo WhatsApp" in m:
            _log_ctx["resumo_final"] = False
            return (
                f"| WhatsApp    : {m.split(':', 1)[1].strip()} |\n"
                "+---------------------------+"
            )
        if "Desenhando gráficos para: Maior que Corte" in m:
            _log_ctx["secao_grafico"] = "acima"
            return "\n──────── ACIMA DO CORTE ────────\n Clientes com atraso acima do corte:"

        if "Desenhando gráficos para: Menor ou Igual ao Corte" in m:
            _log_ctx["secao_grafico"] = "abaixo"
            return "\n──────── ABAIXO DO CORTE ────────\n Clientes com atraso dentro/abaixo do corte:"

        if m == "Gráfico salvo com sucesso.":
            if _log_ctx["secao_grafico"] == "acima":
                return "OK Gráfico do bloco acima do corte salvo."
            if _log_ctx["secao_grafico"] == "abaixo":
                return "OK Gráfico do bloco abaixo do corte salvo."

        if _log_ctx["secao_grafico"] in {"acima", "abaixo"}:
            if (
                m.startswith("Grupo com ")
                or m.startswith("Media de matriculas")
                or m.startswith("Valor total em aberto")
            ):
                return f"   {m}"

        if m.startswith("OK Gráficos gerados com sucesso."):
            _log_ctx["secao_grafico"] = None
            return f"{m}\n"

        return m if m else None

    def _humanize_better(linha: str) -> str | None:
        m = linha.strip()
        if not m:
            return None
        if _SUPRIMIR_RE.search(m):
            return None
        if re.fullmatch(r"[─-]{10,}", m):
            return None
        for pat, rep in _SUBSTITUIR:
            if isinstance(pat, re.Pattern):
                m = pat.sub(rep, m).strip()
            elif pat in m:
                m = m.replace(pat, rep).strip()

        if "o arquivo recebido não é um ZIP válido" in m and "Macro 8091" in m:
            m = re.sub(
                r"Erro ao processar\s+(\d{2}/\d{4}):\s*Macro 8091:.*",
                r"Erro ao processar \1: a macro 8091 retornou um arquivo inválido (não veio ZIP/CSV válido).",
                m,
                flags=re.I,
            )

        if "Resumo 8091:" in m:
            def _faixa_meses(_txt):
                _meses = [x.strip() for x in str(_txt).split(",") if x.strip()]
                if not _meses:
                    return ""
                return f"de {_meses[0]} ate {_meses[-1]}"

        if "Resumo 8091:" in m and "para_baixar=" in m:
            periodo_match = re.search(r"periodo=([^|]+)", m)
            periodo_txt = periodo_match.group(1).strip() if periodo_match else ""
            origem_match = re.search(r"origem=([^|]+)", m)
            origem_txt = origem_match.group(1).strip() if origem_match else "download"
            meses_match = re.search(r"para_baixar=(.+)$", m)
            meses_raw = meses_match.group(1).strip() if meses_match else ""
            faixa_txt = _faixa_meses(meses_raw)
            if faixa_txt:
                return (
                    f" Resumo 8091: status=preparando | periodo={periodo_txt} | "
                    f"origem={origem_txt} | {faixa_txt}"
                )
            return f" Resumo 8091: status=preparando | periodo={periodo_txt} | origem={origem_txt}"

        if "Resumo 8091:" in m and "reaproveitados=" in m and "baixados_agora=" not in m:
            periodo_match = re.search(r"periodo=([^|]+)", m)
            periodo_txt = periodo_match.group(1).strip() if periodo_match else ""
            origem_match = re.search(r"origem=([^|]+)", m)
            origem_txt = origem_match.group(1).strip() if origem_match else "cache"
            meses_match = re.search(r"reaproveitados=(.+)$", m)
            meses_raw = meses_match.group(1).strip() if meses_match else ""
            faixa_txt = _faixa_meses(meses_raw)
            if faixa_txt:
                return (
                    f" Resumo 8091: status=preparando | periodo={periodo_txt} | "
                    f"origem={origem_txt} | {faixa_txt}"
                )
            return f" Resumo 8091: status=preparando | periodo={periodo_txt} | origem={origem_txt}"

        if "Resumo 8091:" in m and "baixados_agora=" in m:
            periodo_match = re.search(r"periodo=([^|]+)", m)
            periodo_txt = periodo_match.group(1).strip() if periodo_match else ""
            origem_match = re.search(r"origem=([^|]+)", m)
            origem_txt = origem_match.group(1).strip() if origem_match else "misto(cache+download)"
            reap_match = re.search(r"reaproveitados=\d+\s*\[([^\]]*)\]", m)
            baix_match = re.search(r"baixados_agora=\d+\s*\[([^\]]*)\]", m)
            reap_faixa = _faixa_meses(reap_match.group(1).strip()) if reap_match else ""
            baix_faixa = _faixa_meses(baix_match.group(1).strip()) if baix_match else ""
            partes = []
            if reap_faixa:
                partes.append(f"cache {reap_faixa}")
            if baix_faixa:
                partes.append(f"download {baix_faixa}")
            sufixo = " | ".join(partes)
            if sufixo:
                return f" Resumo 8091: periodo={periodo_txt} | origem={origem_txt} | {sufixo}"
            return f" Resumo 8091: periodo={periodo_txt} | origem={origem_txt}"

        if "Registros carregados" in m:
            _log_ctx["resumo_final"] = True
            return "\n================ RESUMO FINAL ================\n" + f"[INFO] Registros carregados : {m.split(':', 1)[1].strip()}"
        if _log_ctx["resumo_final"] and "limpeza e filtros" in m:
            return f"[OK] Clientes validos     : {m.split(':', 1)[1].strip()}"
        if _log_ctx["resumo_final"] and "Acima do corte (" in m:
            valor = m.split(':', 1)[1].strip()
            _log_ctx["resumo_acima"] = valor
            return None
        if _log_ctx["resumo_final"] and "Abaixo do corte (" in m:
            valor = m.split(':', 1)[1].strip()
            _log_ctx["resumo_abaixo"] = valor
            return None
        if _log_ctx["resumo_final"] and "WhatsApp" in m:
            _log_ctx["resumo_final"] = False
            valor = m.split(':', 1)[1].strip()
            resumo_acima = _log_ctx.get("resumo_acima", "-")
            resumo_abaixo = _log_ctx.get("resumo_abaixo", "-")
            return (
                "\n+--------------- RESUMO DE CARTEIRA ---------------+\n"
                f"| Acima do corte (30 dias) : {resumo_acima}\n"
                f"| Abaixo do corte (30 dias): {resumo_abaixo}\n"
                f"| So pelo WhatsApp         : {valor}\n"
                "+--------------------------------------------------+"
            )
        if "Desenhando gráficos para: Maior que Corte" in m:
            _log_ctx["secao_grafico"] = "acima"
            return "\n================ ACIMA DO CORTE ================\n[INFO] Clientes com atraso acima do corte:"
        if "Desenhando gráficos para: Menor ou Igual ao Corte" in m:
            _log_ctx["secao_grafico"] = "abaixo"
            return "\n=============== ABAIXO DO CORTE ===============\n[INFO] Clientes com atraso dentro/abaixo do corte:"
        if m == "Gráfico salvo com sucesso.":
            if _log_ctx["secao_grafico"] == "acima":
                return "[OK] Grafico do bloco acima do corte salvo."
            if _log_ctx["secao_grafico"] == "abaixo":
                return "[OK] Grafico do bloco abaixo do corte salvo."
        if _log_ctx["secao_grafico"] in {"acima", "abaixo"}:
            if (
                m.startswith("Grupo com ")
                or m.startswith("Media de matriculas")
                or m.startswith("Valor total em aberto")
            ):
                return f"   • {m}"
        if m.startswith("✅ Gráficos gerados com sucesso.") or m.startswith("OK Gráficos gerados com sucesso."):
            _log_ctx["secao_grafico"] = None
            return "[OK] Graficos gerados com sucesso.\n"
        if m == "Planilha gerada com sucesso.":
            return "[OK] Planilha gerada com sucesso."
        if m in ("🏁 Tudo pronto! Processamento concluido com sucesso.", "[FIM] Tudo pronto! Processamento concluido com sucesso."):
            return "[FIM] Tudo pronto! Processamento concluido com sucesso."
        return m if m else None

    _humanize = _humanize_better

    def _clear_log(e):
        _log_buffer.clear()
        log_list.controls.clear()
        try: log_list.update()
        except Exception: pass

    _copy_btn = ft.Container(
        content=ft.Icon(ic("CONTENT_COPY"), size=14,
                        color=ft.Colors.with_opacity(0.4, TERM_GREEN)),
        tooltip="Copiar tudo",
        border_radius=6, ink=True, width=28, height=28,
        alignment=ft.Alignment(0, 0),
    )

    def _set_copy_icon(icon_name, color):
        _copy_btn.content = ft.Icon(ic(icon_name), size=14, color=color)
        try: _copy_btn.update()
        except: pass

    def _copy_log(e):
        async def _do_copy():
            try:
                if not _log_buffer:
                    _set_copy_icon("WARNING", ft.Colors.with_opacity(0.7, "#FF9800"))
                    await asyncio.sleep(1.5)
                    _set_copy_icon("CONTENT_COPY", ft.Colors.with_opacity(0.4, TERM_GREEN))
                    return

                texto = "\n".join(_log_buffer)
                copied = False

                try:
                    await page.clipboard.set(texto)
                    copied = True
                except Exception:
                    try:
                        import tkinter as tk
                        root = tk.Tk()
                        root.withdraw()
                        root.clipboard_clear()
                        root.clipboard_append(texto)
                        root.update()
                        root.destroy()
                        copied = True
                    except Exception:
                        copied = False

                if not copied:
                    raise RuntimeError("Falha ao copiar para a area de transferencia")

                _set_copy_icon("CHECK", ft.Colors.with_opacity(0.9, TERM_GREEN))
                await asyncio.sleep(2)
                _set_copy_icon("CONTENT_COPY", ft.Colors.with_opacity(0.4, TERM_GREEN))
            except Exception:
                _set_copy_icon("ERROR_OUTLINE", ft.Colors.with_opacity(0.8, RED))
                await asyncio.sleep(2)
                _set_copy_icon("CONTENT_COPY", ft.Colors.with_opacity(0.4, TERM_GREEN))
        page.run_task(_do_copy)

    _copy_btn.on_click = _copy_log

    terminal = ft.Container(
        content=ft.Column([
            ft.Row([
                ft.Icon(ic("COMPUTER"), color=TERM_GREEN, size=13),
                ft.Text("Atividade",
                        color=TERM_GREEN, size=11, weight=ft.FontWeight.BOLD),
                ft.Container(expand=True),
                _copy_btn,
                ft.Container(
                    content=ft.Icon(ic("DELETE_SWEEP"), size=14,
                                    color=ft.Colors.with_opacity(0.4, TERM_GREEN)),
                    tooltip="Limpar",
                    border_radius=6, ink=True, width=28, height=28,
                    alignment=ft.Alignment(0, 0),
                    on_click=_clear_log,
                ),
            ], spacing=8),
            ft.Divider(color=ft.Colors.with_opacity(0.15, WHITE), height=1),
            ft.Container(content=log_list, height=115),
        ], spacing=6),
        bgcolor=TERM_BG,
        border_radius=ft.BorderRadius.only(top_left=12, top_right=12),
        padding=ft.Padding.symmetric(horizontal=18, vertical=10),
        border=ft.Border.only(top=ft.BorderSide(1, ft.Colors.with_opacity(0.12, WHITE))),
    )

    _last_poll_activity = [0.0]   # timestamp da última vez que poll() atualizou a UI

    def poll():
        while True:
            try:
                first = lq.get(timeout=0.1)
            except Exception:
                continue
            batch = [first]
            while True:
                try:
                    batch.append(lq.get_nowait())
                except Exception:
                    break
            try:
                texto = "".join(batch)
                novos = 0
                for linha in texto.split("\n"):
                    raw = linha.strip()
                    if not raw:
                        continue
                    if raw.startswith("__MSG_SENT__"):
                        try: increment_stat("mensagens")
                        except: pass
                        continue
                    if raw.startswith("__PROGRESSO__:"):
                        try:
                            partes = raw.split(":")
                            enviados, falhas, total = int(partes[1]), int(partes[2]), int(partes[3])
                            _fn = getattr(page, "_wpp_counter_fn", None)
                            _st = getattr(page, "_wpp_contagem", None)
                            if _st is not None:
                                _st["enviados"] = enviados
                                _st["falhas"]   = falhas
                                _st["total"]    = total
                            if _fn: _fn()
                        except: pass
                        continue
                    if raw.startswith("__WPP_STAGE__:"):
                        try:
                            payload = raw.split(":", 1)[1]
                            partes = payload.split("|", 2)
                            pct = float(partes[0]) if len(partes) > 0 and partes[0] else 0.0
                            label = partes[1] if len(partes) > 1 else ""
                            detail = partes[2] if len(partes) > 2 else ""
                            _stage_fn = getattr(page, "_wpp_stage_fn", None)
                            if _stage_fn:
                                _stage_fn(pct, label, detail)
                        except: pass
                        continue
                    if raw.startswith("__TIMER__:"):
                        try:
                            secs = int(raw.split(":", 1)[1])
                            _timer_fn = getattr(page, "_wpp_timer_fn", None)
                            if _timer_fn:
                                _timer_fn(secs)
                        except: pass
                        continue
                    amigavel = _humanize(raw)
                    if amigavel is None:
                        continue
                    _log_buffer.append(amigavel)
                    if len(_log_buffer) > _MAX_LOG_BUFFER:
                        del _log_buffer[: len(_log_buffer) - _MAX_LOG_BUFFER]
                    log_list.controls.append(
                        ft.Text(amigavel, color=TERM_GREEN, size=11,
                                selectable=True, font_family="monospace",
                                no_wrap=False, overflow=ft.TextOverflow.VISIBLE)
                    )
                    novos += 1
                if novos:
                    if len(log_list.controls) > _MAX_LOG_LINES:
                        log_list.controls = log_list.controls[-_MAX_LOG_LINES:]
                    log_list.update()
                    _last_poll_activity[0] = time.time()
            except Exception:
                pass

    threading.Thread(target=poll, daemon=True).start()

    # ── Auto-refresh do Dashboard a cada 2 segundos ──────────────────────────
    def _dashboard_refresh():
        async def _update_dashboard():
            _ultimo_stats = {}
            while True:
                await asyncio.sleep(2)
                try:
                    # Evita trabalho de UI desnecessário quando o dashboard não está visível.
                    if content_area.content is not dash_col:
                        continue
                    _s_novo = load_stats()
                    # Só processa se os valores mudaram desde a última leitura
                    if _s_novo == _ultimo_stats:
                        continue
                    _ultimo_stats = _s_novo.copy()
                    mudou = False

                    if _ref_csvs[0] is not None:
                        novo_val = str(_s_novo["csvs"])
                        if _ref_csvs[0].value != novo_val:
                            _ref_csvs[0].value = novo_val
                            _ref_csvs[0].update()
                            mudou = True

                    if _ref_msgs[0] is not None:
                        novo_val = str(_s_novo.get("mensagens", 0))
                        if _ref_msgs[0].value != novo_val:
                            _ref_msgs[0].value = novo_val
                            _ref_msgs[0].update()
                            mudou = True

                    if _ref_props[0] is not None:
                        novo_val = str(_s_novo["propostas"])
                        if _ref_props[0].value != novo_val:
                            _ref_props[0].value = novo_val
                            _ref_props[0].update()
                            mudou = True

                    if _ref_paradas[0] is not None:
                        novo_val = str(_s_novo.get("paradas", 0))
                        if _ref_paradas[0].value != novo_val:
                            _ref_paradas[0].value = novo_val
                            _ref_paradas[0].update()
                            mudou = True

                    if mudou:
                        page.update()
                except Exception:
                    pass

        page.run_task(_update_dashboard)

    _dashboard_refresh()

    # ── sidebar ───────────────────────────────────────────────────────────────
    _role = USER["role"]
    _mods = USER.get("allowed_modules", "Todos")
    def _can(mod):
        return _role == "ADM" or _mods == "Todos" or mod in _mods

    NAV = [
        ("GRID_VIEW",           "Dashboard",     True),
        ("TABLE_CHART",         "Processar CSV", _can("Análise de CSV")),
        ("CHAT",                "WhatsApp",      _can("Disparo WhatsApp")),
        ("ATTACH_MONEY",        "Negociação",    _can("Gestão de Negociação")),
        ("TRAVEL_EXPLORE",      "Logística",     _can("Logística e Rotas")),
        ("MANAGE_ACCOUNTS",     "Gestão de acessos", _role == "ADM"),
        ("EMAIL",               "Disparo de E-mails", True),
        ("HELP",                "Ajuda",         True),
    ]

    sb_items: list[SidebarItem] = []
    pg_map:   list[int]         = []
    EMAIL_PAGE_INDEX = 6
    def on_select(sb_idx: int):
        for i, si in enumerate(sb_items):
            si.set_selected(i == sb_idx)
        abs_idx = pg_map[sb_idx]
        if abs_idx == EMAIL_PAGE_INDEX and not isinstance(pages[abs_idx], ft.Control):
            pages[abs_idx] = build_email_page(page, lq)
        content_area.content = pages[abs_idx]
        try:
            page.update()
        except Exception: pass

    def open_settings_page(e=None):
        settings_overlay.visible = True
        try:
            page.update()
        except Exception:
            pass

    for pg_i, (icon_name, lbl, visible) in enumerate(NAV):
        if visible:
            item = SidebarItem(icon_name, lbl, len(sb_items), on_select, page=page)
            sb_items.append(item)
            pg_map.append(pg_i)

    sb_items[0].set_selected(True)

    W_COLL = 62
    _ui_refs["sidebar_name"] = None

    sidebar_col = ft.Column([
        ft.Container(
            content=ft.Container(
                content=ft.Image(
                    src=resource_path("Logo mathey tk 1.png"),
                    width=28, height=28, fit="contain",
                ),
                width=36, height=36,
                bgcolor=ft.Colors.with_opacity(0.07, WHITE),
                border_radius=10,
                border=ft.Border.all(1, ft.Colors.with_opacity(0.10, WHITE)),
                shadow=ft.BoxShadow(
                    spread_radius=0, blur_radius=8,
                    color=ft.Colors.with_opacity(0.14, ft.Colors.BLACK),
                    offset=ft.Offset(0, 2),
                ),
                alignment=ft.Alignment(0, 0),
            ),
            padding=ft.Padding.symmetric(vertical=18),
            alignment=ft.Alignment(0, 0),
            tooltip="Mathools",
        ),
        ft.Divider(color=ft.Colors.with_opacity(0.1, WHITE), height=1),
        ft.Container(height=6),
        *[si.control for si in sb_items],
        ft.Container(expand=True),
        ft.Divider(color=ft.Colors.with_opacity(0.1, WHITE), height=1),
        ft.Container(
            content=ft.Container(
                content=_build_avatar_content(32),
                width=32, height=32, bgcolor=ORANGE,
                border_radius=16, alignment=ft.Alignment(0, 0),
                clip_behavior=ft.ClipBehavior.ANTI_ALIAS,
            ),
            padding=ft.Padding.symmetric(vertical=14),
            alignment=ft.Alignment(0, 0),
            ink=True,
            on_click=open_settings_page,
            tooltip="Abrir configuracoes",
        ),
    ], spacing=0, expand=True)
    _ui_refs["avatar_container"] = sidebar_col.controls[-1].content
    _ui_refs["avatar_size"] = 32

    sidebar_container = ft.Container(
        content=sidebar_col,
        bgcolor=NAVY_DARK, width=W_COLL,
        expand=False,
        border=ft.Border.only(right=ft.BorderSide(1, ft.Colors.with_opacity(0.07, WHITE))),
    )
    sidebar = sidebar_container

    # ── Banner de atualização (aparece se houver nova versão) ────────────────
    topbar_clock = ft.Text("--:--", size=11,
                           color=ft.Colors.with_opacity(0.82, WHITE),
                           weight=ft.FontWeight.W_500)
    update_status_dot = ft.Container(width=9, height=9, border_radius=5)
    update_status_text = ft.Text("Sistema atualizado", size=11,
                                 color=ft.Colors.with_opacity(0.82, WHITE),
                                 weight=ft.FontWeight.W_500)
    update_status_chip = ft.Container(
        content=ft.Row([
            update_status_dot,
            update_status_text,
        ], spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER),
        padding=ft.Padding.symmetric(horizontal=12, vertical=8),
        border_radius=999,
        margin=ft.Margin.only(right=18),
    )
    _pending_update_info = [_LAST_UPDATE_INFO]

    def _close_app_after_update():
        def _force_exit_later():
            import time as _time
            _time.sleep(2.5)
            os._exit(0)

        import threading as _thr
        _thr.Thread(target=_force_exit_later, daemon=True).start()

        def _try_close_window():
            try:
                page.window_destroy()
                return
            except Exception:
                pass
            try:
                page.window.destroy()
                return
            except Exception:
                pass
            try:
                window_close = getattr(page, "window_close", None)
                if callable(window_close):
                    window_close()
                    return
            except Exception:
                pass

        _thr.Thread(target=_try_close_window, daemon=True).start()

    def _format_file_size(num_bytes):
        try:
            size = float(max(0, int(num_bytes or 0)))
        except Exception:
            size = 0.0
        size_mb = size / (1024.0 * 1024.0)
        return f"{size_mb:.1f} MB"

    def _run_update_flow(info):
        prog_bar = ft.ProgressBar(
            width=380,
            value=0,
            color="#7CFC00",
            bgcolor=ft.Colors.with_opacity(0.2, WHITE),
        )
        prog_bytes = ft.Text(
            "0 B / calculando...",
            size=12,
            color=ft.Colors.with_opacity(0.82, WHITE),
        )
        prog_pct = ft.Text(
            "0%",
            size=12,
            color=WHITE,
            weight=ft.FontWeight.BOLD,
        )
        prog_status = ft.Text(
            "Preparando download da atualizacao...",
            size=12,
            color=ft.Colors.with_opacity(0.7, WHITE),
        )
        dlg_prog = ft.AlertDialog(
            modal=True,
            title=ft.Row([
                ft.Icon(ic("SYSTEM_UPDATE"), color="#7CFC00", size=20),
                ft.Container(width=8),
                ft.Text(f"Baixando v{info.latest_version}...",
                        size=14, weight=ft.FontWeight.BOLD, color=WHITE),
            ], spacing=0),
            content=ft.Column([
                prog_bar,
                ft.Row([
                    prog_bytes,
                    ft.Container(expand=True),
                    prog_pct,
                ], spacing=10),
                prog_status,
            ], spacing=10, tight=True),
            bgcolor=NAVY,
        )
        page.overlay.append(dlg_prog)
        dlg_prog.open = True
        try: page.update()
        except: pass

        def _run():
            def _on_progress(downloaded, total):
                downloaded = max(0, int(downloaded or 0))
                total = max(0, int(total or 0))
                if total > 0:
                    progress_value = min(1.0, downloaded / total)
                    progress_pct = f"{int(round(progress_value * 100))}%"
                    progress_bytes = f"{_format_file_size(downloaded)} / {_format_file_size(total)}"
                    progress_status = "Baixando atualizacao..."
                else:
                    progress_value = None
                    progress_pct = "..."
                    progress_bytes = f"{_format_file_size(downloaded)} / tamanho desconhecido"
                    progress_status = "Baixando atualizacao (calculando tamanho)..."

                prog_bar.value = progress_value
                prog_bytes.value = progress_bytes
                prog_pct.value = progress_pct
                prog_status.value = progress_status
                try:
                    page.update()
                except Exception:
                    pass

            ok = apply_update(info, progress_cb=_on_progress)
            def _finish():
                dlg_prog.open = False
                try: page.update()
                except: pass
                if ok:
                    page.overlay.append(ft.AlertDialog(
                        modal=True,
                        title=ft.Text("Reiniciando...", color=WHITE,
                                      weight=ft.FontWeight.BOLD),
                        content=ft.Text(
                            "Download concluido! O programa sera reiniciado agora.",
                            size=13, color=ft.Colors.with_opacity(0.8, WHITE),
                        ),
                        bgcolor=NAVY,
                        open=True,
                    ))
                    try: page.update()
                    except: pass
                    import time as _time; _time.sleep(1.5)
                    _close_app_after_update()
                else:
                    page.overlay.append(ft.AlertDialog(
                        modal=True,
                        title=ft.Text("Erro no download", color="#FF4B4B",
                                      weight=ft.FontWeight.BOLD),
                        content=ft.Text("Falha ao baixar a atualizacao.\nTente novamente mais tarde.",
                                        size=13, color=ft.Colors.with_opacity(0.8, WHITE)),
                        bgcolor=NAVY,
                        open=True,
                    ))
                    try: page.update()
                    except: pass
            page.run_thread(_finish) if hasattr(page, "run_thread") else page.after(0, _finish)

        import threading as _thr
        _thr.Thread(target=_run, daemon=True).start()

    def _on_update_chip_click(e):
        info = _pending_update_info[0]
        if info and getattr(info, "has_update", False):
            _run_update_flow(info)

    def _refresh_update_chip():
        has_pending = bool(_pending_update_info[0] and getattr(_pending_update_info[0], "has_update", False))
        update_status_dot.bgcolor = "#F59E0B" if has_pending else "#39D98A"
        update_status_dot.shadow = ft.BoxShadow(
            spread_radius=0, blur_radius=8,
            color=ft.Colors.with_opacity(0.40, "#F59E0B" if has_pending else "#39D98A"),
            offset=ft.Offset(0, 0),
        )
        update_status_text.value = "Atualizacao Pendente" if has_pending else "Sistema atualizado"
        update_status_chip.ink = has_pending
        update_status_chip.on_click = _on_update_chip_click if has_pending else None
        update_status_chip.tooltip = "Clique para atualizar agora" if has_pending else "Aplicativo na versao mais recente"
        update_status_chip.bgcolor = ft.Colors.with_opacity(0.12 if has_pending else 0.07, WHITE)
        update_status_chip.border = ft.Border.all(1, ft.Colors.with_opacity(0.18 if has_pending else 0.10, WHITE))
        try:
            update_status_chip.update()
        except Exception:
            pass

    def _set_pending_update_info(info):
        global _LAST_UPDATE_INFO
        _LAST_UPDATE_INFO = info
        _pending_update_info[0] = info
        _refresh_update_chip()

    _refresh_update_chip()

    async def _update_topbar_clock():
        while True:
            try:
                now = time.localtime()
                current_minutes = now.tm_hour * 60 + now.tm_min
                weekday = now.tm_wday
                show_clock = bool(user_prefs.get("show_clock", True))
                alerts_enabled = bool(user_prefs.get("alerts_enabled", False))
                workdays_only = bool(user_prefs.get("workdays_only", True))
                yellow_minutes = int(user_prefs.get("yellow_warning_minutes", 30) or 30)
                red_minutes = int(user_prefs.get("red_warning_minutes", 5) or 5)

                state = "normal"
                label = time.strftime("%d/%m/%Y %H:%M")
                trigger_key = ""

                should_evaluate_clock = (not workdays_only or weekday < 5)
                if should_evaluate_clock:
                    for key, action_label, hh, mm in _clock_targets():
                        target_minutes = hh * 60 + mm
                        delta = target_minutes - current_minutes
                        if delta == 0:
                            state = "exact"
                            label = f"Hora de bater o ponto  {action_label}"
                            trigger_key = f"{key}-{now.tm_year}-{now.tm_yday}-{now.tm_hour}-{now.tm_min}"
                            break
                        if 0 < delta <= red_minutes and state != "exact":
                            state = "red"
                            label = f"{time.strftime('%d/%m/%Y %H:%M')}  {action_label}"
                        elif 0 < delta <= yellow_minutes and state == "normal":
                            state = "yellow"
                            label = f"{time.strftime('%d/%m/%Y %H:%M')}  {action_label}"

                if alerts_enabled and state == "exact" and trigger_key and _clock_alert_state["last_trigger"] != trigger_key:
                    _clock_alert_state["last_trigger"] = trigger_key
                    _play_alert_sound()

                topbar_clock.value = label
                if state == "exact":
                    topbar_clock.color = "#FFE082" if now.tm_sec % 2 == 0 else "#FFCDD2"
                elif state == "red":
                    topbar_clock.color = "#FFCDD2"
                elif state == "yellow":
                    topbar_clock.color = "#FFE082"
                else:
                    topbar_clock.color = ft.Colors.with_opacity(0.82, WHITE)

                if _ui_refs.get("clock_chip"):
                    _ui_refs["clock_chip"].visible = show_clock
                    _ui_refs["clock_chip"].bgcolor = (
                        ft.Colors.with_opacity(0.16, "#B3261E") if state == "red"
                        else ft.Colors.with_opacity(0.16, "#F59E0B") if state in ("yellow", "exact")
                        else ft.Colors.with_opacity(0.07, WHITE)
                    )
                    _ui_refs["clock_chip"].border = ft.Border.all(
                        1,
                        ft.Colors.with_opacity(
                            0.28 if state in ("yellow", "red", "exact") else 0.10,
                            "#F59E0B" if state in ("yellow", "exact") else "#B3261E" if state == "red" else WHITE,
                        ),
                    )
                    _ui_refs["clock_chip"].update()

                topbar_clock.update()
            except Exception:
                pass
            await asyncio.sleep(1)

    page.run_task(_update_topbar_clock)

    update_banner = ft.Container(visible=False, height=0)

    def _show_update_banner_flet(info):
        """Exibe o banner de atualização no topo da janela Flet."""
        nonlocal update_banner

        def _do_update_flet(e):
            """Inicia o download e aplica a atualização."""
            dlg_prog = ft.AlertDialog(
                modal=True,
                title=ft.Row([
                    ft.Icon(ic("SYSTEM_UPDATE"), color="#7CFC00", size=20),
                    ft.Container(width=8),
                    ft.Text(f"Baixando v{info.latest_version}...",
                            size=14, weight=ft.FontWeight.BOLD, color=WHITE),
                ], spacing=0),
                content=ft.Column([
                    ft.ProgressBar(width=380, color="#7CFC00", bgcolor=ft.Colors.with_opacity(0.2, WHITE)),
                    ft.Text("Aguarde, baixando atualização...", size=12,
                            color=ft.Colors.with_opacity(0.7, WHITE)),
                ], spacing=10, tight=True),
                bgcolor=NAVY,
            )
            page.overlay.append(dlg_prog)
            dlg_prog.open = True
            try: page.update()
            except: pass

            def _run():
                ok = apply_update(info)
                def _finish():
                    dlg_prog.open = False
                    try: page.update()
                    except: pass
                    if ok:
                        page.overlay.append(ft.AlertDialog(
                            modal=True,
                            title=ft.Text("Reiniciando...", color=WHITE,
                                          weight=ft.FontWeight.BOLD),
                            content=ft.Text(
                                "Download concluído! O programa será reiniciado agora.",
                                size=13, color=ft.Colors.with_opacity(0.8, WHITE),
                            ),
                            bgcolor=NAVY,
                            open=True,
                        ))
                        try: page.update()
                        except: pass
                        import time as _time; _time.sleep(1.5)
                        _close_app_after_update()
                    else:
                        page.overlay.append(ft.AlertDialog(
                            modal=True,
                            title=ft.Text("Erro no download", color="#FF4B4B",
                                          weight=ft.FontWeight.BOLD),
                            content=ft.Text("Falha ao baixar a atualização.\nTente novamente mais tarde.",
                                            size=13, color=ft.Colors.with_opacity(0.8, WHITE)),
                            bgcolor=NAVY,
                            open=True,
                        ))
                        try: page.update()
                        except: pass
                page.run_thread(_finish) if hasattr(page, "run_thread") else page.after(0, _finish)

            import threading as _thr
            _thr.Thread(target=_run, daemon=True).start()

        def _dismiss_banner(e):
            update_banner.visible = False
            update_banner.height  = 0
            try: page.update()
            except: pass

        update_banner.content = ft.Row([
            ft.Icon(ic("SYSTEM_UPDATE"), color="#7CFC00", size=18),
            ft.Container(width=8),
            ft.Text(
                f"  Nova versão {info.latest_version} disponível!  "
                + (f" {info.release_notes}" if info.release_notes else ""),
                size=12, weight=ft.FontWeight.BOLD, color="#7CFC00", expand=True,
            ),
            ft.TextButton(
                "ATUALIZAR AGORA",
                on_click=lambda e: _run_update_flow(info),
                style=ft.ButtonStyle(
                    color=WHITE,
                    bgcolor={"": "#2E7D32", "hovered": "#1B5E20"},
                    padding=ft.Padding.symmetric(horizontal=14, vertical=6),
                    shape=ft.RoundedRectangleBorder(radius=6),
                ),
            ),
            ft.Container(width=6),
            ft.IconButton(
                icon=ic("CLOSE"),
                icon_color=ft.Colors.with_opacity(0.5, WHITE),
                icon_size=16,
                on_click=_dismiss_banner,
                tooltip="Dispensar",
            ),
        ], vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=0)
        update_banner.bgcolor = "#0d2e0d"
        update_banner.height  = 40
        update_banner.visible = True
        update_banner.padding = ft.Padding.symmetric(horizontal=16, vertical=0)
        update_banner.border  = ft.Border.only(
            bottom=ft.BorderSide(1, ft.Colors.with_opacity(0.3, "#7CFC00"))
        )
        try: page.update()
        except: pass

    _last_update_banner_version = [None]

    # Dispara verificação em background ao abrir o Flet
    if _HAS_UPDATER:
        def _on_update_check_flet(info):
            if info and info.has_update:
                _set_pending_update_info(info)
                return
                # Evita re-render repetido do mesmo banner na checagem periódica.
                if _last_update_banner_version[0] == str(getattr(info, "latest_version", "")):
                    return
                _last_update_banner_version[0] = str(getattr(info, "latest_version", ""))
                try:
                    page.run_thread(lambda: _show_update_banner_flet(info))
                except Exception:
                    import threading as _thr
                    _thr.Thread(
                        target=lambda: _show_update_banner_flet(info),
                        daemon=True
                    ).start()
                return
            _last_update_banner_version[0] = None
            _set_pending_update_info(None)

        def _periodic_update_check():
            # Primeira verificação imediata na abertura.
            check_update_async(_on_update_check_flet)
            # Rechecagem automática para não depender de reiniciar o app.
            while True:
                try:
                    time.sleep(180)
                    check_update_async(_on_update_check_flet)
                except Exception:
                    # Mantém o loop vivo mesmo se houver falha temporária.
                    time.sleep(15)

        import threading as _thr
        _thr.Thread(target=_periodic_update_check, daemon=True, name="UpdatePoller").start()

    topbar = ft.Row([
        ft.Container(
            width=W_COLL,
            height=52,
            bgcolor=NAVY_DARK,
            border=ft.Border.only(
                right=ft.BorderSide(1, ft.Colors.with_opacity(0.07, WHITE)),
                bottom=ft.BorderSide(1, ft.Colors.with_opacity(0.12, WHITE)),
            ),
            content=ft.Container(
                content=ft.Icon(ic("MENU"), size=18,
                                color=ft.Colors.with_opacity(0.92, WHITE)),
                width=40, height=40, border_radius=10, ink=True,
                bgcolor=ft.Colors.with_opacity(0.08, WHITE),
                border=ft.Border.all(1, ft.Colors.with_opacity(0.10, WHITE)),
                shadow=ft.BoxShadow(
                    spread_radius=0, blur_radius=10,
                    color=ft.Colors.with_opacity(0.16, ft.Colors.BLACK),
                    offset=ft.Offset(0, 2),
                ),
                alignment=ft.Alignment(0, 0),
                tooltip=_sidebar_tooltip("Menu lateral"),
            ),
            alignment=ft.Alignment(0, 0),
        ),
        ft.Container(
            expand=True,
            height=52,
            bgcolor=NAVY,
            gradient=ft.LinearGradient(
                begin=ft.Alignment(-1, 0),
                end=ft.Alignment(1, 0),
                colors=[NAVY_DARK, NAVY],
            ),
            border=ft.Border.only(bottom=ft.BorderSide(1, ft.Colors.with_opacity(0.12, WHITE))),
            shadow=ft.BoxShadow(
                spread_radius=0, blur_radius=14,
                color=ft.Colors.with_opacity(0.12, ft.Colors.BLACK),
                offset=ft.Offset(0, 3),
            ),
            content=ft.Row([
                ft.Container(
                    content=ft.Column([
                        ft.Text(
                            "Itapoá Saneamento",
                            size=13, color=WHITE,
                            weight=ft.FontWeight.W_600,
                        ),
                        ft.Text(
                            "Gestão e Processamento RPA",
                            size=11, color=ft.Colors.with_opacity(0.58, WHITE),
                        ),
                    ], spacing=2, alignment=ft.MainAxisAlignment.CENTER),
                    padding=ft.Padding.only(left=18),
                    height=40,
                    alignment=ft.Alignment(0, 0),
                ),
                ft.Container(expand=True),
                ft.Container(
                    content=ft.Row([
                        ft.Icon(ic("PERSON"), size=16, color=ft.Colors.with_opacity(0.88, WHITE)),
                        ft.Text(_display_name(), size=11,
                                color=ft.Colors.with_opacity(0.88, WHITE),
                                weight=ft.FontWeight.W_600),
                    ], spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    padding=ft.Padding.symmetric(horizontal=12, vertical=8),
                    bgcolor=ft.Colors.with_opacity(0.07, WHITE),
                    border_radius=999,
                    border=ft.Border.all(1, ft.Colors.with_opacity(0.10, WHITE)),
                    margin=ft.Margin.only(right=10),
                ),
                ft.Container(
                    content=ft.Row([
                        ft.Icon(ic("BADGE"), size=16, color=ft.Colors.with_opacity(0.88, WHITE)),
                        ft.Text(USER["role"], size=11,
                                color=ft.Colors.with_opacity(0.88, WHITE),
                                weight=ft.FontWeight.W_600),
                    ], spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    padding=ft.Padding.symmetric(horizontal=12, vertical=8),
                    bgcolor=ft.Colors.with_opacity(0.07, WHITE),
                    border_radius=999,
                    border=ft.Border.all(1, ft.Colors.with_opacity(0.10, WHITE)),
                    margin=ft.Margin.only(right=10),
                ),
                ft.Container(
                    content=ft.Row([
                        ft.Icon(ic("SCHEDULE"), size=16, color=ft.Colors.with_opacity(0.88, WHITE)),
                        topbar_clock,
                    ], spacing=8, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    padding=ft.Padding.symmetric(horizontal=12, vertical=8),
                    bgcolor=ft.Colors.with_opacity(0.07, WHITE),
                    border_radius=999,
                    border=ft.Border.all(1, ft.Colors.with_opacity(0.10, WHITE)),
                    margin=ft.Margin.only(right=10),
                ),
                update_status_chip,
            ], vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=0),
        ),
    ], spacing=0)
    _ui_refs["topbar_user"] = topbar.controls[1].content.controls[2].content.controls[1]
    _ui_refs["clock_chip"] = topbar.controls[1].content.controls[4]
    _apply_user_preferences(show_feedback=False)

    # ── Diálogo de confirmação ao fechar ─────────────────────────────────────
    def _destroy():
        try: page.window_destroy()
        except Exception:
            try: page.window.destroy()
            except Exception: pass

    def _on_window_event(e):
        # Flet 0.82: e.data é a string do evento ("close", "minimize", etc)
        data = str(getattr(e, "data", "") or "").strip().lower()
        if data != "close":
            return
        save_fn = getattr(page, "_wpp_save_fn", None)
        if save_fn is None:
            # Nenhum disparo ativo  fecha direto
            _destroy()
            return

        # Disparo em andamento  pergunta o que fazer
        def _salvar_e_sair(ev):
            dlg.open = False
            try: page.update()
            except: pass
            try: save_fn()
            except: pass
            _destroy()

        def _sair_sem_salvar(ev):
            dlg.open = False
            try: page.update()
            except: pass
            _destroy()

        def _cancelar(ev):
            dlg.open = False
            try: page.update()
            except: pass

        dlg = ft.AlertDialog(
            modal=True,
            bgcolor=WHITE,
            title=ft.Row([
                ft.Icon(ic("WARNING_AMBER"), color=ORANGE, size=20),
                ft.Container(width=8),
                ft.Text("Disparo em andamento", size=15,
                        weight=ft.FontWeight.BOLD, color=TEXT_MAIN),
            ], spacing=0),
            content=ft.Text(
                "Há um disparo de WhatsApp em andamento.\n"
                "Deseja salvar o progresso antes de sair?",
                size=13, color=TEXT_MAIN,
            ),
            actions=[
                ft.TextButton(
                    "Cancelar",
                    on_click=_cancelar,
                    style=ft.ButtonStyle(color=NAVY),
                ),
                ft.TextButton(
                    "Sair sem salvar",
                    on_click=_sair_sem_salvar,
                    style=ft.ButtonStyle(color=RED),
                ),
                ft.Container(
                    content=ft.Text("Salvar e Sair", size=13,
                                    weight=ft.FontWeight.BOLD, color=WHITE),
                    bgcolor=ORANGE, border_radius=8,
                    padding=ft.Padding.symmetric(horizontal=16, vertical=8),
                    on_click=_salvar_e_sair, ink=True,
                ),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        page.overlay.append(dlg)
        dlg.open = True
        try: page.update()
        except: pass

    # Flet 0.82 usa page.on_window_event e page.window_prevent_close
    try:
        page.window_prevent_close = True
    except Exception:
        try: page.window.prevent_close = True
        except Exception: pass
    try:
        page.on_window_event = _on_window_event
    except Exception:
        try: page.window.on_event = _on_window_event
        except Exception: pass

    main_body_row = ft.Row([
        sidebar,
        ft.Column([
            content_area,
            terminal,
        ], expand=True, spacing=0),
    ], expand=True, spacing=0,
       vertical_alignment=ft.CrossAxisAlignment.STRETCH)
    page.add(
        ft.Stack([
            ft.Column([
                topbar,
                main_body_row,
            ], spacing=0, expand=True),
            settings_overlay,
        ], expand=True)
    )




def run_with_login():
    """Abre o login Tkinter e, se autenticado, abre o painel Flet."""
    try:
        sys.path.insert(0, str(Path(__file__).parent))
        from LOGIN_GUI import LoginWindow
    except Exception as e:
        # LGPD: erros de importação não exibidos no console
        try:
            ft.run(main=main, assets_dir=os.path.dirname(resource_path("Logo-mathey-tk-3.ico")))
        except Exception as app_error:
            print(f"Erro ao abrir aplicação: {app_error}", file=sys.stderr)
        return

    try:
        login = LoginWindow()
        login.mainloop()  # Bloqueia aqui até o login ser fechado
        
        result = login.login_result
        if not result:
            # Usuário fechou sem logar
            return
        
        # Passa o resultado do login para o painel Flet via variável global
        global _LOGIN_RESULT
        _LOGIN_RESULT = result
        
        # Agora abre a aplicação Flet DEPOIS que o login terminou
        _ico = resource_path("Logo-mathey-tk-3.ico")
        ft.run(main=main, assets_dir=os.path.dirname(_ico))
            
    except Exception as e:
        # Log do erro para diagnóstico
        print(f"Erro em run_with_login: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc(file=sys.stderr)


_LOGIN_RESULT = None

if __name__ == "__main__":
    # Configurar logging (console + arquivo)
    setup_logging_to_file()
    logging.info("╔════════════════════════════════════╗")
    logging.info("║  MATHOOLS 1.0 - INICIANDO          ║")
    logging.info("╚════════════════════════════════════╝")
    run_with_login()