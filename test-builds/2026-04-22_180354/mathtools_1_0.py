print("""


                                                                                    ██╗
                                                                                   ██╔╝
                                                                                  ██╔╝

                                        ██╗████████╗ █████╗ ██████╗  ██████╗  █████╗ 
                                        ██║╚══██╔══╝██╔══██╗██╔══██╗██╔═══██╗██╔══██╗
                                        ██║   ██║   ███████║██████╔╝██║   ██║███████║
                                        ██║   ██║   ██╔══██║██╔═══╝ ██║   ██║██╔══██║
                                        ██║   ██║   ██║  ██║██║     ╚██████╔╝██║  ██║
                                        ╚═╝   ╚═╝   ╚═╝  ╚═╝╚═╝      ╚═════╝ ╚═╝  ╚═╝

                    ███████╗ █████╗ ███╗   ██╗███████╗ █████╗ ███╗   ███╗███████╗███╗   ██╗████████╗ ██████╗ 
                    ██╔════╝██╔══██╗████╗  ██║██╔════╝██╔══██╗████╗ ████║██╔════╝████╗  ██║╚══██╔══╝██╔═══██╗
                    ███████╗███████║██╔██╗ ██║█████╗  ███████║██╔████╔██║█████╗  ██╔██╗ ██║   ██║   ██║   ██║
                    ╚════██║██╔══██║██║╚██╗██║██╔══╝  ██╔══██║██║╚██╔╝██║██╔══╝  ██║╚██╗██║   ██║   ██║   ██║
                    ███████║██║  ██║██║ ╚████║███████╗██║  ██║██║ ╚═╝ ██║███████╗██║ ╚████║   ██║   ╚██████╔╝
                    ╚══════╝╚═╝  ╚═╝╚═╝  ╚═══╝╚══════╝╚═╝  ╚═╝╚═╝     ╚═╝╚══════╝╚═╝  ╚═══╝   ╚═╝    ╚═════╝
      
"""
)

import sys
import io
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
import re
import traceback
import logging
import pandas as pd
from logging.handlers import RotatingFileHandler
import os
import time
import random  
import unicodedata
import concurrent.futures
from pathlib import Path
from typing import Optional, List
PDF_BACKEND = None
try:
    from pypdf import PdfReader, PdfWriter
    HAS_PYPDF = True
    PDF_BACKEND = "pypdf"
except ImportError:
    HAS_PYPDF = False
from decimal import Decimal
from tqdm import tqdm
import matplotlib
matplotlib.use("Agg")          # backend não-interativo — obrigatório em threads
import matplotlib.pyplot as plt
import seaborn as sns
import math

# Integração com download de faturas
try:
    from download_fatura_pdf import (
        fazer_login as fazer_login_waterfy,
        download_pdf,
        parse_curl_file,
        detect_ids_in_payload,
        substitute_ids,
    )
    HAS_FATURA_DOWNLOADER = True
except ImportError:
    HAS_FATURA_DOWNLOADER = False
    fazer_login_waterfy = None
    download_pdf = None

# msvcrt só existe no CPython "oficial" do Windows — não disponível no Python
# da Microsoft Store nem em ambientes empacotados pelo PyInstaller nesse setup.
try:
    import msvcrt as _msvcrt
    _HAS_MSVCRT = True
except ImportError:
    _msvcrt = None
    _HAS_MSVCRT = False

# win32gui para fechar janelas nativas do Windows
try:
    import win32gui
    import win32con
    _HAS_WIN32 = True
except ImportError:
    _HAS_WIN32 = False

# Funcionalidades Excel para o módulo de negociação
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# ============================================================
# MÓDULO: RELATÓRIO CADASTRAL 8162
# Validação de contatos e geração de aba com gráficos
# ============================================================

DDDS_VALIDOS_8162 = {
    11, 12, 13, 14, 15, 16, 17, 18, 19, 21, 22, 24, 27, 28, 31, 32, 33, 34,
    35, 37, 38, 41, 42, 43, 44, 45, 46, 47, 48, 49, 51, 53, 54, 55, 61, 62,
    63, 64, 65, 66, 67, 68, 69, 71, 73, 74, 75, 77, 79, 81, 82, 83, 84, 85,
    86, 87, 88, 89, 91, 92, 93, 94, 95, 96, 97, 98, 99
}

def _8162_checar_movel(numero):
    if pd.isna(numero): return False
    num_limpo = re.sub(r'\D', '', str(numero))
    if len(num_limpo) == 11:
        ddd = int(num_limpo[:2])
        if ddd in DDDS_VALIDOS_8162 and num_limpo[2] == '9':
            return True
    return False

def _8162_checar_fixo(numero):
    if pd.isna(numero): return False
    num_limpo = re.sub(r'\D', '', str(numero))
    if len(num_limpo) == 10:
        ddd = int(num_limpo[:2])
        if ddd in DDDS_VALIDOS_8162 and num_limpo[2] in ['2', '3', '4', '5']:
            return True
    return False

def _8162_checar_email(email):
    if pd.isna(email): return False
    email_str = str(email).strip()
    padrao = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    return bool(re.match(padrao, email_str))

def _8162_tem_dado(valor):
    if pd.isna(valor): return False
    return str(valor).strip() != ''


def _normalizar_chave_matricula(valor) -> str:
    """
    Normaliza a matrícula para cruzamento entre bases.
    Evita falsos negativos como '12345' vs '12345.0' e remove ruído simples.
    """
    if pd.isna(valor):
        return ""
    texto = str(valor).strip()
    if not texto or texto.lower() == "nan":
        return ""
    if re.fullmatch(r"\d+\.0+", texto):
        return texto.split(".", 1)[0]
    return texto


def gerar_aba_relatorio_cadastral(df_8162: pd.DataFrame, writer):
    """
    Recebe o DataFrame bruto da macro 8162 e o ExcelWriter (engine=openpyxl).
    Gera a aba:
      - 'Relatorio_Cadastral': métricas BRUTOS + VALIDADOS com gráficos matplotlib embutidos
    Não depende de xlsxwriter — usa apenas openpyxl e matplotlib.
    """
    import io as _io
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment as OxAlign
    from openpyxl.utils import get_column_letter as _gcl
    from openpyxl.drawing.image import Image as OxImage

    # ── Detectar colunas ──────────────────────────────────────────────────────
    def _find_col(df, candidates):
        for c in candidates:
            if c in df.columns:
                return c
        return None

    col_fixo  = _find_col(df_8162, ['TEL. FIXO CONSUMIDOR', 'TEL_FIXO_CONSUMIDOR',
                                     'TELEFONE_FIXO', 'FIXO', 'TEL_FIXO'])
    col_email = _find_col(df_8162, ['EMAIL CONSUMIDOR', 'EMAIL_CONSUMIDOR',
                                     'EMAIL', 'E_MAIL'])
    col_movel = _find_col(df_8162, ['TEL. MOVEL CONSUMIDOR', 'TEL_MOVEL_CONSUMIDOR',
                                     'TELEFONE_CELULAR', 'CELULAR', 'TEL_CEL',
                                     'TEL_MOVEL', 'MOVEL'])
    col_mat   = _find_col(df_8162, ['MATRICULA', 'LIGACAO', 'MATRICULA',
                                     'LIGACAO', 'ID_LIGACAO'])

    if col_mat is None:
        logger.warning("[CADASTRAL] Coluna MATRICULA nao encontrada na 8162 — aba ignorada.")
        return

    # ── Séries booleanas ─────────────────────────────────────────────────────
    s_vf = df_8162[col_fixo].apply(_8162_checar_fixo)   if col_fixo  else pd.Series(False, index=df_8162.index)
    s_ve = df_8162[col_email].apply(_8162_checar_email)  if col_email else pd.Series(False, index=df_8162.index)
    s_vm = df_8162[col_movel].apply(_8162_checar_movel)  if col_movel else pd.Series(False, index=df_8162.index)

    s_bf = df_8162[col_fixo].apply(_8162_tem_dado)   if col_fixo  else pd.Series(False, index=df_8162.index)
    s_be = df_8162[col_email].apply(_8162_tem_dado)  if col_email else pd.Series(False, index=df_8162.index)
    s_bm = df_8162[col_movel].apply(_8162_tem_dado)  if col_movel else pd.Series(False, index=df_8162.index)

    qtd_v = s_vf.astype(int) + s_ve.astype(int) + s_vm.astype(int)
    qtd_b = s_bf.astype(int) + s_be.astype(int) + s_bm.astype(int)
    total = len(df_8162)

    # ── Regra de cadastro:
    #    Completo   = ao menos 1 e-mail + ao menos 1 telefone
    #    Parcial    = só e-mail OU só telefone
    #    Incompleto = nenhum dado
    tem_tel_b = s_bf | s_bm
    tem_tel_v = s_vf | s_vm

    status_b = pd.Series('Incompleto', index=df_8162.index)
    status_b[(s_be & ~tem_tel_b) | (~s_be & tem_tel_b)] = 'Parcial'
    status_b[s_be & tem_tel_b] = 'Completo'

    status_v = pd.Series('Incompleto', index=df_8162.index)
    status_v[(s_ve & ~tem_tel_v) | (~s_ve & tem_tel_v)] = 'Parcial'
    status_v[s_ve & tem_tel_v] = 'Completo'

    # ── Métricas ──────────────────────────────────────────────────────────────
    bcomp = int((status_b == 'Completo').sum())
    bparc = int((status_b == 'Parcial').sum())
    binc  = int((status_b == 'Incompleto').sum())
    bsum_e = int(s_be.sum()); bsum_m = int(s_bm.sum()); bsum_f = int(s_bf.sum())

    comp = int((status_v == 'Completo').sum())
    parc = int((status_v == 'Parcial').sum())
    inc  = int((status_v == 'Incompleto').sum())
    sum_e = int(s_ve.sum()); sum_m = int(s_vm.sum()); sum_f = int(s_vf.sum())

    dup_e_b = int(df_8162[s_be][col_email].astype(str).str.lower().duplicated().sum()) if col_email else 0
    dup_m_b = int(df_8162[s_bm][col_movel].astype(str).str.replace(r'\D', '', regex=True).duplicated().sum()) if col_movel else 0
    dup_e_v = int(df_8162[s_ve][col_email].astype(str).str.lower().duplicated().sum()) if col_email else 0
    dup_m_v = int(df_8162[s_vm][col_movel].astype(str).str.replace(r'\D', '', regex=True).duplicated().sum()) if col_movel else 0

    def pct(q):
        try:
            return f"{(float(q) / total) * 100:.2f}%" if total > 0 else ''
        except Exception:
            return ''

    # ── Montar tabela de métricas ─────────────────────────────────────────────
    rot_b = [
        '>>> DADOS BRUTOS (campos preenchidos)',
        'Total de Clientes na Base', '---',
        'COM 3 CAMPOS PREENCHIDOS', 'COM 2 CAMPOS PREENCHIDOS', 'COM 1 OU NENHUM CAMPO', '---',
        'APENAS E-mail Preenchido', 'APENAS Celular Preenchido', 'APENAS Fixo Preenchido', '---',
        'TOTAL E-mails Preenchidos', 'TOTAL Celulares Preenchidos', 'TOTAL Fixos Preenchidos', '---',
        'E-mails Repetidos (bruto)', 'Celulares Repetidos (bruto)',
    ]
    qtd_b_vals = [
        '', total, '',
        bcomp, bparc, binc, '',
        int((s_be & ~s_bm & ~s_bf).sum()), int((~s_be & s_bm & ~s_bf).sum()), int((~s_be & ~s_bm & s_bf).sum()), '',
        bsum_e, bsum_m, bsum_f, '',
        dup_e_b, dup_m_b,
    ]
    rot_v = [
        '>>> DADOS VALIDADOS (formato correto)',
        'Total de Clientes na Base', '---',
        'CADASTRO COMPLETO', 'CADASTRO PARCIAL', 'CADASTRO INCOMPLETO', '---',
        'APENAS E-mail Valido', 'APENAS Celular Valido', 'APENAS Fixo Valido', '---',
        'TOTAL E-mails Validos', 'TOTAL Celulares Validos', 'TOTAL Fixos Validos', '---',
        'E-mails Repetidos', 'Celulares Repetidos',
    ]
    qtd_v_vals = [
        '', total, '',
        comp, parc, inc, '',
        int((s_ve & ~s_vm & ~s_vf).sum()), int((~s_ve & s_vm & ~s_vf).sum()), int((~s_ve & ~s_vm & s_vf).sum()), '',
        sum_e, sum_m, sum_f, '',
        dup_e_v, dup_m_v,
    ]

    todos_rot = rot_b + ['', ''] + rot_v
    todos_qtd = qtd_b_vals + ['', ''] + qtd_v_vals
    todos_pct = [pct(q) for q in todos_qtd]
    OFFSET_VALID = len(rot_b) + 2   # = 19

    df_resumo = pd.DataFrame({
        'Metricas Detalhadas': todos_rot,
        'Quantidade':          todos_qtd,
        'Representacao (%)':   todos_pct,
    })

    # ── Escrever abas ─────────────────────────────────────────────────────────
    df_resumo.to_excel(writer, sheet_name='Relatorio_Cadastral', index=False)

    wb       = writer.book
    ws_res   = writer.sheets['Relatorio_Cadastral']

    # ── Helpers de estilo openpyxl ────────────────────────────────────────────
    def _fill(hex_bg):
        return PatternFill("solid", fgColor=hex_bg.lstrip('#'))

    def _font(hex_fg, bold=False, size=11):
        return Font(color=hex_fg.lstrip('#'), bold=bold, size=size)

    def _border(w='thin'):
        s = Side(style=w)
        return Border(left=s, right=s, top=s, bottom=s)

    _align_center = OxAlign(horizontal='center', vertical='center', wrap_text=True)

    fill_header     = _fill('#1F497D')
    font_header     = _font('#FFFFFF', bold=True)
    fill_tit_bruto  = _fill('#D6E4F7')
    font_tit_bruto  = _font('#1F497D', bold=True, size=13)
    fill_tit_valid  = _fill('#E2EFDA')
    font_tit_valid  = _font('#276221', bold=True, size=13)
    fill_sep        = _fill('#808080')
    brd             = _border()

    # ── Cabeçalho aba Relatorio_Cadastral ─────────────────────────────────────
    for col_i, col_nome in enumerate(df_resumo.columns, start=1):
        cell = ws_res.cell(row=1, column=col_i)
        cell.fill      = fill_header
        cell.font      = font_header
        cell.border    = brd
        cell.alignment = _align_center
        ws_res.column_dimensions[_gcl(col_i)].width = 42

    # ── Título seção BRUTOS (linha 2) ─────────────────────────────────────────
    ws_res.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    c = ws_res.cell(row=2, column=1)
    c.value     = '  DADOS BRUTOS  —  campos preenchidos (sem validacao)'
    c.fill      = fill_tit_bruto
    c.font      = font_tit_bruto
    c.border    = brd
    c.alignment = _align_center

    # ── Título seção VALIDADOS ────────────────────────────────────────────────
    row_tit_v = OFFSET_VALID + 2   # +2 = cabeçalho(1) + linha título(1)
    ws_res.merge_cells(start_row=row_tit_v, start_column=1, end_row=row_tit_v, end_column=3)
    c = ws_res.cell(row=row_tit_v, column=1)
    c.value     = '  DADOS VALIDADOS  —  formato correto (DDD + digito + email valido)'
    c.fill      = fill_tit_valid
    c.font      = font_tit_valid
    c.border    = brd
    c.alignment = _align_center

    # ── Linhas de separação cinza ─────────────────────────────────────────────
    rs1 = len(rot_b) + 2    # Excel row (1-based + cabeçalho)
    rs2 = rs1 + 1
    for sep_row in (rs1, rs2):
        for col_i in range(1, 4):
            ws_res.cell(row=sep_row, column=col_i).fill = fill_sep
        ws_res.row_dimensions[sep_row].height = 6

    # ── Gerar gráficos com matplotlib e embutir como PNG ─────────────────────
    COR_COMP = '#107C10'
    COR_PARC = '#FFC000'
    COR_INC  = '#C00000'

    labels_pizza = ['Completo', 'Parcial', 'Incompleto']

    # ── Gerar 1 PNG combinado com os 3 gráficos empilhados verticalmente ──────
    def _gerar_painel_combinado():
        fig = plt.figure(figsize=(9, 15))
        gs  = fig.add_gridspec(3, 1, hspace=0.55)

        # --- Pizza BRUTOS ---
        ax1 = fig.add_subplot(gs[0])
        vals_b = [bcomp, bparc, binc]
        vals_b_plot = [v for v in vals_b if v > 0]
        labs_b_plot = [l for l, v in zip(labels_pizza, vals_b) if v > 0]
        cols_b = [c for c, v in zip([COR_COMP, COR_PARC, COR_INC], vals_b) if v > 0]
        if vals_b_plot:
            ax1.pie(vals_b_plot, labels=labs_b_plot, colors=cols_b,
                    autopct='%1.1f%%', startangle=90,
                    wedgeprops={'edgecolor': 'white', 'linewidth': 1.2})
        ax1.set_title('Cobertura de Campos (Bruto)', fontsize=12, fontweight='bold', pad=10)

        # --- Barra empilhada BRUTOS vs VALIDADOS ---
        ax2 = fig.add_subplot(gs[1])
        grupos = ['Brutos', 'Validados']
        v_comp = [bcomp, comp]; v_parc = [bparc, parc]; v_inc = [binc, inc]
        y = range(len(grupos))
        ax2.barh(y, v_comp, color=COR_COMP, label='Completo', height=0.5)
        ax2.barh(y, v_parc, left=v_comp, color=COR_PARC, label='Parcial', height=0.5)
        ax2.barh(y, v_inc,  left=[a+b for a,b in zip(v_comp, v_parc)],
                 color=COR_INC, label='Incompleto', height=0.5)
        ax2.set_yticks(list(y)); ax2.set_yticklabels(grupos, fontsize=10)
        ax2.set_xlabel('Quantidade de cadastros', fontsize=9)
        ax2.set_title('Distribuicao dos Cadastros sobre o Total da Base',
                      fontsize=12, fontweight='bold', pad=10)
        ax2.legend(loc='upper center', bbox_to_anchor=(0.5, -0.18),
                   ncol=3, fontsize=9, frameon=True)
        for i, (vc, vp, vi) in enumerate(zip(v_comp, v_parc, v_inc)):
            for val, offset in [(vc, 0), (vp, vc), (vi, vc+vp)]:
                if val > 0:
                    ax2.text(offset + val/2, i, str(val), ha='center', va='center',
                             fontsize=9, color='white', fontweight='bold')

        # --- Pizza VALIDADOS ---
        ax3 = fig.add_subplot(gs[2])
        vals_v = [comp, parc, inc]
        vals_v_plot = [v for v in vals_v if v > 0]
        labs_v_plot = [l for l, v in zip(labels_pizza, vals_v) if v > 0]
        cols_v = [c for c, v in zip([COR_COMP, COR_PARC, COR_INC], vals_v) if v > 0]
        if vals_v_plot:
            ax3.pie(vals_v_plot, labels=labs_v_plot, colors=cols_v,
                    autopct='%1.1f%%', startangle=90,
                    wedgeprops={'edgecolor': 'white', 'linewidth': 1.2})
        ax3.set_title('Saude Global da Base (Validado)', fontsize=12, fontweight='bold', pad=10)

        buf = _io.BytesIO()
        plt.savefig(buf, format='png', dpi=120, bbox_inches='tight')
        plt.close(fig)
        buf.seek(0)
        return buf

    buf_painel = _gerar_painel_combinado()

    # Inserir o painel único a partir da coluna E, linha 2
    if buf_painel is not None:
        img = OxImage(buf_painel)
        ws_res.add_image(img, f'{_gcl(5)}2')

    logger.info("[OK] Aba 'Relatorio_Cadastral' gerada com sucesso.")
# ============================================================
# FIM DO MÓDULO RELATÓRIO CADASTRAL 8162
# ============================================================

# Selenium imports
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import (
    NoSuchElementException,
    WebDriverException,
    ElementClickInterceptedException,
)
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Novos módulos integrados (Tarefas 1-10)
try:
    from config_loader import config
    from audit_logger import audit_logger, audit
    from retry_utils import http_request_with_retry, create_session_with_retry
    from circuit_breaker import get_circuit_breaker
    from schema_validator import validate_csv_file
    from html_sanitizer import sanitize_html, sanitize_dashboard_html
    from timezone_utils import tz_manager, format_time
except ImportError as e:
    import sys
    print(f"⚠️  Aviso: Módulo não encontrado: {e}", file=sys.stderr)

# ======== REGEX PRÉ-COMPILADAS (performance) ========
REGEX_TELEFONE_VALIDACAO = re.compile(r"^(?:55)?([1-9]\d{9,10})$")
# =====================================================

# ================= DIRETÓRIOS E DATA =================
DATA_EXEC = datetime.now().strftime("%Y-%m-%d")
DATA_HORA_EXEC = datetime.now().strftime("%d-%m-%Y %Hh%Mm")

# Diretório base — pasta do .exe quando empacotado, pasta do script caso contrário
def _get_output_base() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.abspath(os.path.dirname(__file__))

_BASE = _get_output_base()

LOG_DIR         = os.path.join(_BASE, "logs")
PLANILHA_DIR    = os.path.join(_BASE, "planilhas_filtradas")
DIR_NEGOCIACAO  = os.path.join(_BASE, "Relatorios_Negociacao")
GRAFICOS_DIR    = os.path.join(_BASE, "Graficos")
PASTA_DO_DIA    = os.path.join(GRAFICOS_DIR, DATA_EXEC)
DIR_MAIOR_CORTE = os.path.join(PASTA_DO_DIA, "Maior que corte")
DIR_MENOR_CORTE = os.path.join(PASTA_DO_DIA, "Menor que corte")
DIR_PDD         = os.path.join(PASTA_DO_DIA, "PDD")

# ÚNICA PASTA CRIADA NA INICIALIZAÇÃO (Exigência do sistema de LOGS)
os.makedirs(LOG_DIR, exist_ok=True)

LOG_FILE      = os.path.join(LOG_DIR,      f"processamento_{DATA_EXEC}.log")
SAIDA_XLSX    = os.path.join(PLANILHA_DIR, f"Relatorio Base 8104 {DATA_HORA_EXEC}.xlsx")
FALHAS_XLSX   = os.path.join(PLANILHA_DIR, f"Falhas_Envio_{DATA_HORA_EXEC}.xlsx")
ARQUIVO_ENVIO = os.path.join(PLANILHA_DIR, "Planilha envios WhatsApp.xlsx")
# =====================================================

# ================= LOGGING =================
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)
logger.propagate = False


class AsciiConsoleFormatter(logging.Formatter):
    def format(self, record):
        formatted = super().format(record)
        normalized = unicodedata.normalize("NFKD", formatted)
        return normalized.encode("ascii", "replace").decode("ascii")

formatter = logging.Formatter(
    fmt="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)

if not logger.handlers:
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(
        AsciiConsoleFormatter(
            fmt="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
    )

    file_handler = RotatingFileHandler(
        LOG_FILE, maxBytes=5 * 1024 * 1024, backupCount=5, encoding="utf-8"
    )
    file_handler.setFormatter(formatter)

    logger.addHandler(console_handler)
    logger.addHandler(file_handler)
# =====================================================

# ================= CONFIGURAÇÕES =================
ARQUIVO_CSV = "macro 8104.csv"  
ARQUIVO_MACRO = "macro 8162.csv"

COL_EMAIL = "EMAIL"
COL_CLIENTE = "CLIENTE_DA_FATURA"
COL_ENDERECO = "ENDERECO"
COL_TIPO_IRREGULARIDADE = "TIPO_IRREGULARIDADE"
COL_SITUACAO_LIGACAO_FASE = "SITUACAO LIGACAO / SITUACAO FASE"
COL_MATRICULA = "MATRICULA"
COL_DIAS = "QTDE_DIAS_EM_ATRASO"
COL_TEL  = "TEL_CEL"
COL_FIXO = "TEL_FIXO_CONSUMIDOR"

DIVIDA_COL = "DIVIDA_DO_INQUILINO_ATUAL"

# Mapeamento de aliases para colunas - permite compatibilidade com diferentes cidades
COLUMN_ALIASES = {
    COL_CLIENTE: [
        "CLIENTE_DA_FATURA", "CLIENTE", "NOME_CLIENTE", "NOME_CONSUMIDOR",
        "CONSUMIDOR", "NOME_TITULAR", "TITULAR", "NOME_RESPONSAVEL",
        "RESPONSAVEL", "NOME_PROPRIETARIO", "PROPRIETARIO", "NOME",
        "CLIENTE_DA_FATURA_DA_FATURA", "NOME_FATURA", "RAZAO_SOCIAL",
        "NOME_DO_CLIENTE", "NOME_DO_CONSUMIDOR"
    ],
    COL_EMAIL: [
        "EMAIL", "E_MAIL", "EMAIL_DO_CLIENTE", "EMAIL_CONSUMIDOR",
        "EMAIL_TITULAR", "CORREIO_ELETRONICO", "EMAIL_PARA_CONTATO"
    ],
    COL_ENDERECO: [
        "ENDERECO", "ENDEREÇO", "LOGRADOURO", "LOGRADOURO_IMOVEL",
        "LOGRADOURO_ENTREGA", "ENDERECO_COMPLETO", "ENDEREÇO_COMPLETO",
        "ENDERECO_DO_IMOVEL", "ENDERECO_DO_CLIENTE", "RUA", "AVENIDA",
        "ENDEREÇO_IMOVEL"
    ],
    COL_DIAS: [
        "QTDE_DIAS_EM_ATRASO", "DIAS_EM_ATRASO", "DIAS_ATRASO",
        "QTD_DIAS_ATRASO", "NUMERO_DIAS_ATRASO", "DIAS_VENCIDO",
        "DIAS_EM_ATRASO_NA_FATURA", "TEMPO_ATRASO", "DIAS_ATRASO_FATURA"
    ],
    COL_TEL: [
        "TEL_CEL", "TELEFONE_CELULAR", "CELULAR", "TELEFONE",
        "TEL_CONTATO", "TELEFONE_DO_CLIENTE", "TELEFONE_CONSUMIDOR",
        "CEL", "FONE", "TELEFONE_PARA_CONTATO", "TELEFONE_CEL",
        "TELEFONE_PRINCIPAL", "TELEFONE_CADASTRO"
    ],
    COL_MATRICULA: [
        "MATRICULA", "MATRÍCULA", "LIGACAO", "LIGAÇÃO", "ID_LIGACAO",
        "IDLIGACAO", "CONEXAO", "CONEXÃO", "ID_IMOVEL", "NUMERO_LIGACAO",
        "CODIGO_LIGACAO", "ID_CLIENTE", "IDENTIFICACAO"
    ],
    DIVIDA_COL: [
        "DIVIDA_DO_INQUILINO_ATUAL", "DIVIDA_INQUILINO", "DIVIDA_ATUAL",
        "DIVIDA_DO_CLIENTE", "DIVIDA_CONSUMIDOR", "DIVIDA_ATIVA",
        "DIVIDA_EM_ABERTO", "TEM_DIVIDA", "POSSUI_DIVIDA",
        "DIVIDA_DO_MORADOR", "DIVIDA_MORADOR"
    ],
    COL_TIPO_IRREGULARIDADE: [
        "TIPO_IRREGULARIDADE", "TIPO_DE_IRREGULARIDADE", "IRREGULARIDADE",
        "TIPO_IRREG", "IRREGULARIDADE_TIPO", "DESCRICAO_IRREGULARIDADE"
    ],
    COL_SITUACAO_LIGACAO_FASE: [
        "SITUACAO_LIGACAO", "SITUACAO_FASE", "SITUACAO_LIGACAO_FASE",
        "SITUACAO_DA_LIGACAO", "SITUACAO_DA_FASE",
        "SITUACAO_LIGACAO__SITUACAO_FASE", "SITUACAO_LIGACAO_SITUACAO_FASE"
    ]
}


def normalizar_colunas_para_padrao(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normaliza os nomes das colunas do DataFrame para os nomes padrão,
    identificando aliases comuns usados em diferentes cidades.
    
    Args:
        df: DataFrame com colunas que podem ter nomes variados
        
    Returns:
        DataFrame com colunas renomeadas para os nomes padrão
    """
    colunas_renomeadas = {}
    colunas_ja_mapeadas = set()
    
    for col in df.columns:
        col_upper = str(col).strip().upper()
        # Remove acentos e caracteres especiais para comparação
        col_normalized = unicodedata.normalize("NFKD", col_upper)
        col_normalized = col_normalized.encode("ASCII", "ignore").decode("utf-8")
        col_normalized = re.sub(r"[\s\-]+", "_", col_normalized)
        col_normalized = re.sub(r"[^\w_]+", "", col_normalized)
        
        # Procura por match nos aliases
        for nome_padrao, aliases in COLUMN_ALIASES.items():
            if nome_padrao in colunas_ja_mapeadas:
                continue
                
            for alias in aliases:
                alias_normalized = unicodedata.normalize("NFKD", alias.upper())
                alias_normalized = alias_normalized.encode("ASCII", "ignore").decode("utf-8")
                alias_normalized = re.sub(r"[\s\-]+", "_", alias_normalized)
                alias_normalized = re.sub(r"[^\w_]+", "", alias_normalized)
                
                if col_normalized == alias_normalized:
                    if col != nome_padrao:
                        colunas_renomeadas[col] = nome_padrao
                        colunas_ja_mapeadas.add(nome_padrao)
                        logger.debug(f"   [MAPEAMENTO] Coluna '{col}' -> '{nome_padrao}'")
                    break
    
    if colunas_renomeadas:
        df = df.rename(columns=colunas_renomeadas)
        logger.info(f"   [OK] {len(colunas_renomeadas)} colunas normalizadas para o padrão.")
    
    return df

CORTE_DIAS = 30
POLITICA_NAN_DIAS = "zerar"

DELAY_BETWEEN_SENDS = 5

MAX_MENSAGENS_POR_LOTE = 20
INTERVALO_LOTE_SEGUNDOS = 3600
DELAY_MIN_SEG = 40
DELAY_MAX_SEG = 120
# -----------------------------------------------------

USER_DATA_DIR = os.path.join(_BASE, "chrome-whatsapp-profile")

# =================================================
# FUNÇÃO GPS PARA O PYINSTALLER ACHAR A IMAGEM
# =================================================
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

IMAGE_PATH = resource_path("itapoa_informa.png")
# =================================================

EMAIL_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

def normalize_column_name(col: object) -> str:
    if col is None:
        return ""
    normalized = unicodedata.normalize("NFKD", str(col))
    normalized = normalized.encode("ASCII", "ignore").decode("utf-8")
    normalized = normalized.strip().upper()
    normalized = re.sub(r"[\s\-]+", "_", normalized)
    normalized = re.sub(r"[^\w_]+", "", normalized)
    return normalized


def normalize_df_columns(df: pd.DataFrame) -> None:
    df.columns = [normalize_column_name(c) for c in df.columns]

# Mensagem padrão
TEMPLATE_MENSAGEM = (
    """Olá, {{NOME}}. Esperamos que esteja bem.
 
Somos o Encananelson e a Ambientally, mascotes da Itapoá Saneamento, e viemos trazer uma informação importante sobre o seu imóvel localizado na {{2}} (Matrícula: {{1}}).
 
Conforme verificação em sistema, identificamos uma pendência no valor de R$ {{4}}, com {{3}} dias de atraso. Sabemos que, na correria do dia a dia, alguns prazos podem acabar passando despercebidos, por isso estamos passando aqui para lembrar você e ajudar a manter tudo em dia, evitando possíveis transtornos ou eventual interrupção no fornecimento dos serviços.
Se o pagamento já foi realizado, pode desconsiderar esta mensagem, tudo bem E, se precisar da fatura atualizada ou de qualquer esclarecimento, é só procurar os canais oficiais de atendimento.
 
 
 
⚠️ *_Atenção: Este é um canal exclusivo para envio de avisos e não possui atendentes para receber mensagens ou ligações._*

Para atendimento, dúvidas, emissão de segunda via ou negociação, por favor, utilize um de nossos canais oficiais abaixo:
 0800 643 2750
 WhatsApp: (47) 99278-0310
 Site: www.itapoasaneamento.com.br
 Aplicativo Centro Sul 
 https://play.google.com/store/apps/details?id=com.waterfy.agenciavirtual.centrosul
 Atendimento presencial: de segunda a sexta-feira, das 8h às 16h30.

Atenciosamente,
Equipe Itapoá Saneamento

"""
)

# ---------------------------
# Tutorial de Uso
# ---------------------------
def exibir_tutorial():
    print(r"""
                        
                    __  __   _   _  _ _   _   _   _     ___  ___ 
                    |  \/  | /_\ | \| | | | | /_\ | |   |   \| __|
                    | |\/| |/ _ \| .` | |_| |/ _ \| |__ | |) | _| 
                    |_|  |_/_/ \_\_|\_|\___//_/ \_\____||___/|___|
~          
                     _   _  _  ___  _____  ___  _   _  ___   ___   ___  ___ 
                    | | | \| |/ __||_   _|| _ \| | | |/ __| / _ \ | __|/ __|
                    | | | .  |\__ \  | |  |   /| |_| | (__ | (_) || _| \__ \
                    |_| |_|\_||___/  |_|  |_|_\ \___/ \___| \___/ |___||___/
                                                       ,
                                            ?
                                      _______
                                     [ o   o ]   < "Socorro, qual botão eu aperto?!"
                                     |   ~   |
                                     |_______|
                                    /|       |\
                                   / |_______| \
                                      |     |
                                     _|     |_

=============================================================================
Bem-vindo(a) ao assistente virtual da Itapoá Saneamento!
Este sistema foi criado para automatizar tarefas repetitivas. Veja como usar:

   _______
  |  | |  |   [ 1 ] Processar CSV e gerar planilha WhatsApp
  |__|_|__|   ---------------------------------------------------------------
  |  | |  |   O QUE FAZ: Pega os relatórios brutos (Macro 8104 e 8162) e os
  |__|_|__|   transforma em planilhas limpas para o envio de mensagens. 
              Também gera mapas de calor e gráficos de valores em imagens!
              
              COMO USAR: 
                 - Digite 1 e aperte Enter.
                 - Uma janela vai abrir. Selecione o arquivo da Macro 8104.
                 - Outra janela vai abrir. Selecione o arquivo Macro 8162.
                 - O sistema perguntará a régua de dias para corte (ex: 30).
                 - Aguarde o processamento terminar.
                   
                 📁 PLANILHA: Na pasta 'planilhas_filtradas', será gerado 
                    um arquivo Excel (ex: 'Relatorio Base 8104...xlsx').
                    Dentro deste arquivo estarão as 3 abas. A aba 
                    'emails_invalidos' é a ideal para a Opção 2!
                       
                 📊 GRÁFICOS: O sistema cria uma pasta 'Graficos', separando
                    por data e por cortes (volume de matrículas e valores).


    [0_0]     [ 2 ] Enviar mensagens no WhatsApp
   /|___|\    ---------------------------------------------------------------
    _|_|_     O QUE FAZ: O robô abre o WhatsApp Web e manda as mensagens de 
              cobrança automaticamente para os clientes listados.
              
              COMO USAR:
                 - Digite 2 e aperte Enter.
                 - Selecione a planilha gerada no Passo 1 na janela
                   (ela estará dentro da pasta 'planilhas_filtradas').
                 - Escolha qual aba da planilha você quer enviar.
                 - O Chrome vai abrir sozinho. Leia o QR Code se pedir.
                 - ⚠️ AVISO: Quando o envio começar, LARGUE MOUSE E TECLADO!
                 - 🛑 FREIO DE MÃO: Pressione a tecla "P" no terminal a 
                   qualquer momento para PAUSAR o envio ou PULAR um cliente!
                 - 💡 DICA (Segundo Plano): Para usar o PC enquanto o robô 
                   trabalha, aperte "Windows + TAB", clique em "Nova área 
                   de trabalho" e arraste a janela do Chrome para lá. Volte 
                   para a área principal e trabalhe normalmente!


   .-----.    [ 3 ] Calculo dos descontos macro 8104 (Negociação)
   | 123 |    ---------------------------------------------------------------
   |-----|    O QUE FAZ: Pega o arquivo bruto e cria uma planilha mágica
   | 456 |    com fórmulas de desconto, entrada e parcelamentos calculados.
   | 789 |    
   '-----'    COMO USAR:
                 - Digite 3 e aperte Enter.
                 - Selecione o arquivo Macro 8104 na janela.
                 - O arquivo pronto ficará na pasta 'Relatorios_Negociacao'.


    [  ]     [ 4 ] Ajuda / Tutorial de Uso
              ---------------------------------------------------------------
              O QUE FAZ: Exibe este manual que você está lendo agora!


    [->]      [ 5 ] Sair
              ---------------------------------------------------------------
              O QUE FAZ: Fecha o programa com segurança.

=============================================================================
""")
    input("\nPressione ENTER para fechar o manual e voltar ao menu principal...")

# ---------------------------
# Utilitários de dados
# ---------------------------
def selecionar_arquivo(titulo="Selecione o arquivo", tipos=[("Arquivos CSV", "*.csv"), ("Todos os arquivos", "*.*")]):
    root = tk.Tk()
    root.withdraw() 
    root.attributes('-topmost', True) 
    
    caminho_arquivo = filedialog.askopenfilename(
        title=titulo,
        filetypes=tipos
    )
    return caminho_arquivo

def remover_acentos(s: str) -> str:
    if not isinstance(s, str):
        return ""
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join([c for c in nfkd if not unicodedata.combining(c)])

def contem_palavra_proibida(nome: str) -> bool:
    if pd.isna(nome):
        return False
    s = remover_acentos(str(nome)).upper()
    proibidas = ["PREFEITURA", "PREFEITO", "BOMBEIRO", "BOMBEIROS", "CORPO DE BOMBEIROS"]
    for p in proibidas:
        if p in s:
            logger.warning(f"⛔ Nome bloqueado por palavra proibida: '{nome}' (contém '{p}').")
            return True
    return False

def parse_valor_cell(valor) -> float:
    if pd.isna(valor):
        return 0.0
    s = str(valor).strip()
    if not s:
        return 0.0

    tokens = re.findall(r"\d[\d\.,]*\d|\d+", s)
    total = 0.0
    for t in tokens:
        t = t.strip()
        if not t:
            continue
        try:
            if "." in t and "," in t:
                t_norm = t.replace(".", "").replace(",", ".")
            elif "," in t and "." not in t:
                t_norm = t.replace(",", ".")
            else:
                t_norm = t
            t_norm = re.sub(r"[^\d\.]", "", t_norm)
            if t_norm == "":
                continue
            val = float(t_norm)
            total += val
        except Exception:
            continue
    return round(total, 2)

def email_valido(email: str) -> bool:
    if pd.isna(email):
        return False
    s = str(email).strip().lower()
    if not s or s.count("@") != 1:
        return False
    return bool(EMAIL_REGEX.match(s))

def _to_full_digits(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        try:
            return "{:.0f}".format(value)
        except Exception:
            pass
    s = str(value).strip()
    if not s:
        return ""
    if re.search(r"[eE]\+?\-?\d+", s):
        try:
            d = Decimal(s)
            return str(d.quantize(Decimal(1)))
        except Exception:
            try:
                return "{:.0f}".format(float(s))
            except Exception:
                pass
    s = s.replace("\u200b", "").replace(" ", "")
    s = re.sub(r"[^\d\+]", "", s)
    return s

def normalizar_telefone(numero) -> Optional[str]:
    if pd.isna(numero):
        return None

    s = _to_full_digits(numero)
    if not s:
        return None

    if s.startswith("+"):
        s = s[1:]
    elif s.startswith("00"):
        s = s[2:]

    s = re.sub(r"\D", "", s)
    s = s.lstrip("0")

    m = REGEX_TELEFONE_VALIDACAO.match(s)
    if not m:
        logger.debug(f"Telefone descartado (formato inválido): '{s}'")
        return None

    core = m.group(1)

    if len(set(core)) == 1:
        logger.debug(f"Telefone descartado (dígitos repetidos): '{s}'")
        return None

    if core.startswith("473443"):
        logger.debug(f"Telefone descartado (prefixo bloqueado): '{s}'")
        return None

    return "+55" + core

def formatar_valor(valor) -> str:
    if pd.isna(valor):
        return ""
    try:
        v = float(valor)
        return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        s = str(valor).strip()
        s = re.sub(r"[^\d,\.]", "", s)
        if not s:
            return ""
        try:
            v = float(s.replace(",", "."))
            return f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except Exception:
            return s

def calcular_faixa_pdd(dias_atraso) -> str:
    try:
        dias = int(float(dias_atraso))
    except (TypeError, ValueError):
        dias = 0

    if dias < 0:
        dias = 0

    if dias >= 365:
        return "IN 365 +"

    faixa_superior = ((dias // 30) + 1) * 30
    return f"IN {faixa_superior}"

def _faixas_pdd_ordenadas() -> list:
    return [f"IN {dias}" for dias in range(30, 361, 30)] + ["IN 365 +"]

def log_tabela_pdd(df_validos: pd.DataFrame, df_invalidos: pd.DataFrame = None):
    contagens = {faixa: 0 for faixa in _faixas_pdd_ordenadas()}

    for dataframe in (df_validos, df_invalidos):
        if dataframe is None or dataframe.empty or "PDD" not in dataframe.columns:
            continue
        serie = dataframe["PDD"].fillna("").astype(str).str.strip()
        for faixa, qtd in serie.value_counts().items():
            contagens[faixa] = contagens.get(faixa, 0) + int(qtd)

    faixas_com_valor = [(faixa, contagens.get(faixa, 0)) for faixa in _faixas_pdd_ordenadas() if contagens.get(faixa, 0) > 0]
    if not faixas_com_valor:
        faixas_com_valor = [("IN 30", 0)]

    largura_faixa = max(len("PDD"), max(len(faixa) for faixa, _ in faixas_com_valor), len("TOTAL"))
    largura_qtd = max(len("QTD"), max(len(str(qtd)) for _, qtd in faixas_com_valor), len(str(sum(contagens.values()))))
    topo = f"+-{'-' * largura_faixa}-+-{'-' * largura_qtd}-+"

    logger.info("[INFO] Distribuicao de clientes por faixa PDD:")
    logger.info(topo)
    logger.info(f"| {'PDD'.ljust(largura_faixa)} | {'QTD'.rjust(largura_qtd)} |")
    logger.info(topo)
    for faixa, qtd in faixas_com_valor:
        logger.info(f"| {faixa.ljust(largura_faixa)} | {str(qtd).rjust(largura_qtd)} |")
    logger.info(topo)
    logger.info(f"| {'TOTAL'.ljust(largura_faixa)} | {str(sum(contagens.values())).rjust(largura_qtd)} |")
    logger.info(topo)

def montar_mensagem(row: pd.Series) -> str:
    nome = str(row.get(COL_CLIENTE, "") or "").strip()
    if nome:
        nome_limpo = nome.split()[0].capitalize()
    else:
        nome_limpo = ""

    matricula = str(row.get(COL_MATRICULA, "") or "").strip()
    endereco_full = str(row.get(COL_ENDERECO, "") or "").strip()

    dias = row.get("QTDE_DIAS_EM_ATRASO", row.get(COL_DIAS, ""))
    try:
        dias_int = int(float(dias))
        dias_str = str(dias_int)
    except Exception:
        dias_str = str(dias).strip()

    valor_num = None
    if "VALOR_TOTAL" in row.index and not pd.isna(row.get("VALOR_TOTAL")):
        try:
            valor_num = float(row.get("VALOR_TOTAL"))
        except Exception:
            valor_num = parse_valor_cell(row.get("VALOR_TOTAL"))
    else:
        for cand in ("VALOR", "VALOR_R$", "VALOR_FATURA", "VALOR_DA_FATURA", "VLR"):
            if cand in row.index and not pd.isna(row.get(cand)):
                valor_num = parse_valor_cell(row.get(cand))
                break

    if valor_num is None:
        valor_num = 0.0

    matricula_b = f"*{matricula}*" if matricula else "*-*-*"
    endereco_b = f"*{endereco_full}*" if endereco_full else "*-*-*"
    dias_b = f"*{dias_str}*" if dias_str else "*0*"
    valor_b = f"*{formatar_valor(valor_num)}*"

    if not nome_limpo:
        logger.warning(f"Mensagem montada sem nome — matrícula {matricula or 'N/A'}.")
    if not matricula:
        logger.warning(f"Mensagem montada sem matrícula.")
    if valor_num == 0.0:
        logger.warning(f"Mensagem montada com valor R$ 0,00 — matrícula {matricula or 'N/A'}.")

    msg = TEMPLATE_MENSAGEM
    msg = msg.replace("{{NOME}}", nome_limpo)
    msg = msg.replace("{{1}}", matricula_b)
    msg = msg.replace("{{2}}", endereco_b)
    msg = msg.replace("{{3}}", dias_b)
    msg = msg.replace("{{4}}", valor_b)

    return msg


# ---------------------------
# Leitura CSV com fallback
# ---------------------------
def ler_csv_com_fallback(caminho):
    tentativas = [
        {"sep": ";", "encoding": "ISO-8859-1"},
        {"sep": ";", "encoding": "latin1"},
        {"sep": ";", "encoding": "utf-8"},
        {"sep": ",", "encoding": "utf-8"},
        {"sep": ",", "encoding": "latin1"},
    ]

    ultimo_erro = None
    for cfg in tentativas:
        try:
            # Otimizar: converter apenas colunas de interesse para o tipo apropriado
            df = pd.read_csv(caminho, dtype=str, **cfg)
            logger.info(f"   Arquivo aberto — {len(df)} registros, encoding {cfg['encoding']}, separador '{cfg['sep']}'.")
            return df
        except Exception as e:
            ultimo_erro = e

    logger.critical(f"Falha total ao ler a planilha. Último erro: {ultimo_erro}")
    raise RuntimeError("Não conseguimos ler os dados. Salve a sua planilha novamente como 'CSV UTF-8' ou verifique se o arquivo não está corrompido.")


def gerar_mapa_calor_grupos(df: pd.DataFrame, sufixo_nome: str, pasta_destino: str):
    """Mantido por compatibilidade — substituído por gerar_graficos_combinados."""
    logger.debug("gerar_mapa_calor_grupos chamada — delegando para gerar_graficos_combinados.")

def gerar_grafico_valores_por_grupo(df: pd.DataFrame, sufixo_nome: str, pasta_destino: str):
    """Mantido por compatibilidade — substituído por gerar_graficos_combinados."""
    logger.debug("gerar_grafico_valores_por_grupo chamada — delegando para gerar_graficos_combinados.")

def gerar_graficos_combinados(df: pd.DataFrame, sufixo_nome: str, pasta_destino: str):
    """
    Gera um único PNG com:
      - Linha 1: Heatmap (esq.) + Barplot de valores (dir.)
      - Linha 2: Painel de insights automáticos em texto
    
    Otimizado: libera memória com plt.close(), usa tight_layout, compilação de strings.
    """
    import matplotlib
    matplotlib.use("Agg")      # garante backend não-interativo mesmo em threads
    import matplotlib.gridspec as gridspec
    import matplotlib.patches as mpatches

    tem_grupo = "GRUPO" in df.columns
    tem_valor = "VALOR_TOTAL" in df.columns

    if not tem_grupo:
        logger.warning(f"   [ATENCAO] Coluna GRUPO ausente - grafico de '{sufixo_nome}' ignorado.")
        return

    logger.info(f"[INFO] Desenhando graficos para: {sufixo_nome}...")

    df_temp = df.copy()
    df_temp["GRUPO"] = df_temp["GRUPO"].fillna("NÃO INFORMADO").astype(str)

    # ── Dados: volume de matrículas por grupo ────────────────────────────────
    contagem = df_temp["GRUPO"].value_counts().reset_index()
    contagem.columns = ["GRUPO", "QUANTIDADE"]
    media_matriculas = contagem["QUANTIDADE"].mean() if not contagem.empty else 0

    # ── Dados: valor total e média de dívida por grupo ───────────────────────
    if tem_valor:
        df_temp["VALOR_TOTAL"] = pd.to_numeric(df_temp["VALOR_TOTAL"], errors="coerce").fillna(0)
        agrupado_soma  = df_temp.groupby("GRUPO")["VALOR_TOTAL"].sum().reset_index()
        agrupado_soma  = agrupado_soma.sort_values("VALOR_TOTAL", ascending=False)
        agrupado_media = df_temp.groupby("GRUPO")["VALOR_TOTAL"].mean().reset_index()
        agrupado_media.columns = ["GRUPO", "MEDIA_VALOR"]
        media_geral_valor = agrupado_soma["VALOR_TOTAL"].mean() if not agrupado_soma.empty else 0
        sem_valor = agrupado_soma["VALOR_TOTAL"].sum() == 0
    else:
        agrupado_soma = agrupado_media = pd.DataFrame()
        media_geral_valor = 0
        sem_valor = True

    if contagem.empty:
        logger.warning(f"   [ATENCAO] Sem dados suficientes para gerar o grafico de '{sufixo_nome}'.")
        return

    # ── Formatadores ─────────────────────────────────────────────────────────
    def fmt_val(v):
        return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def fmt_num(v):
        return f"{v:,.0f}".replace(",", ".")

    n_grupos   = max(len(contagem), len(agrupado_soma) if not agrupado_soma.empty else 0)
    n_insights_estimados = 8
    altura_graficos = max(7, n_grupos * 0.65)
    altura_insights = max(4.0, n_insights_estimados * 0.42)

    # ── Layout: 2 linhas — gráficos em cima, insights embaixo ────────────────
    fig = plt.figure(figsize=(24, altura_graficos + altura_insights + 1.5))
    gs  = gridspec.GridSpec(
        2, 2,
        figure=fig,
        height_ratios=[altura_graficos, altura_insights],
        hspace=0.55,
        wspace=0.45
    )

    media_mat_fmt = f"{media_matriculas:.1f}"
    media_val_fmt = fmt_val(media_geral_valor)
    fig.suptitle(
        f"Relatório: {sufixo_nome}",
        fontsize=15, fontweight="bold", y=0.995
    )

    # ── PAINEL ESQUERDO: Heatmap de volume ───────────────────────────────────
    ax_heat = fig.add_subplot(gs[0, 0])
    dados_matriz = contagem.set_index("GRUPO")[["QUANTIDADE"]]
    sns.heatmap(dados_matriz, annot=True, fmt="d", cmap="Reds",
                linewidths=1, cbar=True, ax=ax_heat)
    ax_heat.set_title("Volume de Matrículas por Grupo", fontsize=12, pad=10)
    ax_heat.set_ylabel("Grupo", fontsize=10)
    ax_heat.set_xlabel("Qtd. de Matrículas", fontsize=10)
    # linha de média
    idx_med = sorted(range(len(contagem)),
                     key=lambda i: abs(contagem.iloc[i]["QUANTIDADE"] - media_matriculas))
    ax_heat.axhline(y=idx_med[0] + 0.5, color="orange", lw=1.5, ls="--",
                    label=f"Média: {media_mat_fmt}")
    ax_heat.legend(fontsize=9, loc="lower right")

    # ── PAINEL DIREITO: Barplot de valor total ────────────────────────────────
    ax_bar = fig.add_subplot(gs[0, 1])
    if not sem_valor and not agrupado_soma.empty:
        sns.barplot(x="VALOR_TOTAL", y="GRUPO", data=agrupado_soma,
                    palette="Reds_r", hue="GRUPO", legend=False, ax=ax_bar)
        ax_bar.set_title("Valor Total em Aberto por Grupo", fontsize=12, pad=10)
        ax_bar.set_ylabel("Grupo", fontsize=10)
        ax_bar.set_xlabel("Valor Total (R$)", fontsize=10)
        for p in ax_bar.patches:
            v = p.get_width()
            if v > 0:
                ax_bar.annotate(fmt_val(v),
                                (v, p.get_y() + p.get_height() / 2.),
                                ha="left", va="center",
                                xytext=(5, 0), textcoords="offset points", fontsize=9)
        ax_bar.set_xlim(0, agrupado_soma["VALOR_TOTAL"].max() * 1.3)
        ax_bar.set_xticks([])
        ax_bar.axvline(x=media_geral_valor, color="orange", lw=1.5, ls="--",
                       label=f"Média: {media_val_fmt}")
        ax_bar.legend(fontsize=9, loc="lower right")
        sns.despine(ax=ax_bar, bottom=True)
    else:
        ax_bar.text(0.5, 0.5, "Sem dados de valor", ha="center", va="center",
                    fontsize=12, color="gray", transform=ax_bar.transAxes)
        ax_bar.axis("off")

    # ── PAINEL INFERIOR: Insights automáticos ────────────────────────────────
    ax_ins = fig.add_subplot(gs[1, :])
    ax_ins.axis("off")

    insights = []

    # Volume de matriculas
    grupo_mais_matriculas  = contagem.iloc[0]["GRUPO"]
    qtd_mais               = int(contagem.iloc[0]["QUANTIDADE"])
    grupo_menos_matriculas = contagem.iloc[-1]["GRUPO"]
    qtd_menos              = int(contagem.iloc[-1]["QUANTIDADE"])
    total_matriculas       = int(contagem["QUANTIDADE"].sum())
    perc_dominante         = qtd_mais / total_matriculas * 100

    insights.append("Grupo com MAIS matriculas: " + grupo_mais_matriculas + " -- " + fmt_num(qtd_mais) + " registros (" + f"{perc_dominante:.1f}" + "% do total)")
    insights.append("Grupo com MENOS matriculas: " + grupo_menos_matriculas + " -- " + fmt_num(qtd_menos) + " registros")
    insights.append("Media de matriculas por grupo: " + media_mat_fmt + "  |  Total de matriculas neste corte: " + fmt_num(total_matriculas))

    if not sem_valor and not agrupado_soma.empty and not agrupado_media.empty:
        grupo_maior_valor  = agrupado_soma.iloc[0]["GRUPO"]
        valor_maior        = agrupado_soma.iloc[0]["VALOR_TOTAL"]
        grupo_menor_valor  = agrupado_soma.iloc[-1]["GRUPO"]
        valor_menor        = agrupado_soma.iloc[-1]["VALOR_TOTAL"]
        total_valor        = agrupado_soma["VALOR_TOTAL"].sum()

        insights.append("Grupo com MAIOR divida total: " + grupo_maior_valor + " -- " + fmt_val(valor_maior) + " (" + f"{valor_maior/total_valor*100:.1f}" + "% do total em aberto)")
        insights.append("Grupo com MENOR divida total: " + grupo_menor_valor + " -- " + fmt_val(valor_menor))
        insights.append("Valor total em aberto neste corte: " + fmt_val(total_valor) + "  |  Media por grupo: " + media_val_fmt)

        agrupado_media_ord = agrupado_media.sort_values("MEDIA_VALOR", ascending=False)
        grupo_maior_media  = agrupado_media_ord.iloc[0]["GRUPO"]
        media_maior        = agrupado_media_ord.iloc[0]["MEDIA_VALOR"]
        grupo_menor_media  = agrupado_media_ord.iloc[-1]["GRUPO"]
        media_menor        = agrupado_media_ord.iloc[-1]["MEDIA_VALOR"]

        insights.append("Grupo com MAIOR media de divida por matricula: " + grupo_maior_media + " -- " + fmt_val(media_maior) + " por cliente")
        insights.append("Grupo com MENOR media de divida por matricula: " + grupo_menor_media + " -- " + fmt_val(media_menor) + " por cliente")

        perc_valor_dominante = valor_maior / total_valor * 100
        if perc_valor_dominante >= 40:
            insights.append("ALERTA: O grupo " + grupo_maior_valor + " concentra " + f"{perc_valor_dominante:.1f}" + "% de toda a divida deste corte -- prioridade de cobranca recomendada.")

    texto_completo = "\n".join("  " + ins for ins in insights)

    ax_ins.set_xlim(0, 1)
    ax_ins.set_ylim(0, 1)

    # Fundo cinza claro no painel de insights
    fundo = mpatches.FancyBboxPatch((0.01, 0.02), 0.98, 0.94,
                                     boxstyle="round,pad=0.01",
                                     linewidth=1.2, edgecolor="#cccccc",
                                     facecolor="#f7f7f7", zorder=0)
    ax_ins.add_patch(fundo)

    ax_ins.text(0.5, 0.97, "📋  Insights Automáticos",
                ha="center", va="top", fontsize=12, fontweight="bold",
                color="#333333", transform=ax_ins.transAxes)

    ax_ins.text(0.03, 0.88, texto_completo,
                ha="left", va="top", fontsize=9.5, color="#222222",
                transform=ax_ins.transAxes,
                linespacing=1.75,
                fontfamily="monospace")

    # ── Salva ─────────────────────────────────────────────────────────────────
    os.makedirs(pasta_destino, exist_ok=True)
    nome_seg    = sufixo_nome.replace(" ", "_").lower()
    caminho_img = f"{pasta_destino}/Relatorio_{nome_seg}_{DATA_HORA_EXEC}.png"
    
    # Otimização: usar tight_layout() e liberar memória após salvar
    plt.tight_layout(rect=[0, 0, 1, 0.99], pad=2.0)
    plt.savefig(caminho_img, dpi=150, bbox_inches="tight")
    plt.close(fig)  # Liberar figura específica
    plt.close('all')  # Liberar todas as figuras abertas

    logger.info(f"   Gráfico salvo com sucesso.")
    for ins in insights:
        logger.info(f"   {ins}")

# ---------------------------
# Módulo: Negociação de Dívidas
# ---------------------------
def gerar_grafico_pdd_por_grupo(df: pd.DataFrame, sufixo_nome: str, pasta_destino: str):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.gridspec as gridspec
    import matplotlib.patches as mpatches

    if "GRUPO" not in df.columns:
        logger.warning(f"   [ATENCAO] Coluna GRUPO ausente - grafico de PDD '{sufixo_nome}' ignorado.")
        return

    col_dias_base = COL_DIAS if COL_DIAS in df.columns else "QTDE_DIAS_EM_ATRASO"
    if col_dias_base not in df.columns:
        logger.warning(f"   [ATENCAO] Coluna de dias ausente - grafico de PDD '{sufixo_nome}' ignorado.")
        return

    logger.info(f"[INFO] Desenhando grafico PDD para: {sufixo_nome}...")

    df_temp = df.copy()
    df_temp["GRUPO"] = df_temp["GRUPO"].fillna("NAO INFORMADO").astype(str)
    df_temp["_DIAS_PDD"] = pd.to_numeric(df_temp[col_dias_base], errors="coerce").fillna(0).clip(lower=0)
    df_temp["PDD"] = df_temp["_DIAS_PDD"].apply(calcular_faixa_pdd)

    if df_temp.empty:
        logger.warning(f"   [ATENCAO] Sem dados suficientes para gerar o grafico de PDD '{sufixo_nome}'.")
        return

    ordem_pdd = _faixas_pdd_ordenadas()
    matriz = pd.crosstab(df_temp["GRUPO"], df_temp["PDD"]).reindex(columns=ordem_pdd, fill_value=0)
    matriz = matriz.loc[matriz.sum(axis=1).sort_values(ascending=False).index]
    colunas_ativas = [col for col in ordem_pdd if col in matriz.columns and int(matriz[col].sum()) > 0]
    if not colunas_ativas:
        colunas_ativas = [ordem_pdd[0]]
    matriz = matriz[colunas_ativas]

    media_dias = df_temp.groupby("GRUPO")["_DIAS_PDD"].mean().sort_values(ascending=False).reset_index()
    media_dias.columns = ["GRUPO", "MEDIA_DIAS"]

    total_por_grupo = df_temp["GRUPO"].value_counts()
    media_dias["QTD"] = media_dias["GRUPO"].map(total_por_grupo).fillna(0).astype(int)
    media_dias["PDD_MEDIO"] = media_dias["MEDIA_DIAS"].apply(calcular_faixa_pdd)

    def fmt_num(v):
        return f"{v:,.0f}".replace(",", ".")

    n_grupos = max(len(matriz), len(media_dias))
    n_insights_estimados = 8
    altura_graficos = max(7, n_grupos * 0.65)
    altura_insights = max(4.0, n_insights_estimados * 0.42)

    fig = plt.figure(figsize=(26, altura_graficos + altura_insights + 1.5))
    gs = gridspec.GridSpec(2, 2, figure=fig, height_ratios=[altura_graficos, altura_insights], hspace=0.55, wspace=0.4)
    fig.suptitle(f"Relatorio: {sufixo_nome}", fontsize=15, fontweight="bold", y=0.995)

    ax_heat = fig.add_subplot(gs[0, 0])
    sns.heatmap(matriz, annot=True, fmt="d", cmap="Reds", linewidths=1, cbar=True, ax=ax_heat)
    ax_heat.set_title("Distribuicao de PDD por Grupo", fontsize=12, pad=10)
    ax_heat.set_ylabel("Grupo", fontsize=10)
    ax_heat.set_xlabel("Faixa PDD", fontsize=10)

    ax_bar = fig.add_subplot(gs[0, 1])
    sns.barplot(x="MEDIA_DIAS", y="GRUPO", data=media_dias, palette="Reds_r", hue="GRUPO", legend=False, ax=ax_bar)
    ax_bar.set_title("Media de Dias em Atraso por Grupo", fontsize=12, pad=10)
    ax_bar.set_ylabel("Grupo", fontsize=10)
    ax_bar.set_xlabel("Dias em atraso", fontsize=10)
    for p in ax_bar.patches:
        v = p.get_width()
        if v > 0:
            ax_bar.annotate(
                f"{fmt_num(v)} d | {calcular_faixa_pdd(v)}",
                (v, p.get_y() + p.get_height() / 2.0),
                ha="left", va="center", xytext=(5, 0), textcoords="offset points", fontsize=9
            )
    if not media_dias.empty and media_dias["MEDIA_DIAS"].max() > 0:
        ax_bar.set_xlim(0, media_dias["MEDIA_DIAS"].max() * 1.25)
    ax_bar.set_xticks([])
    sns.despine(ax=ax_bar, bottom=True)

    ax_ins = fig.add_subplot(gs[1, :])
    ax_ins.axis("off")

    insights = []
    total_registros = int(len(df_temp))
    grupo_mais = total_por_grupo.index[0]
    qtd_mais = int(total_por_grupo.iloc[0])
    grupo_menos = total_por_grupo.index[-1]
    qtd_menos = int(total_por_grupo.iloc[-1])
    totais_pdd = matriz.sum(axis=0).sort_values(ascending=False)
    faixa_mais = totais_pdd.index[0]
    qtd_faixa_mais = int(totais_pdd.iloc[0])
    faixa_menos = totais_pdd.index[-1]
    qtd_faixa_menos = int(totais_pdd.iloc[-1])
    grupo_media_alta = media_dias.iloc[0]
    grupo_media_baixa = media_dias.iloc[-1]

    pdds_altos = [faixa for faixa in colunas_ativas if faixa not in {"IN 30", "IN 60", "IN 90"}]
    if pdds_altos:
        grupos_pdd_alto = matriz[pdds_altos].sum(axis=1).sort_values(ascending=False)
        grupo_pdd_alto = grupos_pdd_alto.index[0]
        qtd_pdd_alto = int(grupos_pdd_alto.iloc[0])
    else:
        grupo_pdd_alto = grupo_mais
        qtd_pdd_alto = 0

    insights.append("Grupo com MAIS registros na base bruta: " + grupo_mais + " -- " + fmt_num(qtd_mais) + " linhas")
    insights.append("Grupo com MENOS registros na base bruta: " + grupo_menos + " -- " + fmt_num(qtd_menos) + " linhas")
    insights.append("Faixa PDD mais frequente: " + faixa_mais + " -- " + fmt_num(qtd_faixa_mais) + " linhas")
    insights.append("Faixa PDD menos frequente: " + faixa_menos + " -- " + fmt_num(qtd_faixa_menos) + " linhas")
    insights.append("Grupo com MAIOR media de atraso: " + str(grupo_media_alta["GRUPO"]) + " -- " + fmt_num(grupo_media_alta["MEDIA_DIAS"]) + " dias (" + str(grupo_media_alta["PDD_MEDIO"]) + ")")
    insights.append("Grupo com MENOR media de atraso: " + str(grupo_media_baixa["GRUPO"]) + " -- " + fmt_num(grupo_media_baixa["MEDIA_DIAS"]) + " dias (" + str(grupo_media_baixa["PDD_MEDIO"]) + ")")
    insights.append("Grupo com maior concentracao em PDD alto (>90 dias): " + grupo_pdd_alto + " -- " + fmt_num(qtd_pdd_alto) + " linhas")
    insights.append("Total de linhas analisadas da 8104 bruta: " + fmt_num(total_registros))

    texto_completo = "\n".join("  " + ins for ins in insights)
    ax_ins.set_xlim(0, 1)
    ax_ins.set_ylim(0, 1)
    fundo = mpatches.FancyBboxPatch((0.01, 0.02), 0.98, 0.94, boxstyle="round,pad=0.01", linewidth=1.2, edgecolor="#cccccc", facecolor="#f7f7f7", zorder=0)
    ax_ins.add_patch(fundo)
    ax_ins.text(0.5, 0.97, "Insights Automaticos", ha="center", va="top", fontsize=12, fontweight="bold", color="#333333", transform=ax_ins.transAxes)
    ax_ins.text(0.03, 0.88, texto_completo, ha="left", va="top", fontsize=9.5, color="#222222", transform=ax_ins.transAxes, linespacing=1.75, fontfamily="monospace")

    os.makedirs(pasta_destino, exist_ok=True)
    nome_seg = sufixo_nome.replace(" ", "_").lower()
    caminho_img = f"{pasta_destino}/Relatorio_{nome_seg}_{DATA_HORA_EXEC}.png"
    plt.tight_layout(rect=[0, 0, 1, 0.99], pad=2.0)
    plt.savefig(caminho_img, dpi=150, bbox_inches="tight")
    plt.close(fig)
    plt.close("all")

    logger.info("   Grafico PDD salvo com sucesso.")
    for ins in insights:
        logger.info(f"   {ins}")

def calcular_desconto_proposta(dias_atraso, situacao):
    """
    Calcula as condições de negociação conforme tabela vigente:

    Tipo: Conta Mensal (Água / Esgoto / Serviços)

    Idade    | À Vista          | Parcelamento            | Reparcelamento (após 3 parc.)
             | Multa Juros Desc | Multa Juros Desc Ent Parc| Multa Juros Desc Ent  Parc
    6-12m    | 0%   0%   0%    | 0%   0%   0%  20%  6    | 0%   0%   0%  50%  6
    13-24m   | 100% 0%   0%    | 0%   0%   0%  20%  6    | 0%   0%   0%  50%  6
    25-36m   | 100% 100% 0%    | 100% 100% 0%  20%  6    | 0%   0%   0%  50%  6
    37-48m   | 100% 100% 10%   | 100% 100% 0%  20%  6    | 0%   0%   0%  50%  6
    49-60m   | 100% 100% 20%   | 100% 100% 10% 20%  6    | 0%   0%   0%  50%  6
    >60m     | 100% 100% 30%   | 100% 100% 20% 20%  6    | 0%   0%   0%  50%  6

    Retorna: (desc_avista, desc_parc, perc_entrada, parcelas,
              multa_av, juros_av, multa_parc, juros_parc,
              desc_repar, ent_repar, parc_repar)
    """
    try:
        dias = float(dias_atraso)
    except (ValueError, TypeError):
        dias = 0

    m = math.floor(dias / 30.4166)

    if m < 6:
        return 0, 0, 0.0, 0, 0, 0, 0, 0, 0, 0.50, 6

    entrada    = 0.20   # 20% — padrão para à vista e parcelamento
    ent_repar  = 0.50   # 50% — reparcelamento
    parcelas   = 6
    parc_repar = 6

    # Multa e juros sempre 0 no reparcelamento
    multa_repar = 0; juros_repar = 0; desc_repar = 0

    if 6 <= m <= 12:
        multa_av=0;   juros_av=0;   desc_av=0
        multa_parc=0; juros_parc=0; desc_parc=0
    elif 13 <= m <= 24:
        multa_av=100; juros_av=0;   desc_av=0
        multa_parc=0; juros_parc=0; desc_parc=0
    elif 25 <= m <= 36:
        multa_av=100; juros_av=100; desc_av=0
        multa_parc=100; juros_parc=100; desc_parc=0
    elif 37 <= m <= 48:
        multa_av=100; juros_av=100; desc_av=10
        multa_parc=100; juros_parc=100; desc_parc=0
    elif 49 <= m <= 60:
        multa_av=100; juros_av=100; desc_av=20
        multa_parc=100; juros_parc=100; desc_parc=10
    else:  # > 60 meses
        multa_av=100; juros_av=100; desc_av=30
        multa_parc=100; juros_parc=100; desc_parc=20

    return (desc_av, desc_parc, entrada, parcelas,
            multa_av, juros_av, multa_parc, juros_parc,
            desc_repar, ent_repar, parc_repar)

def gerar_planilha_negociacao(arquivo_entrada=ARQUIVO_CSV, dir_saida=DIR_NEGOCIACAO):
    logger.info("🧮 Iniciando cálculo das propostas de negociação...")
    
    if not os.path.exists(arquivo_entrada): 
        logger.error(f"❌ O arquivo '{arquivo_entrada}' não foi encontrado.")
        return None

    df = ler_csv_com_fallback(arquivo_entrada)
    df.columns = df.columns.str.strip().str.upper().str.replace(" ", "_")

    col_dias  = "QTDE_DIAS_EM_ATRASO"
    col_sit   = next((c for c in df.columns if "SITUACAO" in c or "LIGACAO" in c), None)
    col_valor = next((c for c in df.columns if c in ["VALOR_TOTAL", "VALOR", "VLR"]), None)

    total_faturas = len(df)
    df[col_dias] = pd.to_numeric(df[col_dias], errors='coerce').fillna(0)
    df["SITUACAO_TEMP"]   = df[col_sit].fillna("") if col_sit else ""
    df["VALOR_TOTAL_NUM"] = df[col_valor].apply(parse_valor_cell)

    logger.info("⚙️  Calculando descontos e parcelamentos por perfil de atraso...")
    resultados = df.apply(
        lambda row: calcular_desconto_proposta(row[col_dias], row["SITUACAO_TEMP"]), axis=1
    )

    cols_resultado = [
        'DESC_AVISTA_PERC', 'DESC_PARC_PERC', 'PERC_ENTRADA', 'NUM_PARCELAS',
        'MULTA_AVISTA', 'JUROS_AVISTA', 'MULTA_PARC', 'JUROS_PARC',
        'DESC_REPAR', 'ENT_REPAR', 'PARC_REPAR'
    ]
    df[cols_resultado] = pd.DataFrame(resultados.tolist(), index=df.index)

    # ── Valores calculados ────────────────────────────────────────────────────
    df["VLR_COM_MULTA_JUROS"] = (
        df["VALOR_TOTAL_NUM"] *
        (1 + df["MULTA_AVISTA"] / 100 * (df["MULTA_AVISTA"] > 0)) *
        (1 + df["JUROS_AVISTA"] / 100 * (df["JUROS_AVISTA"] > 0))
    ).round(2)

    df["VALOR_AVISTA"]     = (df["VLR_COM_MULTA_JUROS"] * (1 - df["DESC_AVISTA_PERC"] / 100)).round(2)

    df["VLR_PARC_BASE"]    = (
        df["VALOR_TOTAL_NUM"] *
        (1 + df["MULTA_PARC"] / 100 * (df["MULTA_PARC"] > 0)) *
        (1 + df["JUROS_PARC"] / 100 * (df["JUROS_PARC"] > 0)) *
        (1 - df["DESC_PARC_PERC"] / 100)
    ).round(2)
    df["ENTRADA_PARC"]     = (df["VLR_PARC_BASE"] * df["PERC_ENTRADA"]).round(2)

    relatorio_casos = df.apply(
        lambda row: (
            f"Desc AV: {row['DESC_AVISTA_PERC']}% | "
            f"Desc Parc: {row['DESC_PARC_PERC']}% | "
            f"Multa AV: {row['MULTA_AVISTA']}% | "
            f"Juros AV: {row['JUROS_AVISTA']}%"
        ), axis=1
    ).value_counts()

    # ── Remove colunas internas ───────────────────────────────────────────────
    colunas_remover = (
        ["SITUACAO_TEMP", "VALOR_TOTAL_NUM", "VLR_COM_MULTA_JUROS", "VLR_PARC_BASE"] +
        cols_resultado +
        [c for c in df.columns if 'UNNAMED' in c.upper()]
    )
    df_base = df.drop(columns=[c for c in colunas_remover if c in df.columns]).reset_index(drop=True)

    # ── Adiciona colunas de proposta (sem pagamento à vista) ──────────────────
    df_base["VALOR_AVISTA"]    = df["VALOR_AVISTA"]

    df_base["VALOR_PARC_BASE"] = df["VLR_PARC_BASE"].where(df["NUM_PARCELAS"] > 0, 0).round(2)
    df_base["ENTRADA_PARC"]    = df["ENTRADA_PARC"].where(df["NUM_PARCELAS"] > 0, 0).round(2)
    df_base["SALDO_PARC"]      = ""
    for i in range(1, 7):
        df_base[f"PARCELA_{i}X"] = ""

    df_base["ENT_REPARCELAMENTO"] = (df["VALOR_TOTAL_NUM"] * df["ENT_REPAR"]).round(2).where(df["PARC_REPAR"] > 0, 0)
    df_base["SALDO_REPAR"]        = ""
    for i in range(1, 7):
        df_base[f"REPAR_{i}X"] = ""

    colunas_proposta = (
        ["VALOR_AVISTA", "VALOR_PARC_BASE", "ENTRADA_PARC", "SALDO_PARC"] +
        [f"PARCELA_{i}X" for i in range(1, 7)] +
        ["ENT_REPARCELAMENTO", "SALDO_REPAR"] +
        [f"REPAR_{i}X" for i in range(1, 7)]
    )
    colunas_base = [c for c in df_base.columns if c not in colunas_proposta]
    df_base = df_base[colunas_base + colunas_proposta]

    # ── Monta fórmulas Excel ──────────────────────────────────────────────────
    idx_val_parc   = df_base.columns.get_loc("VALOR_PARC_BASE") + 1
    idx_ent_parc   = df_base.columns.get_loc("ENTRADA_PARC") + 1
    idx_saldo_parc = df_base.columns.get_loc("SALDO_PARC") + 1

    col_valor_saida = col_valor if col_valor in df_base.columns else "VALOR_TOTAL"
    idx_val_total  = df_base.columns.get_loc(col_valor_saida) + 1
    idx_ent_repar  = df_base.columns.get_loc("ENT_REPARCELAMENTO") + 1
    idx_saldo_repar= df_base.columns.get_loc("SALDO_REPAR") + 1

    # Pré-calcula as letras de coluna (são iguais para todas as linhas)
    col_val_parc  = get_column_letter(idx_val_parc)
    col_ent_parc  = get_column_letter(idx_ent_parc)
    col_s_parc    = get_column_letter(idx_saldo_parc)

    col_val_total = get_column_letter(idx_val_total)
    col_ent_repar = get_column_letter(idx_ent_repar)
    col_s_repar   = get_column_letter(idx_saldo_repar)

    for row_idx in range(len(df_base)):
        lin = row_idx + 2


        # SALDO_PARC = valor parcelável total - entrada do parcelamento
        df_base.at[row_idx, "SALDO_PARC"]  = f'=MAX(0,{col_val_parc}{lin}-{col_ent_parc}{lin})'

        # SALDO_REPAR = valor original - entrada do reparcelamento (referência à célula Excel)
        df_base.at[row_idx, "SALDO_REPAR"] = f'=MAX(0,{col_val_total}{lin}-{col_ent_repar}{lin})'

        num_parcelas_total = int(df["NUM_PARCELAS"].iloc[row_idx] or 0)
        num_parcelas_repar = int(df["PARC_REPAR"].iloc[row_idx] or 0)

        for i in range(1, 7):
            if num_parcelas_total > 0 and i <= num_parcelas_total:
                df_base.at[row_idx, f"PARCELA_{i}X"] = f"={col_s_parc}{lin}/{i}"
            else:
                df_base.at[row_idx, f"PARCELA_{i}X"] = ""
            if num_parcelas_repar > 0 and i <= num_parcelas_repar:
                df_base.at[row_idx, f"REPAR_{i}X"] = f"={col_s_repar}{lin}/{i}"
            else:
                df_base.at[row_idx, f"REPAR_{i}X"] = ""

    # ── Salva Excel ───────────────────────────────────────────────────────────
    data_hora     = datetime.now().strftime("%d-%m-%Y_%Hh%Mm")
    arquivo_saida = f"{dir_saida}/Negociacao_{data_hora}.xlsx"

    logger.info("💾 Montando planilha final com fórmulas dinâmicas...")
    try:
        os.makedirs(dir_saida, exist_ok=True)
        with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
            df_base.to_excel(writer, index=False, sheet_name='Negociacao')
            ws = writer.sheets['Negociacao']

            fmt_moeda = '"R$" #,##0.00_-'
            fmt_pct   = '0"%"'
            idx_primeiro_val = df_base.columns.get_loc("VALOR_AVISTA") + 1
            for col_idx in range(idx_primeiro_val, df_base.shape[1] + 1):
                letra = get_column_letter(col_idx)
                for cell in ws[letra]:
                    if cell.row > 1:
                        cell.number_format = fmt_moeda

            alin = Alignment(horizontal='center', vertical='center')
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    cell.alignment = alin
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except: pass
                ws.column_dimensions[col_letter].width = max_len + 3

        logger.info("✅ Planilha de negociação gerada com sucesso!")
        logger.info("--- RESUMO DE ENQUADRAMENTO ---")
        for caso, qtd in relatorio_casos.items():
            logger.info(f"  🔹 {caso}: {qtd} faturas ({qtd/total_faturas*100:.1f}%)")
        logger.info("-------------------------------")
        return arquivo_saida

    except Exception as e:
        logger.error(f"❌ ERRO ao salvar planilha de negociação: {e}")
        return None

# ---------------------------
# Processamento CSV -> Excel
# ---------------------------
def processar_csv(arquivo_8104=None, arquivo_8162=None, corte_dias=None):
    """
    Processa os CSVs e gera a planilha Excel.

    Modo GUI  : chamado com arquivo_8104, arquivo_8162, corte_dias já preenchidos.
    Modo terminal : chamado sem parâmetros — abre diálogos interativos.
    """
    global CORTE_DIAS

    if arquivo_8104 is not None:
        # ── Modo GUI: sem diálogos, sem input() bloqueante ────────────────────
        if not arquivo_8104:
            logger.error("❌ Nenhum arquivo principal fornecido. Operação cancelada.")
            return
        if corte_dias is not None:
            CORTE_DIAS = int(corte_dias)
            logger.info(f"[OK] Regra de corte: {CORTE_DIAS} dias.")
        if not arquivo_8162:
            logger.warning("[ATENCAO] Sem base secundaria - NR_IMOVEL ficara vazio.")
    else:
        # ── Modo terminal: comportamento original com diálogos ────────────────
        print("\n" + "="*50)
        print(" 📂 SELEÇÃO DE ARQUIVOS E PARÂMETROS")
        print("="*50)

        print("Abrindo janela para selecionar a base de dados (Macro 8104)...")
        arquivo_8104 = selecionar_arquivo("Selecione a base principal (Macro 8104)")
        if not arquivo_8104:
            print("❌ Nenhum arquivo principal selecionado. Operação cancelada.")
            return

        print("Abrindo janela para selecionar o cruzamento (Macro 8162)...")
        arquivo_8162 = selecionar_arquivo("Selecione a base secundária (Macro 8162)")
        if not arquivo_8162:
            print("⚠️ Nenhuma base secundária selecionada. O Número do Imóvel ficará em branco.")

        entrada_corte = input(
            f"\nDigite a quantidade de dias para o corte\n"
            f"[Aperte ENTER para usar o padrão de {CORTE_DIAS} dias]: "
        ).strip()
        if entrada_corte.isdigit():
            CORTE_DIAS = int(entrada_corte)
            logger.info(f"[OK] Regra de corte atualizada para: {CORTE_DIAS} dias.")
        else:
            logger.info(f"ℹ️ Nenhum valor digitado. Mantendo o corte padrão de: {CORTE_DIAS} dias.")
        print("="*50 + "\n")

    logger.info("[INFO] Esquentando os motores... preparando o processamento.")

    df0 = ler_csv_com_fallback(arquivo_8104)
    normalize_df_columns(df0)
    # Normaliza colunas para o padrão (suporte a diferentes cidades)
    df0 = normalizar_colunas_para_padrao(df0)

    cpf_main_col = next((c for c in df0.columns if c in ["CPF_CONSUMIDOR", "CPF", "CPF_CNPJ", "CPF_CNPJ_DA_FATURA"]), None)
    cnpj_main_col = next((c for c in df0.columns if c in ["CNPJ_CONSUMIDOR", "CNPJ"]), None)
    cpf_cnpj_main_col = next((c for c in df0.columns if c in ["CPF_CNPJ", "CPF_CNPJ_DA_FATURA"]), None)
    if cpf_main_col:
        df0["CPF CONSUMIDOR"] = (
            df0[cpf_main_col]
            .astype(str)
            .str.strip()
            .str.replace(r"\D", "", regex=True)
        )
        if cpf_main_col != "CPF_CONSUMIDOR":
            df0.drop(columns=[cpf_main_col], inplace=True, errors="ignore")
    if cnpj_main_col:
        df0["CNPJ CONSUMIDOR"] = (
            df0[cnpj_main_col]
            .astype(str)
            .str.strip()
            .str.replace(r"\D", "", regex=True)
        )
        if cnpj_main_col != "CNPJ_CONSUMIDOR":
            df0.drop(columns=[cnpj_main_col], inplace=True, errors="ignore")
    if not cpf_main_col and not cnpj_main_col and cpf_cnpj_main_col:
        df0["CPF CONSUMIDOR"] = (
            df0[cpf_cnpj_main_col]
            .astype(str)
            .str.strip()
            .str.replace(r"\D", "", regex=True)
        )
        logger.info(f"   [OK] CPF/CNPJ misto trazido da base principal ({cpf_cnpj_main_col}).")

    id_col = next(
        (
            c
            for c in df0.columns
            if c in ["MATRICULA", "LIGACAO", "ID_LIGACAO", "IDLIGACAO", "CONEXAO"]
        ),
        None,
    )
    if id_col is None:
        logger.critical("Nenhuma coluna de identificação encontrada (MATRICULA ou LIGACAO).")
        raise KeyError("Coluna obrigatória ausente: MATRÍCULA ou LIGAÇÃO")

    COL_NR_IMOVEL = "NR_IMOVEL"

    if arquivo_8162 and os.path.exists(arquivo_8162):
        logger.info("[INFO] Cruzando as bases de dados... aguarde.")
        try:
            df_macro = ler_csv_com_fallback(arquivo_8162)
            normalize_df_columns(df_macro)
            
            col_macro_ligacao = next(
                (c for c in df_macro.columns if c in ["LIGACAO", "MATRICULA", "ID_LIGACAO", "IDLIGACAO", "CONEXAO"]),
                None,
            )
            col_macro_nr = next((c for c in df_macro.columns if c in ["NR_IMOVEL", "NUMERO_IMOVEL", "NUMERO_DO_IMOVEL"]), None)
            col_macro_logradouro = next((c for c in df_macro.columns if c in ["LOGRADOURO_IMOVEL", "LOGRADOURO_ENTREGA", "LOGRADOURO", "ENDERECO_IMOVEL", "ENDERECO"]), None)
            col_macro_tipo_irregularidade = next((c for c in df_macro.columns if c in ["TIPO_IRREGULARIDADE", "TIPO_DE_IRREGULARIDADE"]), None)
            col_macro_cpf = next((c for c in df_macro.columns if c in ["CPF_CONSUMIDOR", "CPF"]), None)
            col_macro_cnpj = next((c for c in df_macro.columns if c in ["CNPJ_CONSUMIDOR", "CNPJ"]), None)
            col_macro_cidade = next((c for c in df_macro.columns if c in ["CIDADE_IMOVEL", "CIDADE", "MUNICIPIO"]), None)
            col_macro_email = next((c for c in df_macro.columns if c in ["EMAIL_CONSUMIDOR", "EMAIL", "E_MAIL"]), None)
            col_macro_movel = next((c for c in df_macro.columns if c in ["TEL_MOVEL_CONSUMIDOR", "TELEFONE_CELULAR", "CELULAR", "TEL_MOVEL"]), None)
            col_macro_fixo  = next((c for c in df_macro.columns if c in ["TEL_FIXO_CONSUMIDOR", "TEL_FIXO", "TELEFONE_FIXO", "FIXO"]), None)
            
            if col_macro_ligacao and col_macro_nr:
                # Otimização: usar set_index para merge mais rápido
                df_macro_red = df_macro.copy()
                macro_prefix = "MACRO_8162_"
                renamed_columns = {}
                for col in df_macro_red.columns:
                    if col not in {col_macro_ligacao, col_macro_nr}:
                        renamed_columns[col] = f"{macro_prefix}{col}"
                df_macro_red.rename(columns=renamed_columns, inplace=True)

                if col_macro_logradouro:
                    col_macro_logradouro = renamed_columns.get(col_macro_logradouro, col_macro_logradouro)
                if col_macro_tipo_irregularidade:
                    col_macro_tipo_irregularidade = renamed_columns.get(
                        col_macro_tipo_irregularidade,
                        col_macro_tipo_irregularidade,
                    )
                if col_macro_cpf:
                    col_macro_cpf = renamed_columns.get(col_macro_cpf, col_macro_cpf)
                if col_macro_cnpj:
                    col_macro_cnpj = renamed_columns.get(col_macro_cnpj, col_macro_cnpj)
                if col_macro_cidade:
                    col_macro_cidade = renamed_columns.get(col_macro_cidade, col_macro_cidade)
                if col_macro_email:
                    col_macro_email = renamed_columns.get(col_macro_email, col_macro_email)
                if col_macro_movel:
                    col_macro_movel = renamed_columns.get(col_macro_movel, col_macro_movel)
                if col_macro_fixo:
                    col_macro_fixo = renamed_columns.get(col_macro_fixo, col_macro_fixo)

                df_macro_red[col_macro_ligacao] = df_macro_red[col_macro_ligacao].astype(str).str.strip()
                df_macro_red = df_macro_red.drop_duplicates(subset=[col_macro_ligacao], keep='first')
                df_macro_red.set_index(col_macro_ligacao, inplace=True)

                df0_indexed = df0.copy()
                df0_indexed[id_col] = df0_indexed[id_col].astype(str).str.strip()
                df0_indexed = df0_indexed.set_index(id_col)

                # So traz da 8162 o necessario para enriquecer o relatorio final:
                # numero do imovel e logradouro/endereco.
                join_columns = [col_macro_nr]
                if col_macro_logradouro:
                    join_columns.append(col_macro_logradouro)
                if col_macro_tipo_irregularidade:
                    join_columns.append(col_macro_tipo_irregularidade)
                if col_macro_cpf:
                    join_columns.append(col_macro_cpf)
                if col_macro_cnpj:
                    join_columns.append(col_macro_cnpj)
                if col_macro_cidade:
                    join_columns.append(col_macro_cidade)
                if col_macro_email:
                    join_columns.append(col_macro_email)
                if col_macro_movel:
                    join_columns.append(col_macro_movel)
                if col_macro_fixo:
                    join_columns.append(col_macro_fixo)

                df0 = df0_indexed.join(df_macro_red[join_columns], how='left')
                df0 = df0.rename(columns={col_macro_nr: COL_NR_IMOVEL})
                if col_macro_cpf:
                    df0["CPF CONSUMIDOR"] = df0.get(col_macro_cpf, "").fillna("")
                    df0.drop(columns=[col_macro_cpf], inplace=True, errors="ignore")
                    logger.info(f"   [OK] CPF CONSUMIDOR trazido da macro secundaria ({col_macro_cpf}).")

                if col_macro_logradouro:
                    if COL_ENDERECO not in df0.columns:
                        df0[COL_ENDERECO] = ""
                    df0[COL_ENDERECO] = df0[col_macro_logradouro].fillna(df0[COL_ENDERECO])
                    if col_macro_logradouro != COL_ENDERECO:
                        df0.drop(columns=[col_macro_logradouro], inplace=True)
                    logger.info(f"   [OK] Endereco preenchido pela macro secundaria ({col_macro_logradouro}).")

                if col_macro_cnpj:
                    df0["CNPJ CONSUMIDOR"] = df0.get(col_macro_cnpj, "").fillna("")
                    if col_macro_cnpj != "CNPJ_CONSUMIDOR":
                        df0.drop(columns=[col_macro_cnpj], inplace=True, errors="ignore")
                    logger.info(f"   [OK] CNPJ CONSUMIDOR trazido da macro secundaria ({col_macro_cnpj}).")
                    
                if col_macro_cidade:
                    df0["CIDADE_IMOVEL"] = df0.get(col_macro_cidade, "").fillna("")
                    if col_macro_cidade != "CIDADE_IMOVEL":
                        df0.drop(columns=[col_macro_cidade], inplace=True, errors="ignore")
                    logger.info(f"   [OK] CIDADE_IMOVEL trazida da macro secundaria ({col_macro_cidade}).")

                if col_macro_email:
                    if COL_EMAIL not in df0.columns:
                        df0[COL_EMAIL] = ""
                    df0[COL_EMAIL] = df0[COL_EMAIL].replace(['', 'nan', 'NaN', 'None'], pd.NA).fillna(df0[col_macro_email]).fillna('')
                    df0.drop(columns=[col_macro_email], inplace=True, errors="ignore")
                    logger.info(f"   [OK] E-mails vazios preenchidos pela macro 8162.")

                if col_macro_movel:
                    if COL_TEL not in df0.columns:
                        df0[COL_TEL] = ""
                    df0[COL_TEL] = df0[COL_TEL].replace(['', 'nan', 'NaN', 'None'], pd.NA).fillna(df0[col_macro_movel]).fillna('')
                    df0.drop(columns=[col_macro_movel], inplace=True, errors="ignore")
                    logger.info(f"   [OK] Celulares vazios preenchidos pela macro 8162.")
                else:
                    if COL_TEL not in df0.columns:
                        df0[COL_TEL] = ''

                if col_macro_fixo:
                    df0[COL_FIXO] = df0[col_macro_fixo].fillna('')
                    df0.drop(columns=[col_macro_fixo], inplace=True, errors="ignore")
                    logger.info(f"   [OK] Tel. Fixo trazido da macro 8162 ({col_macro_fixo}).")
                else:
                    df0[COL_FIXO] = ''

                if col_macro_tipo_irregularidade:
                    df0 = df0.rename(columns={col_macro_tipo_irregularidade: COL_TIPO_IRREGULARIDADE})

                df0.reset_index(inplace=True)
                logger.info(f"   [OK] {int((df0[COL_NR_IMOVEL].notna()).sum())} imoveis vinculados com sucesso.")
            else:
                logger.warning("   [ATENCAO] Colunas da macro secundaria nao reconhecidas - seguindo sem NR_IMOVEL.")
        except Exception as e:
            logger.exception(f"Erro ao processar a macro secundária: {e}")
    else:
        logger.warning("   [ATENCAO] Sem macro secundaria - campo NR_IMOVEL ficara vazio.")

    if DIVIDA_COL in df0.columns:
        mask_divida = df0[DIVIDA_COL].astype(str).str.strip().str.upper() == "SIM"
        removidos_divida = int((~mask_divida).sum())
        df0 = df0[mask_divida].copy()
        logger.info(f"   {removidos_divida} registros sem divida ativa removidos.")
    else:
        logger.warning(f"[ATENCAO] Coluna '{DIVIDA_COL}' nao encontrada. Nenhum filtro de divida aplicado.")

    colunas_obrigatorias = [COL_EMAIL, COL_CLIENTE, COL_DIAS, COL_TEL, id_col]
    for col in colunas_obrigatorias:
        if col not in df0.columns:
            logger.critical(f"Coluna obrigatória ausente: {col}")
            raise KeyError(f"Coluna obrigatória ausente: {col}")

    if COL_FIXO not in df0.columns:
        df0[COL_FIXO] = ''

    total_bruto = len(df0)
    logger.info(f"[INFO] Base carregada: {total_bruto} registros para analisar.")

    df = df0.copy()

    # Otimização: usar .str.contains() em vez de .apply() para detecção rápida
    proibidas_pattern = "|".join(["PREFEITURA", "PREFEITO", "BOMBEIRO", "BOMBEIROS", "CORPO"])
    mask_proibidos = df[COL_CLIENTE].fillna("").astype(str).str.upper().str.contains(proibidas_pattern, na=False, regex=True)
    removidos_proibidos = int(mask_proibidos.sum())
    if removidos_proibidos > 0:
        logger.info(f"   {removidos_proibidos} registros bloqueados (prefeitura/bombeiros etc).")
    df = df[~mask_proibidos].copy()

    df[COL_EMAIL] = df[COL_EMAIL].astype(str).str.strip().str.lower()
    email_mask = df[COL_EMAIL].str.match(EMAIL_REGEX, na=False)

    df_invalid_email = df[~email_mask].copy()
    logger.info(f"[INFO] Validando e-mails... {len(df_invalid_email)} invalidos encontrados.")

    if not df_invalid_email.empty:
        logger.info(f"   Normalizando {len(df_invalid_email)} telefones (pode demorar)...")
        df_invalid_email["_TEL_ORIGINAL"] = df_invalid_email[COL_TEL].copy()
        df_invalid_email[COL_TEL] = df_invalid_email[COL_TEL].apply(normalizar_telefone)
        antes_invalid = len(df_invalid_email)
        df_invalid_email = df_invalid_email[df_invalid_email[COL_TEL].notna()].copy()
        logger.info(f"   [OK] {len(df_invalid_email)} dos {antes_invalid} tem telefone valido - salvos para WhatsApp.")

    df = df[email_mask].copy()
    removidos_email = int((~email_mask).sum())
    logger.info(f"   {removidos_email} registros redirecionados para aba de inválidos.")

    df[COL_DIAS] = pd.to_numeric(df[COL_DIAS], errors="coerce")
    if not df_invalid_email.empty:
        df_invalid_email[COL_DIAS] = pd.to_numeric(df_invalid_email[COL_DIAS], errors="coerce")

    if POLITICA_NAN_DIAS == "zerar":
        df[COL_DIAS] = df[COL_DIAS].fillna(0)
        if not df_invalid_email.empty:
            df_invalid_email[COL_DIAS] = df_invalid_email[COL_DIAS].fillna(0)
    elif POLITICA_NAN_DIAS == "remover":
        antes = len(df)
        df = df[df[COL_DIAS].notna()]
        logger.info(f"   {antes - len(df)} registros com dias inválidos descartados.")
        if not df_invalid_email.empty:
            antes_inv = len(df_invalid_email)
            df_invalid_email = df_invalid_email[df_invalid_email[COL_DIAS].notna()]
            logger.info(f"   {antes_inv - len(df_invalid_email)} registros com dias inválidos descartados (aba inválidos).")

    df["_TEL_ORIGINAL"] = df[COL_TEL].copy()
    logger.info(f"[INFO] Normalizando {len(df)} telefones (pode demorar)...")
    df[COL_TEL] = df[COL_TEL].apply(normalizar_telefone)

    telefones_invalidos = int(df[COL_TEL].isna().sum())
    telefones_validos = int(df[COL_TEL].notna().sum())
    logger.info(f"   [OK] {telefones_validos} telefones validos prontos.")
    if telefones_invalidos > 0:
        logger.info(f"   [ATENCAO] {telefones_invalidos} telefones invalidos descartados.")

    if COL_NR_IMOVEL not in df.columns:
        df[COL_NR_IMOVEL] = ""
    if COL_NR_IMOVEL not in df_invalid_email.columns:
        df_invalid_email[COL_NR_IMOVEL] = ""
    if COL_TIPO_IRREGULARIDADE not in df.columns:
        df[COL_TIPO_IRREGULARIDADE] = ""
    if COL_TIPO_IRREGULARIDADE not in df_invalid_email.columns:
        df_invalid_email[COL_TIPO_IRREGULARIDADE] = ""
    if COL_SITUACAO_LIGACAO_FASE not in df.columns:
        df[COL_SITUACAO_LIGACAO_FASE] = ""
    if COL_SITUACAO_LIGACAO_FASE not in df_invalid_email.columns:
        df_invalid_email[COL_SITUACAO_LIGACAO_FASE] = ""

    col_valor = None
    for cand in ("VALOR", "VALOR_R$", "VALOR_FATURA", "VALOR_DA_FATURA", "VALOR_TOTAL", "VLR"):
        if cand in df.columns:
            col_valor = cand
            break
    if not col_valor:
        logger.warning("   [ATENCAO] Coluna de valor nao localizada - VALOR_DA_FATURA zerado.")
        df["VALOR_DA_FATURA"] = 0
        if not df_invalid_email.empty:
            df_invalid_email["VALOR_DA_FATURA"] = 0
        col_valor = "VALOR_DA_FATURA"

    df["VALOR_NUM"] = df[col_valor].apply(parse_valor_cell)
    if not df_invalid_email.empty:
        df_invalid_email["VALOR_NUM"] = df_invalid_email[col_valor].apply(parse_valor_cell)

    if not df_invalid_email.empty:
        df_bruta_pdd = pd.concat([df.copy(), df_invalid_email.copy()], ignore_index=True, sort=False)
    else:
        df_bruta_pdd = df.copy()

    def _join_ids(values):
        unique_ids = []
        seen = set()
        for value in values:
            if pd.isna(value):
                continue
            item = str(value).strip()
            if not item:
                continue
            if item not in seen:
                seen.add(item)
                unique_ids.append(item)
        return ";".join(unique_ids)

    def agrupar_clientes(df_base: pd.DataFrame) -> pd.DataFrame:
        logger.info(f"   [INFO] Agrupando {len(df_base)} registros por cliente...")
        agg_dict = {
            COL_CLIENTE: (COL_CLIENTE, "first"),
            "CPF CONSUMIDOR": ("CPF CONSUMIDOR", "first") if "CPF CONSUMIDOR" in df_base.columns else ("CPF CONSUMIDOR", "first"),
            "CNPJ CONSUMIDOR": ("CNPJ CONSUMIDOR", "first") if "CNPJ CONSUMIDOR" in df_base.columns else ("CNPJ CONSUMIDOR", "first"),
            COL_EMAIL: (COL_EMAIL, "first"),
            COL_TEL: (COL_TEL, "first"),
            COL_FIXO: (COL_FIXO, "first") if COL_FIXO in df_base.columns else (COL_FIXO, "first"),
            COL_ENDERECO: (COL_ENDERECO, "first"),
            COL_NR_IMOVEL: (COL_NR_IMOVEL, "first"),
            COL_TIPO_IRREGULARIDADE: (COL_TIPO_IRREGULARIDADE, "first"),
            COL_SITUACAO_LIGACAO_FASE: (COL_SITUACAO_LIGACAO_FASE, "first"),
            DIVIDA_COL: (DIVIDA_COL, "first") if DIVIDA_COL in df_base.columns else (DIVIDA_COL, "first"),
            "QTD_FATURAS": (COL_DIAS, "count"),
            COL_DIAS: (COL_DIAS, "max"),
            "VALOR_TOTAL": ("VALOR_NUM", "sum"),
        }
        if "ID_FATURA" in df_base.columns:
            agg_dict["ID_FATURA"] = ("ID_FATURA", _join_ids)
        
        if "GRUPO" in df_base.columns:
            agg_dict["GRUPO"] = ("GRUPO", "first")

        if "CIDADE_IMOVEL" in df_base.columns:
            agg_dict["CIDADE_IMOVEL"] = ("CIDADE_IMOVEL", "first")

        agrupado = df_base.groupby(id_col, as_index=False).agg(**agg_dict)

        agrupado["VALOR_TOTAL"] = agrupado["VALOR_TOTAL"].fillna(0.0).astype(float).round(2)
        agrupado[COL_DIAS] = agrupado[COL_DIAS].fillna(0).astype(int)

        if COL_DIAS in agrupado.columns:
            agrupado = agrupado.rename(columns={COL_DIAS: "QTDE_DIAS_EM_ATRASO"})

        agrupado["STATUS"] = ""

        # Garante o layout final do relatório, mesmo quando alguma coluna
        # opcional não vier da base original.
        colunas_saida = [
            id_col,
            COL_CLIENTE,
            "CPF CONSUMIDOR",
            "CNPJ CONSUMIDOR",
            "ID_FATURA",
            COL_EMAIL,
            "QTDE_DIAS_EM_ATRASO",
            "PDD",
            COL_SITUACAO_LIGACAO_FASE,
            "QTD_FATURAS",
            COL_TEL,
            COL_FIXO,
            COL_ENDERECO,
            COL_NR_IMOVEL,
            "CIDADE_IMOVEL",
            COL_TIPO_IRREGULARIDADE,
            "GRUPO",
            DIVIDA_COL,
            "VALOR_TOTAL",
            "STATUS",
        ]
        for col in colunas_saida:
            if col not in agrupado.columns:
                agrupado[col] = ""

        agrupado["PDD"] = agrupado["QTDE_DIAS_EM_ATRASO"].apply(calcular_faixa_pdd)

        logger.info(f"   [OK] {len(agrupado)} clientes unicos agrupados.")
        return agrupado[colunas_saida].copy()

    export_df = agrupar_clientes(df)
    export_invalid = agrupar_clientes(df_invalid_email) if not df_invalid_email.empty else pd.DataFrame(columns=export_df.columns)

    log_tabela_pdd(export_df, export_invalid)

    df_maior_corte = export_df[export_df["QTDE_DIAS_EM_ATRASO"] > CORTE_DIAS].copy()
    df_menor_corte = export_df[export_df["QTDE_DIAS_EM_ATRASO"] <= CORTE_DIAS].copy()

    # ── Coluna Cadastro nas sheets da 8104, usando somente os dados finais ───
    # Regra:
    #   Completo   = pelo menos 1 e-mail + pelo menos 1 telefone
    #   Parcial    = só e-mail OU só telefone
    #   Incompleto = nenhum dado
    def _aplicar_cadastro(df_sheet: pd.DataFrame) -> pd.Series:
        if df_sheet.empty:
            return pd.Series(dtype=object)

        tem_email = (
            df_sheet[COL_EMAIL].apply(_8162_tem_dado)
            if COL_EMAIL in df_sheet.columns
            else pd.Series(False, index=df_sheet.index)
        )
        tem_fixo = (
            df_sheet[COL_FIXO].apply(_8162_tem_dado)
            if COL_FIXO in df_sheet.columns
            else pd.Series(False, index=df_sheet.index)
        )
        tem_movel = (
            df_sheet[COL_TEL].apply(_8162_tem_dado)
            if COL_TEL in df_sheet.columns
            else pd.Series(False, index=df_sheet.index)
        )

        tem_telefone = tem_fixo | tem_movel
        status = pd.Series('Incompleto', index=df_sheet.index)
        status[(tem_email & ~tem_telefone) | (~tem_email & tem_telefone)] = 'Parcial'
        status[tem_email & tem_telefone] = 'Completo'
        return status

    df_maior_corte['Cadastro'] = _aplicar_cadastro(df_maior_corte).values
    df_menor_corte['Cadastro'] = _aplicar_cadastro(df_menor_corte).values
    if not export_invalid.empty:
        export_invalid = export_invalid.copy()
        export_invalid['Cadastro'] = _aplicar_cadastro(export_invalid).values

    logger.info("[INFO] Gerando graficos (pode demorar)...")
    gerar_graficos_combinados(df_maior_corte, "Maior que Corte", DIR_MAIOR_CORTE)
    gerar_graficos_combinados(df_menor_corte, "Menor ou Igual ao Corte", DIR_MENOR_CORTE)
    gerar_grafico_pdd_por_grupo(df_bruta_pdd, "PDD por Grupo - Base Bruta 8104", DIR_PDD)
    logger.info("   [OK] Graficos gerados com sucesso.")

    try:
        # === LAZY INITIALIZATION ===
        os.makedirs(PLANILHA_DIR, exist_ok=True)

        with pd.ExcelWriter(SAIDA_XLSX, engine="openpyxl") as writer:
            df_maior_corte.to_excel(writer, sheet_name="maior_que_corte", index=False)
            df_menor_corte.to_excel(writer, sheet_name="menor_ou_igual_corte", index=False)

            if not export_invalid.empty:
                export_invalid.to_excel(writer, sheet_name="emails_invalidos", index=False)
            else:
                pd.DataFrame(columns=export_df.columns).to_excel(writer, sheet_name="emails_invalidos", index=False)

            # ── Colorir MATRÍCULA e coluna Cadastro nas 3 sheets da 8104 ─────
            from openpyxl.styles import PatternFill as _PF, Font as _Ft, Border as _Bd, Side as _Sd
            _s = _Sd(style='thin')
            _brd = _Bd(left=_s, right=_s, top=_s, bottom=_s)
            _estilos = {
                'Completo':   (_PF('solid', fgColor='C6EFCE'), _Ft(color='276221', bold=True)),
                'Parcial':    (_PF('solid', fgColor='FFEB9C'), _Ft(color='9C6500', bold=True)),
                'Incompleto': (_PF('solid', fgColor='FFC7CE'), _Ft(color='9C0006', bold=True)),
            }

            def _colorir_sheet_8104(ws, df_sheet):
                if 'MATRICULA' not in df_sheet.columns and id_col not in df_sheet.columns:
                    return
                col_mat_name = 'MATRICULA' if 'MATRICULA' in df_sheet.columns else id_col
                col_cad_name = 'Cadastro'
                if col_mat_name not in df_sheet.columns or col_cad_name not in df_sheet.columns:
                    return
                mat_idx = df_sheet.columns.get_loc(col_mat_name) + 1
                cad_idx = df_sheet.columns.get_loc(col_cad_name) + 1
                for row_i, status_val in enumerate(df_sheet[col_cad_name], start=2):
                    estilo = _estilos.get(status_val, _estilos['Incompleto'])
                    fill, font = estilo
                    for col_i in (mat_idx, cad_idx):
                        cell = ws.cell(row=row_i, column=col_i)
                        cell.fill   = fill
                        cell.font   = font
                        cell.border = _brd

            for sheet_name, df_sheet in [
                ("maior_que_corte",       df_maior_corte),
                ("menor_ou_igual_corte",  df_menor_corte),
                ("emails_invalidos",      export_invalid if not export_invalid.empty else pd.DataFrame()),
            ]:
                if sheet_name in writer.sheets and not df_sheet.empty:
                    _colorir_sheet_8104(writer.sheets[sheet_name], df_sheet)

            # ── Relatório Cadastral 8162 (coloração de matrícula + gráficos) ──
            if arquivo_8162 and os.path.exists(arquivo_8162):
                try:
                    df_8162_rel = ler_csv_com_fallback(arquivo_8162)
                    normalize_df_columns(df_8162_rel)
                    # Trata alias LIGACAO -> MATRICULA
                    if 'LIGACAO' in df_8162_rel.columns and 'MATRICULA' not in df_8162_rel.columns:
                        df_8162_rel = df_8162_rel.rename(columns={'LIGACAO': 'MATRICULA'})
                    gerar_aba_relatorio_cadastral(df_8162_rel, writer)
                except Exception as _e_cad:
                    logger.warning(f"[CADASTRAL] Falha ao gerar aba cadastral: {_e_cad}")
            else:
                logger.info("[CADASTRAL] Base 8162 nao disponivel — aba cadastral omitida.")

        logger.info("[OK] Planilha Excel montada e salva com sucesso!")
    except PermissionError:
        logger.error("❌ Ops! A planilha de saída parece estar aberta. Por favor, feche o Excel e tente novamente.")
        raise RuntimeError("A planilha de destino está aberta em outro programa. Feche o Excel para podermos salvar os resultados.")
    except Exception as e:
        logger.exception("Erro inesperado ao criar a planilha final.")
        raise RuntimeError("Ocorreu um problema ao tentar salvar a planilha final. Verifique seu espaço em disco e se o arquivo não está corrompido.")

    try:
        total_maior = len(df_maior_corte)
        total_menor = len(df_menor_corte)
        total_invalid = len(export_invalid) if export_invalid is not None else 0

        logger.info("─" * 52)
        logger.info(f"📦 Registros carregados       : {total_bruto}")
        logger.info(f"✅ Após limpeza e filtros      : {len(export_df)} clientes únicos")
        logger.info(f"[ACIMA] Acima do corte ({CORTE_DIAS} dias)    : {total_maior} clientes")
        logger.info(f"[ABAIXO] Abaixo do corte ({CORTE_DIAS} dias)   : {total_menor} clientes")
        logger.info(f"[WHATSAPP] So pelo WhatsApp (sem email) : {total_invalid} clientes")
        logger.info("─" * 52)
    except Exception:
        logger.exception("Erro ao gerar resumo final.")

    logger.info("[FIM] Tudo pronto! Processamento concluido com sucesso.")

# ---------------------------
# Excel sheet selection
# ---------------------------
def listar_abas_excel(path: str):
    if not os.path.exists(path):
        logger.warning(f"Arquivo '{path}' não encontrado.")
        return []
    try:
        xls = pd.ExcelFile(path)
        return xls.sheet_names
    except Exception:
        logger.exception("Falha ao ler abas do Excel.")
        return []

def escolher_sheet_para_envio(caminho_arquivo) -> Optional[List[str]]:
    abas = listar_abas_excel(caminho_arquivo)
    if not abas:
        logger.error(f"Arquivo '{caminho_arquivo}' não encontrado ou inválido.")
        return None

    print("\nAbas disponíveis para envio:")
    for i, s in enumerate(abas, start=1):
        print(f"{i} - {s}")
    print("A - Todas as sheets")
    print("0 - Voltar ao menu")

    escolha = input("Digite o número da aba que deseja usar para envio (ou 'A' para todas): ").strip().lower()
    if escolha == "0":
        return None
    if escolha == "a":
        return abas
    if escolha.isdigit():
        idx = int(escolha) - 1
        if 0 <= idx < len(abas):
            return [abas[idx]]
    print("Escolha inválida. Usando todas as sheets por padrão.")
    return abas

def carregar_planilha_para_envio(sheet_names: List[str], caminho_arquivo: str) -> dict:
    try:
        df_dict = pd.read_excel(caminho_arquivo, sheet_name=sheet_names, dtype=str, engine="openpyxl")
        if isinstance(df_dict, dict):
            for k, v in df_dict.items():
                logger.info(f"   Aba '{k}' carregada: {len(v)} registros.")
            return df_dict
        else:
            logger.info(f"   Aba carregada: {len(df_dict)} registros.")
            return {sheet_names[0]: df_dict}
    except Exception:
        logger.exception("Falha ao carregar planilha para envio.")
        return {}


# ---------------------------
# Selenium helpers for WhatsApp Web
# ---------------------------
def obter_texto_ultimo_balao_saida(driver) -> Optional[str]:
    logger.debug("Verificando último balão de saída no WhatsApp Web.")
    try:
        xpath_candidates = [
            "(//div[@data-testid='msg-container']//div[contains(@class,'message-out')]//span[@dir='ltr'])[last()]",
            "(//div[contains(@class,'message-out')]//span[@dir='ltr'])[last()]",
            "(//div[@data-testid='msg-container']//span[@dir='ltr'])[last()]",
            "(//span[@dir='ltr'])[last()]",
        ]
        for xp in xpath_candidates:
            try:
                el = driver.find_element(By.XPATH, xp)
                text = driver.execute_script("return arguments[0].innerText || arguments[0].textContent;", el)
                if text and text.strip():
                    return text.strip()
            except NoSuchElementException:
                continue
    except Exception as e:
        logger.debug(f"Falha ao ler balão de saída: {e}")
    return None

def clicar_continuar_whatsapp_web(driver, wait) -> bool:
    try:
        cont_btn = driver.find_element(
            By.XPATH, "//a[contains(., 'Continuar para o WhatsApp Web') or contains(., 'Continuar para o WhatsApp Web')]"
        )
        try:
            cont_btn.click()
        except ElementClickInterceptedException:
            driver.execute_script("arguments[0].click();", cont_btn)
        logger.info("🔄 Botão 'Continuar para o WhatsApp Web' clicado com sucesso.")
        return True
    except NoSuchElementException:
        return False
    except Exception as e:
        logger.warning(f"Falha ao clicar em 'Continuar para o WhatsApp Web': {e}")
        return False

def tentar_enviar(driver, wait, mensagem_texto: str) -> bool:
    max_attempts = 3
    msg_limpa = re.sub(r'\W+', '', mensagem_texto).lower()
    logger.debug(f"Iniciando envio de mensagem ({len(mensagem_texto)} chars, até {max_attempts} tentativas).")

    for attempt in range(1, max_attempts + 1):
        if attempt > 1:
            logger.warning(f"   Tentativa {attempt}/{max_attempts} de envio.")
        try:
            js_focus = (
                "(function(){"
                "const els = Array.from(document.querySelectorAll('#main div[contenteditable=\"true\"]'));"
                "if (els.length === 0) return false;"
                "const target = els[els.length - 1];"
                "target.focus();"
                "target.scrollIntoView({block: 'center', inline: 'center'});"
                "return true;"
                "})();"
            )
            try:
                driver.execute_script(js_focus)
                time.sleep(0.15)
            except Exception:
                pass

            input_box = None
            try:
                els = driver.find_elements(By.XPATH, "//*[@id='main']//div[@contenteditable='true']")
                if els:
                    input_box = els[-1]
                    driver.execute_script("arguments[0].click();", input_box)
                    time.sleep(1.5) 
            except Exception:
                input_box = None

            if input_box:
                try:
                    driver.execute_script("arguments[0].focus();", input_box)
                    input_box.send_keys(Keys.CONTROL, "a")
                    input_box.send_keys(Keys.BACKSPACE)
                    time.sleep(0.5)

                    linhas = mensagem_texto.splitlines()
                    for i, linha in enumerate(linhas):
                        input_box.send_keys(linha)
                        if i < len(linhas) - 1:
                            input_box.send_keys(Keys.SHIFT, Keys.ENTER)
                    input_box.send_keys(Keys.ENTER)

                except Exception:
                    try:
                        js_set = (
                            "arguments[0].innerText = arguments[1];"
                            "arguments[0].dispatchEvent(new InputEvent('input', {bubbles: true}));"
                        )
                        driver.execute_script(js_set, input_box, mensagem_texto)
                        try:
                            btn = driver.find_element(By.XPATH, "//button[@data-testid='compose-btn-send']")
                            driver.execute_script("arguments[0].click();", btn)
                        except Exception:
                            pass
                    except Exception:
                        pass

            end_time = time.time() + 6
            while time.time() < end_time:
                last_text = obter_texto_ultimo_balao_saida(driver)
                if last_text:
                    last_limpo = re.sub(r'\W+', '', last_text).lower()
                    if msg_limpa[:40] in last_limpo:
                        return True
                time.sleep(0.4)

            try:
                btn = driver.find_element(By.XPATH, "//button[@data-testid='compose-btn-send']")
                driver.execute_script("arguments[0].click();", btn)
            except NoSuchElementException:
                try:
                    btn_icon = driver.find_element(By.XPATH, "//span[@data-icon='send']")
                    driver.execute_script("arguments[0].click();", btn_icon)
                except NoSuchElementException:
                    pass
            except Exception:
                pass

            end_time = time.time() + 4
            while time.time() < end_time:
                last_text = obter_texto_ultimo_balao_saida(driver)
                if last_text:
                    last_limpo = re.sub(r'\W+', '', last_text).lower()
                    if msg_limpa[:40] in last_limpo:
                        logger.debug(f"Mensagem confirmada no balão de saída (tentativa {attempt}).")
                        return True
                time.sleep(0.4)

        except Exception as e:
            logger.warning(f"   Exceção na tentativa {attempt}: {e}")
            time.sleep(1 + attempt)

    logger.error("❌ Todas as tentativas de envio falharam.")
    return False

def enviar_imagem_para_chat(driver, wait, image_path: str) -> bool:
    if not os.path.exists(image_path):
        logger.warning(f"Imagem não encontrada em: {image_path}. Pulando envio de imagem.")
        return False

    try:
        attach_selectors = [
            "//div[@title='Anexar' or @aria-label='Anexar' or @data-testid='clip']",
            "//span[@data-icon='clip']",
            "//div[@role='button' and contains(@aria-label,'Anexar')]",
            "//div[contains(@class,'_2xkOe') and .//span[@data-icon='clip']]"
        ]
        attached = False
        for xp in attach_selectors:
            try:
                el = driver.find_element(By.XPATH, xp)
                driver.execute_script("arguments[0].click();", el)
                time.sleep(0.4)
                attached = True
                break
            except Exception:
                continue

        file_input = None
        file_selectors = [
            "input[type='file']",
            "//input[@type='file' and contains(@accept,'image')]",
            "//input[@accept and contains(@accept,'image')]",
            "//input[contains(@accept,'image') and @type='file']"
        ]
        for sel in file_selectors:
            try:
                if sel.startswith("//"):
                    file_input = driver.find_element(By.XPATH, sel)
                else:
                    file_input = driver.find_element(By.CSS_SELECTOR, sel)
                if file_input:
                    break
            except Exception:
                file_input = None

        if file_input is None:
            try:
                js = (
                    "var inp = document.createElement('input');"
                    "inp.type = 'file';"
                    "inp.accept = 'image/*,video/*';"
                    "inp.style.display = 'block';"
                    "document.body.appendChild(inp);"
                    "return inp;"
                )
                driver.execute_script(js)
                try:
                    file_input = driver.find_elements(By.CSS_SELECTOR, "input[type='file']")[-1]
                except Exception:
                    file_input = None
            except Exception:
                file_input = None

        if not file_input:
            logger.warning("Campo de upload de arquivo não encontrado no DOM do WhatsApp Web.")
            return False

        try:
            file_input.send_keys(image_path)
        except Exception as e:
            logger.exception(f"Falha ao enviar caminho para input[type=file]: {e}")
            return False

        send_btn_selectors = [
            "//div[@role='button' and @data-testid='attach-image-send']",
            "//button[@data-testid='compose-btn-send']",
            "//span[@data-icon='send']",
            "//div[@role='button' and contains(@aria-label,'Enviar')]"
        ]
        sent_confirmed = False
        end_time = time.time() + 8
        while time.time() < end_time:
            for xp in send_btn_selectors:
                try:
                    btn = driver.find_element(By.XPATH, xp)
                    if btn and btn.is_displayed():
                        try:
                            driver.execute_script("arguments[0].click();", btn)
                        except Exception:
                            try:
                                btn.click()
                            except Exception:
                                pass
                        sent_confirmed = True
                        break
                except Exception:
                    continue
            if sent_confirmed:
                break
            time.sleep(0.4)

        if sent_confirmed:
            time.sleep(1.2)
            return True
        else:
            logger.warning("Não foi possível confirmar envio da imagem (botão de enviar não encontrado).")
            return False
    except Exception as e:
        logger.exception(f"Erro ao enviar imagem: {e}")
        return False


def enviar_documento_para_chat(driver, wait, doc_path: str, doc_type: str = "pdf") -> bool:
    """
    Envia um documento (PDF, DOCX, etc) para o chat do WhatsApp Web de forma robusta.
    
    Args:
        driver: Selenium WebDriver
        wait: WebDriverWait
        doc_path: Caminho completo do arquivo
        doc_type: Tipo de documento (pdf, docx, xlsx, etc) - usado para filtro MIME
    
    Returns:
        bool: True se enviado com sucesso, False caso contrário
    """
    if not os.path.exists(doc_path):
        logger.warning(f"Documento não encontrado em: {doc_path}. Pulando envio de documento.")
        return False

    try:
        abs_path = os.path.abspath(doc_path)
        file_name = os.path.basename(abs_path)

        if not os.path.isfile(abs_path):
            logger.warning(f"Documento inválido para envio: {abs_path}")
            return False

        if os.path.getsize(abs_path) <= 0:
            logger.warning(f"Documento vazio, envio abortado: {abs_path}")
            return False

        def _click_first_visible(selectors, pause: float = 0.4) -> bool:
            for xp in selectors:
                try:
                    elements = driver.find_elements(By.XPATH, xp)
                except Exception:
                    continue
                for el in elements:
                    try:
                        if not el.is_displayed():
                            continue
                        driver.execute_script("arguments[0].click();", el)
                        time.sleep(pause)
                        return True
                    except Exception:
                        try:
                            el.click()
                            time.sleep(pause)
                            return True
                        except Exception:
                            continue
            return False

        def _find_document_input():
            script = """
                const inputs = Array.from(document.querySelectorAll('input[type="file"]'));
                const visible = (el) => {
                    const style = window.getComputedStyle(el);
                    const rect = el.getBoundingClientRect();
                    return style.visibility !== 'hidden' && style.display !== 'none' && rect.width >= 0 && rect.height >= 0;
                };
                const score = (el) => {
                    const accept = (el.getAttribute('accept') || '').toLowerCase();
                    const testid = (el.getAttribute('data-testid') || '').toLowerCase();
                    let score = 0;
                    if (!accept) score += 2;
                    if (accept.includes('*')) score += 4;
                    if (accept.includes('application')) score += 5;
                    if (accept.includes('document')) score += 4;
                    if (accept.includes('image')) score -= 6;
                    if (accept.includes('video')) score -= 6;
                    if (testid.includes('document')) score += 6;
                    if (visible(el)) score += 2;
                    return score;
                };
                return inputs
                    .map((el, idx) => ({ idx, score: score(el) }))
                    .sort((a, b) => b.score - a.score || b.idx - a.idx)
                    .map((item) => item.idx);
            """
            try:
                indices = driver.execute_script(script) or []
            except Exception:
                indices = []
            if not indices:
                return None
            inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='file']")
            for idx in indices:
                try:
                    candidate = inputs[idx]
                    accept = (candidate.get_attribute("accept") or "").lower()
                    if "image" in accept and "application" not in accept and "*" not in accept:
                        continue
                    return candidate
                except Exception:
                    continue
            return None

        attach_selectors = [
            "//div[@title='Anexar' or @aria-label='Anexar' or @data-testid='clip']",
            "//span[@data-icon='clip']/ancestor::*[@role='button'][1]",
            "//span[@data-icon='clip']",
            "//div[@role='button' and contains(@aria-label,'Anexar')]",
            "//button[contains(@aria-label, 'Anexar')]",
            "//*[@data-testid='attach']",
        ]

        document_entry_selectors = [
            "//*[@aria-label='Documento']",
            "//*[@title='Documento']",
            "//span[normalize-space()='Documento']/ancestor::*[@role='button'][1]",
            "//*[@data-testid='attach-menu-document']",
        ]

        send_btn_selectors = [
            "//div[@role='button' and @data-testid='attach-document-send']",
            "//button[@data-testid='compose-btn-send']",
            "//div[@role='button' and contains(@aria-label,'Enviar')]",
            "//button[contains(@aria-label, 'Enviar')]",
            "//*[@data-testid='send']",
        ]
        composer_send_selectors = [
            "//div[@role='button' and @data-testid='attach-document-send']",
        ]

        preview_name_xpath = (
            "//*[contains(@title, \"%s\") or contains(normalize-space(.), \"%s\")]"
            % (file_name.replace('"', ""), file_name.replace('"', ""))
        )
        sent_message_xpath = (
            "//*[@data-testid='msg-container']//div[contains(@class,'message-out')]"
            "//*[contains(@title, \"%s\") or contains(normalize-space(.), \"%s\")]"
            % (file_name.replace('"', ""), file_name.replace('"', ""))
        )

        try:
            wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='main']")))
            wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='main']//div[@contenteditable='true']")))
        except Exception:
            logger.warning("Chat do WhatsApp não está pronto para anexar documento.")
            return False

        if not _click_first_visible(attach_selectors, pause=0.5):
            logger.warning("Não foi possível encontrar o botão de anexar.")
            return False




        # Inicia thread para fechar automaticamente o seletor nativo do Windows
        def _fechar_seletor_windows():
            if not _HAS_WIN32:
                return
            import threading
            deadline = time.time() + 5
            while time.time() < deadline:
                try:
                    hwnd = win32gui.FindWindow('#32770', None)
                    if hwnd:
                        win32gui.PostMessage(hwnd, win32con.WM_CLOSE, 0, 0)
                        return
                except Exception:
                    pass
                time.sleep(0.1)

        import threading
        t = threading.Thread(target=_fechar_seletor_windows, daemon=True)
        t.start()

        # Clica em Documento normalmente
        _click_first_visible(document_entry_selectors, pause=0.4)

        file_input = None
        deadline = time.time() + 8
        while time.time() < deadline and file_input is None:
            file_input = _find_document_input()
            if file_input is not None:
                break
            time.sleep(0.25)

        if file_input is None:
            logger.error('Nao foi possivel localizar o input de documento do WhatsApp Web.')
            return False

        try:
            driver.execute_script(
                'arguments[0].style.display="block";'
                'arguments[0].style.visibility="visible";'
                'arguments[0].style.position="fixed";'
                'arguments[0].style.top="-9999px";'
                'arguments[0].style.width="1px";'
                'arguments[0].style.height="1px";',
                file_input
            )
            file_input.send_keys(abs_path)
        except Exception as e:
            logger.exception(f'Falha ao enviar caminho para input de documento: {e}')
            return False

        preview_ready = False
        preview_deadline = time.time() + 12
        while time.time() < preview_deadline:
            send_ready = False
            for xp in send_btn_selectors:
                try:
                    for btn in driver.find_elements(By.XPATH, xp):
                        if btn.is_displayed():
                            send_ready = True
                            break
                except Exception:
                    continue
                if send_ready:
                    break

            name_visible = False
            try:
                for el in driver.find_elements(By.XPATH, preview_name_xpath):
                    if el.is_displayed():
                        name_visible = True
                        break
            except Exception:
                name_visible = False

            if send_ready and name_visible:
                preview_ready = True
                break
            time.sleep(0.35)

        if not preview_ready:
            logger.warning(
                f"O WhatsApp não confirmou o preview do documento '{file_name}'. "
                "Envio abortado para evitar falso positivo."
            )
            return False

        sent_message_count_before = 0
        try:
            sent_message_count_before = len(driver.find_elements(By.XPATH, sent_message_xpath))
        except Exception:
            sent_message_count_before = 0

        sent_confirmed = _click_first_visible(send_btn_selectors, pause=0.6)
        if not sent_confirmed:
            logger.warning("Não foi possível clicar no botão de enviar do documento.")
            return False

        sent_deadline = time.time() + 12
        while time.time() < sent_deadline:
            sent_message_count_after = sent_message_count_before
            try:
                sent_message_count_after = len(driver.find_elements(By.XPATH, sent_message_xpath))
            except Exception:
                sent_message_count_after = sent_message_count_before

            composer_send_still_visible = False
            for xp in composer_send_selectors:
                try:
                    for btn in driver.find_elements(By.XPATH, xp):
                        if btn.is_displayed():
                            composer_send_still_visible = True
                            break
                except Exception:
                    continue
                if composer_send_still_visible:
                    break

            chat_ready = False
            try:
                chat_inputs = driver.find_elements(By.XPATH, "//*[@id='main']//div[@contenteditable='true']")
                chat_ready = any(el.is_displayed() for el in chat_inputs)
            except Exception:
                chat_ready = False

            if sent_message_count_after > sent_message_count_before:
                return True

            if chat_ready and not composer_send_still_visible:
                return True
            time.sleep(0.4)

        logger.warning(
            f"O clique de envio foi executado, mas o composer do documento '{file_name}' não fechou a tempo."
        )
        return False

    except Exception as e:
        logger.exception(f"Erro ao enviar documento: {e}")
        return False


def enviar_arquivo_para_chat(driver, wait, file_path: str) -> bool:
    """
    Wrapper genérico que envia qualquer tipo de arquivo para o WhatsApp Web.
    Detecta automaticamente se é PDF, imagem ou outro documento e usa a função apropriada.
    
    Args:
        driver: Selenium WebDriver
        wait: WebDriverWait
        file_path: Caminho completo do arquivo
    
    Returns:
        bool: True se enviado com sucesso, False caso contrário
    """
    if not os.path.exists(file_path):
        logger.warning(f"Arquivo não encontrado: {file_path}")
        return False
    
    ext = os.path.splitext(file_path)[1].lower().lstrip('.')
    
    # Tipos de arquivo que devem usar a função robusta de documentos
    document_extensions = {
        "pdf", "docx", "doc", "xlsx", "xls", "pptx", "ppt",
        "zip", "rar", "7z", "txt", "odt", "ods", "odp"
    }
    
    # Tipos de arquivo que são imagens (usa função de imagem otimizada)
    image_extensions = {"jpg", "jpeg", "png", "gif", "webp", "bmp"}
    
    try:
        if ext in image_extensions:
            logger.debug(f"Detectado arquivo de imagem: {os.path.basename(file_path)}")
            return enviar_imagem_para_chat(driver, wait, file_path)
        elif ext in document_extensions or ext == "pdf":
            logger.debug(f"Detectado documento: {os.path.basename(file_path)} (tipo: {ext})")
            return enviar_documento_para_chat(driver, wait, file_path, ext)
        else:
            # Para extensões desconhecidas, usa a função robusta de documentos
            logger.debug(f"Tipo desconhecido, usando envio robusto de documentos: {os.path.basename(file_path)}")
            return enviar_documento_para_chat(driver, wait, file_path, ext)
    except Exception as e:
        logger.exception(f"Erro ao enviar arquivo: {e}")
        return False


# ---------------------------
# Envio via WhatsApp (integração completa)
# ---------------------------

def iniciar_driver_chrome(user_data_dir: str = USER_DATA_DIR, headless: bool = False):
    # === LAZY INITIALIZATION DO PERFIL DO CHROME ===
    os.makedirs(user_data_dir, exist_ok=True)

    chrome_options = Options()
    chrome_options.add_argument(f"--user-data-dir={user_data_dir}")
    chrome_options.add_argument("--profile-directory=Default")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-background-timer-throttling")
    chrome_options.add_argument("--disable-backgrounding-occluded-windows")
    chrome_options.add_argument("--disable-renderer-backgrounding")

    if headless:
        chrome_options.add_argument("--headless=new")

    try:
        # Tenta usar o chromedriver do PATH (funciona no .exe PyInstaller
        # e em máquinas com Chrome instalado normalmente)
        try:
            driver = webdriver.Chrome(options=chrome_options)
        except Exception:
            # Fallback: procura chromedriver ao lado do executável
            import os as _os, sys as _sys
            _base = _os.path.dirname(_sys.executable) if getattr(_sys, "frozen", False)                     else _os.path.dirname(_os.path.abspath(__file__))
            _cd = _os.path.join(_base, "chromedriver.exe")
            service = Service(_cd)
            driver = webdriver.Chrome(service=service, options=chrome_options)
        driver.maximize_window()
        return driver
    except Exception as e:
        logger.exception("Falha ao iniciar o Chrome WebDriver.")
        raise e

def atualizar_planilha_excel(df_dict_full: dict, caminho: str):
    if not os.path.exists(caminho):
        logger.warning(f"Arquivo {caminho} não encontrado. Pulando atualização.")
        return

    temp_caminho = caminho.replace(".xlsx", "_TEMP.xlsx")
    try:
        with pd.ExcelWriter(temp_caminho, engine="openpyxl") as writer:
            for nome_aba, df in df_dict_full.items():
                df.to_excel(writer, sheet_name=nome_aba, index=False)
        
        if os.path.exists(temp_caminho):
            import shutil
            import time
            
            # Tenta mover o ficheiro. Se o OneDrive/Antivírus estiver a bloquear,
            # espera 1.5s e tenta novamente (até 3 tentativas).
            sucesso = False
            for tentativa in range(8):
                try:
                    shutil.move(temp_caminho, caminho)
                    sucesso = True
                    break
                except PermissionError:
                    time.sleep(3) # Compassos de espera para a nuvem libertar o ficheiro
            
            if not sucesso:
                raise PermissionError("O ficheiro continuou bloqueado após retentativas.")
                
    except PermissionError:
        logger.error("❌ Excel aberto ou bloqueado pela nuvem — feche-o para salvar os status de envio.")
        if os.path.exists(temp_caminho):
            try:
                os.remove(temp_caminho)
            except Exception:
                pass
    except Exception as e:
        logger.error(f"Erro ao atualizar planilha: {e}")
        if os.path.exists(temp_caminho):
            try:
                os.remove(temp_caminho)
            except Exception:
                pass

def delay_com_barra(segundos: float, descricao: str):
    segundos_int = int(segundos)
    tqdm.write(f"GUI_TIMER:{segundos_int}")
    for _ in tqdm(range(segundos_int), desc=descricao, unit="s", colour="cyan", leave=False):
        if _HAS_MSVCRT and _msvcrt.kbhit():
            break
        time.sleep(1)

def baixar_fatura_cliente(
    matricula: str,
    id_fatura: str,
    usuario_waterfy: str,
    senha_waterfy: str,
    output_dir: str = None,
    cpf_consumidor: str | None = None,
    encrypt: bool = True,
) -> str | None:
    """
    Baixa a fatura PDF de um cliente específico usando o sistema integrado.
    Retorna o caminho do arquivo baixado ou None se falhar.
    """
    if output_dir is None:
        output_dir = os.path.join(_BASE, "faturas")

    if not HAS_FATURA_DOWNLOADER:
        logger.warning("Sistema de download de faturas não disponível.")
        return None

    try:
        session = fazer_login_waterfy(usuario_waterfy, senha_waterfy)

        curl_path = "cURL emitirFaturaSegVia.txt"
        if not os.path.exists(curl_path):
            logger.error(f"Arquivo curl não encontrado: {curl_path}")
            return None

        parsed = parse_curl_file(Path(curl_path))
        url = parsed["url"]
        headers = parsed["headers"]
        payload_base = parsed["payload"]

        if not payload_base:
            logger.error("Payload não encontrado no curl")
            return None

        cookies_dict = session.cookies.get_dict()
        if cookies_dict:
            headers["Cookie"] = "; ".join(f"{k}={v}" for k, v in cookies_dict.items())

        mat_det, fat_det, idx_mat, idx_fat = detect_ids_in_payload(payload_base)
        payload_modified = substitute_ids(payload_base, idx_mat, idx_fat, matricula, id_fatura)

        if not payload_modified or len(payload_modified) < 10:
            logger.error("Payload modificado inválido")
            return None

        try:
            payload_bytes = payload_modified.encode("utf-8")
        except Exception as enc_err:
            logger.error(f"❌ Erro ao encod payload: {enc_err}")
            return None

        try:
            response = session.post(url, headers=headers, data=payload_bytes, timeout=60)
        except Exception as e:
            logger.error(f"Exceção ao fazer POST: {e}")
            return None

        if not response.ok:
            logger.error(f"❌ Falha ao gerar PDF: {response.status_code} {response.reason}")
            logger.error(f"❌ Response body (primeiros 2000 bytes):\n{response.content[:2000]}")
            logger.error(f"❌ Response text (primeiros 2000 chars):\n{response.text[:2000]}")
            return None

        if not response.text:
            logger.error("❌ Resposta do servidor está vazia! (response.text = '')")
            logger.error(f"❌ Response content não vazio? {len(response.content)} bytes")
            if response.content:
                try:
                    logger.error(
                        f"❌ Tentando decodificar content como string:\n"
                        f"{response.content.decode('utf-8', errors='replace')[:2000]}"
                    )
                except Exception as dec_err:
                    logger.error(f"❌ Erro ao decodificar: {dec_err}")
            return None

        pdf_path = None
        try:
            import json
            response_json = json.loads(response.text)
            if isinstance(response_json, list) and response_json and isinstance(response_json[0], list):
                for elem in response_json[0]:
                    if isinstance(elem, str) and (".pdf" in elem.lower() or "/tmp" in elem.lower()):
                        pdf_path = elem
                        break
            elif isinstance(response_json, dict):
                for value in response_json.values():
                    if isinstance(value, str) and (".pdf" in value.lower() or "/tmp" in value.lower()):
                        pdf_path = value
                        break
        except Exception:
            pass

        if not pdf_path:
            pdf_path_match = re.search(r'"(/tmp/[^"]+\.pdf)"', response.text)
            if pdf_path_match:
                pdf_path = pdf_path_match.group(1)

        if not pdf_path:
            pdf_path_match = re.search(r'(/tmp/[^\s"\'<>]+\.pdf)', response.text)
            if pdf_path_match:
                pdf_path = pdf_path_match.group(1)

        if not pdf_path:
            pdf_path_match = re.search(r'/tmp/[^\s"\'<>]+\.pdf', response.text)
            if pdf_path_match:
                pdf_path = pdf_path_match.group(0)

        if not pdf_path:
            any_pdf_match = re.search(r'["\']?([^"\'<>\s]+\.pdf)["\']?', response.text)
            if any_pdf_match:
                pdf_path = any_pdf_match.group(1)

        if not pdf_path:
            logger.error("❌ Caminho do PDF não encontrado na resposta após análise das estratégias de extração")
            base_url = "https://centrosul.waterfy.net"
            id_fatura_primeiro = id_fatura.split(';')[0] if ';' in id_fatura else id_fatura
            download_url = f"{base_url}/api/report?matricula={matricula}&id_fatura={id_fatura_primeiro}&format=pdf&token=&t={random.random()}"

            try:
                alt_response = session.get(download_url, headers=headers, timeout=60)
                if alt_response.ok and alt_response.headers.get("content-type", "").startswith("application/pdf"):
                    os.makedirs(output_dir, exist_ok=True)
                    pdf_filename = f"tmp_mat{matricula}_fat{id_fatura_primeiro}.pdf"
                    output_path = os.path.join(output_dir, pdf_filename)
                    Path(output_path).write_bytes(alt_response.content)
                    return output_path
                logger.warning(f"❌ URL alternativa não retornou PDF (status: {alt_response.status_code})")
            except Exception as e:
                logger.warning(f"❌ Erro na URL alternativa: {e}")
            return None


        base_url = "https://centrosul.waterfy.net"
        id_fatura_primeiro = id_fatura.split(';')[0] if ';' in id_fatura else id_fatura
        download_url = f"{base_url}/api/report?link={pdf_path}&type=pdf&{random.random()}"

        os.makedirs(output_dir, exist_ok=True)
        pdf_filename = f"tmp_mat{matricula}_fat{id_fatura_primeiro}.pdf"
        output_path = os.path.join(output_dir, pdf_filename)

        output_path_obj = Path(output_path)
        success = download_pdf(session, download_url, headers, output_path_obj, timeout=60)

        if success:
            if cpf_consumidor and encrypt:
                if HAS_PYPDF:
                    try:
                        reader = PdfReader(output_path)
                        writer = PdfWriter()
                        for page in reader.pages:
                            writer.add_page(page)
                        writer.encrypt(cpf_consumidor)
                        with open(output_path, "wb") as f:
                            writer.write(f)
                    except Exception as exc:
                        logger.warning(
                            f"Fatura baixada, mas não foi possível aplicar senha. "
                            f"Arquivo ficou sem senha. Erro: {exc}"
                        )
                else:
                    logger.warning(
                        "Fatura baixada, mas pypdf não está instalado; não foi possível proteger o PDF com senha."
                    )
            return output_path

        logger.error("Falha ao baixar PDF")
        return None

    except Exception as e:
        logger.exception(f"Erro ao baixar fatura: {e}")
        return None

def _split_id_faturas(raw_id_fatura: str) -> list[str]:
    if not raw_id_fatura:
        return []
    return [part.strip() for part in re.split(r"[;,|\s]+", str(raw_id_fatura)) if part.strip()]


def _merge_pdfs(pdf_paths: list[str], output_path: str, password: str | None = None) -> bool:
    if not HAS_PYPDF:
        logger.warning("pypdf não está instalado; não é possível concatenar nem proteger o PDF final.")
        return False
    try:
        writer = PdfWriter()
        for pdf_path in pdf_paths:
            reader = PdfReader(pdf_path)
            for page in reader.pages:
                writer.add_page(page)
        if password:
            writer.encrypt(password)
        with open(output_path, "wb") as f:
            writer.write(f)
        return True
    except Exception as exc:
        logger.warning(f"Falha ao concatenar PDFs com {PDF_BACKEND or 'backend desconhecido'}: {exc}")
        return False


def enviar_whatsapp_sheets(sheet_names: List[str], caminho_arquivo: str, msgs_por_hora: int = 20, baixar_faturas: bool = False, usuario_waterfy: str = "", senha_waterfy: str = ""):
    if not os.path.exists(caminho_arquivo):
        logger.error(f"Arquivo Excel não encontrado: '{caminho_arquivo}'.")
        return

    try:
        df_dict_full = pd.read_excel(caminho_arquivo, sheet_name=None, dtype=str, engine="openpyxl")
    except Exception:
        logger.exception(f"Erro ao ler o arquivo Excel de envio: {caminho_arquivo}")
        return

    if isinstance(df_dict_full, pd.DataFrame):
        df_dict_full = {sheet_names[0]: df_dict_full}

    # --- A MÁGICA MATEMÁTICA ACONTECE AQUI ---
    # Calcula o tempo médio ideal para bater a meta da hora exata
    tempo_medio_por_msg = 3600 / msgs_por_hora
    # Desconta ~15s que é o tempo que o robô gasta clicando e digitando na tela
    espera_media = max(5, tempo_medio_por_msg - 15) 
    
    # Cria a variação humana (30% a mais ou a menos do tempo médio)
    delay_min_dinamico = espera_media * 0.7
    delay_max_dinamico = espera_media * 1.3
    # -----------------------------------------

    # Registra callback de save na page do Flet (para diálogo de fechamento)
    import sys as _sys
    _flet_page = getattr(_sys.modules.get("__main__", None), "_FLET_PAGE", None)
    if _flet_page is not None:
        _flet_page._wpp_save_fn = lambda: atualizar_planilha_excel(df_dict_full, caminho_arquivo)

    backup_caminho = caminho_arquivo.replace(".xlsx", "_BACKUP.xlsx")
    try:
        import shutil
        shutil.copy(caminho_arquivo, backup_caminho)
        logger.info(f"✅ Backup criado: {backup_caminho}")
    except Exception as e:
        logger.warning(f"Não foi possível criar backup: {e}")
        backup_caminho = None

    falhas = []
    driver = None

    # Conta total de linhas (todas) e já enviadas anteriormente
    _total_linhas = 0
    _enviadas = 0  # começa contando as já enviadas anteriormente
    _falhas_count = 0
    for s in sheet_names:
        df_s = df_dict_full.get(s)
        if df_s is None or df_s.empty:
            continue
        _total_linhas += len(df_s)
        if "STATUS" in df_s.columns:
            status_norm = df_s["STATUS"].astype(str).str.strip().str.lower()
            _enviadas += int(((status_norm == "enviado") | (status_norm == "enviado sem imagem")).sum())
            _falhas_count += int(status_norm.str.startswith("falha").sum())
    logger.info(f"[INFO] Contagem inicial - Total: {_total_linhas} | Ja enviadas: {_enviadas}")

    def _emit_progresso():
        tqdm.write(f"__PROGRESSO__:{_enviadas}:{_falhas_count}:{_total_linhas}")

    def _emit_etapa(pct: float, label: str, detail: str = ""):
        pct = max(0.0, min(1.0, float(pct)))
        tqdm.write(f"__WPP_STAGE__:{pct:.3f}|{label}|{detail}")

    def _registrar_falha(df_sheet, idx, row, erro: str):
        nonlocal _falhas_count
        _falhas_count += 1
        falhas.append({**row.to_dict(), "ERRO": erro})
        try:
            df_sheet.at[idx, "STATUS"] = f"Falha: {erro}"
            atualizar_planilha_excel(df_dict_full, caminho_arquivo)
        except Exception:
            pass
        _emit_progresso()

    _emit_progresso()  # emite estado inicial imediatamente

    try:
        _emit_etapa(0.02, "Iniciando navegador...", "carregando Chrome")
        driver = iniciar_driver_chrome()
        wait = WebDriverWait(driver, 300)

        _emit_etapa(0.08, "Conectando ao WhatsApp Web...", "abrindo web.whatsapp.com")
        driver.get("https://web.whatsapp.com/")
        logger.info("[INFO] Abrindo WhatsApp Web... escaneie o QR Code se necessario.")
        indicadores_sucesso = [
            (By.ID, "pane-side"),                    
            (By.XPATH, "//canvas[@aria-label='Scan me!']"), 
            (By.XPATH, "//*[@id='side']"),           
            (By.CSS_SELECTOR, "div[data-testid='chat-list']") 
        ]
        
        login_feito = False
        for _ in range(60): 
            for tipo, seletor in indicadores_sucesso:
                try:
                    if driver.find_elements(tipo, seletor):
                        login_feito = True
                        break
                except:
                    continue
            if login_feito: break
            time.sleep(1)
            
        if login_feito:
            logger.info("WhatsApp Web detectado!")
            tqdm.write(f"[OK] Conexao estabelecida. Operando no ritmo de {msgs_por_hora} msgs/hora.")
        else:
            logger.warning("   Carregamento incerto — tentando prosseguir mesmo assim.")

        mensagens_enviadas_no_lote = 0
        inicio_lote = time.time()
        
        voltar_menu = False 

        for sheet in sheet_names:
            df_sheet = df_dict_full.get(sheet)
            if df_sheet is None or df_sheet.empty:
                tqdm.write(f"[SKIP] Aba '{sheet}' vazia. Pulando.")
                continue

            if "STATUS" not in df_sheet.columns:
                df_sheet["STATUS"] = ""

            logger.info(f"[INFO] Iniciando envios - aba: {sheet}")
            
            loop_linhas = tqdm(df_sheet.iterrows(), total=len(df_sheet), desc=f"Enviando {sheet}", unit="msg", colour="green",file=sys.stdout if sys.stdout else io.StringIO())
            _bloco_inicio = None; _bloco_fim = None; _bloco_count = 0

            for idx, row in loop_linhas:
                if voltar_menu:
                    break
                
                # =======================================================
                # 🛑 SISTEMA DE PAUSA E CANCELAMENTO (FREIO DE MÃO)
                if _HAS_MSVCRT and _msvcrt.kbhit():
                    tecla = _msvcrt.getch().decode('utf-8', errors='ignore').lower()
                    if tecla == 'm':
                        tqdm.write("\n[STOP] Cancelamento solicitado pelo operador. Abortando...")
                        voltar_menu = True
                        break
                    elif tecla == 'p':
                        tqdm.write("\n" + "="*60)
                        pular_cliente = False
                        while True:
                            if _HAS_MSVCRT and _msvcrt.kbhit():
                                acao = _msvcrt.getch().decode('utf-8', errors='ignore').lower()
                                if acao == 'c':
                                    tqdm.write(" [OK] Retomando os envios...")
                                    espera_retomada = random.uniform(delay_min_dinamico, delay_max_dinamico)
                                    delay_com_barra(espera_retomada, "Retomando em")
                                    break
                                elif acao == 's':
                                    tqdm.write(f" [SKIP] Pulando o cliente da linha {idx}...")
                                    pular_cliente = True
                                    break
                                elif acao == 'm':
                                    tqdm.write(" [STOP] Abortando envios e fechando o WhatsApp...")
                                    voltar_menu = True
                                    break
                            time.sleep(0.1)
                        if voltar_menu: break 
                        if pular_cliente:
                            df_sheet.at[idx, "STATUS"] = "Pulado manualmente"
                            atualizar_planilha_excel(df_dict_full, caminho_arquivo)
                            continue 
                # =======================================================
                
                tel_display = row.get(COL_TEL) or row.get("_TEL_ORIGINAL") or "Sem numero"
                loop_linhas.set_postfix_str(f"Linha {idx} | Tel: {tel_display}")
                
                status_atual = str(row.get("STATUS", "")).strip().upper()
                if status_atual in {"ENVIADO", "PULADO MANUALMENTE", "ENVIADO SEM IMAGEM"}:
                    if _bloco_inicio is None:
                        _bloco_inicio = idx
                    _bloco_fim = idx
                    _bloco_count += 1
                    continue
                else:
                    if _bloco_count > 0:
                        tqdm.write(f"[SKIP] Linhas {_bloco_inicio} a {_bloco_fim} ja processadas ({_bloco_count} registros) - pulando.")
                        _bloco_inicio = None; _bloco_fim = None; _bloco_count = 0
                
                # --- SISTEMA DE LOTE BASEADO NA HORA ---
                if mensagens_enviadas_no_lote >= msgs_por_hora:
                    tempo_decorrido = time.time() - inicio_lote
                    tempo_espera_lote = max(0, 3600 - tempo_decorrido)
                    
                    if tempo_espera_lote > 0:
                        tqdm.write(f"[WAIT] Meta de {msgs_por_hora} msgs atingida. Aguardando {tempo_espera_lote/60:.1f} min para fechar a hora...")
                        delay_com_barra(tempo_espera_lote, "Pausa horaria")
                    else:
                        tqdm.write(f"[INFO] Meta de {msgs_por_hora} msgs atingida e ja passou 1 hora. Iniciando proximo lote.")
                        
                    mensagens_enviadas_no_lote = 0
                    inicio_lote = time.time()

                tel = row.get(COL_TEL) or row.get("_TEL_ORIGINAL") or ""
                if pd.isna(tel) or not tel:
                    tqdm.write(f"[AVISO] Pulando linha {idx}: Não encontramos um telefone cadastrado para este cliente.")
                    _registrar_falha(df_sheet, idx, row, "Telefone ausente")
                    continue

                tel_norm = normalizar_telefone(tel)
                if not tel_norm:
                    tqdm.write(f"[AVISO] Pulando linha {idx}: O telefone fornecido parece estar incorreto ({tel}).")
                    _registrar_falha(df_sheet, idx, row, "Telefone inválido")
                    continue

                numero_para_url = tel_norm.replace("+", "")
                mensagem = montar_mensagem(row)
                url = f"https://web.whatsapp.com/send?phone={numero_para_url}"
                
                try:
                    driver.get(url)
                except Exception:
                    tqdm.write(f"[ERRO] Não foi possível iniciar a conversa com o número {numero_para_url} (linha {idx}).")
                    _registrar_falha(df_sheet, idx, row, "Falha ao abrir chat")
                    continue

                try:
                    WebDriverWait(driver, 80).until(EC.presence_of_element_located((By.ID, "pane-side")))
                    time.sleep(3)
                    if not driver.find_elements(By.ID, "main"):
                        clicar_continuar_whatsapp_web(driver, wait)
                        time.sleep(3)
                    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@id='main']//div[@contenteditable='true']")))
                except Exception:
                    tqdm.write(f"[ERRO] O WhatsApp demorou muito para responder e não carregou a conversa de {numero_para_url} (linha {idx}).")
                    _registrar_falha(df_sheet, idx, row, "Chat não carregou")
                    continue

                enviado_texto = tentar_enviar(driver, wait, mensagem)
                
                # ═══════════════════════════════════════════════════════════════════════════
                # ENVIO DE ANEXOS: Prioridade de Envio
                # 1. Arquivo customizado do cliente (coluna: ARQUIVO_PDF, ARQUIVO, PDF, etc)
                # 2. Fatura baixada automaticamente (se baixar_faturas=True)
                # 3. Imagem padrão (fallback)
                # ═══════════════════════════════════════════════════════════════════════════
                enviado_anexo = True
                
                # --- ETAPA 1: Procura por arquivo customizado do cliente ---
                # Suporta colunas: ARQUIVO_PDF, ARQUIVO, PDF, CAMINHO_PDF, DOCUMENTO, etc.
                # Exemplo de uso na planilha Excel:
                #   | Nome | Tel | ARQUIVO_PDF |
                #   | João | ... | C:\docs\fatura.pdf |
                arquivo_customizado = None
                colunas_arquivo = [
                    "ARQUIVO_PDF", "ARQUIVO", "PDF", "CAMINHO_PDF", 
                    "DOCUMENTO", "ARQUIVO_DOC", "FILE", "PATH_PDF",
                    "ANEXO", "ARQUIVO_ENVIO"
                ]
                
                for col_arquivo in colunas_arquivo:
                    if col_arquivo in row.index:
                        caminho_temp = str(row.get(col_arquivo, "")).strip()
                        if caminho_temp and os.path.exists(caminho_temp):
                            arquivo_customizado = caminho_temp
                            tqdm.write(f"[INFO] Arquivo customizado encontrado: {os.path.basename(arquivo_customizado)}")
                            break
                
                if arquivo_customizado and os.path.exists(arquivo_customizado):
                    # Usa arquivo customizado do cliente
                    enviado_anexo = enviar_arquivo_para_chat(driver, wait, arquivo_customizado)
                    if enviado_anexo:
                        logger.info(f"Arquivo customizado enviado: {os.path.basename(arquivo_customizado)}")
                elif baixar_faturas and HAS_FATURA_DOWNLOADER:
                    # Baixa fatura(s) do cliente (método tradicional)
                    matricula = str(row.get(COL_MATRICULA, "")).strip()
                    raw_id_fatura = str(row.get("ID_FATURA", "") or "").strip()
                    id_faturas = _split_id_faturas(raw_id_fatura)
                    cpf_consumidor = re.sub(
                        r"\D",
                        "",
                        (
                            str(row.get("CPF CONSUMIDOR", "") or "").strip()
                            or str(row.get("CPF_CONSUMIDOR", "") or "").strip()
                            or str(row.get("CPF", "") or "").strip()
                        ),
                    )
                    cnpj_consumidor = re.sub(
                        r"\D",
                        "",
                        (
                            str(row.get("CNPJ CONSUMIDOR", "") or "").strip()
                            or str(row.get("CNPJ_CONSUMIDOR", "") or "").strip()
                            or str(row.get("CNPJ", "") or "").strip()
                        ),
                    )
                    senha_consumidor = cpf_consumidor or cnpj_consumidor

                    if matricula and id_faturas:
                        if len(id_faturas) == 1:
                            id_fatura = id_faturas[0]
                            tqdm.write(f"[INFO] Baixando fatura para matrícula {matricula}, ID {id_fatura}...")
                            if not senha_consumidor:
                                tqdm.write("[WARN] Nenhum CPF/CNPJ CONSUMIDOR encontrado; PDF será baixado sem senha.")
                            else:
                                tqdm.write(f"[INFO] Usando documento para senha: {'CPF' if cpf_consumidor else 'CNPJ'}.")
                            fatura_path = baixar_fatura_cliente(
                                matricula,
                                id_fatura,
                                usuario_waterfy,
                                senha_waterfy,
                                os.path.join(_BASE, "faturas"),
                                cpf_consumidor=senha_consumidor or None,
                                encrypt=True,
                            )
                            # Renomeia para o padrão final
                            if fatura_path and os.path.exists(fatura_path):
                                import shutil
                                final_path = os.path.join(_BASE, "faturas", f"Segunda_Via_Faturas_Matricula_{matricula}.pdf")
                                shutil.copy2(fatura_path, final_path)
                                fatura_path = final_path
                        else:
                            tqdm.write(f"[INFO] Baixando {len(id_faturas)} faturas para matrícula {matricula}...")
                            downloaded_paths = []
                            with concurrent.futures.ThreadPoolExecutor(max_workers=min(5, len(id_faturas))) as executor:
                                futures_dict = {
                                    executor.submit(
                                        baixar_fatura_cliente,
                                        matricula,
                                        id_fat,
                                        usuario_waterfy,
                                        senha_waterfy,
                                        "faturas",
                                        cpf_consumidor=None,
                                        encrypt=False,
                                    ): id_fat for id_fat in id_faturas
                                }
                                results_dict = {}
                                for future in concurrent.futures.as_completed(futures_dict):
                                    id_fat = futures_dict[future]
                                    try:
                                        path = future.result()
                                        if path and os.path.exists(path):
                                            tqdm.write(f"  - Download concluído: ID {id_fat}")
                                            results_dict[id_fat] = path
                                        else:
                                            tqdm.write(f"  - Falha no download: ID {id_fat}")
                                    except Exception as exc:
                                        tqdm.write(f"  - Erro no download ID {id_fat}: {exc}")
                                
                                # Respeita a ordem original dos IDs para concatenar as faturas na ordem correta
                                for id_fat in id_faturas:
                                    if id_fat in results_dict:
                                        downloaded_paths.append(results_dict[id_fat])

                            if downloaded_paths:
                                merged_path = os.path.join(_BASE, "faturas", f"Segunda_Via_Faturas_Matricula_{matricula}.pdf")
                                if _merge_pdfs(downloaded_paths, merged_path, password=senha_consumidor or None):
                                    fatura_path = merged_path
                                    tqdm.write(f"[INFO] PDFs concatenados: {os.path.basename(merged_path)}")
                                else:
                                    fallback_path = os.path.join(_BASE, "faturas", f"Segunda_Via_Faturas_Matricula_{matricula}.pdf")
                                    try:
                                        import shutil
                                        shutil.copy2(downloaded_paths[0], fallback_path)
                                        fatura_path = fallback_path
                                    except Exception:
                                        fatura_path = downloaded_paths[0]
                                    tqdm.write("[WARN] Falha ao concatenar todos os PDFs; usando o primeiro arquivo baixado.")
                            else:
                                fatura_path = None

                        if fatura_path and os.path.exists(fatura_path):
                            # Usa a função genérica que detecta automaticamente o tipo
                            enviado_anexo = enviar_arquivo_para_chat(driver, wait, fatura_path)
                            if enviado_anexo:
                                # Envia imagem padrão logo após o PDF
                                if IMAGE_PATH and os.path.exists(IMAGE_PATH):
                                    time.sleep(2)
                                    enviado_imagem = enviar_imagem_para_chat(driver, wait, IMAGE_PATH)
                                    if not enviado_imagem:
                                        tqdm.write(f"[WARN] Falha ao enviar imagem padrão após PDF.")
                        else:
                            tqdm.write(f"[WARN] Falha ao baixar fatura, usando imagem padrão")
                            if IMAGE_PATH and os.path.exists(IMAGE_PATH):
                                enviado_anexo = enviar_imagem_para_chat(driver, wait, IMAGE_PATH)
                    else:
                        tqdm.write(f"[WARN] Matrícula ou ID fatura ausente, usando imagem padrão")
                        if IMAGE_PATH and os.path.exists(IMAGE_PATH):
                            enviado_anexo = enviar_imagem_para_chat(driver, wait, IMAGE_PATH)
                else:
                    # Comportamento original: envia imagem padrão
                    if IMAGE_PATH and os.path.exists(IMAGE_PATH):
                        enviado_anexo = enviar_imagem_para_chat(driver, wait, IMAGE_PATH)

                if enviado_texto and enviado_anexo:
                    _enviadas += 1
                    tqdm.write(f"[OK] Mensagem enviada com sucesso para {numero_para_url} (linha {idx}).")
                    matricula = row.get(COL_MATRICULA, "N/A")
                    logger.info(f"[OK] ENVIADO | Telefone: {tel_norm} | Matricula: {matricula} | Horario: {datetime.now().strftime('%H:%M:%S')} | Anexo: {'PDF' if baixar_faturas else 'Imagem'}")
                    df_sheet.at[idx, "STATUS"] = "Enviado"
                    atualizar_planilha_excel(df_dict_full, caminho_arquivo)
                    _emit_progresso()
                elif enviado_texto and not enviado_anexo:
                    tqdm.write(f"[ERRO] A mensagem de texto foi enviada, mas o anexo falhou para {numero_para_url} (linha {idx}).")
                    _enviadas += 1
                    df_sheet.at[idx, "STATUS"] = "Enviado sem anexo"
                    try:
                        atualizar_planilha_excel(df_dict_full, caminho_arquivo)
                    except Exception:
                        pass
                    falhas.append({**row.to_dict(), "ERRO": "Falha no envio de anexo (texto enviado)"})
                    _emit_progresso()
                else:
                    tqdm.write(f"[ERRO] Não conseguimos enviar a mensagem para {numero_para_url} (linha {idx}).")
                    _registrar_falha(df_sheet, idx, row, "Falha no envio de texto")

                # --- APLICANDO O DELAY RANDÔMICO DINÂMICO ---
                if enviado_texto and enviado_anexo: # ⬅️ Correção feita aqui
                    mensagens_enviadas_no_lote += 1
                    espera_aleatoria = random.uniform(delay_min_dinamico, delay_max_dinamico)
                    delay_com_barra(espera_aleatoria, "Proxima mensagem sera enviada em")
                else:
                    delay_com_barra(5, "Falha. Retentando em")

            # Emite resumo do bloco final caso o loop tenha terminado com linhas já processadas
            if _bloco_count > 0:
                tqdm.write(f"[SKIP] Linhas {_bloco_inicio} a {_bloco_fim} ja processadas ({_bloco_count} registros) - pulando.")
                _bloco_inicio = None; _bloco_fim = None; _bloco_count = 0

    except Exception as e:
        logger.exception(f"Erro crítico: {e}")
    finally:
        if driver:
            try:
                driver.quit() 
            except:
                pass

        try:
            atualizar_planilha_excel(df_dict_full, caminho_arquivo)
        except Exception as e:
            pass

        # Remove callback de save — disparo encerrado
        if _flet_page is not None:
            _flet_page._wpp_save_fn = None

        if falhas:
            try:
                os.makedirs(PLANILHA_DIR, exist_ok=True)
                pd.DataFrame(falhas).to_excel(FALHAS_XLSX, index=False)
            except:
                pass

# ---------------------------
# Menu principal
# ---------------------------
def menu():
    while True:
        print("\n=== MENU PRINCIPAL ===")
        print("1 - Processar CSV e gerar planilha WhatsApp")
        print("2 - Enviar mensagens no WhatsApp (escolher aba(s))")
        print("3 - Calculo dos descontos macro 8104 (Negociação)")
        print("4 - 📘 Ajuda / Tutorial de Uso")
        print("5 - Sair")
        opc = input("Digite o número da opção desejada: ").strip()

        if opc == "1":
            processar_csv()
            
        elif opc == "2":
            print("\n🔄 Iniciando o módulo de Envio WhatsApp...")
            print("Abrindo janela para selecionar a planilha de envios...")
            
            arquivo_envio = selecionar_arquivo(
                "Selecione a planilha de envios (Excel)",
                tipos=[("Planilhas Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
            )
            
            if not arquivo_envio:
                print("❌ Nenhuma planilha selecionada. Operação cancelada.")
                continue

            sheets = escolher_sheet_para_envio(arquivo_envio)
            if sheets:
                df_dict = carregar_planilha_para_envio(sheets, arquivo_envio)
                total_lines = sum(len(df) for df in df_dict.values()) if df_dict else 0
                logger.info(f"Total de linhas carregadas para envio: {total_lines}")
                
                # --- NOVA PERGUNTA DINÂMICA ---
                entrada_msgs = input("\nQuantas mensagens deseja enviar por hora (Recomendado: 20 a 40) [Padrão: 20]: ").strip()
                try:
                    msgs_por_hora = int(entrada_msgs) if entrada_msgs else 20
                except ValueError:
                    print("⚠️ Valor inválido. Usando o padrão de 20 mensagens/hora.")
                    msgs_por_hora = 20

                confirmar = input("\nDeseja iniciar os envios agora? (s/n): ").strip().lower()

                if confirmar == "s":
                    # Pergunta se deve baixar faturas
                    baixar_faturas = False
                    usuario_waterfy = ""
                    senha_waterfy = ""
                    
                    if HAS_FATURA_DOWNLOADER:
                        baixar_resp = input("Deseja baixar e anexar faturas PDF individuais (s/N): ").strip().lower()
                        baixar_faturas = baixar_resp == "s"
                        if baixar_faturas:
                            print("📄 Modo ativado: Faturas PDF serão baixadas e anexadas a cada mensagem.")
                            usuario_waterfy = input("Usuário Waterfy: ").strip()
                            senha_waterfy = input("Senha Waterfy: ").strip()
                        else:
                            print("🖼️ Modo padrão: Imagem padrão será anexada.")
                    else:
                        print("⚠️ Sistema de download de faturas não disponível. Usando imagem padrão.")
                    
                    enviar_whatsapp_sheets(sheets, arquivo_envio, msgs_por_hora, baixar_faturas, usuario_waterfy, senha_waterfy)
                else:
                    logger.info("Envio cancelado pelo usuário.")
                    
        elif opc == "3":
            print("\n🔄 Iniciando o módulo de Negociação...")
            print("Abrindo janela para selecionar a base de faturas...")
            
            arquivo_negociacao = selecionar_arquivo("Selecione a base para Negociação (Macro 8104)")
            
            if arquivo_negociacao:
                gerar_planilha_negociacao(arquivo_negociacao, DIR_NEGOCIACAO)
                print("\n✅ Módulo de negociação finalizado. Voltando ao menu principal...")
            else:
                print("\n❌ Seleção cancelada. Voltando ao menu...")
                
        elif opc == "4":
            exibir_tutorial()
            
        elif opc == "5":
            print("Saindo.")
            break
            
        else:
            print("Opção inválida. Tente novamente.")

if __name__ == "__main__":
    try:
        menu()
    except KeyboardInterrupt:
        print("\n\nPrograma interrompido pelo usuário. Saindo com segurança... Até logo!")
